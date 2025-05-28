import openpyxl

file_path = r"C:\Users\User\Dropbox\DO & INV\Savori Sales Summary (2022 - 2025) - As of 27 May'25.xlsx"
sheet_name = "Format"

print(f"打开文件: {file_path}")

try:
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        print(f"错误：未找到Sheet名为 '{sheet_name}'，实际可用Sheet：{wb.sheetnames}")
    else:
        ws = wb[sheet_name]
        row_count = 0
        modify_count = 0

        for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):  # G列, 跳过表头
            cell = row[0]
            row_count += 1
            old_value = cell.value

            if old_value and isinstance(old_value, str):
                new_value = old_value.strip()
                if new_value != old_value:
                    print(f"第{cell.row}行 G列: 原='{old_value}' → 新='{new_value}'")
                    cell.value = new_value
                    modify_count += 1

        print(f"检查总行数: {row_count}，被修改行数: {modify_count}")

        if modify_count > 0:
            wb.save(file_path)
            print("文件已保存。")
        else:
            print("没有需要修改的内容，无需保存。")

except Exception as e:
    print("处理文件时出错：")
    print(e)
