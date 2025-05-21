import os
import datetime
import sqlite3
import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook

def get_customers():
    conn = sqlite3.connect('sales.db')
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT customer FROM sales ORDER BY customer;")
    customers = [row[0] for row in cursor.fetchall() if row[0] is not None]
    conn.close()
    return customers

def get_outlets(customer):
    conn = sqlite3.connect('sales.db')
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT outlet FROM sales WHERE customer=? ORDER BY outlet;", (customer,))
    outlets = [row[0] for row in cursor.fetchall() if row[0] is not None]
    conn.close()
    return outlets

def get_products(customer, outlet):
    conn = sqlite3.connect('sales.db')
    cursor = conn.cursor()
    cursor.execute("""
    SELECT DISTINCT product_code, product_description 
    FROM sales 
    WHERE customer=? AND outlet=?
    ORDER BY product_code;
    """, (customer, outlet))
    products = cursor.fetchall()
    conn.close()
    return products

def get_latest_record(customer, outlet, product_code):
    conn = sqlite3.connect('sales.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("""
        SELECT * FROM sales
        WHERE customer=? AND outlet=? AND product_code=?
        ORDER BY date DESC
        LIMIT 1;
    """, (customer, outlet, product_code))
    row = cursor.fetchone()
    conn.close()
    if row:
        return dict(row)
    return None

class SalesApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Sales Data Selector (SQLite)")
        self.master.geometry("1200x800")

        # 顶部查询区
        tk.Label(master, text="Customer:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.customer_cmb = ttk.Combobox(master, state="readonly", width=30)
        self.customer_cmb.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.customer_cmb.bind("<<ComboboxSelected>>", self.on_customer_select)

        tk.Label(master, text="Outlet:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.outlet_cmb = ttk.Combobox(master, state="readonly", width=50)
        self.outlet_cmb.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        self.outlet_cmb.bind("<<ComboboxSelected>>", self.on_outlet_select)

        tk.Label(master, text="Product:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.product_cmb = ttk.Combobox(master, state="readonly", width=50)
        self.product_cmb.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        self.add_btn = tk.Button(master, text="Add to List", command=self.add_to_list)
        self.add_btn.grid(row=2, column=2, padx=5, pady=5, sticky="w")

        # 中间区域：已添加记录列表
        self.tree = ttk.Treeview(master, columns=("Customer", "Outlet", "Product", "Date", "Month"), show='headings')
        self.tree.heading("Customer", text="Customer")
        self.tree.heading("Outlet", text="Outlet")
        self.tree.heading("Product", text="Product")
        self.tree.heading("Date", text="Date")
        self.tree.heading("Month", text="Month")
        self.tree.column("Customer", width=150)
        self.tree.column("Outlet", width=200)
        self.tree.column("Product", width=200)
        self.tree.column("Date", width=100)
        self.tree.column("Month", width=50)
        self.tree.grid(row=3, column=0, columnspan=5, padx=5, pady=5, sticky="nsew")
        self.master.grid_rowconfigure(3, weight=1)
        self.master.grid_columnconfigure(1, weight=1)

        # 全局更新的字段(Year, Date, Month, Invoice, Account)
        tk.Label(master, text="Update All Records:").grid(row=4, column=0, padx=5, pady=5, sticky="e")
        
        tk.Label(master, text="Year:").grid(row=5, column=0, padx=5, pady=5, sticky="e")
        self.year_entry = tk.Entry(master, width=10)
        self.year_entry.grid(row=5, column=1, padx=5, pady=5, sticky="w")

        tk.Label(master, text="Date (YYYY-MM-DD):").grid(row=6, column=0, padx=5, pady=5, sticky="e")
        self.date_entry = tk.Entry(master, width=15)
        self.date_entry.grid(row=6, column=1, padx=5, pady=5, sticky="w")
        
        tk.Label(master, text="Month:").grid(row=7, column=0, padx=5, pady=5, sticky="e")
        self.month_entry = tk.Entry(master, width=15)
        self.month_entry.grid(row=7, column=1, padx=5, pady=5, sticky="w")

        tk.Label(master, text="Invoice #:").grid(row=8, column=0, padx=5, pady=5, sticky="e")
        self.invoice_entry = tk.Entry(master, width=15)
        self.invoice_entry.grid(row=8, column=1, padx=5, pady=5, sticky="w")

        tk.Label(master, text="Account:").grid(row=9, column=0, padx=5, pady=5, sticky="e")
        self.account_entry = tk.Entry(master, width=15)
        self.account_entry.grid(row=9, column=1, padx=5, pady=5, sticky="w")

        self.update_all_btn = tk.Button(master, text="Update All", command=self.update_all_records)
        self.update_all_btn.grid(row=10, column=1, padx=5, pady=5, sticky="w")

        # 单条记录更新的字段(Product Code, Qty in Pcs, Qty in Ctns等)
        tk.Label(master, text="Update Selected Record Only:").grid(row=4, column=2, padx=5, pady=5, sticky="e")

        tk.Label(master, text="Product Code:").grid(row=5, column=2, padx=5, pady=5, sticky="e")
        self.product_code_entry = tk.Entry(master, width=15)
        self.product_code_entry.grid(row=5, column=3, padx=5, pady=5, sticky="w")

        tk.Label(master, text="Qty in Pcs:").grid(row=6, column=2, padx=5, pady=5, sticky="e")
        self.qty_pcs_entry = tk.Entry(master, width=15)
        self.qty_pcs_entry.grid(row=6, column=3, padx=5, pady=5, sticky="w")

        tk.Label(master, text="Qty in Ctns:").grid(row=7, column=2, padx=5, pady=5, sticky="e")
        self.qty_ctns_entry = tk.Entry(master, width=15)
        self.qty_ctns_entry.grid(row=7, column=3, padx=5, pady=5, sticky="w")

        self.update_selected_btn = tk.Button(master, text="Update Selected", command=self.update_selected_record)
        self.update_selected_btn.grid(row=8, column=3, padx=5, pady=5, sticky="w")

        # Export按钮
        self.export_btn = tk.Button(master, text="Copy to Excel (tmr sales data)", command=self.export_result)
        self.export_btn.grid(row=10, column=3, padx=5, pady=10, sticky="w")

        self.latest_record = None
        self.product_map = {}
        self.selected_items = []

        self.load_customers()

    def load_customers(self):
        customers = get_customers()
        self.customer_cmb['values'] = customers

    def on_customer_select(self, event):
        customer = self.customer_cmb.get()
        outlets = get_outlets(customer)
        self.outlet_cmb['values'] = outlets
        self.outlet_cmb.set('')
        self.product_cmb.set('')
        self.product_cmb['values'] = []

    def on_outlet_select(self, event):
        customer = self.customer_cmb.get()
        outlet = self.outlet_cmb.get()
        products = get_products(customer, outlet)
        product_list = [f"{p[0]} - {p[1]}" for p in products]
        self.product_map = {f"{p[0]} - {p[1]}": p[0] for p in products}
        self.product_cmb['values'] = product_list
        self.product_cmb.set('')

    def add_to_list(self):
        customer = self.customer_cmb.get()
        outlet = self.outlet_cmb.get()
        product_str = self.product_cmb.get()

        if not (customer and outlet and product_str):
            messagebox.showwarning("Warning", "Please select Customer, Outlet, and Product.")
            return

        product_code = self.product_map[product_str]
        record = get_latest_record(customer, outlet, product_code)

        if record:
            date_val = record.get('date', '')
            month_val = record.get('month', '')
            # 插入treeview行的同时，在selected_items追加一条记录
            self.tree.insert("", "end", values=(customer, outlet, product_str, date_val, month_val))
            self.selected_items.append(record)
        else:
            messagebox.showinfo("Info", "No record found for this selection")

    def update_all_records(self):
        # 只更新5个字段：year, date, month, invoice_no, account
        new_year = self.year_entry.get().strip()
        new_date = self.date_entry.get().strip()
        new_month = self.month_entry.get().strip()
        new_invoice = self.invoice_entry.get().strip()
        new_account = self.account_entry.get().strip()

        for item in self.selected_items:
            if new_year:
                item['year'] = int(new_year)
            if new_date:
                item['date'] = new_date
            if new_month:
                item['month'] = new_month
            if new_invoice:
                item['invoice_no'] = new_invoice
            if new_account:
                item['account'] = new_account

        messagebox.showinfo("Info", "All selected records updated.")

    def update_selected_record(self):
        # 从Tree中获取选中的行
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a record in the list to update.")
            return
        item_id = selection[0]
        index = self.tree.index(item_id)  # 取得该行在selected_items中的索引（假设添加顺序与selected_items一致）

        new_product_code = self.product_code_entry.get().strip()
        new_qty_pcs = self.qty_pcs_entry.get().strip()
        new_qty_ctns = self.qty_ctns_entry.get().strip()

        rec = self.selected_items[index]
        if new_product_code:
            rec['product_code'] = new_product_code
        if new_qty_pcs:
            rec['qty_in_pcs'] = int(new_qty_pcs)
        if new_qty_ctns:
            rec['qty_in_ctns'] = int(new_qty_ctns)

        messagebox.showinfo("Info", "Selected record updated.")

    def export_result(self):
        if not self.selected_items:
            messagebox.showinfo("Info", "No items to export.")
            return

        export_path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                                   initialfile="tmr sales data.xlsx",
                                                   filetypes=[("Excel Files", "*.xlsx")])
        if not export_path:
            return

        # 使用原Excel列名
        headers = [
            "Year", "Date", "Month", "Brand/Category", "Supplier",
            "Product Code", "Product Description", "Carton Packing", "Customer", "Outlet",
            "Qty in Pcs", "Qty in Ctns", "Total Qty in Pcs", "Total Qty in Ctns",
            "Invoice #", "Total Value", "GST", "Total Value Inclusive GST",
            "Account", "Customer PO#", "Remarks", "Selling Price"
        ]

        wb = Workbook()
        ws = wb.active
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, item in enumerate(self.selected_items, start=2):
            ws.cell(row=row_idx, column=1, value=item.get('year'))
            ws.cell(row=row_idx, column=2, value=item.get('date'))
            ws.cell(row=row_idx, column=3, value=item.get('month'))
            ws.cell(row=row_idx, column=4, value=item.get('brand_category'))
            ws.cell(row=row_idx, column=5, value=item.get('supplier'))
            ws.cell(row=row_idx, column=6, value=item.get('product_code'))
            ws.cell(row=row_idx, column=7, value=item.get('product_description'))
            ws.cell(row=row_idx, column=8, value=item.get('carton_packing'))
            ws.cell(row=row_idx, column=9, value=item.get('customer'))
            ws.cell(row=row_idx, column=10, value=item.get('outlet'))
            ws.cell(row=row_idx, column=11, value=item.get('qty_in_pcs'))
            ws.cell(row=row_idx, column=12, value=item.get('qty_in_ctns'))
            ws.cell(row=row_idx, column=13, value=item.get('total_qty_in_pcs'))
            ws.cell(row=row_idx, column=14, value=item.get('total_qty_in_ctns'))
            ws.cell(row=row_idx, column=15, value=item.get('invoice_no'))
            ws.cell(row=row_idx, column=16, value=item.get('total_value'))
            ws.cell(row=row_idx, column=17, value=item.get('gst'))

            # 公式复制，如同excel填充
            tv_col = 16  # Total Value 列(P)
            gst_col = 17 # GST 列(Q)
            tvi_col = 18 # Total Value Inclusive GST 列(R)
            ws.cell(row=row_idx, column=tvi_col, value=f"={chr(64+tv_col)}{row_idx}+{chr(64+gst_col)}{row_idx}")

            ws.cell(row=row_idx, column=19, value=item.get('account'))
            ws.cell(row=row_idx, column=20, value=item.get('customer_po'))
            ws.cell(row=row_idx, column=21, value=item.get('remarks'))
            ws.cell(row=row_idx, column=22, value=item.get('selling_price'))

        wb.save(export_path)
        messagebox.showinfo("Info", f"Exported to {export_path}")


if __name__ == "__main__":
    root = tk.Tk()
    app = SalesApp(root)
    root.mainloop()
