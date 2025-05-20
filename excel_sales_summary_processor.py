#!/usr/bin/env python3
"""Clean sales summary Excel files.

This script scans the directory ``C:\\Users\\User\\Dropbox\\DO & INV\\`` for
``.xlsx`` files whose filenames contain ``sales summary`` (case-insensitive) and
trims whitespace from columns F and G in matching sheets. The special file
``Savori Sales Summary (2022 - 2025) - As of 20 May'25.xlsx`` is processed only
on the sheet named ``delivery details`` (case-insensitive).
"""

import logging
import os
import re
from typing import List

from openpyxl import load_workbook

ROOT_DIR = r"C:\Users\User\Dropbox\DO & INV"
TARGET_FILE = os.path.join(
    ROOT_DIR,
    "Savori Sales Summary (2022 - 2025) - As of 20 May'25.xlsx",
)

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")


def list_target_files() -> List[str]:
    """Return all sales summary Excel files under ``ROOT_DIR``."""
    pattern = re.compile(r"sales summary", re.IGNORECASE)
    files = []
    try:
        for name in os.listdir(ROOT_DIR):
            if not name.lower().endswith(".xlsx"):
                continue
            if pattern.search(name):
                files.append(os.path.join(ROOT_DIR, name))
    except FileNotFoundError:
        logging.error(f"Directory not found: {ROOT_DIR}")
        return []
    if os.path.exists(TARGET_FILE) and TARGET_FILE not in files:
        files.append(TARGET_FILE)
    return files


def target_sheets(workbook, path: str) -> List[str]:
    """Return list of sheet names to process for *path*."""
    if os.path.normcase(path) == os.path.normcase(TARGET_FILE):
        return [s for s in workbook.sheetnames if s.lower() == "delivery details"]
    return [s for s in workbook.sheetnames if "sales summary" in s.lower()]


def trim_columns(path: str) -> None:
    """Open workbook at *path*, trim columns F and G, and save."""
    try:
        wb = load_workbook(path)
    except Exception as exc:  # noqa: BLE001
        logging.warning(f"Unable to open {path}: {exc}")
        return

    sheets = target_sheets(wb, path)
    if not sheets:
        logging.warning(f"No target sheets in {path}")
        return

    changed = 0
    for sheet_name in sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_col=6, max_col=7):
            for cell in row:
                val = cell.value
                if isinstance(val, str):
                    trimmed = val.strip()
                    if trimmed != val:
                        cell.value = trimmed
                        changed += 1
                        logging.info(
                            f"{os.path.basename(path)} - {sheet_name}!{cell.coordinate} trimmed"
                        )

    try:
        wb.save(path)
        logging.info(f"Saved {path} ({changed} cells changed)")
    except Exception as exc:  # noqa: BLE001
        logging.error(f"Failed to save {path}: {exc}")


def main() -> None:
    """Entry point."""
    files = list_target_files()
    if not files:
        logging.info("No matching Excel files found.")
        return
    for file_path in files:
        logging.info(f"Processing {file_path}")
        trim_columns(file_path)


if __name__ == "__main__":
    main()

