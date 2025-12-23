# new_year_safe_fast_final_v2.py
# 重点修正：
# - 2025 -> 2026（档名 + 内容）
# - xx25 -> xx26（允许大小写，档名 + 内容）
#   使用更稳的 token 边界：前后不是字母或数字，避免误截断
#
# 依赖：pip install openpyxl

import os
import re
import shutil
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    openpyxl = None


# =========================
# 配置区
# =========================
SRC_ROOT = Path(r"C:\Users\jhunj\Dropbox\DO & INV\DO & INV 2025")

DEST_PARENT = Path(r"C:\Users\jhunj\Dropbox\for jj")
DEST_FOLDER_NAME = "DO & INV 2026"

BUILD_ROOT = Path(r"C:\Users\jhunj\Desktop\DO & INV 2026_build")

MONTH_REQUIRED = "12. Dec"
KEEP_MONTH = "1. Jan"

EMPTY_ONLY_FOLDERS = {
    "1. Airfreight",
    "1. Old pricing history",
    "1. Order Summary & Sales Summary(Mth end)",
    "All - cash sales",
    "All - Staff Purchase",
}

YEAR_OLD = "2025"
YEAR_NEW = "2026"

DRY_RUN = False
REBUILD_BUILD_DIR = True
OVERWRITE_DEST_IF_EXISTS = False

EXCEL_EXTS_CAN_EDIT = {".xlsx", ".xlsm", ".xltx", ".xltm"}
EXCEL_EXTS_ALL = EXCEL_EXTS_CAN_EDIT | {".xls"}
# =========================


MONTH_FOLDER_REGEX = re.compile(r"^\d{1,2}\.\s*[A-Za-z]{3}\b")
HISTORY_REGEX = re.compile(r"history", re.IGNORECASE)

# ✅ 允许大小写 xx25 -> xx26，但要求是完整 token（前后不是字母数字），且至少 2 个字母
ALPHA_LETTERS_25 = re.compile(
    r"(?<![A-Za-z0-9])([A-Za-z]{2,12})25(?![A-Za-z0-9])")
# ✅ 两位数字+25：0225 -> 0226
TWO_DIGITS_25 = re.compile(r"\b(\d{2})25\b")


def log(msg: str, fh=None) -> None:
    print(msg)
    if fh:
        fh.write(msg + "\n")


def ensure_dir(p: Path, fh=None) -> None:
    if p.exists():
        return
    log(f"[MKDIR] {p}", fh)
    if not DRY_RUN:
        p.mkdir(parents=True, exist_ok=True)


def remove_tree(p: Path, fh=None) -> None:
    if not p.exists():
        return
    log(f"[RM TREE] {p}", fh)
    if not DRY_RUN:
        shutil.rmtree(p, ignore_errors=True)


def copy_file(src: Path, dst: Path, fh=None) -> None:
    ensure_dir(dst.parent, fh)
    log(f"[COPY FILE] {src} -> {dst}", fh)
    if not DRY_RUN:
        shutil.copy2(src, dst)


def ignore_month_history(_dir: str, names: list[str]) -> set[str]:
    ignored = set()
    for n in names:
        if HISTORY_REGEX.search(n):
            ignored.add(n)
            continue
        if MONTH_FOLDER_REGEX.match(n):
            ignored.add(n)
            continue
    return ignored


def copy_tree_full(src: Path, dst: Path, fh=None, ignore=None) -> None:
    if dst.exists():
        log(f"[SKIP] 目标已存在，跳过复制: {dst}", fh)
        return
    log(f"[COPY TREE] {src} -> {dst}", fh)
    if not DRY_RUN:
        shutil.copytree(src, dst, ignore=ignore)


def should_copy_normal_folder(folder: Path) -> bool:
    for _root, dirs, _files in os.walk(folder):
        dirs[:] = [d for d in dirs if not HISTORY_REGEX.search(d)]
        if MONTH_REQUIRED in dirs:
            return True
    return False


def clear_folder_contents(folder: Path, fh=None) -> None:
    if not folder.exists() or not folder.is_dir():
        return
    for item in folder.iterdir():
        if item.is_dir():
            log(f"[DEL DIR] {item}", fh)
            if not DRY_RUN:
                shutil.rmtree(item, ignore_errors=True)
        else:
            log(f"[DEL FILE] {item}", fh)
            if not DRY_RUN:
                try:
                    item.unlink()
                except Exception as e:
                    log(f"[WARN] 删除失败: {item} , {e}", fh)


def transform_text(s: str) -> str:
    """
    同时用于：Excel 档名 + Excel 内容
    """
    out = s.replace(YEAR_OLD, YEAR_NEW)
    out = ALPHA_LETTERS_25.sub(r"\g<1>26", out)   # group1 + 26
    out = TWO_DIGITS_25.sub(r"\g<1>26", out)      # group1 + 26
    return out


def rename_file_if_needed(path: Path, fh=None) -> Path:
    new_name = transform_text(path.name)
    if new_name == path.name:
        return path

    new_path = path.with_name(new_name)
    if new_path.exists():
        log(f"[WARN] 改名目标已存在，跳过: {new_path}", fh)
        return path

    log(f"[RENAME] {path} -> {new_path}", fh)
    if not DRY_RUN:
        path.rename(new_path)
    return new_path


def build_normal_folder(src_folder: Path, dst_folder: Path, fh=None) -> None:
    ensure_dir(dst_folder, fh)

    for item in src_folder.iterdir():
        name = item.name

        if HISTORY_REGEX.search(name):
            log(f"[SKIP] history 项目不复制: {item}", fh)
            continue
        if MONTH_FOLDER_REGEX.match(name):
            log(f"[SKIP] 月份资料夹不复制: {item}", fh)
            continue

        dst_item = dst_folder / name

        if item.is_file():
            copy_file(item, dst_item, fh)
        elif item.is_dir():
            copy_tree_full(item, dst_item, fh, ignore=ignore_month_history)

    create_keep_month_at_dec_level(src_folder, dst_folder, fh)


def create_keep_month_at_dec_level(src_folder: Path, dst_folder: Path, fh=None) -> None:
    for root, dirs, _files in os.walk(src_folder):
        if MONTH_REQUIRED in dirs:
            src_root = Path(root)
            rel = src_root.relative_to(src_folder)
            jan = (dst_folder / rel) / KEEP_MONTH
            ensure_dir(jan, fh)
            log(f"[CLEAN] keep {KEEP_MONTH} folder but clear contents: {jan}", fh)
            clear_folder_contents(jan, fh)

        dirs[:] = [d for d in dirs if not HISTORY_REGEX.search(d)]


def replace_in_excel(file_path: Path, fh=None) -> None:
    ext = file_path.suffix.lower()
    if ext not in EXCEL_EXTS_CAN_EDIT:
        if ext == ".xls":
            log(f"[SKIP] .xls 旧格式不改内容，只改档名: {file_path}", fh)
        return

    if openpyxl is None:
        log("[ERROR] 未安装 openpyxl，无法修改 Excel 内容。请先 pip install openpyxl", fh)
        return

    log(f"[EDIT] 替换内容字样: {file_path}", fh)
    if DRY_RUN:
        return

    try:
        wb = openpyxl.load_workbook(
            file_path, data_only=False, keep_vba=(ext == ".xlsm"))
        changed = False

        for ws in wb.worksheets:
            max_r = ws.max_row or 0
            max_c = ws.max_column or 0
            for r in range(1, max_r + 1):
                for c in range(1, max_c + 1):
                    cell = ws.cell(row=r, column=c)
                    v = cell.value
                    if isinstance(v, str):
                        new_v = transform_text(v)
                        if new_v != v:
                            cell.value = new_v
                            changed = True

        if changed:
            wb.save(file_path)
        else:
            log(f"[SKIP] 内容无可替换字样: {file_path}", fh)

    except Exception as e:
        log(f"[WARN] 修改失败: {file_path} , {e}", fh)


def main() -> None:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ensure_dir(BUILD_ROOT.parent)

    log_path = BUILD_ROOT.parent / f"new_year_build_log_{timestamp}.txt"
    fh = open(log_path, "w", encoding="utf-8")

    try:
        log("=== new_year_safe_fast_final_v2 开始 ===", fh)
        log(f"[INFO] DRY_RUN={DRY_RUN}", fh)
        log(f"[INFO] SRC_ROOT={SRC_ROOT}", fh)
        log(f"[INFO] BUILD_ROOT={BUILD_ROOT}", fh)

        if not SRC_ROOT.exists():
            log(f"[ERROR] 找不到来源路径: {SRC_ROOT}", fh)
            return

        if BUILD_ROOT.exists() and REBUILD_BUILD_DIR:
            log("[INFO] REBUILD_BUILD_DIR=True，删除旧 build 再重建", fh)
            remove_tree(BUILD_ROOT, fh)

        ensure_dir(BUILD_ROOT, fh)

        log("\n=== Step 2: 构建 2026_build（一般资料夹选择性复制 + 指定5个空资料夹）===", fh)

        for item in SRC_ROOT.iterdir():
            if not item.is_dir():
                continue

            dst_item = BUILD_ROOT / item.name

            if item.name in EMPTY_ONLY_FOLDERS:
                log(f"[EMPTY FOLDER] 只创建空资料夹: {dst_item}", fh)
                ensure_dir(dst_item, fh)
                continue

            if not should_copy_normal_folder(item):
                log(f"[SKIP] 不符合条件(无 {MONTH_REQUIRED}): {item.name}", fh)
                continue

            log(f"[BUILD] 一般资料夹（符合 {MONTH_REQUIRED}）: {item.name}", fh)
            build_normal_folder(item, dst_item, fh)

        log("\n=== Step 3: build 内 Excel 改名与内容替换（2025->2026, xx25->xx26）===", fh)

        excel_files = [p for p in BUILD_ROOT.rglob(
            "*") if p.is_file() and p.suffix.lower() in EXCEL_EXTS_ALL]

        renamed = []
        for f in excel_files:
            renamed.append(rename_file_if_needed(f, fh))

        for f in renamed:
            replace_in_excel(f, fh)

        dest_root = DEST_PARENT / DEST_FOLDER_NAME
        log("\n=== Step 4: commit 到 Dropbox 目标目录（一次性 move）===", fh)
        log(f"[INFO] DEST={dest_root}", fh)

        if dest_root.exists():
            has_any = any(dest_root.iterdir())
            if has_any and not OVERWRITE_DEST_IF_EXISTS:
                log("[ERROR] 目标 DO & INV 2026 已存在且非空，为安全停止。", fh)
                log("        如你确定要覆盖，请将 OVERWRITE_DEST_IF_EXISTS=True 再运行。", fh)
                return

            if OVERWRITE_DEST_IF_EXISTS:
                log("[WARN] OVERWRITE_DEST_IF_EXISTS=True，将删除现有目标目录再写入", fh)
                remove_tree(dest_root, fh)

        ensure_dir(DEST_PARENT, fh)

        log(f"[MOVE] {BUILD_ROOT} -> {dest_root}", fh)
        if not DRY_RUN:
            if dest_root.exists():
                shutil.rmtree(dest_root, ignore_errors=True)
            shutil.move(str(BUILD_ROOT), str(dest_root))

        log("\n=== 完成 ===", fh)
        log(f"[LOG] {log_path}", fh)

        if DRY_RUN:
            log("\n[DRY_RUN] 你现在是测试模式，没有真的复制/删除/改档/move。", fh)
            log("         确认日志动作正确后，把 DRY_RUN=False 再跑一次。", fh)

    finally:
        fh.close()


if __name__ == "__main__":
    main()
