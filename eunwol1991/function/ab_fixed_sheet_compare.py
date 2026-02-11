import os
import threading
import queue
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None


START_SHEET_NAME = "SC Details"
MAX_EMPTY_ROWS = 20
COL_START = 1  # A
COL_END = 7    # G
KEY_SEPARATOR = "|||"


def cell_is_empty(value) -> bool:
    return value is None or value == ""


def make_key(values: list) -> str:
    parts = ["" if v is None else str(v) for v in values]
    return KEY_SEPARATOR.join(parts)


def scan_sheet(ws) -> tuple[int, list[tuple[int, list, str]]]:
    empty_run = 0
    effective_rows = 0
    records = []

    for row_idx in range(1, ws.max_row + 1):
        row_values = [
            ws.cell(row=row_idx, column=col_idx).value
            for col_idx in range(COL_START, COL_END + 1)
        ]
        if all(cell_is_empty(v) for v in row_values):
            empty_run += 1
        else:
            empty_run = 0
            effective_rows += 1
            key = make_key(row_values)
            records.append((row_idx, row_values, key))
        if empty_run >= MAX_EMPTY_ROWS:
            break

    return effective_rows, records


def scan_workbook(path: str) -> dict:
    if not openpyxl:
        raise RuntimeError("openpyxl not installed")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        return {name: wb[name] for name in wb.sheetnames}
    finally:
        wb.close()


def build_results(a_path: str, b_path: str, on_progress=None):
    if not openpyxl:
        raise RuntimeError("openpyxl not installed")

    wb_a = openpyxl.load_workbook(a_path, data_only=True, read_only=True)
    wb_b = openpyxl.load_workbook(b_path, data_only=True, read_only=True)
    try:
        if START_SHEET_NAME not in wb_a.sheetnames:
            raise ValueError(f"Sheet '{START_SHEET_NAME}' not found in A.")

        start_index = wb_a.sheetnames.index(START_SHEET_NAME)
        a_sheet_names = wb_a.sheetnames[start_index:]
        a_sheet_set = set(a_sheet_names)
        if START_SHEET_NAME in wb_b.sheetnames:
            b_start_index = wb_b.sheetnames.index(START_SHEET_NAME)
            b_sheet_names = wb_b.sheetnames[b_start_index:]
        else:
            b_sheet_names = wb_b.sheetnames
        b_sheet_set = set(b_sheet_names)
        b_only_sheets = [name for name in b_sheet_names if name not in a_sheet_set]

        summary_rows = []
        detail_rows = []

        total = len(a_sheet_names) + len(b_only_sheets)
        progress_idx = 0

        for sheet_name in a_sheet_names:
            progress_idx += 1
            if on_progress:
                on_progress(total, progress_idx, f"Scanning {sheet_name}")

            ws_a = wb_a[sheet_name]
            a_effective, a_records = scan_sheet(ws_a)
            a_keys = {rec[2] for rec in a_records}

            if sheet_name not in b_sheet_set:
                summary_rows.append({
                    "SheetName": sheet_name,
                    "A_EffectiveRows": a_effective,
                    "B_EffectiveRows": 0,
                    "MissingCount": 0,
                    "Status": "OK",
                    "Note": "SHEET_ONLY_IN_A_IGNORED",
                })
                continue

            ws_b = wb_b[sheet_name]
            b_effective, b_records = scan_sheet(ws_b)
            b_key_first = {}
            for row_num, row_values, key in b_records:
                if key not in b_key_first:
                    b_key_first[key] = (row_num, row_values)

            missing_keys = [k for k in b_key_first.keys() if k not in a_keys]
            for key in missing_keys:
                row_num, row_values = b_key_first[key]
                detail_rows.append({
                    "IssueType": "ROW_MISSING_IN_A",
                    "SheetName": sheet_name,
                    "B_RowNumber": row_num,
                    "ColA": row_values[0],
                    "ColB": row_values[1],
                    "ColC": row_values[2],
                    "ColD": row_values[3],
                    "ColE": row_values[4],
                    "ColF": row_values[5],
                    "ColG": row_values[6],
                    "Key": key,
                })

            status = "OK" if not missing_keys else "ISSUE"
            summary_rows.append({
                "SheetName": sheet_name,
                "A_EffectiveRows": a_effective,
                "B_EffectiveRows": b_effective,
                "MissingCount": len(missing_keys),
                "Status": status,
                "Note": "",
            })

        for sheet_name in b_only_sheets:
            progress_idx += 1
            if on_progress:
                on_progress(total, progress_idx, f"Scanning {sheet_name} (B only)")

            ws_b = wb_b[sheet_name]
            b_effective, b_records = scan_sheet(ws_b)
            b_key_first = {}
            for row_num, row_values, key in b_records:
                if key not in b_key_first:
                    b_key_first[key] = (row_num, row_values)

            for key, (row_num, row_values) in b_key_first.items():
                detail_rows.append({
                    "IssueType": "SHEET_MISSING_IN_A",
                    "SheetName": sheet_name,
                    "B_RowNumber": row_num,
                    "ColA": row_values[0],
                    "ColB": row_values[1],
                    "ColC": row_values[2],
                    "ColD": row_values[3],
                    "ColE": row_values[4],
                    "ColF": row_values[5],
                    "ColG": row_values[6],
                    "Key": key,
                })

            summary_rows.append({
                "SheetName": sheet_name,
                "A_EffectiveRows": 0,
                "B_EffectiveRows": b_effective,
                "MissingCount": len(b_key_first),
                "Status": "ISSUE",
                "Note": "MISSING_SHEET_IN_A",
            })

        return summary_rows, detail_rows
    finally:
        wb_a.close()
        wb_b.close()


class FixedCompareApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("A vs B Fixed Sheet Compare")
        self.geometry("1200x760")
        self.minsize(1000, 600)

        self.a_path_var = tk.StringVar()
        self.b_path_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Ready")
        self.progress_var = tk.DoubleVar(value=0)

        self._queue = queue.Queue()
        self._running = False

        self._build_ui()
        self.after(150, self._process_queue)

    def _build_ui(self):
        top = ttk.Frame(self, padding=10)
        top.pack(fill=tk.X)

        ttk.Label(top, text="A Excel:").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(top, textvariable=self.a_path_var, width=90).grid(
            row=0, column=1, sticky=tk.W, padx=6
        )
        ttk.Button(top, text="Browse...", command=self._choose_a).grid(
            row=0, column=2, padx=4
        )

        ttk.Label(top, text="B Excel:").grid(
            row=1, column=0, sticky=tk.W, pady=(6, 0))
        ttk.Entry(top, textvariable=self.b_path_var, width=90).grid(
            row=1, column=1, sticky=tk.W, padx=6, pady=(6, 0)
        )
        ttk.Button(top, text="Browse...", command=self._choose_b).grid(
            row=1, column=2, padx=4, pady=(6, 0)
        )

        top.columnconfigure(1, weight=1)

        action_bar = ttk.Frame(self, padding=(10, 4, 10, 6))
        action_bar.pack(fill=tk.X)
        self.run_btn = ttk.Button(action_bar, text="Run", command=self._start)
        self.run_btn.pack(side=tk.LEFT)
        self.progress = ttk.Progressbar(
            action_bar, orient="horizontal", length=320, mode="determinate", variable=self.progress_var
        )
        self.progress.pack(side=tk.LEFT, padx=10)
        ttk.Label(action_bar, textvariable=self.status_var).pack(side=tk.LEFT)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        self._build_summary_tab()
        self._build_detail_tab()

    def _build_summary_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Summary")
        columns = ("SheetName", "A_EffectiveRows",
                   "B_EffectiveRows", "MissingCount", "Status", "Note")
        self.summary_tree = ttk.Treeview(
            frame, columns=columns, show="headings")
        for col in columns:
            self.summary_tree.heading(col, text=col)
            self.summary_tree.column(
                col, width=160 if col == "SheetName" else 130, anchor=tk.W)

        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL,
                            command=self.summary_tree.yview)
        hsb = ttk.Scrollbar(frame, orient=tk.HORIZONTAL,
                            command=self.summary_tree.xview)
        self.summary_tree.configure(
            yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.summary_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

    def _build_detail_tab(self):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text="Details")
        columns = (
            "IssueType", "SheetName", "B_RowNumber", "ColA", "ColB", "ColC", "ColD", "ColE", "ColF", "ColG", "Key"
        )
        self.detail_tree = ttk.Treeview(
            frame, columns=columns, show="headings")
        for col in columns:
            self.detail_tree.heading(col, text=col)
            width = 140 if col in ("IssueType", "SheetName", "Key") else 100
            self.detail_tree.column(col, width=width, anchor=tk.W)

        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL,
                            command=self.detail_tree.yview)
        hsb = ttk.Scrollbar(frame, orient=tk.HORIZONTAL,
                            command=self.detail_tree.xview)
        self.detail_tree.configure(
            yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.detail_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

    def _choose_a(self):
        path = filedialog.askopenfilename(
            title="Select A Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.a_path_var.set(path)

    def _choose_b(self):
        path = filedialog.askopenfilename(
            title="Select B Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.b_path_var.set(path)

    def _start(self):
        if self._running:
            return
        if not openpyxl:
            messagebox.showerror("Error", "openpyxl is not installed.")
            return
        a_path = self.a_path_var.get().strip()
        b_path = self.b_path_var.get().strip()
        if not a_path or not os.path.isfile(a_path):
            messagebox.showerror(
                "Error", "Please select a valid A Excel file.")
            return
        if not b_path or not os.path.isfile(b_path):
            messagebox.showerror(
                "Error", "Please select a valid B Excel file.")
            return

        self._running = True
        self.run_btn.configure(state=tk.DISABLED)
        self.progress_var.set(0)
        self.status_var.set("Running...")
        self._clear_results()

        worker = threading.Thread(
            target=self._run_job, args=(a_path, b_path), daemon=True)
        worker.start()

    def _run_job(self, a_path: str, b_path: str):
        try:
            def on_progress(total, current, note):
                self._queue.put(("progress", total, current, note))

            summary, details = build_results(
                a_path, b_path, on_progress=on_progress)
            self._queue.put(("done", summary, details))
        except Exception as exc:
            self._queue.put(("error", str(exc)))

    def _process_queue(self):
        try:
            while True:
                msg = self._queue.get_nowait()
                kind = msg[0]
                if kind == "progress":
                    total, current, note = msg[1], msg[2], msg[3]
                    self.progress.configure(maximum=total)
                    self.progress_var.set(current)
                    self.status_var.set(note)
                elif kind == "done":
                    self._running = False
                    self.run_btn.configure(state=tk.NORMAL)
                    summary, details = msg[1], msg[2]
                    self._load_results(summary, details)
                    self.status_var.set(
                        f"Done. Summary: {len(summary)}, Details: {len(details)}")
                elif kind == "error":
                    self._running = False
                    self.run_btn.configure(state=tk.NORMAL)
                    self.status_var.set("Error")
                    messagebox.showerror("Error", msg[1])
        except queue.Empty:
            pass
        self.after(150, self._process_queue)

    def _clear_results(self):
        for item in self.summary_tree.get_children():
            self.summary_tree.delete(item)
        for item in self.detail_tree.get_children():
            self.detail_tree.delete(item)

    def _load_results(self, summary_rows: list[dict], detail_rows: list[dict]):
        for row in summary_rows:
            values = (
                row.get("SheetName", ""),
                row.get("A_EffectiveRows", 0),
                row.get("B_EffectiveRows", 0),
                row.get("MissingCount", 0),
                row.get("Status", ""),
                row.get("Note", ""),
            )
            self.summary_tree.insert("", tk.END, values=values)

        for row in detail_rows:
            values = (
                row.get("IssueType", ""),
                row.get("SheetName", ""),
                row.get("B_RowNumber", 0),
                row.get("ColA", ""),
                row.get("ColB", ""),
                row.get("ColC", ""),
                row.get("ColD", ""),
                row.get("ColE", ""),
                row.get("ColF", ""),
                row.get("ColG", ""),
                row.get("Key", ""),
            )
            self.detail_tree.insert("", tk.END, values=values)


if __name__ == "__main__":
    app = FixedCompareApp()
    app.mainloop()
