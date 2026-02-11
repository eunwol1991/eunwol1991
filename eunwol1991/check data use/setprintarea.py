import os
import threading
import traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
from typing import List, Optional

import openpyxl

# ========= 常量与工具 =========
DO_SHEET_NAME = "DO"
INVOICE_SHEET_NAME = "Invoice"

TARGET_DO = "authorised signature & stamp"
TARGET_INV = "this is a computer generated invoice. no signature is required."

def norm_text(v) -> str:
    """统一文本：去首尾空白、压缩多空格、转小写。"""
    if v is None:
        return ""
    s = str(v).strip()
    s = " ".join(s.split())
    return s.lower()

def iter_excel_files_excluding_history(directory: str, filename_substring: str) -> List[str]:
    """遍历目录，忽略含 'history' 的文件夹，筛选 .xlsx/.xlsm，且文件名包含 filename_substring（忽略大小写）。"""
    out = []
    key = (filename_substring or "").lower()
    for root, dirs, files in os.walk(directory):
        dirs[:] = [d for d in dirs if "history" not in d.lower()]
        for f in files:
            low = f.lower()
            if low.endswith((".xlsx", ".xlsm")) and key in low:
                out.append(os.path.join(root, f))
    return out

def find_last_row_contains_text(ws, target_norm: str) -> Optional[int]:
    """在整张表里查找“包含 target_norm 的单元格”，返回最后一次出现的行号；找不到返回 None。"""
    last_row = None
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell_txt = norm_text(ws.cell(row=r, column=c).value)
            if cell_txt and target_norm in cell_txt:
                last_row = r
    return last_row

def set_print_area(ws, end_col_letter: str, end_row: int):
    """把打印区域设为 A1:{end_col_letter}{end_row}。"""
    if end_row < 1:
        end_row = 1
    ws.print_area = f"A1:{end_col_letter}{end_row}"

# ========= 主逻辑 =========
def define_print_areas_for_files(
    directory: str,
    filename_substring: str,
    process_do: bool,
    process_invoice: bool,
    do_extra_rows: int,
    inv_extra_rows: int,
    debug: bool,
    log_func=print,
):
    files = iter_excel_files_excluding_history(directory, filename_substring)
    total = len(files)
    log_func(f"共发现 {total} 个候选文件。")

    for i, path in enumerate(files, start=1):
        if debug:
            log_func(f"🔄 [{i}/{total}] 打开：{path}")

        wb = None
        modified = False
        try:
            keep_vba = path.lower().endswith(".xlsm")
            wb = openpyxl.load_workbook(path, keep_vba=keep_vba, data_only=False)

            # 处理 DO：A1:K{终点+1}
            if process_do and DO_SHEET_NAME in wb.sheetnames:
                ws = wb[DO_SHEET_NAME]
                row_do = find_last_row_contains_text(ws, TARGET_DO)
                if row_do is not None:
                    end_row = row_do + int(do_extra_rows)
                    set_print_area(ws, "K", end_row)
                    modified = True
                    if debug:
                        log_func(f"   📄 DO：定位在第 {row_do} 行；打印区域 → A1:K{end_row}")
                else:
                    if debug:
                        log_func(f"   ⚠️ DO：未找到 “{TARGET_DO}”，不改打印区域（现有={ws.print_area}).")

            # 处理 Invoice：A1:I{终点+2}
            if process_invoice and INVOICE_SHEET_NAME in wb.sheetnames:
                ws = wb[INVOICE_SHEET_NAME]
                row_inv = find_last_row_contains_text(ws, TARGET_INV)
                if row_inv is not None:
                    end_row = row_inv + int(inv_extra_rows)
                    set_print_area(ws, "I", end_row)
                    modified = True
                    if debug:
                        log_func(f"   📄 Invoice：定位在第 {row_inv} 行；打印区域 → A1:I{end_row}")
                else:
                    if debug:
                        log_func(f"   ⚠️ Invoice：未找到 “{TARGET_INV}”，不改打印区域（现有={ws.print_area}).")

            if modified:
                wb.save(path)
                if debug:
                    log_func(f"   💾 已保存：{path}")

        except Exception as e:
            traceback.print_exc()
            log_func(f"❌ 处理失败：{path}，错误：{e}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

        yield i, total

# ========= GUI =========
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("批量设置打印区域（DO / Invoice）")
        self.geometry("860x640")

        # 顶部目录
        top = ttk.Frame(self, padding=10)
        top.pack(fill=tk.X)
        ttk.Label(top, text="根目录：").pack(side=tk.LEFT)
        self.dir_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.dir_var, width=70).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="浏览…", command=self.choose_dir).pack(side=tk.LEFT)

        # 参数
        frm = ttk.LabelFrame(self, text="参数设置", padding=10)
        frm.pack(fill=tk.X, padx=10, pady=6)

        self.filename_substring_var = tk.StringVar(value="")  # 留空=全部
        self.debug_var = tk.BooleanVar(value=True)
        self.process_do_var = tk.BooleanVar(value=True)
        self.process_invoice_var = tk.BooleanVar(value=True)
        self.do_extra_var = tk.StringVar(value="1")  # DO 多给一行
        self.inv_extra_var = tk.StringVar(value="2") # Invoice 多给两行

        r = 0
        ttk.Label(frm, text="文件名包含（忽略大小写；留空=全部）：").grid(row=r, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.filename_substring_var, width=28).grid(row=r, column=1, padx=6, pady=2, sticky="w"); r += 1

        ttk.Checkbutton(frm, text="处理 DO 表（定位 Authorised Signature & Stamp → A1:K{行+1}）", variable=self.process_do_var).grid(row=r, column=0, columnspan=2, sticky="w"); r += 1
        ttk.Label(frm, text="DO 额外行数：").grid(row=r, column=0, sticky="e")
        ttk.Entry(frm, textvariable=self.do_extra_var, width=10).grid(row=r, column=1, padx=6, pady=2, sticky="w"); r += 1

        ttk.Checkbutton(frm, text="处理 Invoice 表（定位 'This is a computer generated invoice. No signature is required.' → A1:I{行+2}）", variable=self.process_invoice_var).grid(row=r, column=0, columnspan=2, sticky="w"); r += 1
        ttk.Label(frm, text="Invoice 额外行数：").grid(row=r, column=0, sticky="e")
        ttk.Entry(frm, textvariable=self.inv_extra_var, width=10).grid(row=r, column=1, padx=6, pady=2, sticky="w"); r += 1

        ttk.Checkbutton(frm, text="Debug 日志", variable=self.debug_var).grid(row=r, column=0, sticky="w"); r += 1

        # 操作区
        ops = ttk.Frame(self, padding=10)
        ops.pack(fill=tk.X)
        self.btn_run = ttk.Button(ops, text="开始设置打印区域", command=self.on_run)
        self.btn_run.pack(side=tk.LEFT)

        self.prog = ttk.Progressbar(ops, mode="determinate", length=420)
        self.prog.pack(side=tk.LEFT, padx=10)
        self.prog_label = ttk.Label(ops, text="0/0")
        self.prog_label.pack(side=tk.LEFT)

        # 日志
        frm_log = ttk.LabelFrame(self, text="日志", padding=8)
        frm_log.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        self.log = ScrolledText(frm_log, height=18)
        self.log.pack(fill=tk.BOTH, expand=True)

        self.worker = None

    def choose_dir(self):
        path = filedialog.askdirectory(title="选择根目录")
        if path:
            self.dir_var.set(path)

    def append_log(self, msg: str):
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.update_idletasks()

    def on_run(self):
        if self.worker and self.worker.is_alive():
            messagebox.showinfo("提示", "任务正在进行中…")
            return

        directory = self.dir_var.get().strip()
        if not directory or not os.path.isdir(directory):
            messagebox.showerror("错误", "请选择有效的根目录")
            return

        filename_substring = self.filename_substring_var.get().strip()
        process_do = self.process_do_var.get()
        process_invoice = self.process_invoice_var.get()
        debug = self.debug_var.get()

        try:
            do_extra = int(self.do_extra_var.get().strip() or "1")
            inv_extra = int(self.inv_extra_var.get().strip() or "2")
        except ValueError:
            messagebox.showerror("错误", "额外行数必须是整数")
            return

        if not (process_do or process_invoice):
            messagebox.showerror("错误", "请至少勾选 DO 或 Invoice 其中之一")
            return

        # 清空进度与日志
        self.log.delete("1.0", tk.END)
        self.prog["value"] = 0
        self.prog_label.config(text="0/0")

        def run_task():
            try:
                files = iter_excel_files_excluding_history(directory, filename_substring)
                total = len(files)
                self.prog["maximum"] = max(total, 1)
                self.prog_label.config(text=f"0/{total}")

                gen = define_print_areas_for_files(
                    directory=directory,
                    filename_substring=filename_substring,
                    process_do=process_do,
                    process_invoice=process_invoice,
                    do_extra_rows=do_extra,
                    inv_extra_rows=inv_extra,
                    debug=debug,
                    log_func=self.append_log,
                )
                for done, tot in gen:
                    self.prog["value"] = done
                    self.prog_label.config(text=f"{done}/{tot}")
            except Exception as e:
                self.append_log(f"[GUI] 任务异常: {e}")
            finally:
                self.btn_run.config(state=tk.NORMAL)

        self.btn_run.config(state=tk.DISABLED)
        self.worker = threading.Thread(target=run_task, daemon=True)
        self.worker.start()

if __name__ == "__main__":
    app = App()
    app.mainloop()
