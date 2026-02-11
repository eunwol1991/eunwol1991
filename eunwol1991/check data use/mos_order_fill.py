"""
MOS / LTO Order → Excel 自动填表脚本（支持三行版 + 单行表格版）
依赖：pip install pymupdf openpyxl regex
"""

from pathlib import Path
import re
import sys
import shutil
from difflib import get_close_matches, SequenceMatcher
import fitz                     # PyMuPDF
import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

# === 1. 配置区 ============================================================ #
EXCEL_PATH = Path(
    r"C:\Users\jhunj\Dropbox\for jj\mos order\Order Summary For MOS - JJ.xlsx")
SHEET_NAME = "MOS Format"
PDF_DIR = Path(r"C:\Users\jhunj\Dropbox\for jj\mos_pdfs")  # 待处理 PDF 文件夹
PO_COL = "AI"
ROW_START = 39
ROW_END = 73
COL_START = "C"
COL_END = "AD"
HIGHLIGHT = PatternFill(fill_type="solid", fgColor="00FFFCD7")
ARCHIVE_DIR = Path(r"C:\Users\jhunj\Dropbox\for jj\mos order")

# —— PO 号下限（含）——
PO_MIN = 600000
ENABLE_ITEM_FUZZY = True
ITEM_FUZZY_MIN = 0.88
ITEM_FUZZY_GAP = 0.05
INTERACTIVE_ITEM_SELECT = True

# ======== 门店别名（键已全部小写） ======================================= #
STORE_ALIAS = {
    "mos burger 100 am (38)":                 "MOS - 100AM",
    "mos burger 18 tai seng (67)":            "MOS - 18 Tai Seng",
    "mos burger amk hub (24)":                "MOS - Ang Mo Kio Hub",
    "mos burger bedok mall (42)":             "MOS - Bedok Mall",
    "mos burger jewel changi airport (59)":   "MOS - Café Jewel Changi Airport*",
    "mos burger causeway point (33)":         "MOS - Causeway Point",
    "mos burger changi city point (80)":      "MOS - Changi City Point",
    "mos burger city square mall (69)":       "MOS - City Square Mall",
    "mos burger compass one (49)":           "MOS - Compass One",
    "mos burger eastpoint (64)":             "MOS - Eastpoint Mall",
    "mos burger bukit gombak mrt (71)":       "MOS - Express Bukit Gombak MRT",
    "mos burger holland village mrt (70)":    "MOS - Express Holland Village MRT",
    "mos burger fusionopolis one (55)":       "MOS - Fusionopolis One",
    "mos burger harbourfront centre (31)":    "MOS - Harbourfront Centre",
    "mos burger heartland mall kovan (58)":   "MOS - Heartland Mall Kovan",
    "mos burger hillion mall (56)":           "MOS - Hillion Mall",
    "mos burger hougang mall (63)":           "MOS - Hougang Mall",
    "mos burger ion orchard (79)":            "MOS - ION Orchard",
    "mos burger jem (39)":                    "MOS - JEM",
    "mos burger jurong point (14)":           "MOS - Jurong Point",
    "mos burger kampung admiralty (53)":      "MOS - Kampung Admiralty",
    "mos burger marina bay link mall (81)":   "MOS - Marina Bay Link Mall",
    "mos burger merlion park (72)":           "MOS - Merlion Park",
    "mos burger millenia walk (30)":          "MOS - Millenia Walk",
    "mos burger nex (36)":                    "MOS - NEX",
    "mos burger northpoint city (50)":        "MOS - Northpoint",
    "mos burger our tampines hub (57)":       "MOS - Our Tampines Hub",
    "mos burger raffles city (20)":           "MOS - Raffles City",
    "mos burger seletar mall (75)":           "MOS - Seletar Mall",
    "mos burger tampines mall (35)":          "MOS - Tampines Mall",
    "mos burger the centrepoint (51)":        "MOS - The Centrepoint",
    "mos burger tiong bahru plaza (48)":      "MOS - Tiong Bahru Plaza",
    "mos burger toa payoh hdb hub (12)":      "MOS - Toa Payoh HDB Hub",
    "mos burger waterway point (62)":         "MOS - Waterway Point",
    "mos burger white sands (47)":            "MOS - White Sands",
}

# ======== 物品别名（键已全部小写） ======================================= #
ITEM_ALIAS = {
    "ikeda japanese chicken cutlet":                "Ikeda Japanese Chicken Cutlet (6 x 1.1kg)",
    "japanese chicken katsu 70g":                   "CS TAY Japanese Chicken Katsu 70G (6 x 1kg)",
    "tays golden crispy chicken wing":              "Golden Crispy Chicken Wing (5 pkts x 10 pcs x 1.15kg)",
    "cs tay crispy chicken patty":                  "CS Tay Crispy Chicken Patty (4 x 2.5kg)",
    "eb kranch alaska pollock fingers":             "Kranch Alaska Pollock Finger (5 x 900g)",
    "battered natural onion rings":                 "Salud Battered Natural Onion Ring (6 x 1kg)",
    "rosti hashbrown":                              "Rosti Hashbrown (10 x 1kg)",
    "tomato capsicum soup":                         "Tomato Capsicum Soup (10 x 1kg)",
    "par cooked beef patty 120g":                   "Cooked Hamburg Beef Patty 120G (10 x 720g)",
    "demi glace sauce":                             "Demi Glace Sauce (10 x 1kg)",
    "chilli crab sauce":                            "Chilli Crab Sauce (10 x 1kg)",
    "anchor minidish unsalted butter 7g":           "Anchor Minidish Unsalted Butter (1 x 144 x 7g)",
    "fried chicken tulip":                          "Fried Chicken Tulip (10 x 1kg)",
    "tsukune japanese minced chicken with potato":  "Tsukune Japanese Minced Chicken with Potato (20 x 480g)",
    "mini mantou, plain":                           "Mini Mantou Plain (24 x 180g)",
    "bobo sliced fish cake 1kg":                    "BoBo Sliced Fish Cake (1kg)",
    "chicken honey baked ham sliced":               "Chicken Honey Baked Ham Sliced (12 x 1kg)",
    "mozzarella cheese 5g pearl iqf":               "Mozzerella Cheese 5G Pearls IQF (8 x 1kg)",
    "premium smoked salmon sliced r trout":         "Premium Smoked Salmon Sliced (R Trout) (10 x 1kg)",
    "hard boiled egg 10pcs only":                   "Hard Boiled Egg (18 x 10pcs)",
    "fried shallot 1kg":                            "Fried Shallot (10 x 1kg)",
    "jalapenos nachos sliced":                      "Jalapeno Pepper Sliced (6 x 2.95kg)",
    "parsley flakes":                               "Parsley Leaves (6 x 5gm)",
    "toasted coconut flakes":                       "Toasted Coconut Flakes (36 x 200g)",
    "red sauce":                               "Red Sauce (10 x 1kg)",
    "j-lek sriracha sauce":                         "J-Lek Sriracha Hot Chilli Sauce (12 x 455ml)",
    "uht coconut water 1l":                         "UHT Coconut Water 1L (12 x 1L)",
}


# ========================================================================== #
UNIT_RE = r"(?:ctn|ctns|pkt|pkts|tin|tins|can|cans|box|boxes|btl|btls|pc|pcs)"

_PLURAL_MAP = {
    "ctn": "ctns",
    "pkt": "pkts",
    "tin": "tins",
    "can": "cans",
    "box": "boxes",
    "btl": "btls",
    "pc": "pcs",
}
_SINGULAR_MAP = {v: k for k, v in _PLURAL_MAP.items()}


def _normalize_uom(uom: str) -> str:
    u = (uom or "").strip().lower()
    if u in _SINGULAR_MAP:
        return _SINGULAR_MAP[u]
    return u


def _format_qty_uom(qty: int, uom: str) -> str:
    base = _normalize_uom(uom)
    if qty > 1:
        plural = _PLURAL_MAP.get(base, base + "s")
        return f"{qty} {plural}"
    return f"{qty} {base}"


def _norm_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()


def _looks_like_date_fragment(s: str) -> bool:
    if not s:
        return False
    return ("/" in s) or bool(re.search(r"20\d{2}", s))


# ---------- Item / Store normalization ---------- #

def _norm_item_key(desc: str) -> str:
    s = desc.lower()
    s = re.sub(r"\(.*?\)", "", s)
    s = re.sub(r"(?<=\d)\s*[xX]\s*(?=\d)", " ", s)
    s = re.sub(r"(?<=\d)\s*[xX]\s*(?=[a-z])", " ", s)
    s = re.sub(r"(?<=[a-z])\s*[xX]\s*(?=\d)", " ", s)
    s = re.sub(r"([a-z])(\d)", r"\1 \2", s)
    s = re.sub(r"(\d)([a-z])", r"\1 \2", s)
    s = re.sub(r"[^a-z0-9\s,.-]+", " ", s)
    return _norm_spaces(s)


def _map_item(desc: str) -> str:
    key = _norm_item_key(desc)
    if key in ITEM_ALIAS:
        return ITEM_ALIAS[key]
    k2 = key.replace(",", "")
    if k2 in ITEM_ALIAS:
        return ITEM_ALIAS[k2]
    for k in ITEM_ALIAS:
        if k in key or key in k:
            return ITEM_ALIAS[k]
    return desc


def _suggest_store_candidates(pdf_store_line: str, store_row_keys: list[str], limit: int = 3) -> list[str]:
    if not pdf_store_line:
        return []
    norm_line = _norm_store_text(pdf_store_line)
    norm_map = {}
    for k in store_row_keys:
        nk = _norm_store_text(k)
        if nk not in norm_map:
            norm_map[nk] = k
    matches = get_close_matches(norm_line, list(norm_map.keys()), n=limit, cutoff=0.4)
    return [norm_map[m] for m in matches]


def _suggest_item_candidates(item_desc: str, item_col_keys: list[str], limit: int = 3) -> list[str]:
    if not item_desc:
        return []
    norm_desc = _norm_item_key(item_desc)
    norm_map = {}
    for k in item_col_keys:
        nk = _norm_item_key(k)
        if nk not in norm_map:
            norm_map[nk] = k
    matches = get_close_matches(norm_desc, list(norm_map.keys()), n=limit, cutoff=0.5)
    return [norm_map[m] for m in matches]


def _build_item_norm_map(item_col_keys: list[str]) -> dict[str, str]:
    norm_map = {}
    for k in item_col_keys:
        nk = _norm_item_key(k)
        if nk and nk not in norm_map:
            norm_map[nk] = k
    return norm_map


def _fuzzy_match_item(item_desc: str, item_norm_map: dict[str, str]) -> dict | None:
    if not item_desc:
        return None
    norm_desc = _norm_item_key(item_desc)
    if not norm_desc:
        return None

    best_score = 0.0
    best_key = None
    second_score = 0.0
    for nk, orig in item_norm_map.items():
        score = SequenceMatcher(None, norm_desc, nk).ratio()
        if score > best_score:
            second_score = best_score
            best_score = score
            best_key = orig
        elif score > second_score:
            second_score = score

    if best_key and best_score >= ITEM_FUZZY_MIN and (best_score - second_score) >= ITEM_FUZZY_GAP:
        return {"item": best_key, "score": round(best_score, 3)}
    return None


def _prompt_use_item_suggestion(pdf_name: str, store: str | None, po: str | None, desc: str,
                                suggestions: list[str]) -> bool:
    print("\n[匹配建议] 发现可能的 Excel 列名，请确认是否自动匹配：")
    print(f"- PDF: {pdf_name}, 门店: {store}, PO: {po}, 品名: {desc}")
    print(f"  可能列名: {', '.join(suggestions)}")
    while True:
        ans = input("  是否使用第 1 个候选列并临时加入别名? (y/n): ").strip().lower()
        if ans in {"y", "yes"}:
            return True
        if ans in {"n", "no"}:
            return False
        print("  请输入 y 或 n。")


def _collect_item_hint_lines(lines: list[str], limit: int = 5) -> list[str]:
    hints = []
    for ln in lines:
        if UNIT_WORD_RE.search(ln):
            hints.append(ln)
            if len(hints) >= limit:
                break
    return hints


def _norm_store_text(s: str) -> str:
    s = s.lower()
    s = re.sub(r"\(.*?\)", "", s)
    s = s.replace("mos", " ").replace("burger", " ").replace("mosburger", " ")
    s = s.replace("mos -", " ").replace("mos–", " ").replace("-", " ")
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return _norm_spaces(s)


def resolve_store(pdf_store_line: str, store_row_keys: list[str]) -> str | None:
    alias_hit = STORE_ALIAS.get(pdf_store_line.lower())
    if alias_hit:
        return alias_hit

    pdf_norm = _norm_store_text(pdf_store_line)
    excel_norm_map = {_norm_store_text(k): k for k in store_row_keys}

    for e, orig in excel_norm_map.items():
        if pdf_norm in e or e in pdf_norm:
            return orig

    words = [w for w in pdf_norm.split() if len(w) >= 3]
    for e, orig in excel_norm_map.items():
        if all(w in e for w in words):
            return orig

    return None


def _ensure_dir(p: Path):
    try:
        p.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass


def _safe_move_file(src: Path, dest_dir: Path):
    try:
        _ensure_dir(dest_dir)
        target = dest_dir / src.name
        if target.exists():
            stem, suf, i = src.stem, src.suffix, 1
            while (dest_dir / f"{stem} ({i}){suf}").exists():
                i += 1
            target = dest_dir / f"{stem} ({i}){suf}"
        shutil.move(str(src), str(target))
    except Exception as e:
        print(f"[!] 无法移动文件 {src} -> {dest_dir}: {e}", file=sys.stderr)


# ---------- LTO 单行结构解析（新增） ---------- #

LINE_ITEM_RE = re.compile(
    rf"^\s*\d+\s+(.+?)\s+(\d+)\s+({UNIT_RE})\b", re.I
)
UNIT_WORD_RE = re.compile(rf"\b{UNIT_RE}\b", re.I)


def parse_single_line_items(lines: list[str]) -> list:
    items = []
    for ln in lines:
        m = LINE_ITEM_RE.match(ln)
        if not m:
            continue

        desc = _norm_spaces(m.group(1))
        qty = int(m.group(2))
        uom = m.group(3).lower()

        items.append({"desc": desc, "qty": qty, "uom": uom})

    return items


# ---------- PDF basic parsing ---------- #

def parse_pdf(pdf_path: Path):
    pages = []
    with fitz.open(pdf_path) as doc:
        for page in doc:
            raw = page.get_text("text") or ""
            pages.append([ln.strip() for ln in raw.splitlines() if ln.strip()])
    return pages


# ---------- 解析 PO（保留你的 PO_MIN 逻辑） ---------- #

def _extract_po_from_lines(lines: list[str]) -> str:
    joined = " \n ".join(lines)

    m = re.search(r"(?is)approved\s*by[^\d]{0,80}?(\d{6})", joined)
    if m and int(m.group(1)) >= PO_MIN:
        return m.group(1)

    for m in re.finditer(r"\b\d{6}\b", joined):
        if int(m.group(0)) >= PO_MIN:
            return m.group(0)

    return ""


# ---------- 主解析：同时支持三行结构 + 单行结构 ---------- #

def extract_orders_from_lines(lines: list[str]) -> dict:
    store_line = None
    for ln in lines:
        if re.fullmatch(r"(?i)mos burger.+\(\d+\)", ln):
            store_line = ln
            break

    po = _extract_po_from_lines(lines)

    # 先试 LTO 单行结构
    items = parse_single_line_items(lines)

    # 若单行结构抓不到，再尝试三行结构（旧格式兼容）
    if not items:
        items = []
        qty_re = re.compile(rf"^(\d+)\s+({UNIT_RE})$", re.I)
        i = 0
        while i < len(lines) - 2:
            desc = lines[i]
            m_qty = qty_re.fullmatch(lines[i + 1])
            if m_qty:
                items.append({
                    "desc": desc,
                    "qty": int(m_qty.group(1)),
                    "uom": m_qty.group(2).lower()
                })
                i += 2
                continue
            i += 1

    return {"store_line": store_line, "po": po, "items": items}


# ---------- Excel 表头映射 ---------- #

def build_mappings(ws):
    store_row = {}
    for r in range(ROW_START, ROW_END + 1):
        v = ws[f"B{r}"].value
        if v:
            store_row[_norm_spaces(str(v))] = r

    c1 = column_index_from_string(COL_START)
    c2 = column_index_from_string(COL_END)

    item_col = {}
    for c in range(c1, c2 + 1):
        v = ws.cell(38, c).value
        if v:
            item_col[_norm_spaces(str(v))] = c

    return store_row, item_col


def clear_target_cells(ws):
    c1 = column_index_from_string(COL_START)
    c2 = column_index_from_string(COL_END)

    for r in range(ROW_START, ROW_END + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).value = None
            ws.cell(r, c).fill = PatternFill()
        ws[f"{PO_COL}{r}"].value = None


# ---------- 主流程 ---------- #

def main():
    wb = xl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    clear_target_cells(ws)

    store_row, item_col = build_mappings(ws)

    not_found_store = []
    not_found_store_seen = set()
    # 记录找不到物品的详细信息：每一笔是一个 dict
    # [{"pdf": xxx, "store": xxx, "po": xxx, "desc": xxx, "qty": n, "uom": "ctn"}, ...]
    not_found_item = []
    no_item_pdf = []
    missing_po = []
    missing_po_seen = set()
    fuzzy_item_matches = []
    user_added_aliases = []
    archive_skipped = []
    declined_item_keys = set()
    updated = []

    pdf_list = sorted(PDF_DIR.glob("*.pdf"))
    store_row_keys = list(store_row.keys())
    item_col_keys = list(item_col.keys())
    item_norm_map = _build_item_norm_map(item_col_keys)

    for pdf_file in pdf_list:
        pdf_had_update = False
        pdf_found_items = False
        pdf_hint_lines = []
        parse_ok = True
        pdf_block_archive = False

        try:
            pages = parse_pdf(pdf_file)
            for lines in pages:
                info = extract_orders_from_lines(lines)
                if not info["items"]:
                    if len(pdf_hint_lines) < 5:
                        pdf_hint_lines.extend(
                            _collect_item_hint_lines(lines, 5 - len(pdf_hint_lines))
                        )
                    continue
                pdf_found_items = True

                resolved_store = None
                store_line = info["store_line"]
                if store_line:
                    resolved_store = resolve_store(store_line, store_row_keys)
                else:
                    for ln in lines:
                        if "mos" in ln.lower():
                            store_line = ln
                            resolved_store = resolve_store(ln, store_row_keys)
                            if resolved_store:
                                break

                if not resolved_store:
                    reason = "门店行未找到或无法匹配到 Excel 门店行"
                    store_key = (pdf_file.name, store_line, reason)
                    if store_key not in not_found_store_seen:
                        not_found_store_seen.add(store_key)
                        not_found_store.append({
                            "pdf": pdf_file.name,
                            "store_line": store_line,
                            "store_line_norm": _norm_store_text(store_line or ""),
                            "reason": reason,
                            "candidates": _suggest_store_candidates(store_line, store_row_keys),
                            "mos_lines": [ln for ln in lines if "mos" in ln.lower()][:3],
                        })
                    continue

                r = store_row.get(resolved_store)
                if not r:
                    reason = "门店已解析但在 Excel 中找不到对应行"
                    store_key = (pdf_file.name, resolved_store, reason)
                    if store_key not in not_found_store_seen:
                        not_found_store_seen.add(store_key)
                        not_found_store.append({
                            "pdf": pdf_file.name,
                            "store_line": resolved_store,
                            "store_line_norm": _norm_store_text(resolved_store),
                            "reason": reason,
                            "candidates": _suggest_store_candidates(resolved_store, store_row_keys),
                            "mos_lines": [ln for ln in lines if "mos" in ln.lower()][:3],
                        })
                    continue

                for it in info["items"]:
                    item_std = _map_item(it["desc"])
                    c = item_col.get(item_std)
                    fuzzy_used = None
                    if not c and ENABLE_ITEM_FUZZY:
                        fuzzy_used = _fuzzy_match_item(item_std, item_norm_map)
                        if fuzzy_used:
                            item_std = fuzzy_used["item"]
                            c = item_col.get(item_std)

                    if not c:
                        suggestions = _suggest_item_candidates(item_std, item_col_keys)
                        norm_desc = _norm_item_key(it["desc"])
                        if suggestions and INTERACTIVE_ITEM_SELECT and norm_desc not in declined_item_keys:
                            use_suggest = _prompt_use_item_suggestion(
                                pdf_file.name, resolved_store, info["po"], it["desc"], suggestions)
                            if use_suggest:
                                item_std = suggestions[0]
                                c = item_col.get(item_std)
                                if c:
                                    if norm_desc and norm_desc not in ITEM_ALIAS:
                                        ITEM_ALIAS[norm_desc] = item_std
                                        user_added_aliases.append({
                                            "alias": it["desc"],
                                            "mapped": item_std,
                                        })
                                else:
                                    c = None
                            else:
                                declined_item_keys.add(norm_desc)
                                pdf_block_archive = True

                    if not c:
                        # 记录：是哪一个 PDF、门店、PO、原始品名、数量、单位
                        not_found_item.append({
                            "pdf": pdf_file.name,
                            "store": resolved_store,
                            "po": info["po"],
                            "desc": it["desc"],
                            "mapped": item_std,
                            "qty": it.get("qty"),
                            "uom": it.get("uom"),
                            "reason": "Excel 表头没有对应的品名列",
                            "suggestions": _suggest_item_candidates(item_std, item_col_keys),
                        })
                        continue

                    ws.cell(r, c).value = _format_qty_uom(
                        it["qty"], it["uom"])
                    ws.cell(r, c).fill = HIGHLIGHT
                    pdf_had_update = True
                    if fuzzy_used:
                        fuzzy_item_matches.append({
                            "pdf": pdf_file.name,
                            "store": resolved_store,
                            "po": info["po"],
                            "desc": it["desc"],
                            "matched": item_std,
                            "score": fuzzy_used["score"],
                        })

                # 写入 PO
                if info["po"]:
                    po_cell = ws[f"{PO_COL}{r}"]
                    po_cell.value = int(info["po"])
                    pdf_had_update = True
                else:
                    po_key = (pdf_file.name, resolved_store)
                    if po_key not in missing_po_seen:
                        missing_po_seen.add(po_key)
                        missing_po.append({
                            "pdf": pdf_file.name,
                            "store": resolved_store,
                            "reason": f"未找到 PO（需 >= {PO_MIN}）",
                        })

                updated.append((resolved_store, info["po"]))

        except Exception as e:
            parse_ok = False
            print(f"[!] 解析出错 {pdf_file.name}: {e}")

        if parse_ok and not pdf_found_items:
            no_item_pdf.append({
                "pdf": pdf_file.name,
                "reason": "未匹配到单行或三行的商品结构",
                "hints": pdf_hint_lines,
            })

        if pdf_had_update and not pdf_block_archive:
            _safe_move_file(pdf_file, ARCHIVE_DIR)
        elif pdf_had_update and pdf_block_archive:
            archive_skipped.append({
                "pdf": pdf_file.name,
                "reason": "用户选择不自动匹配，暂不移动文件",
            })

    wb.save(EXCEL_PATH)
    wb.close()

    print(f"已更新 {len(updated)} 条记录, 处理 {len(pdf_list)} 份 PDF → {EXCEL_PATH}")

    if not_found_store:
        print("\n[警告] 以下门店无法匹配到 Excel 门店行，请检查门店名称或 STORE_ALIAS：\n")
        for rec in not_found_store:
            store_line = rec.get("store_line") or "(未找到门店行)"
            print(f"- PDF: {rec['pdf']}, 门店原文: {store_line}, 原因: {rec['reason']}")
            if rec.get("store_line_norm"):
                print(f"  规范化: {rec['store_line_norm']}")
            if rec.get("candidates"):
                print(f"  可能门店: {', '.join(rec['candidates'])}")
            if rec.get("mos_lines"):
                print(f"  PDF 中疑似门店行: {', '.join(rec['mos_lines'])}")

    if no_item_pdf:
        print("\n[警告] 以下 PDF 没有解析到任何商品行，可能格式不匹配：\n")
        for rec in no_item_pdf:
            print(f"- PDF: {rec['pdf']}, 原因: {rec['reason']}")
            if rec.get("hints"):
                print(f"  参考行: {', '.join(rec['hints'])}")

    if missing_po:
        print("\n[提示] 以下记录未找到 PO（可能格式变化或 PO 小于阈值）：\n")
        for rec in missing_po:
            print(f"- PDF: {rec['pdf']}, 门店: {rec['store']}, 原因: {rec['reason']}")

    if user_added_aliases:
        print("\n[提示] 以下品名已临时加入别名（仅本次运行有效）：\n")
        for rec in user_added_aliases:
            print(f"- 别名: {rec['alias']} → {rec['mapped']}")

    if archive_skipped:
        print("\n[提示] 以下文件未移动到归档目录：\n")
        for rec in archive_skipped:
            print(f"- PDF: {rec['pdf']}, 原因: {rec['reason']}")

    if fuzzy_item_matches:
        print("\n[提示] 以下物品使用了模糊匹配并已写入 Excel：\n")
        for rec in fuzzy_item_matches:
            print(
                f"- PDF: {rec['pdf']}, 门店: {rec['store']}, PO: {rec['po']}, "
                f"原始品名: {rec['desc']} → 匹配列: {rec['matched']} (相似度 {rec['score']})"
            )

    # 若有 PO 里的 items 没有在 Excel 找到对应列，提醒用户
    if not_found_item:
        print("\n[警告] 以下物品在 Excel 中找不到对应的列，请检查 ITEM_ALIAS 或表头：\n")
        seen = set()
        for rec in not_found_item:
            key = (rec["pdf"], rec["store"], rec["po"], rec["desc"])
            if key in seen:
                continue
            seen.add(key)
            desc = rec["desc"]
            mapped = rec.get("mapped")
            if mapped and mapped != desc:
                desc = f"{desc} (映射后: {mapped})"
            print(
                f"- PDF: {rec['pdf']}, 门店: {rec['store']}, PO: {rec['po']}, "
                f"品名: {desc}, 数量: {rec['qty']} {rec['uom']}"
            )
            if rec.get("suggestions"):
                print(f"  可能列名: {', '.join(rec['suggestions'])}")
        print("\n请根据以上资讯，\n1) 确认该物品是否应加入 ITEM_ALIAS；\n2) 或确认 Excel 第 38 行的品名是否与标准描述一致。")
    else:
        print("\n已解析到的物品均已成功匹配到 Excel 列，没有遗漏的品项。")


if __name__ == "__main__":
    main()
