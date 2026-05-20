import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from streamlit.testing.v1 import AppTest


SCRIPT_PATH = Path(__file__).with_name("stock_datagrid.py")


def _create_workbook(path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Stocks report"
    ws1.append([])
    ws1.append([])
    ws1.append(
        [
            "Supplier",
            "Brand",
            "Product Code",
            "Description",
            "Pack Size",
            "Unit",
            "Expiry Date",
            "Relabel To Date",
            "Daily Update",
            "Stock Qty",
        ]
    )
    ws1.append(
        [
            "Acme",
            "Sauces",
            "P001",
            "Aioli Sauce (Cold)",
            "24 x 50g",
            "ctn",
            "2026-04-01",
            "2026-03-20",
            "",
            10,
        ]
    )
    ws1.append(
        [
            "Acme",
            "Frozen",
            "P002",
            "Bobo Sliced Fish Cake",
            "12 x 100g",
            "ctn",
            "2026-05-01",
            "2026-03-22",
            "",
            8,
        ]
    )
    ws2 = wb.create_sheet("Lai Hock Whse")
    ws2.append([])
    ws2.append([])
    ws2.append(
        [
            "Supplier",
            "Brand",
            "Product Code",
            "Description",
            "Unused",
            "Pack Size",
            "Unit",
            "Expiry Date",
            "Relabel To Date",
            "Stocks Balance",
            "Stock Qty",
        ]
    )
    ws2.append(
        [
            "Acme",
            "Sauces",
            "P001",
            "Aioli Sauce (Cold)",
            "",
            "24 x 50g",
            "ctn",
            "2026-04-01",
            "2026-03-20",
            "",
            5,
        ]
    )
    ws2.append(
        [
            "Acme",
            "Sauces",
            "P003",
            "Brown Sauce (Hot)",
            "",
            "24 x 50g",
            "ctn",
            "2026-06-01",
            "2026-03-24",
            "",
            12,
        ]
    )
    wb.save(path)


def _create_sales_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Delivery details"
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append(
        [
            "Year",
            "Date",
            "Month",
            "Brand/Category",
            "Supplier",
            "Product Code",
            "Product Description",
            "Carton Packing",
            "Customer",
            "Outlet",
            "Qty in Pcs",
            "Qty in Ctns",
            "Total Qty in Pcs",
            "Total Qty in Ctns",
            "Invoice #",
            "Total Value",
            "GST",
            "Total Value Inclusive GST",
            "Account",
            "Customer PO#",
            "Remarks",
        ]
    )
    ws.append(
        [
            "2026",
            "2026-04-01",
            "Apr",
            "Sauces",
            "Acme",
            "P001",
            "Aioli Sauce",
            "24 x 50g",
            "Customer A",
            "Outlet A",
            5,
            13,
            0,
            0,
            "INV-001",
            100,
            0,
            100,
            "Account A",
            "",
            "",
        ]
    )
    wb.save(path)


def _create_wrapper(path: Path, workbook_path: Path) -> None:
    path.write_text(
        "import sys\n"
        "from pathlib import Path\n"
        "import streamlit as st\n"
        f"script = Path(r'{SCRIPT_PATH}')\n"
        "sys.path.insert(0, str(script.parent))\n"
        "import stock_datagrid\n"
        f"file_path = Path(r'{workbook_path}')\n"
        "st.session_state['stock_file_name'] = file_path.name\n"
        "st.session_state['stock_file_bytes'] = file_path.read_bytes()\n"
        "stock_datagrid.run_stock_page()\n",
        encoding="utf-8",
    )


def _create_reconciliation_wrapper(
    path: Path, stock_workbook_path: Path, sales_workbook_path: Path
) -> None:
    path.write_text(
        "import sys\n"
        "from pathlib import Path\n"
        "import streamlit as st\n"
        f"script = Path(r'{SCRIPT_PATH}')\n"
        "sys.path.insert(0, str(script.parent))\n"
        "import stock_datagrid\n"
        f"stock_path = Path(r'{stock_workbook_path}')\n"
        f"sales_path = Path(r'{sales_workbook_path}')\n"
        "st.session_state['reconciliation_stock_file_name'] = stock_path.name\n"
        "st.session_state['reconciliation_stock_file_bytes'] = stock_path.read_bytes()\n"
        "st.session_state['reconciliation_sales_files_payload'] = [{'name': sales_path.name, 'bytes': sales_path.read_bytes()}]\n"
        "stock_datagrid.run_reconciliation_page()\n",
        encoding="utf-8",
    )


def _create_reconciliation_wrapper_with_old_date_state(
    path: Path, stock_workbook_path: Path, sales_workbook_path: Path
) -> None:
    path.write_text(
        "import sys\n"
        "from pathlib import Path\n"
        "import streamlit as st\n"
        f"script = Path(r'{SCRIPT_PATH}')\n"
        "sys.path.insert(0, str(script.parent))\n"
        "import stock_datagrid\n"
        f"stock_path = Path(r'{stock_workbook_path}')\n"
        f"sales_path = Path(r'{sales_workbook_path}')\n"
        "st.session_state['reconciliation_stock_file_name'] = stock_path.name\n"
        "st.session_state['reconciliation_stock_file_bytes'] = stock_path.read_bytes()\n"
        "st.session_state['reconciliation_sales_files_payload'] = [{'name': sales_path.name, 'bytes': sales_path.read_bytes()}]\n"
        "st.session_state['reconciliation_filter_date'] = ['01-Apr-2026']\n"
        "stock_datagrid.run_reconciliation_page()\n",
        encoding="utf-8",
    )


class StockFilterUiTests(unittest.TestCase):
    def test_stock_filters_render_as_multiselects_like_supplier(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            workbook_path = temp_path / "stock.xlsx"
            wrapper_path = temp_path / "wrapper.py"
            _create_workbook(workbook_path)
            _create_wrapper(wrapper_path, workbook_path)

            sys.path.insert(0, str(SCRIPT_PATH.parent))
            at = AppTest.from_file(str(wrapper_path), default_timeout=60)
            at.run()

            multiselect_labels = [x.label for x in at.multiselect]
            text_input_labels = [x.label for x in at.text_input]

            self.assertIn("Description（去括号后）", multiselect_labels)
            self.assertIn("Product Code", multiselect_labels)
            self.assertIn("Remark（来自描述括号）", multiselect_labels)
            self.assertNotIn("Description（去括号后）", text_input_labels)
            self.assertNotIn("Product Code", text_input_labels)
            self.assertNotIn("Remark（来自描述括号）", text_input_labels)

    def test_reconciliation_page_renders_mismatch_summary(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            stock_workbook_path = temp_path / "stock.xlsx"
            sales_workbook_path = temp_path / "sales.xlsx"
            wrapper_path = temp_path / "wrapper_reconciliation.py"

            _create_workbook(stock_workbook_path)
            stock_wb = Workbook()
            stock_ws = stock_wb.active
            if stock_ws is None:
                self.fail("Expected active worksheet")
            stock_ws.title = "Stocks report"
            stock_ws.append([None] * 12)
            stock_ws.append([None] * 10 + ["1-Apr"])
            stock_ws.append(
                [
                    "Supplier",
                    "Brand",
                    "Product Code",
                    "Description",
                    "Pack Size",
                    "Unit",
                    "Expiry Date",
                    "Relabel To Date",
                    "Daily Update",
                    "Stock Qty",
                    "",
                    "",
                ]
            )
            stock_ws.append(
                [
                    "Acme",
                    "Sauces",
                    "P001",
                    "Aioli Sauce",
                    "24 x 50g",
                    "ctn",
                    "2026-04-01",
                    "2026-03-20",
                    "",
                    10,
                    "13 ctns 5 pkts",
                    "",
                ]
            )
            stock_wb.save(stock_workbook_path)
            _create_sales_workbook(sales_workbook_path)
            _create_reconciliation_wrapper(
                wrapper_path, stock_workbook_path, sales_workbook_path
            )

            sys.path.insert(0, str(SCRIPT_PATH.parent))
            at = AppTest.from_file(str(wrapper_path), default_timeout=60)
            at.run()

            text_values = [getattr(node, "value", "") for node in at.markdown] + [
                getattr(node, "value", "") for node in at.caption
            ]
            joined_text = "\n".join(str(value) for value in text_values)
            multiselect_labels = [x.label for x in at.multiselect]
            date_input_labels = [x.label for x in at.date_input]
            checkbox_labels = [x.label for x in at.checkbox]

            self.assertIn("Reconciliation", joined_text)
            self.assertIn("Mismatch rows", joined_text)
            self.assertIn("Product Code", multiselect_labels)
            self.assertIn("Description", multiselect_labels)
            self.assertIn("Year", multiselect_labels)
            self.assertIn("Month", multiselect_labels)
            self.assertIn("Use Specific Date", checkbox_labels)
            self.assertIn("Date Range", date_input_labels)
            self.assertIn("Specific Date Override", date_input_labels)
            self.assertEqual(len(at.date_input), 2)

    def test_reconciliation_page_accepts_old_list_date_state(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            stock_workbook_path = temp_path / "stock.xlsx"
            sales_workbook_path = temp_path / "sales.xlsx"
            wrapper_path = temp_path / "wrapper_reconciliation_old_date.py"

            stock_wb = Workbook()
            stock_ws = stock_wb.active
            if stock_ws is None:
                self.fail("Expected active worksheet")
            stock_ws.title = "Stocks report"
            stock_ws.append([None] * 12)
            stock_ws.append([None] * 10 + ["1-Apr"])
            stock_ws.append([None] * 12)
            stock_ws.append(
                [
                    "Acme",
                    "Sauces",
                    "P001",
                    "Aioli Sauce",
                    "24 x 50g",
                    "ctn",
                    "2026-04-01",
                    "2026-03-20",
                    "",
                    10,
                    "13 ctns 5 pkts",
                    "",
                ]
            )
            stock_wb.save(stock_workbook_path)
            _create_sales_workbook(sales_workbook_path)
            _create_reconciliation_wrapper_with_old_date_state(
                wrapper_path, stock_workbook_path, sales_workbook_path
            )

            sys.path.insert(0, str(SCRIPT_PATH.parent))
            at = AppTest.from_file(str(wrapper_path), default_timeout=60)
            at.run()

            self.assertFalse(at.exception)


if __name__ == "__main__":
    unittest.main()
