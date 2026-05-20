import importlib.util
import sys
import unittest
import datetime
import io
from pathlib import Path

import pandas as pd
from openpyxl import Workbook


MODULE_PATH = Path(__file__).with_name("stock_datagrid.py")


def _load_module():
    sys.path.insert(0, str(MODULE_PATH.parent))
    spec = importlib.util.spec_from_file_location("stock_datagrid", MODULE_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load module from {MODULE_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class _DummyContext:
    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False


class _FakeStreamlit:
    def __init__(self, session_state=None):
        self.session_state = session_state if session_state is not None else {}

    def button(self, *args, **kwargs):
        return False

    def form(self, *args, **kwargs):
        return _DummyContext()

    def toggle(self, label, key=None, value=False, **kwargs):
        if key is None:
            return value
        return self.session_state.setdefault(key, value)

    def multiselect(self, label, options, key=None, **kwargs):
        if key is not None:
            self.session_state.setdefault(key, [])
            return self.session_state[key]
        return []

    def columns(self, count):
        return [_DummyContext() for _ in range(count)]

    def expander(self, *args, **kwargs):
        return _DummyContext()

    def text_input(self, label, key=None, value="", **kwargs):
        if key is not None:
            return self.session_state.setdefault(key, value)
        return value

    def selectbox(self, label, options, key=None, **kwargs):
        default = options[0] if options else None
        if key is not None:
            return self.session_state.setdefault(key, default)
        return default

    def date_input(self, label, value=None, key=None, **kwargs):
        if key is not None:
            if key not in self.session_state:
                self.session_state[key] = value
            return self.session_state[key]
        return value

    def form_submit_button(self, *args, **kwargs):
        return False


class StockFilterQueryTests(unittest.TestCase):
    def test_format_quantity_breakdown_preserves_original_unit_labels(self):
        module = _load_module()

        text = module._format_quantity_breakdown(
            {
                "ctn": 2.0,
                "btl": 4.0,
                "box": 3.0,
                "tin": 1.0,
            }
        )

        self.assertEqual(text, "2 ctns 4 btl 3 box 1 tin")

    def test_aggregate_summary_counts_box_btl_and_tin_quantities_in_packet_totals(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {
                    "supplier": "Supplier Alpha",
                    "brand": "Brand A",
                    "product_code": "SKU-01",
                    "description": "Chili Sauce",
                    "pack_size": "12 x 500ml",
                    "warehouse": "Savori Whse",
                    "unit": "btl",
                    "stock_qty": 4,
                },
                {
                    "supplier": "Supplier Alpha",
                    "brand": "Brand A",
                    "product_code": "SKU-01",
                    "description": "Chili Sauce",
                    "pack_size": "12 x 500ml",
                    "warehouse": "Savori Whse",
                    "unit": "box",
                    "stock_qty": 3,
                },
                {
                    "supplier": "Supplier Alpha",
                    "brand": "Brand A",
                    "product_code": "SKU-01",
                    "description": "Chili Sauce",
                    "pack_size": "12 x 500ml",
                    "warehouse": "Lai Hock Whse",
                    "unit": "tin",
                    "stock_qty": 2,
                },
            ]
        )

        summary_df, _detail_map = module.aggregate_summary(df)

        self.assertEqual(len(summary_df), 1)
        self.assertEqual(
            summary_df.loc[0, "savori_qty_breakdown"],
            {"btl": 4.0, "box": 3.0},
        )
        self.assertEqual(summary_df.loc[0, "lai_hock_qty_breakdown"], {"tin": 2.0})
        self.assertEqual(
            summary_df.loc[0, "total_qty_breakdown"],
            {"btl": 4.0, "box": 3.0, "tin": 2.0},
        )
        self.assertEqual(summary_df.loc[0, "savori_pkt"], 7.0)
        self.assertEqual(summary_df.loc[0, "lai_hock_pkt"], 2.0)
        self.assertEqual(summary_df.loc[0, "total_pkt"], 9.0)

    def test_split_by_expiry_counts_box_btl_and_tin_quantities_in_subtotals(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {
                    "warehouse": "Savori Whse",
                    "unit": "btl",
                    "stock_qty": 4,
                    "expiry_date": "2026-05-01",
                    "description": "Chili Sauce (Cold)",
                },
                {
                    "warehouse": "Savori Whse",
                    "unit": "box",
                    "stock_qty": 3,
                    "expiry_date": "2026-05-01",
                    "description": "Chili Sauce (Cold)",
                },
                {
                    "warehouse": "Lai Hock Whse",
                    "unit": "tin",
                    "stock_qty": 2,
                    "expiry_date": "2026-05-01",
                    "description": "Chili Sauce (Cold)",
                },
            ]
        )

        result = module.split_by_expiry(df)

        self.assertEqual(len(result), 1)
        self.assertEqual(
            result.loc[0, "subtotal_qty_breakdown"],
            {"btl": 4.0, "box": 3.0, "tin": 2.0},
        )
        self.assertEqual(result.loc[0, "subtotal_pkt"], 9.0)
        self.assertEqual(result.loc[0, "Savori Whse"], "4 btl 3 box")
        self.assertEqual(result.loc[0, "Lai Hock Whse"], "2 tin")
        self.assertEqual(result.loc[0, "Subtotal"], "4 btl 3 box 2 tin")

    def test_description_selection_filters_by_base_description(self):
        module = _load_module()
        df = pd.DataFrame(
            {
                "description": [
                    "Aioli Sauce (Cold)",
                    "Bobo Sliced Fish Cake",
                    "Brown Sauce",
                ]
            }
        )

        filtered = module._apply_filter_state(df, {"f_desc": ["Brown Sauce"]})

        self.assertEqual(filtered["description"].tolist(), ["Brown Sauce"])

    def test_product_code_selection_filters_by_exact_choice(self):
        module = _load_module()
        df = pd.DataFrame({"product_code": ["AB-001", "AC-002", "ZX-100"]})

        filtered = module._apply_filter_state(df, {"f_code": ["AC-002"]})

        self.assertEqual(filtered["product_code"].tolist(), ["AC-002"])

    def test_remark_selection_can_exclude_matching_rows(self):
        module = _load_module()
        df = pd.DataFrame(
            {
                "description": [
                    "Aioli Sauce (Cold)",
                    "Brown Sauce (Hot)",
                    "Fish Cake",
                ]
            }
        )

        filtered = module._apply_filter_state(
            df,
            {"f_remark": ["Hot"], "f_remark_ex": True},
        )

        self.assertEqual(
            filtered["description"].tolist(), ["Aioli Sauce (Cold)", "Fish Cake"]
        )

    def test_brand_selection_can_exclude_matching_rows(self):
        module = _load_module()
        df = pd.DataFrame(
            {
                "brand": ["Alpha", "Beta", "Gamma"],
                "product_code": ["A-1", "B-1", "G-1"],
            }
        )

        filtered = module._apply_filter_state(
            df,
            {"f_brand": ["Beta"], "f_brand_ex": True},
        )

        self.assertEqual(filtered["brand"].tolist(), ["Alpha", "Gamma"])

    def test_normalize_sales_pack_group_key_treats_1l_and_1kg_as_equivalent(self):
        module = _load_module()

        self.assertEqual(
            module._normalize_sales_pack_group_key("10 x 1l"),
            module._normalize_sales_pack_group_key("10 x 1kg"),
        )

    def test_normalize_sales_pack_group_key_keeps_different_gram_sizes_separate(self):
        module = _load_module()

        self.assertNotEqual(
            module._normalize_sales_pack_group_key("6 x 850g"),
            module._normalize_sales_pack_group_key("6 x 900g"),
        )

    def test_build_product_monthly_summary_merges_equivalent_pack_sizes(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {
                    "Date": "2026-01-05",
                    "Supplier": "Supplier Alpha",
                    "Product Description": "Hickory BBQ Sauce",
                    "Product Code": "HBQ-01",
                    "Carton Packing": "10 x 1l",
                    "Qty in Ctns": 1,
                    "Qty in Pcs": 0,
                    "Total Value": 100,
                    "carton_packing_numeric": 10,
                },
                {
                    "Date": "2026-01-20",
                    "Supplier": "Supplier Alpha",
                    "Product Description": "Hickory BBQ Sauce",
                    "Product Code": "HBQ-01",
                    "Carton Packing": "10 x 1kg",
                    "Qty in Ctns": 0,
                    "Qty in Pcs": 5,
                    "Total Value": 50,
                    "carton_packing_numeric": 10,
                },
            ]
        )

        result = module.build_product_monthly_summary(df)

        self.assertEqual(len(result), 1)
        self.assertEqual(result.loc[0, "年月"], "2026-Jan")
        self.assertEqual(result.loc[0, "总销量"], "1 ctn 5 pkts")

    def test_build_sales_usage_views_formats_month_as_year_mon(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {
                    "Year": "2026",
                    "Month": "January",
                    "Customer": "Customer One",
                    "Qty in Ctns": 1,
                    "Qty in Pcs": 2,
                    "Total Value": 30,
                    "carton_packing_numeric": 10,
                    "total_ctn_equivalent": 1.2,
                }
            ]
        )

        monthly_usage, _customer_usage, customer_month_matrix = (
            module.build_sales_usage_views(df)
        )

        self.assertEqual(monthly_usage.loc[0, "Month"], "2026-Jan")
        self.assertIn("2026-Jan", customer_month_matrix.columns.tolist())

    def test_build_daily_sales_timeline_aggregates_by_date_not_customer(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {
                    "Date": "2026-01-05 09:00:00",
                    "Customer": "Customer One",
                    "Qty in Ctns": 1,
                    "Qty in Pcs": 2,
                    "Total Value": 30,
                    "carton_packing_numeric": 10,
                },
                {
                    "Date": "2026-01-05 15:30:00",
                    "Customer": "Customer Two",
                    "Qty in Ctns": 0,
                    "Qty in Pcs": 3,
                    "Total Value": 15,
                    "carton_packing_numeric": 10,
                },
                {
                    "Date": "2026-01-06",
                    "Customer": "Customer One",
                    "Qty in Ctns": 0,
                    "Qty in Pcs": 4,
                    "Total Value": 20,
                    "carton_packing_numeric": 10,
                },
            ]
        )

        result = module.build_daily_sales_timeline(df)

        self.assertEqual(result["年月"].tolist(), ["2026-Jan", "2026-Jan"])
        self.assertEqual(result["日期"].tolist(), ["05-Jan-2026", "06-Jan-2026"])
        self.assertEqual(result["总销量"].tolist(), ["1 ctn 5 pkts", "4 pkts"])
        self.assertEqual(result["总销售额"].tolist(), ["45.00", "20.00"])

    def test_build_daily_sales_timeline_uses_packet_total_for_mixed_pack_sizes(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {
                    "Date": "2026-01-05",
                    "Customer": "Customer One",
                    "Qty in Ctns": 1,
                    "Qty in Pcs": 0,
                    "Total Value": 100,
                    "carton_packing_numeric": 10,
                },
                {
                    "Date": "2026-01-05",
                    "Customer": "Customer Two",
                    "Qty in Ctns": 1,
                    "Qty in Pcs": 0,
                    "Total Value": 80,
                    "carton_packing_numeric": 12,
                },
            ]
        )

        result = module.build_daily_sales_timeline(df)

        self.assertEqual(result.loc[0, "日期"], "05-Jan-2026")
        self.assertEqual(result.loc[0, "总销量"], "22 pkts")

    def test_load_stock_reconciliation_data_extracts_dated_outbound_quantities(self):
        module = _load_module()
        wb = Workbook()
        ws = wb.active
        if ws is None:
            self.fail("Expected active worksheet")
        ws.title = "Stocks report"
        ws.append([None] * 12)
        ws.append([None] * 10 + ["1-Apr", "2-Apr"])
        ws.append(
            [
                "Supplier",
                "Brand",
                "Product Code",
                "Description",
                "Pack Size",
                "Unit",
                "Expiry Date",
                "Relabel To Date",
                "Daily Update",
                "Stock Qty",
                "",
                "",
            ]
        )
        ws.append(
            [
                "Acme",
                "Sauces",
                "P001",
                "Aioli Sauce (Cold)",
                "24 x 50g",
                "ctn",
                "2026-04-01",
                "2026-03-20",
                "",
                10,
                "13 ctns 5 pkts",
                "2 pkts",
            ]
        )

        payload = io.BytesIO()
        wb.save(payload)

        stock_df, timeline_df, warns = module.load_stock_reconciliation_data_from_bytes(
            payload.getvalue(),
            preferred_year=2026,
        )

        self.assertFalse(warns)
        self.assertEqual(stock_df.loc[0, "product_code"], "P001")
        self.assertEqual(
            timeline_df["date"].dt.strftime("%Y-%m-%d").tolist(),
            ["2026-04-01", "2026-04-02"],
        )
        self.assertEqual(timeline_df["stock_ctn"].tolist(), [13, 0])
        self.assertEqual(timeline_df["stock_pkt"].tolist(), [5, 2])
        self.assertEqual(
            timeline_df["stock_qty_text"].tolist(),
            ["13 ctns 5 pkts", "2 pkts"],
        )

    def test_load_stock_reconciliation_data_uses_fixed_positions_not_header_labels(
        self,
    ):
        module = _load_module()
        wb = Workbook()
        ws = wb.active
        if ws is None:
            self.fail("Expected active worksheet")
        ws.title = "Stocks report"
        ws.append([None] * 12)
        ws.append([None] * 10 + ["1-Apr"])
        ws.append([None] * 12)
        ws.append([None] * 12)
        ws.append(
            [
                "Acme",
                "Sauces",
                "P001",
                "Aioli Sauce",
                "24 x 50g",
                "ctn",
                "2026-04-01",
                "2026-03-20",
                "",
                10,
                "13 ctns 5 pkts",
                "",
            ]
        )

        payload = io.BytesIO()
        wb.save(payload)

        stock_df, timeline_df, warns = module.load_stock_reconciliation_data_from_bytes(
            payload.getvalue(),
            preferred_year=2026,
        )

        self.assertFalse(warns)
        self.assertEqual(stock_df.loc[0, "product_code"], "P001")
        self.assertEqual(timeline_df.loc[0, "stock_qty_text"], "13 ctns 5 pkts")

    def test_load_stock_reconciliation_data_detects_date_headers_on_excel_row_3(self):
        module = _load_module()
        wb = Workbook()
        ws = wb.active
        if ws is None:
            self.fail("Expected active worksheet")
        ws.title = "Stocks report"
        ws.append([None] * 12)
        ws.append([None] * 12)
        ws.append([None] * 10 + ["1-Apr", None])
        ws.append([None] * 12)
        ws.append(
            [
                "Acme",
                "Sauces",
                "P001",
                "Aioli Sauce",
                "24 x 50g",
                "ctn",
                "2026-04-01",
                "2026-03-20",
                "",
                10,
                "13 ctns 5 pkts",
                "",
            ]
        )

        payload = io.BytesIO()
        wb.save(payload)

        _stock_df, timeline_df, warns = (
            module.load_stock_reconciliation_data_from_bytes(
                payload.getvalue(),
                preferred_year=2026,
            )
        )

        self.assertFalse(warns)
        self.assertEqual(len(timeline_df), 1)
        self.assertEqual(timeline_df.loc[0, "stock_qty_text"], "13 ctns 5 pkts")

    def test_load_stock_reconciliation_data_treats_numeric_cell_as_row_unit_quantity(
        self,
    ):
        module = _load_module()
        wb = Workbook()
        ws = wb.active
        if ws is None:
            self.fail("Expected active worksheet")
        ws.title = "Stocks report"
        ws.append([None] * 12)
        ws.append([None] * 12)
        ws.append([None] * 10 + ["1-Apr-2026"])
        ws.append(
            [
                "Acme",
                "Cheese",
                "CH-001",
                "Cheddar Cheese",
                "1 x 1kg",
                "ctn",
                "2026-04-30",
                None,
                None,
                20,
                5,
                None,
            ]
        )

        payload = io.BytesIO()
        wb.save(payload)

        _stock_df, timeline_df, warns = (
            module.load_stock_reconciliation_data_from_bytes(
                payload.getvalue(),
                preferred_year=2026,
            )
        )

        self.assertFalse(warns)
        self.assertEqual(len(timeline_df), 1)
        self.assertEqual(timeline_df.loc[0, "stock_ctn"], 5)
        self.assertEqual(timeline_df.loc[0, "stock_pkt"], 0)
        self.assertEqual(timeline_df.loc[0, "stock_qty_text"], "5 ctns")

    def test_load_stock_reconciliation_data_uses_second_duplicate_date_as_outbound(
        self,
    ):
        module = _load_module()
        wb = Workbook()
        ws = wb.active
        if ws is None:
            self.fail("Expected active worksheet")
        ws.title = "Stocks report"
        ws.append([None] * 13)
        ws.append([None] * 10 + ["1-Apr-2026", "1-Apr-2026", None])
        ws.append([None] * 13)
        ws.append(
            [
                "Acme",
                "Cheese",
                "CH-001",
                "Cheddar Cheese",
                "1 x 1kg",
                "ctn",
                "2026-04-30",
                None,
                None,
                20,
                99,
                5,
                None,
            ]
        )

        payload = io.BytesIO()
        wb.save(payload)

        _stock_df, timeline_df, warns = (
            module.load_stock_reconciliation_data_from_bytes(
                payload.getvalue(),
                preferred_year=2026,
            )
        )

        self.assertFalse(warns)
        self.assertEqual(len(timeline_df), 1)
        self.assertEqual(timeline_df.loc[0, "stock_ctn"], 5)
        self.assertEqual(timeline_df.loc[0, "stock_qty_text"], "5 ctns")

    def test_build_sales_reconciliation_timeline_aggregates_by_item_and_date(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {
                    "Date": "2026-04-01 10:00:00",
                    "Product Code": "P001",
                    "Product Description": "Aioli Sauce",
                    "Carton Packing": "24 x 50g",
                    "Qty in Ctns": 1,
                    "Qty in Pcs": 3,
                },
                {
                    "Date": "2026-04-01 15:00:00",
                    "Product Code": "P001",
                    "Product Description": "Aioli Sauce",
                    "Carton Packing": "24 x 50g",
                    "Qty in Ctns": 2,
                    "Qty in Pcs": 1,
                },
            ]
        )

        result = module.build_sales_reconciliation_timeline(df)

        self.assertEqual(len(result), 1)
        self.assertEqual(result.loc[0, "product_code"], "P001")
        self.assertEqual(result.loc[0, "sales_ctn"], 3)
        self.assertEqual(result.loc[0, "sales_pkt"], 4)
        self.assertEqual(result.loc[0, "sales_qty_text"], "3 ctns 4 pkts")

    def test_build_reconciliation_result_flags_matches_and_mismatches(self):
        module = _load_module()
        stock_timeline = pd.DataFrame(
            [
                {
                    "match_key": "code::P001",
                    "match_basis": "product_code",
                    "product_code": "P001",
                    "description": "Aioli Sauce",
                    "pack_size": "24 x 50g",
                    "date": pd.Timestamp("2026-04-01"),
                    "stock_ctn": 1,
                    "stock_pkt": 2,
                    "stock_qty_text": "1 ctn 2 pkts",
                },
                {
                    "match_key": "code::P002",
                    "match_basis": "product_code",
                    "product_code": "P002",
                    "description": "Brown Sauce",
                    "pack_size": "12 x 100g",
                    "date": pd.Timestamp("2026-04-01"),
                    "stock_ctn": 2,
                    "stock_pkt": 0,
                    "stock_qty_text": "2 ctns",
                },
            ]
        )
        sales_timeline = pd.DataFrame(
            [
                {
                    "match_key": "code::P001",
                    "match_basis": "product_code",
                    "product_code": "P001",
                    "description": "Aioli Sauce",
                    "pack_size": "24 x 50g",
                    "date": pd.Timestamp("2026-04-01"),
                    "sales_ctn": 1,
                    "sales_pkt": 2,
                    "sales_qty_text": "1 ctn 2 pkts",
                },
                {
                    "match_key": "code::P003",
                    "match_basis": "product_code",
                    "product_code": "P003",
                    "description": "Fish Cake",
                    "pack_size": "10 x 200g",
                    "date": pd.Timestamp("2026-04-01"),
                    "sales_ctn": 1,
                    "sales_pkt": 1,
                    "sales_qty_text": "1 ctn 1 pkt",
                },
            ]
        )

        result = module.build_reconciliation_result(stock_timeline, sales_timeline)

        status_map = {
            (row.product_code, row.date.strftime("%Y-%m-%d")): row.mismatch_reason
            for row in result.itertuples()
        }
        self.assertEqual(status_map[("P001", "2026-04-01")], "match")
        self.assertEqual(status_map[("P002", "2026-04-01")], "missing_in_sales")
        self.assertEqual(status_map[("P003", "2026-04-01")], "missing_in_stock")

    def test_build_reconciliation_result_aggregates_duplicate_rows_before_merging(self):
        module = _load_module()
        stock_timeline = pd.DataFrame(
            [
                {
                    "match_key": "code::p001",
                    "match_basis": "product_code",
                    "product_code": "P001",
                    "description": "Aioli Sauce",
                    "pack_size": "24 x 50g",
                    "supplier": "Acme",
                    "brand": "Sauces",
                    "date": pd.Timestamp("2026-04-01"),
                    "stock_ctn": 1,
                    "stock_pkt": 2,
                    "stock_qty_text": "1 ctn 2 pkts",
                },
                {
                    "match_key": "code::p001",
                    "match_basis": "product_code",
                    "product_code": "P001",
                    "description": "Aioli Sauce (Cold)",
                    "pack_size": "24 x 50g",
                    "supplier": "Acme",
                    "brand": "Sauces",
                    "date": pd.Timestamp("2026-04-01"),
                    "stock_ctn": 2,
                    "stock_pkt": 3,
                    "stock_qty_text": "2 ctns 3 pkts",
                },
            ]
        )
        sales_timeline = pd.DataFrame(
            [
                {
                    "match_key": "code::p001",
                    "match_basis": "product_code",
                    "product_code": "P001",
                    "description": "Aioli Sauce",
                    "pack_size": "24 x 50g",
                    "supplier": "Acme",
                    "brand": "Sauces",
                    "date": pd.Timestamp("2026-04-01"),
                    "sales_ctn": 3,
                    "sales_pkt": 5,
                    "sales_qty_text": "3 ctns 5 pkts",
                }
            ]
        )

        result = module.build_reconciliation_result(stock_timeline, sales_timeline)

        self.assertEqual(len(result), 1)
        self.assertTrue(bool(result.loc[0, "is_match"]))
        self.assertEqual(result.loc[0, "stock_qty_text"], "3 ctns 5 pkts")

    def test_reconciliation_match_key_treats_parenthetical_remark_as_same_item(self):
        module = _load_module()

        stock_key, stock_basis = module._build_reconciliation_match_key(
            "",
            "13294632 (ER.38.17.20E) Kartoffel Instant Mashed Potato",
            "5 x 2kg",
            "Acme",
            "Potato",
        )
        sales_key, sales_basis = module._build_reconciliation_match_key(
            "",
            "13294632 Kartoffel Instant Mashed Potato",
            "5 x 2kg",
            "Acme",
            "Potato",
        )

        self.assertEqual(stock_basis, sales_basis)
        self.assertEqual(stock_key, sales_key)

    def test_reconciliation_match_key_normalizes_parenthetical_product_code(self):
        module = _load_module()

        stock_key, stock_basis = module._build_reconciliation_match_key(
            "13294632  (ER.38.17.20E)",
            "Kartoffel Instant Mashed Potato",
            "5 x 2kg",
            "Acme",
            "Potato",
        )
        sales_key, sales_basis = module._build_reconciliation_match_key(
            "13294632",
            "Kartoffel Instant Mashed Potato",
            "5 x 2kg",
            "Acme",
            "Potato",
        )

        self.assertEqual(stock_basis, sales_basis)
        self.assertEqual(stock_key, sales_key)

    def test_reconciliation_match_key_ignores_placeholder_product_code(self):
        module = _load_module()

        key, basis = module._build_reconciliation_match_key(
            "NA",
            "Highway BBQ Sauce",
            "12 x 1kg",
            "Hock Seng",
            "Savori",
        )

        self.assertEqual(basis, "fallback_supplier_description_pack")
        self.assertEqual(key, "descpack::hock seng::highway bbq sauce::12 x 1kg")

    def test_reconciliation_match_key_fallback_ignores_blank_sales_brand(self):
        module = _load_module()

        stock_key, stock_basis = module._build_reconciliation_match_key(
            "NA",
            "Highway BBQ Sauce",
            "12 x 1kg",
            "Hock Seng",
            "Savori",
        )
        sales_key, sales_basis = module._build_reconciliation_match_key(
            pd.NA,
            "Highway BBQ Sauce",
            "12 x 1kg",
            "Hock Seng",
            pd.NA,
        )

        self.assertEqual(stock_basis, sales_basis)
        self.assertEqual(stock_key, sales_key)

    def test_strip_html_preserves_missing_values_before_sales_normalization(self):
        module = _load_module()
        df = pd.DataFrame({"Product Code": [pd.NA, "NA", "<b>P001</b>"]})

        result = module._strip_html_df(df)

        self.assertTrue(pd.isna(result.loc[0, "Product Code"]))
        self.assertEqual(result.loc[1, "Product Code"], "NA")
        self.assertEqual(result.loc[2, "Product Code"], "P001")

    def test_build_reconciliation_result_handles_empty_stock_timeline(self):
        module = _load_module()
        sales_timeline = pd.DataFrame(
            [
                {
                    "match_key": "code::p001",
                    "match_basis": "product_code",
                    "product_code": "P001",
                    "description": "Aioli Sauce",
                    "pack_size": "24 x 50g",
                    "supplier": "Acme",
                    "brand": "Sauces",
                    "date": pd.Timestamp("2026-04-01"),
                    "sales_ctn": 3,
                    "sales_pkt": 5,
                    "sales_qty_text": "3 ctns 5 pkts",
                }
            ]
        )

        result = module.build_reconciliation_result(pd.DataFrame(), sales_timeline)

        self.assertEqual(len(result), 1)
        self.assertEqual(result.loc[0, "mismatch_reason"], "missing_in_stock")

    def test_build_reconciliation_result_adds_difference_quantity_text(self):
        module = _load_module()
        stock_timeline = pd.DataFrame(
            [
                {
                    "match_key": "code::p001",
                    "match_basis": "product_code",
                    "product_code": "P001",
                    "description": "Aioli Sauce",
                    "pack_size": "24 x 50g",
                    "supplier": "Acme",
                    "brand": "Sauces",
                    "date": pd.Timestamp("2026-04-01"),
                    "stock_ctn": 3,
                    "stock_pkt": 5,
                    "stock_qty_text": "3 ctns 5 pkts",
                }
            ]
        )
        sales_timeline = pd.DataFrame(
            [
                {
                    "match_key": "code::p001",
                    "match_basis": "product_code",
                    "product_code": "P001",
                    "description": "Aioli Sauce",
                    "pack_size": "24 x 50g",
                    "supplier": "Acme",
                    "brand": "Sauces",
                    "date": pd.Timestamp("2026-04-01"),
                    "sales_ctn": 1,
                    "sales_pkt": 2,
                    "sales_qty_text": "1 ctn 2 pkts",
                }
            ]
        )

        result = module.build_reconciliation_result(stock_timeline, sales_timeline)

        self.assertEqual(result.loc[0, "difference_ctn"], 2)
        self.assertEqual(result.loc[0, "difference_pkt"], 3)
        self.assertEqual(result.loc[0, "difference_qty_text"], "2 ctns 3 pkts")

    def test_normalize_ctn_pkt_by_pack_size_rolls_packet_overflow_into_cartons(self):
        module = _load_module()

        cartons, packets = module._normalize_ctn_pkt_by_pack_size(4, 24, "12 x 1kg")

        self.assertEqual((cartons, packets), (6, 0))

    def test_default_reconciliation_year_prefers_2026(self):
        module = _load_module()

        self.assertEqual(
            module._default_reconciliation_year_selection(["2025", "2026", "2027"]),
            ["2026"],
        )
        self.assertEqual(
            module._default_reconciliation_year_selection(["2024", "2025"]),
            ["2025"],
        )

    def test_default_reconciliation_month_prefers_apr(self):
        module = _load_module()

        self.assertEqual(
            module._default_reconciliation_month_selection(["Jan", "Apr", "May"]),
            ["Apr"],
        )
        self.assertEqual(
            module._default_reconciliation_month_selection(["Jan", "May"]),
            ["May"],
        )

    def test_apply_reconciliation_filters_matches_code_description_year_and_date(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {
                    "supplier": "Acme",
                    "date": pd.Timestamp("2026-04-01"),
                    "month": "Apr",
                    "product_code": "P001",
                    "description": "Aioli Sauce (Cold)",
                    "pack_size": "24 x 50g",
                    "sales_qty_text": "1 ctn",
                    "stock_qty_text": "2 ctns",
                    "difference_qty_text": "1 ctn",
                },
                {
                    "supplier": "Acme",
                    "date": pd.Timestamp("2025-04-01"),
                    "month": "Apr",
                    "product_code": "P002",
                    "description": "Brown Sauce",
                    "pack_size": "12 x 100g",
                    "sales_qty_text": "1 pkt",
                    "stock_qty_text": "1 pkt",
                    "difference_qty_text": "0",
                },
                {
                    "supplier": "Acme",
                    "date": pd.Timestamp("2026-05-01"),
                    "month": "May",
                    "product_code": "P001",
                    "description": "Aioli Sauce (Cold)",
                    "pack_size": "24 x 50g",
                    "sales_qty_text": "1 ctn",
                    "stock_qty_text": "2 ctns",
                    "difference_qty_text": "1 ctn",
                },
            ]
        )

        result = module._apply_reconciliation_filters(
            df,
            {
                "product_code": ["P001"],
                "description": ["Aioli Sauce"],
                "year": ["2026"],
                "month": ["Apr"],
                "date": ["01-Apr-2026"],
            },
        )

        self.assertEqual(len(result), 1)
        self.assertEqual(result.loc[0, "product_code"], "P001")

    def test_get_reconciliation_date_bounds_narrows_to_selected_year_and_month(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {"date": pd.Timestamp("2026-04-01")},
                {"date": pd.Timestamp("2026-04-15")},
                {"date": pd.Timestamp("2026-05-01")},
                {"date": pd.Timestamp("2025-04-01")},
            ]
        )

        start_date, end_date = module._get_reconciliation_date_bounds(
            df,
            {"year": ["2026"], "month": ["Apr"]},
        )

        self.assertEqual(start_date, datetime.date(2026, 4, 1))
        self.assertEqual(end_date, datetime.date(2026, 4, 30))

    def test_apply_reconciliation_filters_matches_calendar_date_value(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {
                    "supplier": "Acme",
                    "date": pd.Timestamp("2026-04-01"),
                    "product_code": "P001",
                    "description": "Aioli Sauce",
                    "pack_size": "24 x 50g",
                },
                {
                    "supplier": "Acme",
                    "date": pd.Timestamp("2026-04-02"),
                    "product_code": "P001",
                    "description": "Aioli Sauce",
                    "pack_size": "24 x 50g",
                },
            ]
        )

        result = module._apply_reconciliation_filters(
            df,
            {
                "product_code": [],
                "description": [],
                "year": [],
                "month": [],
                "date": datetime.date(2026, 4, 1),
            },
        )

        self.assertEqual(len(result), 1)
        self.assertEqual(result.loc[0, "date"].strftime("%Y-%m-%d"), "2026-04-01")

    def test_apply_reconciliation_filters_matches_calendar_date_range(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {"date": pd.Timestamp("2026-04-01"), "product_code": "P001"},
                {"date": pd.Timestamp("2026-04-15"), "product_code": "P002"},
                {"date": pd.Timestamp("2026-05-01"), "product_code": "P003"},
            ]
        )

        result = module._apply_reconciliation_filters(
            df,
            {
                "product_code": [],
                "description": [],
                "year": [],
                "month": [],
                "date": (datetime.date(2026, 4, 1), datetime.date(2026, 4, 30)),
            },
        )

        self.assertEqual(result["product_code"].tolist(), ["P001", "P002"])

    def test_apply_reconciliation_filters_range_and_specific_date_work_together(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {"date": pd.Timestamp("2026-04-01"), "product_code": "P001"},
                {"date": pd.Timestamp("2026-04-15"), "product_code": "P002"},
                {"date": pd.Timestamp("2026-04-20"), "product_code": "P003"},
            ]
        )

        result = module._apply_reconciliation_filters(
            df,
            {
                "product_code": [],
                "description": [],
                "year": [],
                "month": [],
                "date": (datetime.date(2026, 4, 1), datetime.date(2026, 4, 30)),
                "specific_date": datetime.date(2026, 4, 15),
                "use_specific_date": True,
            },
        )

        self.assertEqual(result["product_code"].tolist(), ["P002"])

    def test_apply_reconciliation_filters_uses_date_range_when_specific_date_disabled(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {"date": pd.Timestamp("2026-04-01"), "product_code": "P001"},
                {"date": pd.Timestamp("2026-04-15"), "product_code": "P002"},
                {"date": pd.Timestamp("2026-05-15"), "product_code": "P003"},
            ]
        )

        result = module._apply_reconciliation_filters(
            df,
            {
                "product_code": [],
                "description": [],
                "year": [],
                "month": [],
                "date": (datetime.date(2026, 4, 1), datetime.date(2026, 4, 30)),
                "specific_date": datetime.date(2026, 5, 15),
                "use_specific_date": False,
            },
        )

        self.assertEqual(result["product_code"].tolist(), ["P001", "P002"])

    def test_apply_reconciliation_filters_specific_date_overrides_other_date_filters(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {"date": pd.Timestamp("2026-04-01"), "product_code": "P001"},
                {"date": pd.Timestamp("2026-05-15"), "product_code": "P002"},
                {"date": pd.Timestamp("2025-05-15"), "product_code": "P003"},
            ]
        )

        result = module._apply_reconciliation_filters(
            df,
            {
                "product_code": [],
                "description": [],
                "year": ["2026"],
                "month": ["Apr"],
                "date": (datetime.date(2026, 4, 1), datetime.date(2026, 4, 30)),
                "specific_date": datetime.date(2026, 5, 15),
                "use_specific_date": True,
            },
        )

        self.assertEqual(result["product_code"].tolist(), ["P002"])

    def test_coerce_reconciliation_date_input_value_handles_old_list_state(self):
        module = _load_module()

        self.assertEqual(
            module._coerce_reconciliation_date_input_value(["01-Apr-2026"]),
            datetime.date(2026, 4, 1),
        )
        self.assertEqual(
            module._coerce_reconciliation_date_input_value(["2026-04-01"]),
            datetime.date(2026, 4, 1),
        )
        self.assertEqual(
            module._coerce_reconciliation_date_input_value(
                ["2026-04-01", "2026-04-30"]
            ),
            (datetime.date(2026, 4, 1), datetime.date(2026, 4, 30)),
        )
        self.assertIsNone(module._coerce_reconciliation_date_input_value([]))

    def test_build_reconciliation_result_creates_possible_match_for_similar_descriptions(
        self,
    ):
        module = _load_module()
        stock_timeline = pd.DataFrame(
            [
                {
                    "match_key": "descpack::acme::dairy::cheddar cheese::5 x 2kg",
                    "match_basis": "fallback_supplier_brand_description_pack",
                    "product_code": "",
                    "description": "cheddar cheese",
                    "pack_size": "5 x 2kg",
                    "supplier": "Acme",
                    "brand": "Dairy",
                    "date": pd.Timestamp("2026-04-01"),
                    "stock_ctn": 3,
                    "stock_pkt": 0,
                    "stock_qty_text": "3 ctns",
                }
            ]
        )
        sales_timeline = pd.DataFrame(
            [
                {
                    "match_key": "descpack::acme::dairy::cheddar cheese sauce mix::5 x 2kg",
                    "match_basis": "fallback_supplier_brand_description_pack",
                    "product_code": "",
                    "description": "cheddar cheese sauce mix",
                    "pack_size": "5 x 2kg",
                    "supplier": "Acme",
                    "brand": "Dairy",
                    "date": pd.Timestamp("2026-04-01"),
                    "sales_ctn": 2,
                    "sales_pkt": 0,
                    "sales_qty_text": "2 ctns",
                }
            ]
        )

        result = module.build_reconciliation_result(stock_timeline, sales_timeline)

        self.assertEqual(len(result), 1)
        self.assertEqual(result.loc[0, "mismatch_reason"], "possible_match")
        self.assertIn("cheddar cheese", result.loc[0, "description"].lower())


    def test_build_reconciliation_result_flags_possible_match_for_pack_size_typo(self):
        module = _load_module()
        stock_timeline = pd.DataFrame(
            [
                {
                    "match_key": "descpack::3g pte ltd::pizza sauce campagna::6 x 2.55g",
                    "match_basis": "fallback_supplier_description_pack",
                    "product_code": "NA",
                    "description": "Pizza Sauce Campagna",
                    "pack_size": "6 x 2.55g",
                    "supplier": "3G Pte Ltd",
                    "brand": "Savori",
                    "date": pd.Timestamp("2026-04-01"),
                    "stock_ctn": 54,
                    "stock_pkt": 0,
                    "stock_qty_text": "54 ctns",
                }
            ]
        )
        sales_timeline = pd.DataFrame(
            [
                {
                    "match_key": "code::001195",
                    "match_basis": "product_code",
                    "product_code": "001195",
                    "description": "Pizza Sauce / Campagna",
                    "pack_size": "6 x 2.55kg",
                    "supplier": "3G Pte Ltd",
                    "brand": "",
                    "date": pd.Timestamp("2026-04-01"),
                    "sales_ctn": 54,
                    "sales_pkt": 0,
                    "sales_qty_text": "54 ctns",
                }
            ]
        )

        result = module.build_reconciliation_result(stock_timeline, sales_timeline)

        self.assertEqual(len(result), 1)
        self.assertEqual(result.loc[0, "mismatch_reason"], "possible_match")
        self.assertEqual(result.loc[0, "difference_qty_text"], "0")

    def test_build_reconciliation_result_matches_supplier_pack_core_description_equivalent_items(self):
        module = _load_module()
        stock_timeline = pd.DataFrame(
            [
                {
                    "match_key": "descpack::yong wen food::jalapeno pepper sliced::6 x 2.9kg",
                    "match_basis": "fallback_supplier_description_pack",
                    "product_code": "NA",
                    "description": "Jalapeno Pepper Sliced",
                    "pack_size": "6 x 2.9kg",
                    "supplier": "Yong Wen Food",
                    "brand": "Savori",
                    "date": pd.Timestamp("2026-04-01"),
                    "stock_ctn": 3,
                    "stock_pkt": 0,
                    "stock_qty_text": "3 ctns",
                }
            ]
        )
        sales_timeline = pd.DataFrame(
            [
                {
                    "match_key": "descpack::yongwen::saporito jalapeno pepper sliced::6 x 2.95kg",
                    "match_basis": "fallback_supplier_description_pack",
                    "product_code": "",
                    "description": "Saporito Jalapeno Pepper Sliced",
                    "pack_size": "6 x 2.95kg",
                    "supplier": "YongWen",
                    "brand": "",
                    "date": pd.Timestamp("2026-04-01"),
                    "sales_ctn": 3,
                    "sales_pkt": 0,
                    "sales_qty_text": "3 ctns",
                }
            ]
        )

        result = module.build_reconciliation_result(stock_timeline, sales_timeline)

        self.assertEqual(len(result), 1)
        self.assertTrue(bool(result.loc[0, "is_match"]))
        self.assertEqual(result.loc[0, "mismatch_reason"], "match")
        self.assertEqual(result.loc[0, "difference_qty_text"], "0")
        self.assertIn("Jalapeno Pepper Sliced", result.loc[0, "description"])

    def test_reconciliation_same_item_path_rejects_different_supplier(self):
        module = _load_module()
        self.assertFalse(
            module._is_high_confidence_reconciliation_same_item(
                "Different Supplier",
                "Jalapeno Pepper Sliced",
                "6 x 2.9kg",
                "YongWen",
                "Saporito Jalapeno Pepper Sliced",
                "6 x 2.95kg",
            )
        )

    def test_reconciliation_same_item_path_accepts_supplier_pack_core_description_equivalent(self):
        module = _load_module()
        self.assertTrue(
            module._is_high_confidence_reconciliation_same_item(
                "Yong Wen Food",
                "Jalapeno Pepper Sliced",
                "6 x 2.9kg",
                "YongWen",
                "Saporito Jalapeno Pepper Sliced",
                "6 x 2.95kg",
            )
        )

    def test_reconciliation_same_item_path_rejects_different_core_product(self):
        module = _load_module()
        self.assertFalse(
            module._is_high_confidence_reconciliation_same_item(
                "Yong Wen Food",
                "Whole Jalapeno Pepper",
                "6 x 2.9kg",
                "YongWen",
                "Saporito Jalapeno Pepper Sliced",
                "6 x 2.95kg",
            )
        )

    def test_reconciliation_same_item_path_rejects_materially_different_pack_size(self):
        module = _load_module()
        self.assertFalse(
            module._is_high_confidence_reconciliation_same_item(
                "Yong Wen Food",
                "Jalapeno Pepper Sliced",
                "12 x 2.9kg",
                "YongWen",
                "Saporito Jalapeno Pepper Sliced",
                "6 x 2.95kg",
            )
        )

    def test_reconciliation_same_item_path_rejects_different_pack_units(self):
        module = _load_module()
        self.assertFalse(
            module._are_reconciliation_pack_sizes_close("6 x 2.55g", "6 x 2.55kg")
        )
        self.assertFalse(
            module._is_high_confidence_reconciliation_same_item(
                "3G Pte Ltd",
                "Pizza Sauce Campagna",
                "6 x 2.55g",
                "3G Pte Ltd",
                "Pizza Sauce / Campagna",
                "6 x 2.55kg",
            )
        )

    def test_reconciliation_possible_match_uses_similar_descriptions_only(self):
        module = _load_module()

        self.assertTrue(
            module._is_possible_reconciliation_description_match(
                "cheddar cheese",
                "cheddar cheese sauce mix",
            )
        )
        self.assertFalse(
            module._is_possible_reconciliation_description_match(
                "cheddar cheese",
                "tomato ketchup",
            )
        )

    def test_parse_stock_header_date_only_accepts_intended_date_formats(self):
        module = _load_module()

        self.assertEqual(
            module._parse_stock_header_date("Apr-1", 2026).strftime("%Y-%m-%d"),
            "2026-04-01",
        )
        self.assertEqual(
            module._parse_stock_header_date("1 Apr", 2026).strftime("%Y-%m-%d"),
            "2026-04-01",
        )
        self.assertIsNone(module._parse_stock_header_date("Total", 2026))
        self.assertIsNone(module._parse_stock_header_date("Notes", 2026))

    def test_render_sales_business_dashboard_sorts_customer_month_detail_chronologically(
        self,
    ):
        module = _load_module()

        class _FakeMetric:
            def metric(self, *_args, **_kwargs):
                return None

        class _FakeStreamlit:
            def __init__(self):
                self._current_subheader = None
                self.month_view = None

            def info(self, *_args, **_kwargs):
                return None

            def columns(self, spec):
                count = spec if isinstance(spec, int) else len(spec)
                return [_FakeMetric() for _ in range(count)]

            def subheader(self, label):
                self._current_subheader = label

            def dataframe(self, df, **_kwargs):
                if self._current_subheader == "客户月度采购明细":
                    self.month_view = df.copy()

        fake_st = _FakeStreamlit()
        original_st = getattr(module, "st")
        setattr(module, "st", fake_st)
        try:
            df = pd.DataFrame(
                [
                    {
                        "Date": "2026-03-05",
                        "Customer": "Customer One",
                        "Account": "Account A",
                        "Supplier": "Supplier Alpha",
                        "Product Description": "Hickory BBQ Sauce",
                        "Product Code": "HBQ-01",
                        "Carton Packing": "10 x 1kg",
                        "Qty in Ctns": 1,
                        "Qty in Pcs": 0,
                        "Total Value": 100,
                        "carton_packing_numeric": 10,
                        "total_packets": 10,
                    },
                    {
                        "Date": "2026-01-05",
                        "Customer": "Customer One",
                        "Account": "Account A",
                        "Supplier": "Supplier Alpha",
                        "Product Description": "Hickory BBQ Sauce",
                        "Product Code": "HBQ-01",
                        "Carton Packing": "10 x 1kg",
                        "Qty in Ctns": 1,
                        "Qty in Pcs": 0,
                        "Total Value": 100,
                        "carton_packing_numeric": 10,
                        "total_packets": 10,
                    },
                    {
                        "Date": "2026-02-05",
                        "Customer": "Customer One",
                        "Account": "Account A",
                        "Supplier": "Supplier Alpha",
                        "Product Description": "Hickory BBQ Sauce",
                        "Product Code": "HBQ-01",
                        "Carton Packing": "10 x 1kg",
                        "Qty in Ctns": 1,
                        "Qty in Pcs": 0,
                        "Total Value": 100,
                        "carton_packing_numeric": 10,
                        "total_packets": 10,
                    },
                ]
            )

            module._render_sales_business_dashboard(df)
        finally:
            setattr(module, "st", original_st)

        month_view = fake_st.month_view
        if month_view is None:
            self.fail("Expected 客户月度采购明细 dataframe to be rendered")
        self.assertEqual(
            month_view["年月"].tolist(),
            ["2026-Jan", "2026-Feb", "2026-Mar"],
        )

    def test_render_sales_business_dashboard_shows_non_customer_daily_timeline(self):
        module = _load_module()

        class _FakeMetric:
            def metric(self, *_args, **_kwargs):
                return None

        class _FakeStreamlit:
            def __init__(self):
                self._current_subheader = None
                self.daily_timeline_view = None

            def info(self, *_args, **_kwargs):
                return None

            def columns(self, spec):
                count = spec if isinstance(spec, int) else len(spec)
                return [_FakeMetric() for _ in range(count)]

            def subheader(self, label):
                self._current_subheader = label

            def dataframe(self, df, **_kwargs):
                if self._current_subheader == "每日销量时间线（不分客户）":
                    self.daily_timeline_view = df.copy()

        fake_st = _FakeStreamlit()
        original_st = getattr(module, "st")
        setattr(module, "st", fake_st)
        try:
            df = pd.DataFrame(
                [
                    {
                        "Date": "2026-01-05",
                        "Customer": "Customer One",
                        "Account": "Account A",
                        "Supplier": "Supplier Alpha",
                        "Product Description": "Hickory BBQ Sauce",
                        "Product Code": "HBQ-01",
                        "Carton Packing": "10 x 1kg",
                        "Qty in Ctns": 1,
                        "Qty in Pcs": 0,
                        "Total Value": 100,
                        "carton_packing_numeric": 10,
                        "total_packets": 10,
                    },
                    {
                        "Date": "2026-01-05",
                        "Customer": "Customer Two",
                        "Account": "Account B",
                        "Supplier": "Supplier Alpha",
                        "Product Description": "Hickory BBQ Sauce",
                        "Product Code": "HBQ-01",
                        "Carton Packing": "10 x 1kg",
                        "Qty in Ctns": 0,
                        "Qty in Pcs": 5,
                        "Total Value": 50,
                        "carton_packing_numeric": 10,
                        "total_packets": 5,
                    },
                ]
            )

            module._render_sales_business_dashboard(df)
        finally:
            setattr(module, "st", original_st)

        timeline_view = fake_st.daily_timeline_view
        if timeline_view is None:
            self.fail("Expected 每日销量时间线（不分客户） dataframe to be rendered")

        self.assertEqual(timeline_view["日期"].tolist(), ["05-Jan-2026"])
        self.assertEqual(timeline_view["总销量"].tolist(), ["1 ctn 5 pkts"])

    def test_render_sales_business_dashboard_shows_base_price_and_price_per_kg(self):
        module = _load_module()

        class _FakeMetric:
            def metric(self, *_args, **_kwargs):
                return None

        class _FakeStreamlit:
            def __init__(self):
                self._current_subheader = None
                self.month_view = None
                self.date_view = None

            def info(self, *_args, **_kwargs):
                return None

            def columns(self, spec):
                count = spec if isinstance(spec, int) else len(spec)
                return [_FakeMetric() for _ in range(count)]

            def subheader(self, label):
                self._current_subheader = label

            def dataframe(self, df, **_kwargs):
                if self._current_subheader == "客户月度采购明细":
                    self.month_view = df.copy()
                if self._current_subheader == "客户采购时间线":
                    self.date_view = df.copy()

        fake_st = _FakeStreamlit()
        original_st = getattr(module, "st")
        setattr(module, "st", fake_st)
        try:
            df = pd.DataFrame(
                [
                    {
                        "Date": "2026-01-05",
                        "Customer": "Customer One",
                        "Account": "Account A",
                        "Supplier": "Supplier Alpha",
                        "Product Description": "Hickory BBQ Sauce",
                        "Product Code": "HBQ-01",
                        "Carton Packing": "10 x 1kg",
                        "Qty in Ctns": 1,
                        "Qty in Pcs": 0,
                        "Total Value": 100,
                        "carton_packing_numeric": 10,
                        "total_packets": 10,
                    }
                ]
            )

            module._render_sales_business_dashboard(df)
        finally:
            setattr(module, "st", original_st)

        month_view = fake_st.month_view
        date_view = fake_st.date_view
        if month_view is None:
            self.fail("Expected 客户月度采购明细 dataframe to be rendered")
        if date_view is None:
            self.fail("Expected 客户采购时间线 dataframe to be rendered")

        self.assertIn("基础卖价", month_view.columns)
        self.assertIn("每公斤卖价", month_view.columns)
        self.assertNotIn("总销售额", month_view.columns)
        self.assertEqual(month_view.loc[0, "基础卖价"], "100.00")
        self.assertEqual(month_view.loc[0, "每公斤卖价"], "10.00")

        self.assertIn("基础卖价", date_view.columns)
        self.assertIn("每公斤卖价", date_view.columns)
        self.assertNotIn("销售额", date_view.columns)
        self.assertEqual(date_view.loc[0, "基础卖价"], "100.00")
        self.assertEqual(date_view.loc[0, "每公斤卖价"], "10.00")

    def test_stock_module_does_not_expose_sales_account_conflict_checker(self):
        module = _load_module()

        self.assertFalse(
            hasattr(module, "_find_sales_month_customer_account_conflicts")
        )

    def test_build_account_price_list_merges_equivalent_pack_sizes(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {
                    "Account": "Account A",
                    "Date": "2026-01-05",
                    "Customer": "Customer One",
                    "Supplier": "Supplier Alpha",
                    "Product Code": "HBQ-01",
                    "Product Description": "Hickory BBQ Sauce",
                    "Carton Packing": "10 x 1l",
                    "Qty in Pcs": 0,
                    "Qty in Ctns": 1,
                    "Total Value": 100,
                },
                {
                    "Account": "Account A",
                    "Date": "2026-01-20",
                    "Customer": "Customer One",
                    "Supplier": "Supplier Alpha",
                    "Product Code": "HBQ-01",
                    "Product Description": "Hickory BBQ Sauce",
                    "Carton Packing": "10 x 1kg",
                    "Qty in Pcs": 5,
                    "Qty in Ctns": 0,
                    "Total Value": 50,
                },
            ]
        )

        result = module.build_account_price_list(df, "Account A")

        self.assertEqual(len(result), 1)
        self.assertEqual(result.loc[0, "Last Selling Date"], "2026-01-20")

    def test_cache_stock_upload_persists_name_bytes_and_browse_time(self):
        module = _load_module()

        class _UploadedFile:
            name = "stock.xlsx"

            @staticmethod
            def getvalue():
                return b"excel-bytes"

        session_state = {}
        browse_time = datetime.datetime(2026, 4, 13, 9, 5)

        module._cache_stock_upload(
            _UploadedFile(),
            session_state,
            now=browse_time,
        )

        self.assertEqual(session_state["stock_file_name"], "stock.xlsx")
        self.assertEqual(session_state["stock_file_bytes"], b"excel-bytes")
        self.assertEqual(session_state["stock_file_browsed_at"], "09:05 AM 13/04/2026")

    def test_cache_stock_upload_ignores_missing_upload(self):
        module = _load_module()
        session_state = {"stock_file_name": "existing.xlsx"}

        module._cache_stock_upload(None, session_state)

        self.assertEqual(session_state, {"stock_file_name": "existing.xlsx"})

    def test_clear_stock_upload_cache_removes_cached_upload_and_browse_time(self):
        module = _load_module()
        session_state = {
            "stock_file_name": "existing.xlsx",
            "stock_file_bytes": b"abc",
            "stock_file_browsed_at": "09:05 AM 13/04/2026",
        }

        module._clear_stock_upload_cache(session_state)

        self.assertEqual(session_state, {})

    def test_stock_persist_keys_include_browse_timestamp(self):
        module = _load_module()

        self.assertIn("stock_file_browsed_at", module._STOCK_PERSIST_KEYS)

    def test_stock_issue_focus_index_accepts_stale_display_label(self):
        module = _load_module()

        focus_index = module._coerce_stock_issue_focus_index(
            "1. Near-Expiry | TDF Food | Chicken Salt | 20 x 1kg | -",
            max_index=4,
        )

        self.assertEqual(focus_index, 0)

    def test_cache_sales_upload_persists_payload_and_browse_time(self):
        module = _load_module()

        class _UploadedFile:
            def __init__(self, name, content):
                self.name = name
                self._content = content

            def getvalue(self):
                return self._content

        session_state = {}
        browse_time = datetime.datetime(2026, 4, 13, 10, 45)

        module._cache_sales_upload(
            [
                _UploadedFile("sales-1.xlsx", b"one"),
                _UploadedFile("sales-2.xlsx", b"two"),
            ],
            session_state,
            now=browse_time,
        )

        self.assertEqual(
            session_state["sales_files_payload"],
            [
                {"name": "sales-1.xlsx", "bytes": b"one"},
                {"name": "sales-2.xlsx", "bytes": b"two"},
            ],
        )
        self.assertEqual(session_state["sales_files_browsed_at"], "10:45 AM 13/04/2026")

    def test_clear_sales_upload_cache_removes_cached_files_and_browse_time(self):
        module = _load_module()
        session_state = {
            "sales_files_payload": [{"name": "sales.xlsx", "bytes": b"abc"}],
            "sales_files_browsed_at": "10:45 AM 13/04/2026",
        }

        module._clear_sales_upload_cache(session_state)

        self.assertEqual(session_state, {})

    def test_sales_persist_keys_include_browse_timestamp(self):
        module = _load_module()

        self.assertIn("sales_files_browsed_at", module._SALES_PERSIST_KEYS)

    def test_clear_sales_filters_state_resets_exact_date_keys(self):
        module = _load_module()
        session_state = {
            "sales_filter_use_exact_date": True,
            "sales_filter_exact_date": datetime.date(2026, 4, 15),
        }

        module._clear_sales_filters_state(session_state)

        self.assertFalse(session_state["sales_filter_use_exact_date"])
        self.assertIsNone(session_state["sales_filter_exact_date"])

    def test_sales_persist_keys_include_exact_date_keys(self):
        module = _load_module()

        self.assertIn("sales_filter_use_exact_date", module._SALES_PERSIST_KEYS)
        self.assertIn("sales_filter_exact_date", module._SALES_PERSIST_KEYS)

    def test_apply_sales_filters_uses_exact_date_in_basic_mode(self):
        module = _load_module()
        original_st = getattr(module, "st")
        session_state = {
            "sales_basic_mode": True,
            "sales_filter_year": [],
            "sales_filter_month": [],
            "sales_filter_customer": [],
            "sales_filter_customer_exclude": [],
            "sales_filter_outlet": [],
            "sales_filter_product_description": [],
            "sales_filter_supplier": [],
            "sales_filter_brand": [],
            "sales_filter_product_code": [],
            "sales_filter_account": [],
            "sales_filter_invoice": "",
            "sales_filter_date_from": datetime.date(2026, 4, 1),
            "sales_filter_date_to": datetime.date(2026, 4, 30),
            "sales_filter_use_exact_date": True,
            "sales_filter_exact_date": datetime.date(2026, 4, 15),
        }
        setattr(module, "st", _FakeStreamlit(session_state))
        df = pd.DataFrame(
            {
                "Year": ["2026", "2026", "2026"],
                "Month": ["Apr", "Apr", "Apr"],
                "Customer": ["A", "A", "B"],
                "Outlet": ["Outlet 1", "Outlet 1", "Outlet 2"],
                "Product Description": ["Prod 1", "Prod 1", "Prod 2"],
                "Supplier": ["Supp", "Supp", "Supp"],
                "Brand/Category": ["Brand", "Brand", "Brand"],
                "Product Code": ["P1", "P1", "P2"],
                "Account": ["ACC", "ACC", "ACC2"],
                "Invoice #": ["INV-1", "INV-2", "INV-3"],
                "Date": pd.to_datetime(["2026-04-15", "2026-04-16", "2026-04-15"]),
            }
        )

        try:
            filtered_df, filter_state = module.apply_sales_filters(df)
        finally:
            setattr(module, "st", original_st)

        self.assertEqual(filtered_df["Invoice #"].tolist(), ["INV-1", "INV-3"])
        self.assertEqual(filter_state["Exact date"], "2026-04-15")
        self.assertIsNone(session_state["sales_filter_date_from"])
        self.assertIsNone(session_state["sales_filter_date_to"])

    def test_append_forecast_total_row_adds_combined_total(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {
                    "客户": "Customer One",
                    "账户": "Account A",
                    "供应商": "Supplier Alpha",
                    "产品描述": "Hickory BBQ Sauce",
                    "产品编码": "HBQ-01",
                    "箱规": "10 x 1kg",
                    "下月预测销量": "1 ctn 5 pkts",
                    "下月预测销售额": "150.00",
                    "_forecast_packets": 15.0,
                    "_forecast_value": 150.0,
                },
                {
                    "客户": "Customer Two",
                    "账户": "Account B",
                    "供应商": "Supplier Beta",
                    "产品描述": "Hot Sauce",
                    "产品编码": "HS-01",
                    "箱规": "5 x 500g",
                    "下月预测销量": "2 ctns",
                    "下月预测销售额": "80.00",
                    "_forecast_packets": 10.0,
                    "_forecast_value": 80.0,
                },
            ]
        )

        result = module._append_sales_forecast_total_row(df)

        self.assertEqual(len(result), 3)
        self.assertEqual(result.iloc[-1]["客户"], "Total")
        self.assertEqual(result.iloc[-1]["下月预测销量"], "25 pkts")
        self.assertEqual(result.iloc[-1]["下月预测销售额"], "230.00")

    def test_append_forecast_total_row_uses_ctns_for_consistent_pack_size(self):
        module = _load_module()
        df = pd.DataFrame(
            [
                {
                    "客户": "Customer One",
                    "账户": "Account A",
                    "供应商": "Supplier Alpha",
                    "产品描述": "Mozzarella",
                    "产品编码": "FROZEN",
                    "箱规": "1 x 6.81kg",
                    "下月预测销量": "15 ctns",
                    "下月预测销售额": "870.00",
                    "_forecast_packets": 15.0,
                    "_forecast_value": 870.0,
                }
            ]
        )

        result = module._append_sales_forecast_total_row(df)

        self.assertEqual(result.iloc[-1]["客户"], "Total")
        self.assertEqual(result.iloc[-1]["下月预测销量"], "15 ctns")
        self.assertEqual(result.iloc[-1]["下月预测销售额"], "870.00")


if __name__ == "__main__":
    unittest.main()
