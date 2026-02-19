import os
import re
import threading
import traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
from typing import List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Side

# ===== 列号常量 =====
COL_F = 6
COL_G = 7
COL_H = 8
COL_I = 9
COL_J = 10
COL_K = 11

# ========= 扫描：忽略 history 文件夹 =========
def iter_excel_files_excluding_history(directory: str, filename_substring: str) -> List[str]:
    matches = []
    sub = (filename_substring or "").lower()
    for root, dirs, files in os.walk(directory):
        dirs[:] = [d for d in dirs if "history" not in d.lower()]
        for file in files:
            low = file.lower()
            if low.endswith((".xlsx", ".xlsm")) and sub in low:
                matches.append(os.path.join(root, file))
    return matches

# ========= 工具：列字母 -> 索引 =========
def col_letters_to_index(col_letter: str) -> Optional[int]:
    col_letter = (col_letter or "").strip().upper()
    if not col_letter:
        return None
    idx_val = 0
    for ch in col_letter:
        if "A" <= ch <= "Z":
            idx_val = idx_val * 26 + (ord(ch) - ord("A") + 1)
        else:
            return None
    return idx_val or None

# ========= 删除区间 & 归并 =========
def normalize_delete_ranges(ranges: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
    if not ranges:
        return []
    ranges.sort()
    merged = [list(ranges[0])]
    for s, e in ranges[1:]:
        ls, le = merged[-1]
        if s <= le + 1:
            merged[-1][1] = max(le, e)
        else:
            merged.append([s, e])
    return [(s, e) for s, e in merged]

# ========= 将“G列引用”改为当前行号（只处理本表引用）=========
LOCAL_G_REF_RE = re.compile(r'(?<![A-Za-z0-9_])(\$?G)(\$?)(\d+)\b')

def rewrite_formula_G_refs_to_row(formula: str, row_index: int) -> str:
    """
    把公式中所有“本表 G 列”的行号重写为当前行号 row_index；跨表引用（含 !）不改。
    例：在第 38 行，'=3.8*G31 + SUM(G10:G40)' -> '=3.8*G38 + SUM(G38:G38)'
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    if '!' in formula:
        return formula  # 跳过跨表，避免 repair
    def repl(m: re.Match) -> str:
        col_abs = m.group(1)   # G 或 $G
        row_abs = m.group(2)   # '' 或 $
        return f"{col_abs}{row_abs}{row_index}"
    return LOCAL_G_REF_RE.sub(repl, formula)

# ========= 规范化文本（去空格、大小写不敏感）=========
def norm_text(val) -> str:
    if val is None:
        return ""
    s = str(val).strip().lower()
    s = " ".join(s.split())
    return s

# ========= 合并 H:K（仅 DO）+ 画横线 + 调行高 =========
TARGET_TEXTS = {
    "received in good order",
    "authorised signature & stamp",
}

def merge_H_to_K_on_rows_with_labels_DO(sheet):
    """
    在 DO 表里：
    1) 找到包含 TARGET_TEXTS 的行，在该行合并 H:K；
    2) 若两行文案都存在，则先清除它们之间（含 Received 下一行、到 Authorised 上一行）的所有边框；
       再在“Authorised Signature & Stamp”那一行的 H:K 画一条上边框；
       然后把这条线的上一行行高设为 42.75，但要确保该上一行不是“Received In Good Order”，若是则再上移一行。
    3) 若那段文字不在 H 列，会复制到 H 列（合并后的起始格）并居中。
    """
    max_row = sheet.max_row or 1
    max_col = sheet.max_column or 1

    # 收集两个目标行
    r_received = None
    r_authorised = None
    target_rows = set()

    for r in range(1, max_row + 1):
        found_norm = None
        for c in range(1, max_col + 1):
            val = sheet.cell(row=r, column=c).value
            nv = norm_text(val)
            if nv in TARGET_TEXTS:
                found_norm = nv
                break
        if found_norm:
            target_rows.add(r)
            if found_norm == "received in good order":
                r_received = r
            elif found_norm == "authorised signature & stamp":
                r_authorised = r

    if not target_rows:
        return

    # 清除与目标行 H:K 相交的旧合并
    to_remove = []
    for mr in list(sheet.merged_cells.ranges):
        for r in target_rows:
            if mr.min_row <= r <= mr.max_row and not (mr.max_col < COL_H or mr.min_col > COL_K):
                to_remove.append(mr)
                break
    for mr in to_remove:
        sheet.merged_cells.ranges.remove(mr)

    # 合并 H:K，复制文字到 H 并居中
    for r in sorted(target_rows):
        sheet.merge_cells(start_row=r, start_column=COL_H, end_row=r, end_column=COL_K)
        hv = sheet.cell(row=r, column=COL_H).value
        if hv is None or str(hv).strip() == "":
            moved = None
            for c in range(1, max_col + 1):
                v = sheet.cell(row=r, column=c).value
                if norm_text(v) in TARGET_TEXTS:
                    moved = v
                    break
            if moved is not None:
                sheet.cell(row=r, column=COL_H).value = moved
        sheet.cell(row=r, column=COL_H).alignment = Alignment(horizontal="center", vertical="center")

    # 若两行都存在，清除中间所有边框
    if r_received and r_authorised and r_authorised > r_received + 1:
        for r in range(r_received + 1, r_authorised):
            for c in range(1, max_col + 1):
                sheet.cell(row=r, column=c).border = Border()  # 清空边框

    # 画“横线”+ 调行高
    if r_authorised:
        thin = Side(style="thin")
        for c in range(COL_H, COL_K + 1):
            cell = sheet.cell(row=r_authorised, column=c)
            cell.border = Border(top=thin)  # 上边框为横线，其他清空

        # 把“横线上一行”的行高设为 42.75，若上一行是 Received 再上移一行
        target_height_row = r_authorised - 1
        if target_height_row >= 1:
            h_text = norm_text(sheet.cell(row=target_height_row, column=COL_H).value)
            if h_text == "received in good order":
                target_height_row -= 1
        if target_height_row >= 1:
            sheet.row_dimensions[target_height_row].height = 42.75


# ========= 主逻辑：删除（多关键字；命中行+下一行） + I列特别规则 + 行内G修正 + DO合并H:K =========
def delete_rows_in_excel_files(
    directory: str,
    debug: bool,
    filename_substring: str,
    find_keywords: List[str],
    process_do: bool,
    process_invoice: bool,
    start_from_row: int,
    scan_columns: Optional[List[str]],
    log_func=print,
    skip_delete_rows: bool = False,
    delete_row_count: int = 2,   # ← 新增，可输入要删除的行数（默认2）
):

    processed_count = 0
    error_files = []

    matched_files = iter_excel_files_excluding_history(directory, filename_substring)
    total = len(matched_files)
    log_func(f"共发现 {total} 个候选文件。")

    kws = [k.strip().lower() for k in (find_keywords or []) if k.strip()]
    if not kws:
        log_func("⚠️ 查找关键字列表为空，将不进行删除操作。")
        for i in range(total):
            yield i + 1, total
        return

    def row_hits_keywords(sheet, r, col_indexes) -> bool:
        # 命中：任一单元格包含任一关键字
        if col_indexes:
            for c in col_indexes:
                v = sheet.cell(row=r, column=c).value
                s = str(v).strip().lower() if v is not None else ""
                for kw in kws:
                    if kw and kw in s:
                        return True
        else:
            for cell in sheet[r]:
                v = cell.value
                s = str(v).strip().lower() if v is not None else ""
                for kw in kws:
                    if kw and kw in s:
                        return True
        return False

    for idx, file_path in enumerate(matched_files, start=1):
        if debug:
            log_func(f"🔄 正在处理 [{idx}/{total}]: {file_path}")

        wb = None
        try:
            keep_vba = file_path.lower().endswith(".xlsm")
            wb = openpyxl.load_workbook(file_path, keep_vba=keep_vba, data_only=False)
            if debug:
                log_func(f"✅ 已加载: {os.path.basename(file_path)}")
                log_func(f"📄 工作表: {wb.sheetnames}")

            sheets_to_process = []
            if process_do and "DO" in wb.sheetnames:
                sheets_to_process.append("DO")
            if process_invoice and "Invoice" in wb.sheetnames:
                sheets_to_process.append("Invoice")

            if not sheets_to_process:
                if debug:
                    log_func("ℹ️ 无 DO/Invoice 可处理（或未勾选），跳过此文件。")
            else:
                file_modified = False

                for sheet_name in sheets_to_process:
                    sh = wb[sheet_name]
                    max_row = sh.max_row or 1

                    # 限定扫描列
                    col_indexes = None
                    if scan_columns:
                        col_indexes = set()
                        for col_letter in scan_columns:
                            idx_val = col_letters_to_index(col_letter)
                            if idx_val:
                                col_indexes.add(idx_val)

                    # 1) 收集“匹配行 + N-1 行”的删除区间（多关键字命中）
                    raw_ranges: List[Tuple[int, int]] = []
                    start_r = max(1, int(start_from_row))
                    for r in range(start_r, max_row + 1):
                        if row_hits_keywords(sh, r, col_indexes):
                            end_r = min(max_row, r + delete_row_count - 1)  # 删除 N 行
                            raw_ranges.append((r, end_r))



                    # 2) 合并区间并执行删除（自底向上）
                    merged_del = normalize_delete_ranges(raw_ranges)
                    if merged_del and not skip_delete_rows:
                        for a, b in reversed(merged_del):
                            sh.delete_rows(a, b - a + 1)

                    # 3) 删除后：逐行应用 I 列特别规则与通用 G 修正
                    #    Subtotal 范围：从“本区块里 I 列出现的第一个公式行”到“Subtotal 上一行”
                    first_formula_row_in_block: Optional[int] = None
                    last_subtotal_row: Optional[int] = None
                    last_gst_row: Optional[int] = None

                    for row in sh.iter_rows():
                        r_idx = row[0].row
                        f_text = norm_text(sh.cell(row=r_idx, column=COL_F).value)

                        # 统计“本区块里 I 列的第一个公式行”
                        cell_I = sh.cell(row=r_idx, column=COL_I).value
                        if first_formula_row_in_block is None and isinstance(cell_I, str) and cell_I.startswith("="):
                            first_formula_row_in_block = r_idx

                        if f_text == "subtotal":
                            if first_formula_row_in_block is not None and r_idx - 1 >= first_formula_row_in_block:
                                sh.cell(row=r_idx, column=COL_I).value = f"=SUM(I{first_formula_row_in_block}:I{r_idx-1})"
                            else:
                                sh.cell(row=r_idx, column=COL_I).value = "=0"
                            last_subtotal_row = r_idx
                            continue

                        if f_text == "add gst 9%":
                            if last_subtotal_row:
                                sh.cell(row=r_idx, column=COL_I).value = f"=I{last_subtotal_row}*0.09"
                                last_gst_row = r_idx
                            else:
                                sh.cell(row=r_idx, column=COL_I).value = "=0"
                            continue

                        if f_text == "total":
                            if last_subtotal_row and last_gst_row:
                                sh.cell(row=r_idx, column=COL_I).value = f"=I{last_subtotal_row}+I{last_gst_row}"
                            elif last_subtotal_row:
                                sh.cell(row=r_idx, column=COL_I).value = f"=I{last_subtotal_row}"
                            else:
                                sh.cell(row=r_idx, column=COL_I).value = "=0"
                            # 一个区块结束：重置，下一段重新寻找“第一个公式”
                            first_formula_row_in_block = None
                            last_subtotal_row = None
                            last_gst_row = None
                            continue

                        # 普通行：把“本行里所有公式”的 G 引用改成当前行号
                        for cell in row:
                            v = cell.value
                            if isinstance(v, str) and v.startswith('='):
                                new_v = rewrite_formula_G_refs_to_row(v, r_idx)
                                if new_v != v:
                                    cell.value = new_v

                    # 4) 仅对 DO：查找两段文字的行，并合并 H:K + 画横线 + 调行高
                    if sheet_name == "DO":
                        merge_H_to_K_on_rows_with_labels_DO(sh)

                    file_modified = True
                    if debug:
                        if merged_del:
                            if skip_delete_rows:
                                log_func(f"📝 {sheet_name}: 识别到可删区间 {len(merged_del)} 处，但已按设置跳过删除。")
                            else:
                                cnt = sum(b - a + 1 for a, b in merged_del)
                                ex = ", ".join([f"[{a}-{b}]" for a, b in merged_del[:3]])
                                log_func(f"🗑️ {sheet_name}: 删除 {cnt} 行（区间示例：{ex}{' ...' if len(merged_del)>3 else ''}）")
                        log_func(f"🧩 {sheet_name}: 已应用 I 列小计/税/总计（Subtotal 起点=本区块 I 列第一个公式），G 引用重写；DO 已合并 H:K 并加横线/行高。")

                if file_modified:
                    wb.save(file_path)
                    processed_count += 1
                    if debug:
                        log_func(f"💾 已保存修改: {file_path}")

        except Exception as e:
            traceback.print_exc()
            error_files.append(file_path)
            log_func(f"❌ 处理失败: {file_path}，错误: {e}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

        yield idx, total

    log_func(f"✅ 完成：共修改 {processed_count} 个文件。")
    if error_files:
        log_func("⚠️ 以下文件处理失败：")
        for ef in error_files:
            log_func(f"  ❌ {ef}")

# ========= GUI =========
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel 批量处理：多关键字删行 + I列小计/税/总计（起点=区块首个I列公式） + 行内G修正 + DO合并H:K+横线+行高")
        self.geometry("1080x780")

        # 顶部
        frm_top = ttk.Frame(self, padding=10)
        frm_top.pack(fill=tk.X)
        ttk.Label(frm_top, text="根目录：").pack(side=tk.LEFT)
        self.dir_var = tk.StringVar()
        ttk.Entry(frm_top, textvariable=self.dir_var, width=92).pack(side=tk.LEFT, padx=6)
        ttk.Button(frm_top, text="浏览…", command=self.choose_dir).pack(side=tk.LEFT)

        # 参数设置
        frm_cfg = ttk.LabelFrame(self, text="参数设置", padding=10)
        frm_cfg.pack(fill=tk.X, padx=10, pady=6)

        left = ttk.Frame(frm_cfg)
        left.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.filename_substring_var = tk.StringVar(value="xx25")
        self.find_keywords_var = tk.StringVar(value="abcd, efgh")  # 多关键字，逗号分隔
        self.delete_count_var = tk.StringVar(value="2")  # 默认删除2行
        ttk.Label(left, text="命中后删除几行：").grid(row=4, column=0, sticky="w", pady=2)
        ttk.Entry(left, textvariable=self.delete_count_var, width=10).grid(row=4, column=1, sticky="w", pady=2)
        self.debug_var = tk.BooleanVar(value=True)
        self.process_do_var = tk.BooleanVar(value=True)
        self.process_invoice_var = tk.BooleanVar(value=True)
        self.skip_delete_var = tk.BooleanVar(value=False)  # 新增：只执行修正，不删行

        ttk.Label(left, text="文件名包含（忽略大小写）：").grid(row=0, column=0, sticky="w")
        ttk.Entry(left, textvariable=self.filename_substring_var, width=30).grid(row=0, column=1, padx=6, pady=2, sticky="w")

        ttk.Label(left, text="多个关键字（逗号分隔）：").grid(row=1, column=0, sticky="w")
        ttk.Entry(left, textvariable=self.find_keywords_var, width=30).grid(row=1, column=1, padx=6, pady=2, sticky="w")

        ttk.Checkbutton(left, text="处理 DO 表", variable=self.process_do_var).grid(row=2, column=0, sticky="w", pady=2)
        ttk.Checkbutton(left, text="处理 Invoice 表", variable=self.process_invoice_var).grid(row=2, column=1, sticky="w", pady=2)

        ttk.Checkbutton(left, text="Debug 日志", variable=self.debug_var).grid(row=3, column=0, sticky="w", pady=2)
        ttk.Checkbutton(left, text="只执行修正，不删行", variable=self.skip_delete_var).grid(row=3, column=1, sticky="w", pady=2)

        right = ttk.LabelFrame(frm_cfg, text="扫描/删除控制", padding=10)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))

        self.start_row_var = tk.StringVar(value="2")
        self.columns_var = tk.StringVar(value="")  # 如 "B,C,F,G"

        r = 0
        ttk.Label(right, text="从第几行开始检查：").grid(row=r, column=0, sticky="e")
        ttk.Entry(right, textvariable=self.start_row_var, width=10).grid(row=r, column=1, padx=6, pady=2, sticky="w"); r += 1

        ttk.Label(right, text="限定扫描列（可留空）：").grid(row=r, column=0, sticky="e")
        ttk.Entry(right, textvariable=self.columns_var, width=28).grid(row=r, column=1, padx=6, pady=2, sticky="w")
        ttk.Label(right, text="示例：B,C,F,G（留空=整行扫描）").grid(row=r, column=2, sticky="w"); r += 1

        # 操作与进度
        frm_ops = ttk.Frame(self, padding=10)
        frm_ops.pack(fill=tk.X)

        self.btn_run = ttk.Button(frm_ops, text="开始处理", command=self.on_run)
        self.btn_run.pack(side=tk.LEFT)

        self.prog = ttk.Progressbar(frm_ops, mode="determinate", length=560)
        self.prog.pack(side=tk.LEFT, padx=10)

        self.prog_label = ttk.Label(frm_ops, text="0/0")
        self.prog_label.pack(side=tk.LEFT)

        # 日志
        frm_log = ttk.LabelFrame(self, text="日志", padding=8)
        frm_log.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        self.log = ScrolledText(frm_log, height=18)
        self.log.pack(fill=tk.BOTH, expand=True)

        self.worker_thread = None

    def choose_dir(self):
        path = filedialog.askdirectory(title="选择根目录")
        if path:
            self.dir_var.set(path)

    def append_log(self, msg: str):
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.update_idletasks()

    def on_run(self):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showinfo("提示", "任务正在进行中…")
            return

        directory = self.dir_var.get().strip()
        if not directory or not os.path.isdir(directory):
            messagebox.showerror("错误", "请选择有效的根目录")
            return
        
        try:
            delete_row_count = int(self.delete_count_var.get().strip() or "2")
        except ValueError:
            messagebox.showerror("错误", "删除行数必须是整数")
            return
        if delete_row_count < 1:
            messagebox.showerror("错误", "删除行数必须 ≥ 1")
            return


        filename_substring = self.filename_substring_var.get().strip()
        # 多关键字解析
        find_keywords = [k.strip() for k in (self.find_keywords_var.get() or "").split(",") if k.strip()]
        process_do = self.process_do_var.get()
        process_invoice = self.process_invoice_var.get()
        debug = self.debug_var.get()
        skip_delete = self.skip_delete_var.get()

        try:
            start_from_row = int(self.start_row_var.get().strip() or "2")
        except ValueError:
            messagebox.showerror("错误", "起始行号必须是整数")
            return
        if start_from_row < 1:
            messagebox.showerror("错误", "起始行号必须 ≥ 1")
            return

        cols_raw = (self.columns_var.get() or "").strip()
        scan_columns = [c.strip().upper() for c in cols_raw.split(",") if c.strip()] if cols_raw else None

        if not (process_do or process_invoice):
            messagebox.showerror("错误", "请至少勾选一种工作表（DO 或 Invoice）")
            return

        self.log.delete("1.0", tk.END)
        self.prog["value"] = 0
        self.prog_label.config(text="0/0")

        def run_task():
            try:
                matched_files = iter_excel_files_excluding_history(directory, filename_substring)
                total = len(matched_files)
                self.prog["maximum"] = max(total, 1)
                self.prog_label.config(text=f"0/{total}")

                gen = delete_rows_in_excel_files(
                    directory=directory,
                    debug=debug,
                    filename_substring=filename_substring,
                    find_keywords=find_keywords,
                    process_do=process_do,
                    process_invoice=process_invoice,
                    start_from_row=start_from_row,
                    scan_columns=scan_columns,
                    log_func=self.append_log,
                    skip_delete_rows=skip_delete,
                    delete_row_count=delete_row_count,  # ← 传入
                )

                for done, tot in gen:
                    self.prog["value"] = done
                    self.prog_label.config(text=f"{done}/{tot}")
            except Exception as e:
                self.append_log(f"[GUI] 任务异常: {e}")
            finally:
                self.btn_run.config(state=tk.NORMAL)

        self.btn_run.config(state=tk.DISABLED)
        self.worker_thread = threading.Thread(target=run_task, daemon=True)
        self.worker_thread.start()

if __name__ == "__main__":
    app = App()
    app.mainloop()
