import streamlit as st
import streamlit.components.v1 as components
from typing import Any, Optional, Tuple, List, Dict
import datetime
import math
import pandas as pd
import re
import json
import os
import sys
import io
import calendar
import platform
from difflib import SequenceMatcher
from pathlib import Path
from openpyxl import load_workbook

# 尝试检测当前是否已经在 streamlit 运行环境中
try:
    from streamlit.runtime.scriptrunner import get_script_run_ctx  # type: ignore
except Exception:

    def get_script_run_ctx():
        return None


def _short_path(p: str) -> str:
    """在 Windows 下把带空格的长路径转换成短路径，避免 cmd 解析问题。"""
    if os.name != "nt" or " " not in p:
        return p
    try:
        import ctypes
        from ctypes import wintypes

        GetShortPathNameW = ctypes.windll.kernel32.GetShortPathNameW
        GetShortPathNameW.argtypes = [wintypes.LPCWSTR, wintypes.LPWSTR, wintypes.DWORD]
        GetShortPathNameW.restype = wintypes.DWORD
        buf = ctypes.create_unicode_buffer(260)
        res = GetShortPathNameW(p, buf, 260)
        return buf.value if res else p
    except Exception:
        return p


def _is_wsl(
    env: Optional[Dict[str, str]] = None,
    *,
    platform_name: Optional[str] = None,
    proc_version: Optional[str] = None,
) -> bool:
    env = env or dict(os.environ)
    platform_name = platform_name or sys.platform
    if not str(platform_name).startswith("linux"):
        return False
    if env.get("WSL_INTEROP") or env.get("WSL_DISTRO_NAME"):
        return True
    text = (proc_version if proc_version is not None else platform.version()).lower()
    return ("microsoft" in text) or ("wsl" in text)


def _should_force_streamlit_headless(
    env: Optional[Dict[str, str]] = None,
    *,
    platform_name: Optional[str] = None,
    proc_version: Optional[str] = None,
) -> bool:
    env = env or dict(os.environ)
    platform_name = platform_name or sys.platform
    if not str(platform_name).startswith("linux"):
        return False
    if _is_wsl(env, platform_name=platform_name, proc_version=proc_version):
        return True
    return not bool(env.get("DISPLAY") or env.get("WAYLAND_DISPLAY"))


def _build_streamlit_command(
    *,
    python_executable: str,
    script_path: str,
    passthrough_args: List[str],
    force_headless: bool,
) -> List[str]:
    cmd: List[str] = [python_executable, "-m", "streamlit", "run"]
    if force_headless:
        cmd.append("--server.headless=true")
    cmd.append(script_path)
    if passthrough_args:
        cmd += ["--"] + passthrough_args
    return cmd


def _run_streamlit_command(cmd: List[str], runner=None) -> int:
    if runner is None:
        import subprocess

        runner = subprocess.run
    try:
        result = runner(cmd, check=False)
    except KeyboardInterrupt:
        return 130
    return int(getattr(result, "returncode", 0))


# 如果是直接 python xxx.py 启动，并且当前不在 streamlit 环境里，就自动帮你转成 streamlit run
if (
    __name__ == "__main__"
    and os.environ.get("ST_REDIRECTED", "0") != "1"
    and (get_script_run_ctx() is None)
):
    os.environ["ST_REDIRECTED"] = "1"

    script_path = str(Path(__file__).resolve())
    script_path = _short_path(script_path)

    # 把扩展名统一改成小写 .py，避免 Streamlit 讨厌 .PY
    if script_path.lower().endswith(".py"):
        script_path = script_path[:-3] + ".py"

    cmd = _build_streamlit_command(
        python_executable=sys.executable,
        script_path=script_path,
        passthrough_args=sys.argv[1:],
        force_headless=_should_force_streamlit_headless(),
    )

    if os.environ.get("ST_DEBUG_REDIRECT") == "1":
        print("[streamlit-redirect] ", cmd)

    exit_code = _run_streamlit_command(cmd)
    sys.exit(exit_code)


st.set_page_config(page_title="Warehouse Suite", layout="wide")

PRIMARY_WAREHOUSES = ["Savori Whse", "Lai Hock Whse"]

_HOTKEY_ACTION_PARAM = "__hotkey_action"
_HOTKEY_ACTION_CLEAR_STOCK = "clear_stock_filters"
_HOTKEY_ACTION_CLEAR_SALES = "clear_sales_filters"

# ==== 新增：预编译正则 ====
HTML_TAG_RE = re.compile(r"<.*?>")
PAREN_CONTENT_RE = re.compile(r"\s*\([^)]*\)")


def _inject_delete_hotkey_listener(action: str) -> None:
    script = """
<script>
(function() {
  const action = __ACTION__;
  const actionParam = "__hotkey_action";
  function resolveHostWindow() {
    try {
      if (window.parent && window.parent !== window) {
        void window.parent.location.href;
        return window.parent;
      }
    } catch (e) {
      // Ignore cross-origin parent access and fall back to current window.
    }
    return window;
  }

  const hostWindow = resolveHostWindow();
  hostWindow.__savoriDeleteHotkeyAction = action;

  if (hostWindow.__savoriDeleteHotkeyBound) {
    return;
  }
  hostWindow.__savoriDeleteHotkeyBound = true;

  let hostDocument = null;
  try {
    hostDocument = hostWindow.document || document;
  } catch (e) {
    hostDocument = document;
  }

  function isEditable(target) {
    if (!target) {
      return false;
    }
    const tag = String(target.tagName || "").toUpperCase();
    if (tag === "INPUT" || tag === "TEXTAREA" || tag === "SELECT") {
      return true;
    }
    return Boolean(target.isContentEditable);
  }

  function isDeleteKey(event) {
    const key = String(event.key || "");
    const code = String(event.code || "");
    return key === "Delete" || key === "Del" || code === "Delete";
  }

  function triggerClear(currentAction) {
    try {
      const currentHref = String(hostWindow.location.href || "");
      const url = new URL(currentHref, window.location.origin);
      if (url.searchParams.get(actionParam) === currentAction) {
        return;
      }
      url.searchParams.set(actionParam, currentAction);
      hostWindow.location.assign(url.toString());
    } catch (e) {
      try {
        const fallbackUrl = new URL(window.location.href);
        if (fallbackUrl.searchParams.get(actionParam) === currentAction) {
          return;
        }
        fallbackUrl.searchParams.set(actionParam, currentAction);
        window.location.assign(fallbackUrl.toString());
      } catch (err) {
        // No-op fallback.
      }
    }
  }

  function onKeydown(event) {
    const currentAction = hostWindow.__savoriDeleteHotkeyAction || action;
    if (!isDeleteKey(event) || event.repeat) {
      return;
    }
    const activeEl = hostDocument ? hostDocument.activeElement : null;
    if (isEditable(event.target) || isEditable(activeEl)) {
      return;
    }
    triggerClear(currentAction);
  }

  hostWindow.addEventListener("keydown", onKeydown, true);
  if (hostDocument) {
    hostDocument.addEventListener("keydown", onKeydown, true);
  }
})();
</script>
"""
    components.html(
        script.replace("__ACTION__", json.dumps(action)),
        height=0,
    )


def _consume_hotkey_action() -> Optional[str]:
    action: Optional[str] = None
    if hasattr(st, "query_params"):
        raw = st.query_params.get(_HOTKEY_ACTION_PARAM)
        if isinstance(raw, list):
            action = raw[0] if raw else None
        else:
            action = raw
    else:
        get_query_params = getattr(st, "experimental_get_query_params", None)
        params_raw = get_query_params() if callable(get_query_params) else {}
        params = params_raw if isinstance(params_raw, dict) else {}
        raw = params.get(_HOTKEY_ACTION_PARAM)
        if isinstance(raw, list):
            action = raw[0] if raw else None
        else:
            action = raw

    if not action:
        return None

    if hasattr(st, "query_params"):
        try:
            del st.query_params[_HOTKEY_ACTION_PARAM]
        except Exception:
            pass
    else:
        get_query_params = getattr(st, "experimental_get_query_params", None)
        set_query_params = getattr(st, "experimental_set_query_params", None)
        params_raw = get_query_params() if callable(get_query_params) else {}
        params = params_raw if isinstance(params_raw, dict) else {}
        params.pop(_HOTKEY_ACTION_PARAM, None)
        if callable(set_query_params):
            set_query_params(**params)

    return str(action)


def _clear_stock_filters_state(ss: Dict[str, Any]) -> None:
    ss["f_wh"] = []
    ss["f_sup"] = []
    ss["f_sup_ex"] = False
    ss["f_brand"] = []
    ss["f_brand_ex"] = False
    ss["f_desc"] = []
    ss["f_code"] = []
    ss["f_remark"] = []
    ss["f_remark_ex"] = False
    ss["use_date_filters"] = False
    ss["expiry_range"] = None
    ss["relabel_date_range"] = None
    ss["__sig_wo_wh"] = None
    ss["__focus_target"] = None


# ----------------------------- 基础工具函数 -----------------------------


def _normalize_warehouse_name(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _normalize_pack_size_value(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    return text


def _format_pack_size_label(normalized: str) -> str:
    return normalized if normalized else "-"


def _canonical_unit(u: Optional[str]) -> str:
    if u is None:
        return ""
    s = str(u).strip().lower()
    if not s:
        return ""
    synonyms = {
        "ctn": {"ctn", "ctns", "carton", "cartons"},
        "pkt": {"pkt", "pkts", "pack", "packs", "package", "packages"},
        "btl": {"btl", "btls", "bottle", "bottles"},
        "box": {"box", "boxes"},
        "tin": {"tin", "tins"},
        "can": {"can", "cans"},
        "bag": {"bag", "bags"},
        "pc": {"pc", "pcs", "piece", "pieces"},
    }
    for key, vals in synonyms.items():
        if s in vals:
            return key
    return s


def _summary_qty_bucket(unit: Optional[str]) -> str:
    canonical = _canonical_unit(unit)
    if canonical == "ctn":
        return "ctn"
    if canonical in {"pkt", "btl", "box", "tin", "can", "bag", "pc"}:
        return "pkt"
    return ""


_DISPLAY_UNIT_ORDER = ["ctn", "pkt", "btl", "box", "tin", "can", "bag", "pc"]


def _add_quantity_to_breakdown(
    breakdown: Optional[Dict[str, float]], unit: Optional[str], qty: Optional[float]
) -> Dict[str, float]:
    result = dict(breakdown or {})
    canonical = _canonical_unit(unit)
    if not canonical:
        return result
    if qty is None or pd.isna(qty):
        return result
    qty_value = float(qty)
    if math.isclose(qty_value, 0.0, abs_tol=1e-9):
        return result
    result[canonical] = float(result.get(canonical, 0.0)) + qty_value
    return result


def _merge_quantity_breakdowns(items: Any) -> Dict[str, float]:
    merged: Dict[str, float] = {}
    if items is None:
        return merged
    for item in items:
        if not isinstance(item, dict):
            continue
        for unit, qty in item.items():
            merged = _add_quantity_to_breakdown(merged, unit, qty)
    return merged


def _format_quantity_breakdown(breakdown: Optional[Dict[str, float]]) -> str:
    data = {
        _canonical_unit(unit): float(qty)
        for unit, qty in dict(breakdown or {}).items()
        if _canonical_unit(unit)
        and qty is not None
        and not pd.isna(qty)
        and not math.isclose(float(qty), 0.0, abs_tol=1e-9)
    }
    if not data:
        return "0"

    ordered_units = [unit for unit in _DISPLAY_UNIT_ORDER if unit in data]
    ordered_units += sorted(unit for unit in data.keys() if unit not in ordered_units)

    parts: List[str] = []
    for unit in ordered_units:
        qty_text = _format_qty_number(data[unit])
        if not qty_text:
            continue
        if unit == "ctn":
            label = "ctn" if math.isclose(data[unit], 1.0, abs_tol=1e-9) else "ctns"
        elif unit == "pkt":
            label = "pkt" if math.isclose(data[unit], 1.0, abs_tol=1e-9) else "pkts"
        else:
            label = unit
        parts.append(f"{qty_text} {label}")
    return " ".join(parts) if parts else "0"


def _format_qty_number(value: Optional[float]) -> Optional[str]:
    if value is None or pd.isna(value):
        return None
    try:
        num = float(value)
    except (TypeError, ValueError):
        return None
    if math.isclose(num, 0.0, abs_tol=1e-9):
        return "0"
    if math.isclose(num, round(num), rel_tol=1e-9, abs_tol=1e-9):
        return f"{int(round(num))}"
    text_val = f"{num:,.2f}"
    return text_val.rstrip("0").rstrip(".")


def _format_quantity_pair(ctn: Optional[float], pkt: Optional[float]) -> str:
    parts: List[str] = []
    if ctn is not None and not pd.isna(ctn):
        ctn_val = float(ctn)
        if not math.isclose(ctn_val, 0.0, abs_tol=1e-9):
            qty_text = _format_qty_number(ctn_val)
            if qty_text:
                unit = "ctn" if math.isclose(ctn_val, 1.0, abs_tol=1e-9) else "ctns"
                parts.append(f"{qty_text} {unit}")
    if pkt is not None and not pd.isna(pkt):
        pkt_val = float(pkt)
        if not math.isclose(pkt_val, 0.0, abs_tol=1e-9):
            qty_text = _format_qty_number(pkt_val)
            if qty_text:
                unit = "pkt" if math.isclose(pkt_val, 1.0, abs_tol=1e-9) else "pkts"
                parts.append(f"{qty_text} {unit}")
    return " ".join(parts) if parts else "0"


def _strip_html_df(df: pd.DataFrame) -> pd.DataFrame:
    clean = df.copy()
    for col in clean.select_dtypes(include="object").columns:
        clean[col] = clean[col].where(
            clean[col].isna(),
            clean[col].astype(str).str.replace(HTML_TAG_RE, "", regex=True),
        )
    return clean


def _build_copy_values_tsv(df: pd.DataFrame) -> str:
    output = df.to_csv(sep="\t", index=False, lineterminator="\n")
    return output[:-1] if output.endswith("\n") else output


STOCK_SUMMARY_COPY_COLUMNS = [
    "Supplier",
    "Brand",
    "Product",
    "Pack Size",
    "Product Code",
    "Savori Whse",
    "Lai Hock Whse",
    "Total",
    "Status",
]


def _build_stock_summary_copy_df(df: Optional[pd.DataFrame]) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=pd.Index(STOCK_SUMMARY_COPY_COLUMNS))
    columns = [col for col in STOCK_SUMMARY_COPY_COLUMNS if col in df.columns]
    return _strip_html_df(df.reindex(columns=columns).copy())


def _build_stock_summary_with_expiry_copy_tsv(
    summary_df: pd.DataFrame | None, expiry_df: pd.DataFrame | None
) -> str:
    summary_copy_df = _build_stock_summary_copy_df(summary_df)
    payload_parts = [_build_copy_values_tsv(summary_copy_df)]
    if expiry_df is not None and not expiry_df.empty:
        expiry_copy_df = _strip_html_df(expiry_df.copy())
        payload_parts.append(
            "Expiry Breakdown\n" + _build_copy_values_tsv(expiry_copy_df)
        )
    return "\n\n".join(payload_parts)


def _render_copy_values_button(
    table_payload: str,
    key: str,
    *,
    has_rows: bool,
    label: str = "复制全部数值",
) -> None:
    container_id = f"copy-values-{key}"
    status_id = f"{container_id}-status"
    disabled = not has_rows
    button_label = label if has_rows else "无可复制数据"
    script = """
<div id=__CONTAINER_ID__ class="copy-values-root">
  <button id=__BUTTON_ID__ class="copy-values-button" type="button" __DISABLED__>
    <span class="copy-values-icon">⧉</span>
    <span>__LABEL_TEXT__</span>
  </button>
  <span id=__STATUS_ID__ class="copy-values-status" aria-live="polite"></span>
</div>
<style>
  .copy-values-root {
    --copy-bg: #f8fafc;
    --copy-border: #cbd5e1;
    --copy-text: #0f172a;
    --copy-muted: #64748b;
    --copy-accent: #0f766e;
    --copy-disabled: #e2e8f0;
    display: flex;
    align-items: center;
    gap: 8px;
    min-height: 32px;
    font-family: "Noto Sans SC", "Microsoft YaHei", sans-serif;
  }
  .copy-values-button {
    appearance: none;
    border: 1px solid var(--copy-border);
    border-radius: 999px;
    background: var(--copy-bg);
    color: var(--copy-text);
    cursor: pointer;
    display: inline-flex;
    align-items: center;
    gap: 6px;
    font-size: 13px;
    font-weight: 600;
    line-height: 1;
    padding: 7px 12px;
    transition: border-color 140ms ease, color 140ms ease, transform 140ms ease;
  }
  .copy-values-button:hover:not(:disabled) {
    border-color: var(--copy-accent);
    color: var(--copy-accent);
    transform: translateY(-1px);
  }
  .copy-values-button:disabled {
    background: var(--copy-disabled);
    color: var(--copy-muted);
    cursor: not-allowed;
  }
  .copy-values-icon {
    font-size: 12px;
  }
  .copy-values-status {
    color: var(--copy-muted);
    font-size: 12px;
  }
</style>
<script>
(function() {
  const payload = __PAYLOAD__;
  const button = document.getElementById(__BUTTON_ID__);
  const status = document.getElementById(__STATUS_ID__);
  if (!button || button.disabled) {
    return;
  }
  button.addEventListener("click", async function() {
    try {
      await navigator.clipboard.writeText(payload);
      if (status) {
        status.textContent = "已复制";
      }
    } catch (error) {
      if (status) {
        status.textContent = "复制失败";
      }
    }
  });
})();
</script>
"""
    html = (
        script.replace("__CONTAINER_ID__", json.dumps(container_id))
        .replace("__BUTTON_ID__", json.dumps(f"{container_id}-button"))
        .replace("__STATUS_ID__", json.dumps(status_id))
        .replace("__DISABLED__", "disabled" if disabled else "")
        .replace("__LABEL_TEXT__", button_label)
        .replace("__PAYLOAD__", json.dumps(table_payload))
    )
    _ = components.html(html, height=42)


MONTH_MAP: Dict[str, int] = {}
for index, name in enumerate(calendar.month_name):
    if name:
        MONTH_MAP[name.lower()] = index
for index, abbr in enumerate(calendar.month_abbr):
    if abbr:
        MONTH_MAP[abbr.lower()] = index


def _month_sort_key(value) -> int:
    if pd.isna(value):
        return 999
    text = str(value).strip()
    if not text:
        return 999
    if text.isdigit():
        num = int(text)
        if 1 <= num <= 12:
            return num
    lower = text.lower()
    if lower in MONTH_MAP:
        return MONTH_MAP[lower]
    if lower[:3] in MONTH_MAP:
        return MONTH_MAP[lower[:3]]
    return 999


def _format_month_label(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text.isdigit():
        num = int(text)
        if 1 <= num <= 12:
            return calendar.month_abbr[num]
    lower = text.lower()
    if lower in MONTH_MAP:
        num = MONTH_MAP[lower]
        return calendar.month_abbr[num] if num else text
    if lower[:3] in MONTH_MAP:
        num = MONTH_MAP[lower[:3]]
        return calendar.month_abbr[num] if num else text
    return text


def _format_date_with_month_label(value: Any, pattern: str = "%d-%b-%Y") -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return ""
    return parsed.strftime(pattern)


def _format_qty_display(value) -> str:
    if pd.isna(value):
        return ""
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    if math.isclose(num, round(num), rel_tol=1e-9, abs_tol=1e-9):
        return f"{int(round(num)):,}"
    formatted = f"{num:,.2f}".rstrip("0").rstrip(".")
    return formatted


def _format_ctn_equivalent_display(value) -> str:
    if pd.isna(value):
        return ""
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    if math.isclose(num, 0.0, abs_tol=1e-9):
        return "0"
    formatted = f"{num:,.2f}".rstrip("0").rstrip(".")
    return formatted


def _format_price_display(value) -> str:
    if pd.isna(value):
        return ""
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    if math.isclose(num, 0.0, abs_tol=1e-9):
        return "0.00"
    return f"{num:,.2f}"


def _parse_qty_text(value: Any) -> Tuple[int, int]:
    text = str(value or "").strip().lower()
    if not text or text == "-":
        return 0, 0
    text = text.replace(",", "")
    ctn_match = re.search(r"(-?\d+)\s*ctn(?:s)?\b", text)
    pkt_match = re.search(r"(-?\d+)\s*pkt(?:s)?\b", text)
    ctn = int(ctn_match.group(1)) if ctn_match else 0
    pkt = int(pkt_match.group(1)) if pkt_match else 0
    return ctn, pkt


def _format_ctn_pkt_total(ctn: int, pkt: int) -> str:
    if ctn == 0 and pkt == 0:
        return "0"
    return _format_quantity_pair(float(ctn), float(pkt))


SALES_COLUMNS = [
    "Year",
    "Date",
    "Month",
    "Brand/Category",
    "Supplier",
    "Product Code",
    "Product Description",
    "Carton Packing",
    "Customer",
    "Outlet",
    "Qty in Pcs",
    "Qty in Ctns",
    "Total Qty in Pcs",
    "Total Qty in Ctns",
    "Invoice #",
    "Total Value",
    "GST",
    "Total Value Inclusive GST",
    "Account",
    "Customer PO#",
    "Remarks",
]

SALES_NUMERIC_COLUMNS = [
    "Qty in Pcs",
    "Qty in Ctns",
    "Total Value",
    "GST",
    "Total Value Inclusive GST",
]

SALES_DATE_COLUMNS = ["Date"]

SALES_FILTER_SESSION_KEYS = {
    "Year": "sales_filter_year",
    "Month": "sales_filter_month",
    "Customer": "sales_filter_customer",
    "Outlet": "sales_filter_outlet",
    "Product Description": "sales_filter_product_description",
    "Supplier": "sales_filter_supplier",
    "Brand/Category": "sales_filter_brand",
    "Product Code": "sales_filter_product_code",
    "Account": "sales_filter_account",
}

SALES_FILTER_CUSTOMER_EXCLUDE_KEY = "sales_filter_customer_exclude"
SALES_FILTER_SUPPLIER_EXCLUDE_KEY = "sales_filter_supplier_exclude"


def _clear_sales_filters_state(ss: Dict[str, Any]) -> None:
    for key in SALES_FILTER_SESSION_KEYS.values():
        ss[key] = []
    ss[SALES_FILTER_CUSTOMER_EXCLUDE_KEY] = []
    ss[SALES_FILTER_SUPPLIER_EXCLUDE_KEY] = False
    ss["sales_filter_invoice"] = ""
    ss["sales_filter_use_exact_date"] = False
    ss["sales_filter_exact_date"] = None
    ss["sales_filter_date_from"] = None
    ss["sales_filter_date_to"] = None


def _safe_divide_series(numerator, denominator) -> pd.Series:
    numerator = pd.to_numeric(numerator, errors="coerce")
    denominator = pd.to_numeric(denominator, errors="coerce")
    result = numerator / denominator
    mask = denominator.isna() | (denominator == 0)
    return result.where(~mask, other=pd.NA)


def _highlight_missing_cell(value) -> str:
    if isinstance(value, str):
        if not value.strip():
            return "background-color: rgba(255, 235, 205, 0.6);"
    if pd.isna(value):
        return "background-color: rgba(255, 235, 205, 0.6);"
    return ""


_WEIGHT_UNIT_PATTERN = re.compile(
    r"\d+(?:\.\d+)?\s*(?:kg|kgs|g|grams?|gm|lt|ltr|l|ml|milliliter|millilitre|liter|litre|oz|lb|lbs)\b"
)


def _parse_carton_packing_numeric(value) -> Optional[float]:
    if pd.isna(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    lowered = text.lower()

    match = re.search(r"(\d+)\s*[x×]", lowered)
    if not match:
        match = re.search(
            r"(\d+)(?=\s*(?:kg|g|ltr|lt|l|ml|litre|litres|liter|liters))",
            lowered,
        )
    if not match:
        match = re.search(r"(\d+)", lowered)
    if not match:
        return None
    try:
        return float(match.group(1))
    except ValueError:
        return None


def _infer_carton_pack_size(value) -> Optional[float]:
    if pd.isna(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    lowered = text.lower()
    segments = [seg.strip() for seg in re.split(r"[x×*]", lowered) if seg.strip()]

    numeric_segments: List[Tuple[float, bool]] = []
    for seg in segments:
        match = re.search(r"(\d+(?:\.\d+)?)", seg)
        if not match:
            continue
        try:
            number_value = float(match.group(1))
        except ValueError:
            continue
        is_weight = bool(_WEIGHT_UNIT_PATTERN.search(seg))
        numeric_segments.append((number_value, is_weight))

    non_weight_numbers = [
        number for number, is_weight in numeric_segments if not is_weight and number > 0
    ]
    if non_weight_numbers:
        product = 1.0
        for number in non_weight_numbers:
            product *= number
        return product
    if numeric_segments:
        first_value, first_is_weight = numeric_segments[0]
        if first_is_weight:
            return 1.0
        if first_value > 0:
            return first_value
    return None


def _format_total_qty_text(total_pieces: float, carton_size: Optional[float]) -> str:
    if carton_size is None or carton_size <= 0 or pd.isna(total_pieces):
        return "-"
    try:
        pieces = int(round(total_pieces))
    except (TypeError, ValueError):
        return "-"
    if pieces == 0:
        return "0"
    cartons = pieces // int(carton_size)
    packets = pieces - cartons * int(carton_size)
    parts: List[str] = []
    if cartons > 0:
        parts.append(f"{cartons} ctn{'s' if cartons != 1 else ''}")
    if packets > 0:
        parts.append(f"{packets} pkt{'s' if packets != 1 else ''}")
    return " ".join(parts) if parts else "0"


def _coerce_qty_int(value) -> int:
    if value is None or pd.isna(value):
        return 0
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return 0


def _format_summary_total_qty(ctn_value, pkt_value, pack_size_value) -> str:
    cartons = _coerce_qty_int(ctn_value)
    packets = _coerce_qty_int(pkt_value)
    pack_size = _coerce_qty_int(pack_size_value)
    if pack_size > 0:
        total_packets = cartons * pack_size + packets
        formatted = _format_total_qty_text(total_packets, pack_size)
        if formatted and formatted != "-":
            return formatted
    return f"{cartons} ctns {packets} pkts"


def _format_stock_scope_total_qty(scope_df: pd.DataFrame) -> str:
    if scope_df is None or scope_df.empty:
        return "0"

    total_ctns = pd.to_numeric(
        scope_df.get("total_ctn", pd.Series(0.0, index=scope_df.index)),
        errors="coerce",
    ).fillna(0.0)
    total_pkts = pd.to_numeric(
        scope_df.get("total_pkt", pd.Series(0.0, index=scope_df.index)),
        errors="coerce",
    ).fillna(0.0)

    pack_sizes = {
        int(round(float(pack_size)))
        for pack_size in scope_df.get("Pack Size", pd.Series(dtype="object"))
        .map(_infer_carton_pack_size)
        .dropna()
        .tolist()
        if float(pack_size) > 0
    }

    if len(pack_sizes) == 1:
        pack_size = next(iter(pack_sizes))
        total_packets = float((total_ctns * pack_size + total_pkts).sum())
        return _format_total_qty_text(total_packets, pack_size)

    return _format_quantity_breakdown(
        _merge_quantity_breakdowns(scope_df.get("total_qty_breakdown", []))
    )


def build_norm_desc(df: pd.DataFrame) -> pd.Series:
    if df is None or df.empty:
        return pd.Series(dtype="string", name="norm_desc")

    desc_series = (
        df.get("description", pd.Series(dtype="string")).astype("string").fillna("")
    )
    base = desc_series.str.replace(PAREN_CONTENT_RE, "", regex=True).str.strip()

    code_series = (
        df.get("product_code", pd.Series(dtype="string"))
        .astype("string")
        .fillna("")
        .str.strip()
    )
    fallback_desc = desc_series.str.strip()

    norm = base.copy()
    mask_empty = norm.isna() | (norm.str.len() == 0)
    if mask_empty.any():
        norm = norm.mask(mask_empty, code_series)
        mask_empty = norm.isna() | (norm.str.len() == 0)
    if mask_empty.any():
        norm = norm.mask(mask_empty, fallback_desc)
        mask_empty = norm.isna() | (norm.str.len() == 0)
    if mask_empty.any():
        norm = norm.mask(mask_empty, "Unnamed product")

    norm = norm.fillna("").str.strip().replace("", "Unnamed product")
    norm.name = "norm_desc"
    return norm


def _extract_reorder_points(df_group: pd.DataFrame) -> Dict[str, Optional[float]]:
    result: Dict[str, Optional[float]] = {"ctn": None, "pkt": None}
    if df_group is None or df_group.empty:
        return result
    candidate_cols = [
        c
        for c in df_group.columns
        if isinstance(c, str) and ("reorder" in c.lower() or "rop" in c.lower())
    ]
    for col in candidate_cols:
        values = pd.to_numeric(df_group[col], errors="coerce")
        valid = values.dropna()
        if valid.empty:
            continue
        value = float(valid.iloc[0])
        lower = col.lower()
        if "pkt" in lower:
            result["pkt"] = value
        elif "ctn" in lower:
            result["ctn"] = value
        else:
            if result["ctn"] is None:
                result["ctn"] = value
            if result["pkt"] is None:
                result["pkt"] = value
    return result


# ----------------------------- 业务聚合 -----------------------------


def aggregate_summary(
    df: pd.DataFrame,
    warehouses: Optional[List[str]] = None,
) -> Tuple[pd.DataFrame, Dict[Tuple[str, str, str, str, str], pd.DataFrame]]:
    warehouses = warehouses or PRIMARY_WAREHOUSES
    columns = [
        "Supplier",
        "Brand",
        "Product",
        "Pack Size",
        "Product Code",
        "savori_ctn",
        "savori_pkt",
        "lai_hock_ctn",
        "lai_hock_pkt",
        "total_ctn",
        "total_pkt",
        "savori_qty_breakdown",
        "lai_hock_qty_breakdown",
        "total_qty_breakdown",
        "reorder_point_ctn",
        "reorder_point_pkt",
        "group_key",
    ]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns), {}

    work = df.copy()
    if "warehouse" in work.columns:
        work["_warehouse_norm"] = work["warehouse"].map(_normalize_warehouse_name)
        work = work[work["_warehouse_norm"].isin(warehouses)]
    else:
        work["_warehouse_norm"] = ""
    if work.empty:
        return pd.DataFrame(columns=columns), {}

    work["norm_desc"] = build_norm_desc(work)
    work["_unit_norm"] = work.get("unit", pd.Series(dtype="string")).apply(
        _canonical_unit
    )
    work["_qty"] = pd.to_numeric(
        work.get("stock_qty", pd.Series(dtype=float)), errors="coerce"
    ).fillna(0.0)
    work["_product_code_norm"] = (
        work.get("product_code", pd.Series(dtype="string"))
        .astype("string")
        .fillna("")
        .str.strip()
    )
    if "supplier" not in work.columns:
        work["supplier"] = pd.NA

    if "pack_size" in work.columns:
        work["_pack_size_norm"] = work["pack_size"].map(_normalize_pack_size_value)
    else:
        work["_pack_size_norm"] = ""
    if "brand" in work.columns:
        work["_brand_norm"] = work["brand"].astype("string").fillna("").str.strip()
    else:
        work["_brand_norm"] = ""

    rows = []
    detail_map: Dict[Tuple[str, str, str, str, str], pd.DataFrame] = {}

    for key, grp in work.groupby(
        [
            "supplier",
            "_brand_norm",
            "norm_desc",
            "_pack_size_norm",
            "_product_code_norm",
        ],
        dropna=False,
    ):
        if not isinstance(key, tuple):
            key = (key,)
        supplier_val = key[0]
        brand_val = key[1] if len(key) > 1 else ""
        norm_desc_val = key[2] if len(key) > 2 else ""
        code_val = key[4] if len(key) > 4 else ""
        supplier_name = (
            str(supplier_val).strip()
            if pd.notna(supplier_val) and str(supplier_val).strip()
            else "Unknown supplier"
        )
        product_label = (
            str(norm_desc_val).strip()
            if pd.notna(norm_desc_val) and str(norm_desc_val).strip()
            else "Unnamed product"
        )
        raw_pack_val = key[3] if len(key) > 3 else ""
        pack_norm_val = _normalize_pack_size_value(raw_pack_val)
        pack_size_label = _format_pack_size_label(pack_norm_val)
        brand_text = (
            str(brand_val).strip()
            if pd.notna(brand_val) and str(brand_val).strip()
            else ""
        )
        brand_label = brand_text if brand_text else "Unknown brand"
        code_text = str(code_val).strip() if pd.notna(code_val) else ""
        product_code_label = code_text if code_text else "-"
        group_key = (
            supplier_name,
            brand_label,
            product_label,
            pack_norm_val,
            code_text,
        )
        detail_map[group_key] = grp.copy()

        wh_totals = {wh: {"ctn": 0.0, "pkt": 0.0} for wh in warehouses}
        wh_qty_breakdowns = {wh: {} for wh in warehouses}
        wh_unit = (
            grp.groupby(["_warehouse_norm", "_unit_norm"], dropna=False)["_qty"]
            .sum()
            .reset_index()
        )
        for _, r in wh_unit.iterrows():
            wh = r["_warehouse_norm"]
            unit_raw = _canonical_unit(r["_unit_norm"])
            unit = _summary_qty_bucket(unit_raw)
            qty = float(r["_qty"])
            if wh in wh_qty_breakdowns:
                wh_qty_breakdowns[wh] = _add_quantity_to_breakdown(
                    wh_qty_breakdowns[wh], unit_raw, qty
                )
            if wh in wh_totals and unit in ("ctn", "pkt"):
                wh_totals[wh][unit] += qty

        total_ctn = sum(wh_totals.get(wh, {}).get("ctn", 0.0) for wh in warehouses)
        total_pkt = sum(wh_totals.get(wh, {}).get("pkt", 0.0) for wh in warehouses)
        total_qty_breakdown = _merge_quantity_breakdowns(wh_qty_breakdowns.values())

        reorder_points = _extract_reorder_points(grp)

        rows.append(
            {
                "Supplier": supplier_name,
                "Brand": brand_label,
                "Product": product_label,
                "Pack Size": pack_size_label,
                "Product Code": product_code_label,
                "savori_ctn": wh_totals.get("Savori Whse", {}).get("ctn", 0.0),
                "savori_pkt": wh_totals.get("Savori Whse", {}).get("pkt", 0.0),
                "lai_hock_ctn": wh_totals.get("Lai Hock Whse", {}).get("ctn", 0.0),
                "lai_hock_pkt": wh_totals.get("Lai Hock Whse", {}).get("pkt", 0.0),
                "total_ctn": total_ctn,
                "total_pkt": total_pkt,
                "savori_qty_breakdown": dict(wh_qty_breakdowns.get("Savori Whse", {})),
                "lai_hock_qty_breakdown": dict(
                    wh_qty_breakdowns.get("Lai Hock Whse", {})
                ),
                "total_qty_breakdown": dict(total_qty_breakdown),
                "reorder_point_ctn": reorder_points.get("ctn"),
                "reorder_point_pkt": reorder_points.get("pkt"),
                "group_key": group_key,
            }
        )

    summary_df = pd.DataFrame(rows, columns=columns if rows else columns)
    return summary_df, detail_map


# ----------------------------- 判定规则（批次层/产品层） -----------------------------


def classify_batch_status(
    expiry_date: Optional[datetime.date], ctn: float, pkt: float, expiry_days: int
):
    total = (ctn or 0.0) + (pkt or 0.0)
    if not isinstance(expiry_date, datetime.date):
        return ("Depleted" if total == 0 else "OK", None)
    days = (expiry_date - datetime.date.today()).days
    if total > 0:
        if days < 0:
            return ("Expired", days)
        if 0 <= days <= int(expiry_days):
            return ("Near-Expiry", days)
        return ("OK", days)
    else:
        return ("Depleted", days)


def classify_product_status(
    has_expired: bool, has_near: bool, is_low_stock: bool
) -> str:
    if has_expired:
        return "Expired"
    if has_near:
        return "Near-Expiry"
    if is_low_stock:
        return "Low-Stock"
    return "OK"


def split_by_expiry(
    df_row_scope: pd.DataFrame,
    warehouses: Optional[List[str]] = None,
    *,
    expiry_days: int = 30,
    show_depleted: bool = True,
    batch_mode: str = "expiry",
) -> pd.DataFrame:
    warehouses = warehouses or PRIMARY_WAREHOUSES
    columns = [
        "Expiry",
        "Remark",
        "Savori Whse",
        "Lai Hock Whse",
        "Subtotal",
        "subtotal_ctn",
        "subtotal_pkt",
        "subtotal_qty_breakdown",
        "expiry_date",
        "status_batch",
        "days_to_expiry",
        "Info",
    ]
    if df_row_scope is None or df_row_scope.empty:
        return pd.DataFrame(columns=columns)

    work = df_row_scope.copy()
    if "warehouse" not in work.columns:
        return pd.DataFrame(columns=columns)

    work["_warehouse_norm"] = work["warehouse"].map(_normalize_warehouse_name)
    work = work[work["_warehouse_norm"].isin(warehouses)]
    if work.empty:
        return pd.DataFrame(columns=columns)

    work["_unit_norm"] = work.get("unit", pd.Series(dtype="string")).apply(
        _canonical_unit
    )
    work["_qty"] = pd.to_numeric(
        work.get("stock_qty", pd.Series(dtype=float)), errors="coerce"
    ).fillna(0.0)

    work["_expiry_norm"] = pd.to_datetime(
        work.get("expiry_date"), errors="coerce", format="mixed"
    )
    work["_expiry_label"] = work["_expiry_norm"].apply(
        lambda x: x.date().isoformat() if pd.notna(x) else "No Expiry"
    )
    work["_expiry_sort_key"] = work["_expiry_norm"].apply(
        lambda x: x.date() if pd.notna(x) else datetime.date.max
    )

    def _extract_remark(s: str) -> str:
        lst = re.findall(r"\(([^)]*)\)", str(s) or "")
        vals = [t.strip() for t in lst if t and t.strip()]
        return ", ".join(vals) if vals else "No Remark"

    work["_remark_label"] = work.get("description", pd.Series(dtype="string")).apply(
        _extract_remark
    )

    if batch_mode == "remark":
        group_keys = ["_remark_label"]
    elif batch_mode == "both":
        group_keys = ["_expiry_label", "_remark_label"]
    else:
        group_keys = ["_expiry_label"]

    rows = []
    for key, grp in work.groupby(group_keys, dropna=False):
        if not isinstance(key, tuple):
            key = (key,)
        exp_label = None
        remark_label = None
        if batch_mode == "remark":
            remark_label = key[0]
        elif batch_mode == "both":
            exp_label = key[0] if len(key) > 0 else None
            remark_label = key[1] if len(key) > 1 else None
        else:
            exp_label = key[0]

        sort_key = grp["_expiry_sort_key"].min()
        expiry_val = grp["_expiry_norm"].dropna().min()

        per_wh = {wh: {"ctn": 0.0, "pkt": 0.0} for wh in warehouses}
        per_wh_breakdown = {wh: {} for wh in warehouses}
        wh_unit = (
            grp.groupby(["_warehouse_norm", "_unit_norm"], dropna=False)["_qty"]
            .sum()
            .reset_index()
        )
        for _, r in wh_unit.iterrows():
            wh = r["_warehouse_norm"]
            unit_raw = _canonical_unit(r["_unit_norm"])
            unit = _summary_qty_bucket(unit_raw)
            qty = float(r["_qty"])
            if wh in per_wh_breakdown:
                per_wh_breakdown[wh] = _add_quantity_to_breakdown(
                    per_wh_breakdown[wh], unit_raw, qty
                )
            if wh in per_wh and unit in ("ctn", "pkt"):
                per_wh[wh][unit] += qty

        subtotal_ctn = sum(per_wh.get(wh, {}).get("ctn", 0.0) for wh in warehouses)
        subtotal_pkt = sum(per_wh.get(wh, {}).get("pkt", 0.0) for wh in warehouses)
        subtotal_qty_breakdown = _merge_quantity_breakdowns(per_wh_breakdown.values())

        exp_date = expiry_val.date() if pd.notna(expiry_val) else None
        status_batch, d2e = classify_batch_status(
            exp_date, subtotal_ctn, subtotal_pkt, expiry_days
        )

        if (status_batch == "Depleted") and (not show_depleted):
            continue

        qty_text = _format_quantity_breakdown(subtotal_qty_breakdown)
        info = ""
        if status_batch == "Expired":
            info = f"已过期 {abs(d2e or 0)} 天，数量 {qty_text}"
        elif status_batch == "Near-Expiry":
            info = f"到期日 {exp_date}，距离到期 {d2e} 天，数量 {qty_text}"
        elif status_batch == "Depleted":
            info = "该批次已用完（0），等待仓库处理"

        rows.append(
            {
                "Expiry": exp_label if exp_label is not None else "",
                "Remark": remark_label if remark_label is not None else "",
                "Savori Whse": _format_quantity_breakdown(
                    per_wh_breakdown.get("Savori Whse", {})
                ),
                "Lai Hock Whse": _format_quantity_breakdown(
                    per_wh_breakdown.get("Lai Hock Whse", {})
                ),
                "Subtotal": qty_text,
                "subtotal_ctn": subtotal_ctn,
                "subtotal_pkt": subtotal_pkt,
                "subtotal_qty_breakdown": dict(subtotal_qty_breakdown),
                "expiry_date": exp_date,
                "status_batch": status_batch,
                "days_to_expiry": d2e,
                "Info": info,
                "_sort_key": sort_key,
            }
        )

    if not rows:
        return pd.DataFrame(columns=columns)

    result = (
        pd.DataFrame(rows)
        .sort_values(by=["_sort_key", "Expiry", "Remark"], kind="stable")
        .drop(columns="_sort_key")
        .reset_index(drop=True)
    )
    return result


# ----------------------------- 侧边栏筛选 -----------------------------


def _desc_base_series(series: pd.Series) -> pd.Series:
    return series.astype(str).str.replace(r"\s*\([^)]*\)", "", regex=True).str.strip()


def _extract_remarks_from_desc(series: pd.Series) -> list:
    vals = series.astype(str).str.findall(r"\(([^)]*)\)").dropna().tolist()
    out = set()
    for lst in vals:
        for r in lst or []:
            s = str(r).strip()
            if s:
                out.add(s)
    return sorted(out)


def _parse_text_query_terms(value: Any) -> List[str]:
    parts = re.split(r"[,;\n]", str(value or ""))
    return [part.strip().lower() for part in parts if part and part.strip()]


def _build_stock_description_suggestion_options(df: pd.DataFrame) -> List[str]:
    if df is None or df.empty or "description" not in df.columns:
        return []
    base_values = _desc_base_series(df["description"])
    unique_values = {
        str(value).strip() for value in base_values.tolist() if str(value).strip()
    }
    return sorted(unique_values)


def _build_stock_product_code_suggestion_options(df: pd.DataFrame) -> List[str]:
    if df is None or df.empty or "product_code" not in df.columns:
        return []
    unique_values = {
        str(value).strip()
        for value in df["product_code"].dropna().astype(str).tolist()
        if str(value).strip()
    }
    return sorted(unique_values)


def _build_stock_remark_suggestion_options(df: pd.DataFrame) -> List[str]:
    if df is None or df.empty or "description" not in df.columns:
        return []
    return _extract_remarks_from_desc(df["description"])


def _apply_filter_state(
    df_in: pd.DataFrame, filter_state: Dict[str, Any], exclude: str = ""
) -> pd.DataFrame:
    d = df_in

    def _include(series: pd.Series, selected: list):
        if not selected:
            return pd.Series(True, index=series.index)
        return series.isin(selected)

    if exclude != "warehouse" and "warehouse" in d.columns:
        sel = list(filter_state.get("f_wh", []))
        if sel:
            d = d[_include(d["warehouse"], sel)]

    if exclude != "supplier" and "supplier" in d.columns:
        sel = list(filter_state.get("f_sup", []))
        exm = bool(filter_state.get("f_sup_ex", False))
        if sel:
            m = d["supplier"].isin(sel)
            d = d[~m] if exm else d[m]

    if exclude != "brand" and "brand" in d.columns:
        sel = list(filter_state.get("f_brand", []))
        exm = bool(filter_state.get("f_brand_ex", False))
        if sel:
            m = d["brand"].isin(sel)
            d = d[~m] if exm else d[m]

    if exclude != "desc" and "description" in d.columns:
        base_ser = _desc_base_series(d["description"])
        sel = list(filter_state.get("f_desc", []))
        if sel:
            d = d[base_ser.isin(sel)]

    if exclude != "code" and "product_code" in d.columns:
        sel = list(filter_state.get("f_code", []))
        if sel:
            d = d[_include(d["product_code"], sel)]

    if exclude != "remark" and "description" in d.columns:
        sel = list(filter_state.get("f_remark", []))
        exm = bool(filter_state.get("f_remark_ex", False))
        if sel:
            matches = d["description"].astype(str).str.findall(r"\(([^)]*)\)")
            has_any = matches.apply(
                lambda lst: any((str(x).strip() in sel) for x in (lst or []))
            )
            d = d[~has_any] if exm else d[has_any]

    return d


def apply_filters_v2(df: pd.DataFrame):
    base = df.copy()

    ss = st.session_state

    def _set_focus_target(target: str) -> None:
        ss["__focus_target"] = target

    def _ensure_multiselect_key(key: str, options: list, init: list):
        if key not in ss:
            ss[key] = list(init)
        else:
            current = ss.get(key, [])
            if isinstance(current, (str, int, float)):
                current = [current]
            ss[key] = [v for v in current if v in options]

    sel_wh = list(ss.get("f_wh", []))
    sel_sup = list(ss.get("f_sup", []))
    sel_brand = list(ss.get("f_brand", []))
    sel_desc = list(ss.get("f_desc", []))
    sel_code = list(ss.get("f_code", []))
    sel_remark = list(ss.get("f_remark", []))

    with st.sidebar:
        st.header("筛选条件")

        if st.button("清除 Stock 筛选", key="stock_clear_filters_button"):
            _clear_stock_filters_state(ss)

        sig_wo_wh = (
            tuple(ss.get("f_sup", [])),
            bool(ss.get("f_sup_ex", False)),
            tuple(ss.get("f_brand", [])),
            bool(ss.get("f_brand_ex", False)),
            tuple(ss.get("f_desc", [])),
            tuple(ss.get("f_code", [])),
            tuple(ss.get("f_remark", [])),
            bool(ss.get("f_remark_ex", False)),
            bool(ss.get("use_date_filters", False)),
            ss.get("expiry_range"),
            ss.get("relabel_date_range"),
        )
        prev_sig = ss.get("__sig_wo_wh")
        sig_changed = sig_wo_wh != prev_sig

        if "warehouse" in base.columns:
            d = _apply_filter_state(base, ss, exclude="warehouse")
            wh_options = [
                x for x in d["warehouse"].dropna().astype(str).unique().tolist()
            ]
            ordered = [w for w in ["Savori Whse", "Lai Hock Whse"] if w in wh_options]
            ordered += [w for w in wh_options if w not in ordered]
            default_selection = [
                w for w in ["Savori Whse", "Lai Hock Whse"] if w in ordered
            ] or list(ordered)

            cur = [w for w in ss.get("f_wh", []) if w in ordered]
            need_reset = sig_changed or (not cur and bool(ordered))
            ss["f_wh"] = (
                list(default_selection)
                if need_reset
                else (cur or list(default_selection))
            )

            st.multiselect(
                "仓库",
                options=ordered,
                key="f_wh",
                placeholder="选择一个或多个仓库",
            )
            sel_wh = list(ss.get("f_wh", []))

            ss["__sig_wo_wh"] = sig_wo_wh

        if "supplier" in base.columns:
            d = _apply_filter_state(base, ss, exclude="supplier")
            sup_options = sorted([x for x in d["supplier"].dropna().unique().tolist()])
            _ensure_multiselect_key("f_sup", sup_options, [])
            st.multiselect(
                "Supplier", sup_options, key="f_sup", placeholder="选择供应商"
            )
            st.checkbox("Exclude selected (Supplier)", key="f_sup_ex", value=False)
            sel_sup = list(ss.get("f_sup", []))

        if "brand" in base.columns:
            d = _apply_filter_state(base, ss, exclude="brand")
            brand_options = sorted([x for x in d["brand"].dropna().unique().tolist()])
            _ensure_multiselect_key("f_brand", brand_options, [])
            st.multiselect(
                "Brand", brand_options, key="f_brand", placeholder="选择品牌"
            )
            st.checkbox("Exclude selected (Brand)", key="f_brand_ex", value=False)
            sel_brand = list(ss.get("f_brand", []))

        if "description" in base.columns:
            d = _apply_filter_state(base, ss, exclude="desc")
            desc_options = _build_stock_description_suggestion_options(d)
            _ensure_multiselect_key("f_desc", desc_options, [])
            st.multiselect(
                "Description（去括号后）",
                desc_options,
                key="f_desc",
                placeholder="选择描述",
            )
            sel_desc = list(ss.get("f_desc", []))

        if "product_code" in base.columns:
            d = _apply_filter_state(base, ss, exclude="code")
            code_options = _build_stock_product_code_suggestion_options(d)
            _ensure_multiselect_key("f_code", code_options, [])
            st.multiselect(
                "Product Code",
                code_options,
                key="f_code",
                placeholder="选择产品编码",
            )
            sel_code = list(ss.get("f_code", []))

        if "description" in base.columns:
            d = _apply_filter_state(base, ss, exclude="remark")
            remark_options = _build_stock_remark_suggestion_options(d)
            _ensure_multiselect_key("f_remark", remark_options, [])
            st.multiselect(
                "Remark（来自描述括号）",
                remark_options,
                key="f_remark",
                placeholder="选择 Remark",
            )
            sel_remark = list(ss.get("f_remark", []))
            st.checkbox("Exclude matches (Remark)", key="f_remark_ex", value=False)

        use_date_filters = st.checkbox(
            "启用日期范围筛选",
            key="use_date_filters",
            value=bool(ss.get("use_date_filters", False)),
        )
        start = end = None
        r_start = r_end = None

        def _clamp(cur_range, min_d, max_d):
            if pd.isna(min_d) or pd.isna(max_d):
                return None
            if not cur_range or len(cur_range) != 2:
                return (min_d.date(), max_d.date())
            a, b = cur_range
            a = max(min_d.date(), min(a, max_d.date()))
            b = max(min_d.date(), min(b, max_d.date()))
            if a > b:
                a, b = (min_d.date(), max_d.date())
            return (a, b)

        if use_date_filters:
            if "expiry_date" in base.columns:
                min_d, max_d = base["expiry_date"].min(), base["expiry_date"].max()
                if pd.notna(min_d) and pd.notna(max_d):
                    st.session_state["expiry_range"] = _clamp(
                        st.session_state.get("expiry_range"), min_d, max_d
                    )
                    st.date_input(
                        "有效期范围",
                        key="expiry_range",
                        min_value=min_d.date(),
                        max_value=max_d.date(),
                    )
                    start, end = st.session_state.get("expiry_range")

            if "relabel_to_date" in base.columns:
                min_r, max_r = (
                    base["relabel_to_date"].min(),
                    base["relabel_to_date"].max(),
                )
                if pd.notna(min_r) and pd.notna(max_r):
                    st.session_state["relabel_date_range"] = _clamp(
                        st.session_state.get("relabel_date_range"), min_r, max_r
                    )
                    st.date_input(
                        "Relabel To 日期范围",
                        key="relabel_date_range",
                        min_value=min_r.date(),
                        max_value=max_r.date(),
                    )
                    r_start, r_end = st.session_state.get("relabel_date_range")

        if ss.get("__focus_target") in {"desc", "remark"}:
            ss["__focus_target"] = None

    work = _apply_filter_state(base, ss)
    if use_date_filters and "expiry_date" in work.columns and start and end:
        mask_exp = (
            work["expiry_date"].notna()
            & (work["expiry_date"].dt.date >= start)
            & (work["expiry_date"].dt.date <= end)
        )
        work = work[mask_exp]
    if use_date_filters and "relabel_to_date" in work.columns and r_start and r_end:
        mask_rel = (
            work["relabel_to_date"].notna()
            & (work["relabel_to_date"].dt.date >= r_start)
            & (work["relabel_to_date"].dt.date <= r_end)
        )
        work = work[mask_rel]

    selections = {
        "warehouse": list(sel_wh),
        "supplier": list(sel_sup),
        "brand": list(sel_brand),
        "description": list(sel_desc),
        "product_code": list(sel_code),
        "remark": list(sel_remark),
        "use_date_filters": use_date_filters,
        "expiry_range": (start, end) if use_date_filters else None,
        "relabel_range": (r_start, r_end) if use_date_filters else None,
    }
    return work, sel_desc, st.session_state.get("summary_expiry_days", 30), selections


# ----------------------------- Excel 读入与规范化 -----------------------------


def _find_sheet_name_in_list(sheet_names: List[str], desired: str) -> Optional[str]:
    names = sheet_names or []
    for n in names:
        if n == desired:
            return n
    for n in names:
        if n.lower() == desired.lower():
            return n
    d_norm = desired.replace(" ", "").lower()
    for n in names:
        if n.replace(" ", "").lower() == d_norm:
            return n
    return None


def _find_sheet_name(file_like, desired: str) -> Optional[str]:
    try:
        xls = pd.ExcelFile(file_like, engine="openpyxl")
    except Exception:
        return None
    return _find_sheet_name_in_list(xls.sheet_names, desired)


@st.cache_data(show_spinner=False)
def load_and_normalize_cached(file_bytes: bytes) -> Tuple[pd.DataFrame, list]:
    warns = []
    if not file_bytes:
        return pd.DataFrame(), ["空文件内容，无法读取。"]

    try:
        xls = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
    except Exception as e:
        return pd.DataFrame(), [f"读取 Excel 失败：{e}"]

    name_sr = _find_sheet_name_in_list(xls.sheet_names, "Stocks report")
    name_lh = _find_sheet_name_in_list(xls.sheet_names, "Lai Hock Whse")

    df_sr = None
    df_lh = None
    try:
        if name_sr:
            df_sr = xls.parse(name_sr, header=2, dtype=str).dropna(axis=0, how="all")
    except Exception as e:
        warns.append(f"读取工作表 '{name_sr}' 失败：{e}")
    try:
        if name_lh:
            df_lh = xls.parse(name_lh, header=2, dtype=str).dropna(axis=0, how="all")
    except Exception as e:
        warns.append(f"读取工作表 '{name_lh}' 失败：{e}")

    frames = []
    if df_sr is not None:
        n_sr, w_sr = _normalize_stocks_report(df_sr)
        if w_sr:
            warns.append(f"Stocks report: {w_sr}")
        frames.append(n_sr)
    else:
        warns.append(
            f"未找到工作表：Stocks report（实际存在：{name_sr if name_sr else '无'}）"
        )

    if df_lh is not None:
        n_lh, w_lh = _normalize_lai_hock_whse(df_lh)
        if w_lh:
            warns.append(f"Lai Hock Whse: {w_lh}")
        frames.append(n_lh)
    else:
        warns.append(
            f"未找到工作表：Lai Hock Whse（实际存在：{name_lh if name_lh else '无'}）"
        )

    if frames:
        cols = [
            "supplier",
            "brand",
            "product_code",
            "description",
            "pack_size",
            "unit",
            "expiry_date",
            "relabel_to_date",
            "stock_qty",
            "warehouse",
        ]
        valid_frames = [f for f in frames if f is not None and not f.empty]
        combined = (
            pd.concat(valid_frames, ignore_index=True, sort=False)
            if valid_frames
            else pd.DataFrame(columns=cols)
        )
        for c in cols:
            if c not in combined.columns:
                combined[c] = pd.NA
        combined["stock_qty"] = pd.to_numeric(combined["stock_qty"], errors="coerce")
        for col in ["expiry_date", "relabel_to_date"]:
            combined[col] = pd.to_datetime(
                combined[col], errors="coerce", format="mixed"
            )
        if "unit" in combined.columns:
            combined["unit"] = combined["unit"].apply(_canonical_unit)
        combined = combined[cols]
        return combined, warns

    return pd.DataFrame(), warns


def _normalize_stocks_report(df: pd.DataFrame) -> Tuple[pd.DataFrame, Optional[str]]:
    expected_map = {
        0: "supplier",
        1: "brand",
        2: "product_code",
        3: "description",
        4: "pack_size",
        5: "unit",
        6: "expiry_date",
        7: "relabel_to_date",
        8: "daily_update_stock_i",
        9: "stock_qty",
    }
    warning = None
    pos_renames = {}
    for idx, name in expected_map.items():
        if idx < df.shape[1]:
            pos_renames[df.columns[idx]] = name

    norm = df.rename(columns=pos_renames).copy()
    norm["warehouse"] = "Savori Whse"

    required = [
        "supplier",
        "brand",
        "product_code",
        "description",
        "pack_size",
        "unit",
        "expiry_date",
        "relabel_to_date",
        "stock_qty",
    ]
    missing = [c for c in required if c not in norm.columns]
    if missing:
        warning = (
            "根据预期位置缺少列，或表头不在第 3 行："
            + ", ".join(missing)
            + "。请确保 Excel 的 'Stocks report' 工作表表头位于 A3:J3。"
        )

    for col in ["expiry_date", "relabel_to_date"]:
        if col in norm.columns:
            norm[col] = pd.to_datetime(norm[col], errors="coerce", format="mixed")

    if "stock_qty" in norm.columns:
        norm["stock_qty"] = pd.to_numeric(
            norm["stock_qty"].astype(str).str.replace(",", "", regex=False),
            errors="coerce",
        )

    for col in [
        "supplier",
        "brand",
        "product_code",
        "description",
        "pack_size",
        "unit",
    ]:
        if col in norm.columns:
            norm[col] = norm[col].astype("string").str.strip()

    return norm, warning


def _normalize_lai_hock_whse(df: pd.DataFrame) -> Tuple[pd.DataFrame, Optional[str]]:
    col_texts = [str(c).strip().lower() for c in df.columns]

    idx_stock_qty = next(
        (
            idx
            for idx, text in enumerate(col_texts)
            if ("stock" in text) and (("qty" in text) or ("quantity" in text))
        ),
        None,
    )
    if idx_stock_qty is None:
        if df.shape[1] > 10:
            idx_stock_qty = 10
        elif df.shape[1] > 9:
            idx_stock_qty = 9

    idx_stocks_balance = next(
        (
            idx
            for idx, text in enumerate(col_texts)
            if ("stock" in text) and ("balance" in text)
        ),
        None,
    )
    if (
        idx_stocks_balance is None
        and df.shape[1] > 9
        and (idx_stock_qty is None or idx_stock_qty != 9)
    ):
        idx_stocks_balance = 9

    expected_map = {
        0: "supplier",
        1: "brand",
        2: "product_code",
        3: "description",
        5: "pack_size",
        6: "unit",
        7: "expiry_date",
        8: "relabel_to_date",
    }
    if idx_stocks_balance is not None:
        expected_map[idx_stocks_balance] = "stocks_balance"
    if idx_stock_qty is not None:
        expected_map[idx_stock_qty] = "stock_qty"

    warn_msgs = []
    pos_renames = {}
    for idx, name in expected_map.items():
        if idx < df.shape[1]:
            pos_renames[df.columns[idx]] = name

    norm = df.rename(columns=pos_renames).copy()
    norm["warehouse"] = "Lai Hock Whse"

    required = [
        "supplier",
        "brand",
        "product_code",
        "description",
        "pack_size",
        "unit",
        "expiry_date",
        "relabel_to_date",
        "stock_qty",
    ]
    missing = [c for c in required if c not in norm.columns]
    if missing:
        warn_msgs.append(
            "根据预期位置缺少列，或表头不在第 3 行："
            + ", ".join(missing)
            + "。请确保 Excel 的 'Lai Hock Whse' 工作表表头位于 A3:K3。"
        )

    for col in ["expiry_date", "relabel_to_date"]:
        if col in norm.columns:
            norm[col] = pd.to_datetime(norm[col], errors="coerce", format="mixed")

    if "stock_qty" in norm.columns:
        norm["stock_qty"] = pd.to_numeric(
            norm["stock_qty"].astype(str).str.replace(",", "", regex=False),
            errors="coerce",
        )

    if "unit" not in norm.columns:
        norm["unit"] = "CTN"
        warn_msgs.append("缺少 Unit 列，已将缺失处默认填为 CTN")
    else:
        norm["unit"] = norm["unit"].astype("string").str.strip()
        norm.loc[norm["unit"].isna() | (norm["unit"] == ""), "unit"] = "CTN"

    for col in ["supplier", "brand", "product_code", "description", "pack_size"]:
        if col in norm.columns:
            norm[col] = norm[col].astype("string").str.strip()

    warning = "；".join(warn_msgs) if warn_msgs else None
    return norm, warning


def load_and_normalize(file) -> Tuple[pd.DataFrame, list]:
    if file is None:
        return pd.DataFrame(), ["未提供文件。"]
    try:
        file_bytes = file.getvalue() if hasattr(file, "getvalue") else file.read()
    except Exception as e:
        return pd.DataFrame(), [f"读取文件内容失败：{e}"]
    return load_and_normalize_cached(file_bytes)


def _normalize_reconciliation_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _normalize_reconciliation_description(value: Any) -> str:
    text = _normalize_reconciliation_text(value)
    if not text:
        return ""
    return re.sub(PAREN_CONTENT_RE, "", text).strip()


def _normalize_reconciliation_product_code(value: Any) -> str:
    text = _normalize_reconciliation_description(value)
    if not text:
        return ""
    if text.strip().lower() in {"na", "n/a", "nan", "none", "-"}:
        return ""
    return re.sub(r"\s+", " ", text).strip()


def _simplify_reconciliation_description(value: Any) -> str:
    text = _normalize_reconciliation_description(value).lower()
    if not text:
        return ""
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


_RECONCILIATION_SUPPLIER_NOISE_TOKENS = {
    "food",
    "foods",
    "pte",
    "ltd",
    "trading",
    "trade",
    "co",
    "company",
}

_RECONCILIATION_DESCRIPTION_CONTEXT_TOKENS = {
    "savori",
    "saporito",
    "sapporito",
}


def _reconciliation_tokens(value: Any) -> List[str]:
    text = _simplify_reconciliation_description(value)
    if not text:
        return []
    return [token for token in text.split() if token]


def _normalize_reconciliation_supplier_tokens(value: Any) -> List[str]:
    return [
        token
        for token in _reconciliation_tokens(value)
        if token not in _RECONCILIATION_SUPPLIER_NOISE_TOKENS
    ]


def _are_reconciliation_suppliers_equivalent(left: Any, right: Any) -> bool:
    left_tokens = _normalize_reconciliation_supplier_tokens(left)
    right_tokens = _normalize_reconciliation_supplier_tokens(right)
    if not left_tokens or not right_tokens:
        return False

    left_joined = "".join(left_tokens)
    right_joined = "".join(right_tokens)
    if left_joined == right_joined:
        return True

    left_set = set(left_tokens)
    right_set = set(right_tokens)
    return left_set.issubset(right_set) or right_set.issubset(left_set)


def _normalize_reconciliation_pack_text(value: Any) -> str:
    text = _normalize_reconciliation_text(value)
    if not text:
        return ""
    normalized = re.sub(r"\s+", " ", text.strip().lower().replace("×", "x"))
    normalized = normalized.replace("kgs", "kg").replace("kilograms", "kg").replace(
        "kilogram", "kg"
    )
    normalized = normalized.replace("grams", "g").replace("gram", "g")
    normalized = normalized.replace("litres", "l").replace("litre", "l")
    normalized = normalized.replace("ltrs", "l").replace("ltr", "l")
    normalized = normalized.replace("liters", "l").replace("liter", "l")
    normalized = normalized.replace("kgs", "kg")
    normalized = normalized.replace("millilitres", "ml").replace("millilitre", "ml")
    normalized = normalized.replace("milliliters", "ml").replace("milliliter", "ml")
    return normalized


def _parse_reconciliation_pack_components(
    value: Any,
) -> Optional[Tuple[float, float, str]]:
    text = _normalize_reconciliation_pack_text(value)
    if not text:
        return None
    if text in {"na", "n/a", "nan", "none", "-"}:
        return None

    match = re.match(
        r"^(\d+(?:\.\d+)?)\s*[x*]\s*(\d+(?:\.\d+)?)\s*([a-z]+)$",
        text.replace(" ", ""),
    )
    if not match:
        return None

    unit = match.group(3)
    unit_aliases = {
        "kgs": "kg",
        "kg": "kg",
        "grams": "g",
        "gram": "g",
        "g": "g",
        "litre": "l",
        "litres": "l",
        "liter": "l",
        "liters": "l",
        "ltr": "l",
        "ltrs": "l",
        "l": "l",
        "millilitre": "ml",
        "millilitres": "ml",
        "milliliter": "ml",
        "milliliters": "ml",
        "ml": "ml",
    }
    canonical_unit = unit_aliases.get(unit, unit)

    try:
        return float(match.group(1)), float(match.group(2)), canonical_unit
    except ValueError:
        return None


def _are_reconciliation_pack_sizes_close(left: Any, right: Any) -> bool:
    left_text = _normalize_reconciliation_pack_text(left)
    right_text = _normalize_reconciliation_pack_text(right)
    if not left_text or not right_text:
        return False
    if left_text == right_text:
        return True

    left_components = _parse_reconciliation_pack_components(left)
    right_components = _parse_reconciliation_pack_components(right)
    if left_components is None or right_components is None:
        return False

    left_multiplier, left_amount, left_unit = left_components
    right_multiplier, right_amount, right_unit = right_components
    if not math.isclose(left_multiplier, right_multiplier, rel_tol=0.0, abs_tol=1e-9):
        return False
    if left_unit != right_unit:
        return False

    tolerance = max(0.1, max(abs(left_amount), abs(right_amount)) * 0.03)
    return abs(left_amount - right_amount) <= tolerance


def _core_reconciliation_description_tokens(value: Any) -> List[str]:
    return [
        token
        for token in _reconciliation_tokens(value)
        if token not in _RECONCILIATION_DESCRIPTION_CONTEXT_TOKENS
    ]


def _are_reconciliation_core_descriptions_equivalent(left: Any, right: Any) -> bool:
    left_tokens = _core_reconciliation_description_tokens(left)
    right_tokens = _core_reconciliation_description_tokens(right)
    if len(left_tokens) < 2 or len(right_tokens) < 2:
        return False

    left_set = set(left_tokens)
    right_set = set(right_tokens)
    if len(left_set & right_set) < 2:
        return False
    return left_set == right_set or left_set.issubset(right_set) or right_set.issubset(
        left_set
    )


def _is_high_confidence_reconciliation_same_item(
    stock_supplier: Any,
    stock_description: Any,
    stock_pack_size: Any,
    sales_supplier: Any,
    sales_description: Any,
    sales_pack_size: Any,
) -> bool:
    return (
        _are_reconciliation_suppliers_equivalent(stock_supplier, sales_supplier)
        and _are_reconciliation_pack_sizes_close(stock_pack_size, sales_pack_size)
        and _are_reconciliation_core_descriptions_equivalent(
            stock_description, sales_description
        )
    )


def _is_possible_reconciliation_description_match(left: Any, right: Any) -> bool:
    left_text = _simplify_reconciliation_description(left)
    right_text = _simplify_reconciliation_description(right)
    if not left_text or not right_text or left_text == right_text:
        return False

    left_tokens = left_text.split()
    right_tokens = right_text.split()
    if len(left_tokens) < 2 or len(right_tokens) < 2:
        return False

    shorter, longer = (
        (left_text, right_text)
        if len(left_text) <= len(right_text)
        else (right_text, left_text)
    )
    if shorter in longer:
        return True

    left_set = set(left_tokens)
    right_set = set(right_tokens)
    overlap = len(left_set & right_set)
    if overlap >= max(2, min(len(left_set), len(right_set))):
        return True

    return SequenceMatcher(None, left_text, right_text).ratio() >= 0.72


def _parse_stock_outbound_cell(raw_value: Any, unit_value: Any) -> Tuple[int, int]:
    text = _normalize_reconciliation_text(raw_value)
    if not text:
        return 0, 0
    parsed_ctn, parsed_pkt = _parse_qty_text(text)
    if parsed_ctn != 0 or parsed_pkt != 0:
        return parsed_ctn, parsed_pkt

    try:
        numeric_value = float(text.replace(",", ""))
    except ValueError:
        return 0, 0

    qty = int(round(numeric_value))
    if qty == 0:
        return 0, 0

    canonical_unit = _canonical_unit(unit_value)
    if canonical_unit == "ctn":
        return qty, 0
    return 0, qty


def _normalize_ctn_pkt_by_pack_size(
    ctn_value: Any, pkt_value: Any, pack_size_value: Any
) -> Tuple[int, int]:
    cartons = _coerce_qty_int(ctn_value)
    packets = _coerce_qty_int(pkt_value)
    pack_size = _coerce_qty_int(_infer_carton_pack_size(pack_size_value))
    if pack_size <= 0:
        return cartons, packets
    if packets >= 0:
        cartons += packets // pack_size
        packets = packets % pack_size
        return cartons, packets
    borrowed_cartons = math.ceil(abs(packets) / pack_size)
    cartons -= borrowed_cartons
    packets += borrowed_cartons * pack_size
    return cartons, packets


def _build_reconciliation_match_key(
    product_code: Any,
    description: Any,
    pack_size: Any,
    supplier: Any = None,
    brand: Any = None,
) -> Tuple[str, str]:
    code = _normalize_reconciliation_product_code(product_code)
    if code:
        return f"code::{code.lower()}", "product_code"
    supplier_text = _normalize_reconciliation_text(supplier).lower()
    desc = _normalize_reconciliation_description(description).lower()
    pack = _normalize_pack_size_value(pack_size).lower()
    if supplier_text and desc and pack:
        return (
            f"descpack::{supplier_text}::{desc}::{pack}",
            "fallback_supplier_description_pack",
        )
    return "", "unmatched"


def _resolve_reconciliation_year(sales_df: Optional[pd.DataFrame]) -> Optional[int]:
    if sales_df is None or sales_df.empty or "Date" not in sales_df.columns:
        return None
    dates = pd.to_datetime(sales_df["Date"], errors="coerce").dropna()
    if dates.empty:
        return None
    years = sorted({int(value.year) for value in dates.tolist()})
    if len(years) == 1:
        return years[0]
    return years[-1]


def _default_reconciliation_year_selection(options: List[str]) -> List[str]:
    if "2026" in options:
        return ["2026"]
    if options:
        return [options[-1]]
    return []


def _default_reconciliation_month_selection(options: List[str]) -> List[str]:
    if "Apr" in options:
        return ["Apr"]
    if options:
        return [options[-1]]
    return []


def _get_reconciliation_date_bounds(
    df: pd.DataFrame, filter_state: Dict[str, Any]
) -> Tuple[Optional[datetime.date], Optional[datetime.date]]:
    selected_years = [
        str(value).strip()
        for value in filter_state.get("year", [])
        if str(value).strip()
    ]
    selected_months = [
        str(value).strip()
        for value in filter_state.get("month", [])
        if str(value).strip()
    ]

    if len(selected_years) == 1 and len(selected_months) == 1:
        year_value = selected_years[0]
        month_value = selected_months[0]
        if year_value.isdigit():
            month_number = _month_sort_key(month_value)
            if 1 <= month_number <= 12:
                start_date = datetime.date(int(year_value), month_number, 1)
                end_day = calendar.monthrange(int(year_value), month_number)[1]
                end_date = datetime.date(int(year_value), month_number, end_day)
                return start_date, end_date

    if df is None or df.empty:
        return None, None
    narrowed = _apply_reconciliation_filters(
        df,
        {
            "product_code": filter_state.get("product_code", []),
            "description": filter_state.get("description", []),
            "year": filter_state.get("year", []),
            "month": filter_state.get("month", []),
            "date": None,
        },
    )
    if narrowed.empty:
        return None, None
    date_series = pd.to_datetime(narrowed.get("date"), errors="coerce").dropna()
    if date_series.empty:
        return None, None
    return date_series.min().date(), date_series.max().date()


def _format_reconciliation_date_label(value: Any) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return ""
    return pd.Timestamp(parsed).strftime("%d-%b-%Y")


def _coerce_reconciliation_date_input_value(
    value: Any,
) -> Optional[datetime.date | Tuple[datetime.date, datetime.date]]:
    if value in (None, "", []):
        return None
    if isinstance(value, tuple):
        coerced_values: List[datetime.date] = []
        for item in value:
            if item in (None, ""):
                continue
            coerced_item = _coerce_reconciliation_date_input_value(item)
            if isinstance(coerced_item, tuple):
                coerced_values.extend(list(coerced_item))
            elif coerced_item is not None:
                coerced_values.append(coerced_item)
        coerced_values = [item for item in coerced_values if item is not None]
        if len(coerced_values) >= 2:
            return (coerced_values[0], coerced_values[1])
        return coerced_values[0] if coerced_values else None
    if isinstance(value, list):
        coerced_values: List[datetime.date] = []
        for item in value:
            if item in (None, ""):
                continue
            coerced_item = _coerce_reconciliation_date_input_value(item)
            if isinstance(coerced_item, tuple):
                coerced_values.extend(list(coerced_item))
            elif coerced_item is not None:
                coerced_values.append(coerced_item)
        coerced_values = [item for item in coerced_values if item is not None]
        if len(coerced_values) >= 2:
            return (coerced_values[0], coerced_values[1])
        return coerced_values[0] if coerced_values else None

    parsed = pd.to_datetime(value, errors="coerce", format="mixed")
    if pd.isna(parsed):
        return None
    return pd.Timestamp(parsed).date()


def _build_reconciliation_filter_options(df: pd.DataFrame) -> Dict[str, List[str]]:
    if df is None or df.empty:
        return {
            "product_code": [],
            "description": [],
            "year": [],
            "month": [],
            "date": [],
        }

    product_code_options = sorted(
        {
            _normalize_reconciliation_text(value)
            for value in df.get("product_code", pd.Series(dtype="object")).tolist()
            if _normalize_reconciliation_text(value)
        }
    )
    description_options = sorted(
        {
            _normalize_reconciliation_description(value)
            for value in df.get("description", pd.Series(dtype="object")).tolist()
            if _normalize_reconciliation_description(value)
        }
    )
    date_series = pd.to_datetime(
        df.get("date", pd.Series(dtype="object")), errors="coerce"
    )
    year_options = sorted(
        {str(int(value.year)) for value in date_series.dropna().tolist()}
    )
    month_options = sorted(
        {pd.Timestamp(value).strftime("%b") for value in date_series.dropna().tolist()},
        key=_month_sort_key,
    )
    date_options = sorted(
        {
            _format_reconciliation_date_label(value)
            for value in date_series.dropna().tolist()
            if _format_reconciliation_date_label(value)
        }
    )
    return {
        "product_code": product_code_options,
        "description": description_options,
        "year": year_options,
        "month": month_options,
        "date": date_options,
    }


def _apply_reconciliation_filters(
    df: pd.DataFrame, filter_state: Dict[str, Any]
) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=list(df.columns) if df is not None else [])

    work = df.copy()
    date_series = pd.to_datetime(
        work.get("date", pd.Series(index=work.index)), errors="coerce"
    )

    selected_codes = [
        _normalize_reconciliation_text(value)
        for value in filter_state.get("product_code", [])
        if _normalize_reconciliation_text(value)
    ]
    if selected_codes:
        code_series = work.get(
            "product_code", pd.Series(index=work.index, dtype="object")
        ).apply(_normalize_reconciliation_text)
        work = work[code_series.isin(selected_codes)].copy()
        date_series = pd.to_datetime(
            work.get("date", pd.Series(index=work.index)), errors="coerce"
        )

    selected_descriptions = [
        _normalize_reconciliation_description(value)
        for value in filter_state.get("description", [])
        if _normalize_reconciliation_description(value)
    ]
    if selected_descriptions:
        description_series = work.get(
            "description", pd.Series(index=work.index, dtype="object")
        ).apply(_normalize_reconciliation_description)
        work = work[description_series.isin(selected_descriptions)].copy()
        date_series = pd.to_datetime(
            work.get("date", pd.Series(index=work.index)), errors="coerce"
        )

    use_specific_date = bool(filter_state.get("use_specific_date", False))
    specific_date_value = filter_state.get("specific_date")
    specific_date = pd.to_datetime(specific_date_value, errors="coerce")
    if use_specific_date and pd.notna(specific_date):
        specific_norm = pd.Timestamp(specific_date).normalize()
        work = work[date_series == specific_norm].copy()
        return work.reset_index(drop=True)

    selected_years = [
        str(value) for value in filter_state.get("year", []) if str(value).strip()
    ]
    if selected_years:
        year_series = date_series.dt.year.astype("Int64").astype("string")
        work = work[year_series.isin(selected_years)].copy()
        date_series = pd.to_datetime(
            work.get("date", pd.Series(index=work.index)), errors="coerce"
        )

    selected_months = [
        str(value).strip()
        for value in filter_state.get("month", [])
        if str(value).strip()
    ]
    if selected_months:
        month_series = date_series.dt.strftime("%b")
        work = work[month_series.isin(selected_months)].copy()
        date_series = pd.to_datetime(
            work.get("date", pd.Series(index=work.index)), errors="coerce"
        )

    selected_date_value = filter_state.get("date")
    if isinstance(selected_date_value, tuple):
        start_raw, end_raw = (
            selected_date_value if len(selected_date_value) == 2 else (None, None)
        )
        start_date = pd.to_datetime(start_raw, errors="coerce")
        end_date = pd.to_datetime(end_raw, errors="coerce")
        if pd.notna(start_date) and pd.notna(end_date):
            start_norm = pd.Timestamp(start_date).normalize()
            end_norm = pd.Timestamp(end_date).normalize()
            work = work[(date_series >= start_norm) & (date_series <= end_norm)].copy()
    elif isinstance(selected_date_value, list):
        selected_dates = [
            pd.to_datetime(value, errors="coerce") for value in selected_date_value
        ]
        normalized_dates = {
            pd.Timestamp(value).normalize()
            for value in selected_dates
            if pd.notna(value)
        }
        if normalized_dates:
            work = work[date_series.isin(list(normalized_dates))].copy()
    else:
        selected_date = pd.to_datetime(selected_date_value, errors="coerce")
        if pd.notna(selected_date):
            selected_norm = pd.Timestamp(selected_date).normalize()
            work = work[date_series == selected_norm].copy()

    return work.reset_index(drop=True)


def _build_reconciliation_display_df(df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Supplier",
        "Date",
        "Product Code",
        "Description",
        "Pack Size",
        "Sales Data",
        "Stock Data",
        "Difference",
        "mismatch_reason",
    ]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)

    display_df = pd.DataFrame(
        {
            "Supplier": df.get(
                "supplier", pd.Series(index=df.index, dtype="object")
            ).apply(_normalize_reconciliation_text),
            "Date": pd.to_datetime(df.get("date"), errors="coerce").dt.strftime(
                "%d-%b-%Y"
            ),
            "Product Code": df.get(
                "product_code", pd.Series(index=df.index, dtype="object")
            ).apply(_normalize_reconciliation_text),
            "Description": df.get(
                "description", pd.Series(index=df.index, dtype="object")
            ).apply(_normalize_reconciliation_description),
            "Pack Size": df.get(
                "pack_size", pd.Series(index=df.index, dtype="object")
            ).apply(_normalize_pack_size_value),
            "Sales Data": df.get(
                "sales_qty_text", pd.Series(index=df.index, dtype="object")
            ).fillna("0"),
            "Stock Data": df.get(
                "stock_qty_text", pd.Series(index=df.index, dtype="object")
            ).fillna("0"),
            "Difference": df.get(
                "difference_qty_text", pd.Series(index=df.index, dtype="object")
            ).fillna("0"),
            "mismatch_reason": df.get(
                "mismatch_reason", pd.Series(index=df.index, dtype="object")
            ).fillna(""),
        }
    )
    display_df["Date"] = display_df["Date"].fillna("")
    return display_df


def _parse_stock_header_date(
    value: Any, preferred_year: Optional[int]
) -> Optional[pd.Timestamp]:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, datetime.datetime):
        return pd.Timestamp(value.date())
    if isinstance(value, datetime.date):
        return pd.Timestamp(value)

    text = str(value).strip()
    if not text:
        return None

    yearless_patterns = [
        r"^\d{1,2}[-/\s][A-Za-z]{3,9}$",
        r"^[A-Za-z]{3,9}[-/\s]\d{1,2}$",
        r"^\d{1,2}[-/]\d{1,2}$",
    ]
    full_patterns = [
        r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$",
        r"^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$",
        r"^\d{1,2}[-/\s][A-Za-z]{3,9}[-/\s]\d{2,4}$",
        r"^[A-Za-z]{3,9}[-/\s]\d{1,2}[-/\s]\d{2,4}$",
    ]
    if any(re.fullmatch(pattern, text) for pattern in yearless_patterns):
        target_year = preferred_year or datetime.date.today().year
        parsed = pd.to_datetime(
            f"{text}-{target_year}",
            errors="coerce",
            format="mixed",
        )
        if pd.isna(parsed):
            return None
        return pd.Timestamp(parsed).normalize()
    if any(re.fullmatch(pattern, text) for pattern in full_patterns):
        parsed = pd.to_datetime(text, errors="coerce", format="mixed")
        if pd.isna(parsed):
            return None
        return pd.Timestamp(parsed).normalize()
    return None


def _first_non_empty_text(series: pd.Series) -> str:
    for value in series.tolist():
        text = _normalize_reconciliation_text(value)
        if text:
            return text
    return ""


def _aggregate_reconciliation_rows(
    df: pd.DataFrame,
    *,
    qty_ctn_col: str,
    qty_pkt_col: str,
    qty_text_col: str,
) -> pd.DataFrame:
    base_columns = [
        "match_key",
        "match_basis",
        "product_code",
        "description",
        "pack_size",
        "supplier",
        "brand",
        "date",
        qty_ctn_col,
        qty_pkt_col,
        qty_text_col,
    ]
    if df is None or df.empty:
        return pd.DataFrame(columns=base_columns)

    work = df.copy()
    for field in [
        "match_basis",
        "product_code",
        "description",
        "pack_size",
        "supplier",
        "brand",
    ]:
        if field not in work.columns:
            work[field] = ""
    work["match_key"] = work.get(
        "match_key", pd.Series(index=work.index, dtype="object")
    ).fillna("")
    work = work[work["match_key"].astype(str).str.strip() != ""].copy()
    if work.empty:
        return pd.DataFrame(columns=base_columns)

    work["date"] = pd.to_datetime(work.get("date"), errors="coerce").dt.normalize()
    work = work.dropna(subset=["date"]).copy()
    if work.empty:
        return pd.DataFrame(columns=base_columns)

    work[qty_ctn_col] = pd.to_numeric(work.get(qty_ctn_col), errors="coerce").fillna(0)
    work[qty_pkt_col] = pd.to_numeric(work.get(qty_pkt_col), errors="coerce").fillna(0)
    grouped = (
        work.groupby(["match_key", "date"], as_index=False, dropna=False)
        .agg(
            {
                "match_basis": _first_non_empty_text,
                "product_code": _first_non_empty_text,
                "description": _first_non_empty_text,
                "pack_size": _first_non_empty_text,
                "supplier": _first_non_empty_text,
                "brand": _first_non_empty_text,
                qty_ctn_col: "sum",
                qty_pkt_col: "sum",
            }
        )
        .copy()
    )
    normalized_qty = grouped.apply(
        lambda row: _normalize_ctn_pkt_by_pack_size(
            row[qty_ctn_col], row[qty_pkt_col], row.get("pack_size", "")
        ),
        axis=1,
        result_type="expand",
    )
    grouped[qty_ctn_col] = normalized_qty[0].astype(int)
    grouped[qty_pkt_col] = normalized_qty[1].astype(int)
    grouped[qty_text_col] = grouped.apply(
        lambda row: _format_ctn_pkt_total(row[qty_ctn_col], row[qty_pkt_col]),
        axis=1,
    )
    return grouped


def _select_outbound_date_columns(
    candidates: List[Tuple[int, pd.Timestamp]],
) -> List[Tuple[int, pd.Timestamp]]:
    grouped: Dict[pd.Timestamp, List[Tuple[int, pd.Timestamp]]] = {}
    ordered_dates: List[pd.Timestamp] = []
    for col_idx, parsed_date in candidates:
        date_key = pd.Timestamp(parsed_date).normalize()
        if date_key not in grouped:
            grouped[date_key] = []
            ordered_dates.append(date_key)
        grouped[date_key].append((col_idx, parsed_date))

    selected: List[Tuple[int, pd.Timestamp]] = []
    for date_key in ordered_dates:
        occurrences = grouped[date_key]
        if len(occurrences) >= 2:
            selected.append(occurrences[1])
        else:
            selected.append(occurrences[0])
    return selected


def _normalize_reconciliation_stock_rows(ws) -> Tuple[pd.DataFrame, Optional[str]]:
    rows: List[Dict[str, Any]] = []
    carry_fields = {
        "supplier": "",
        "brand": "",
        "product_code": "",
        "description": "",
        "pack_size": "",
        "unit": "",
    }
    for row_idx in range(4, ws.max_row + 1):
        raw = [ws.cell(row=row_idx, column=idx).value for idx in range(1, 11)]
        if not any(value not in (None, "") for value in raw):
            continue

        current = {
            "supplier": _normalize_reconciliation_text(raw[0]),
            "brand": _normalize_reconciliation_text(raw[1]),
            "product_code": _normalize_reconciliation_text(raw[2]),
            "description": _normalize_reconciliation_text(raw[3]),
            "pack_size": _normalize_reconciliation_text(raw[4]),
            "unit": _normalize_reconciliation_text(raw[5]),
            "expiry_date": raw[6],
            "relabel_to_date": raw[7],
            "daily_update_stock_i": raw[8],
            "stock_qty": raw[9],
        }
        for field in carry_fields:
            if current[field]:
                carry_fields[field] = current[field]
            else:
                current[field] = carry_fields[field]

        if not any(
            current[field]
            for field in [
                "supplier",
                "brand",
                "product_code",
                "description",
                "pack_size",
            ]
        ):
            continue
        rows.append(current)

    norm = pd.DataFrame(rows)
    if norm.empty:
        return _normalize_stocks_report(pd.DataFrame(columns=list(range(10))))

    norm["warehouse"] = "Savori Whse"
    for col in ["expiry_date", "relabel_to_date"]:
        if col in norm.columns:
            norm[col] = pd.to_datetime(norm[col], errors="coerce", format="mixed")
    if "stock_qty" in norm.columns:
        norm["stock_qty"] = pd.to_numeric(
            norm["stock_qty"].astype(str).str.replace(",", "", regex=False),
            errors="coerce",
        )
    for col in [
        "supplier",
        "brand",
        "product_code",
        "description",
        "pack_size",
        "unit",
    ]:
        if col in norm.columns:
            norm[col] = norm[col].astype("string").str.strip()

    required = [
        "supplier",
        "brand",
        "product_code",
        "description",
        "pack_size",
        "unit",
        "expiry_date",
        "relabel_to_date",
        "stock_qty",
        "warehouse",
    ]
    for col in required:
        if col not in norm.columns:
            norm[col] = pd.NA
    return norm[required], None


def _extract_stock_outbound_timeline_from_sheet(
    ws,
    *,
    preferred_year: Optional[int],
) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    date_candidates: List[Tuple[int, pd.Timestamp]] = []
    header_rows = [2, 3]
    carry_fields = {
        "supplier": "",
        "brand": "",
        "product_code": "",
        "description": "",
        "pack_size": "",
        "unit": "",
    }
    data_start_row = 4
    for header_row in header_rows:
        for col_idx in range(11, ws.max_column + 1):
            parsed_date = _parse_stock_header_date(
                ws.cell(row=header_row, column=col_idx).value, preferred_year
            )
            if parsed_date is not None:
                date_candidates.append((col_idx, parsed_date))
                data_start_row = max(data_start_row, header_row + 1)

    date_columns = _select_outbound_date_columns(date_candidates)

    if not date_columns:
        return pd.DataFrame(
            columns=[
                "match_key",
                "match_basis",
                "product_code",
                "description",
                "pack_size",
                "supplier",
                "brand",
                "date",
                "stock_ctn",
                "stock_pkt",
                "stock_qty_text",
            ]
        )

    for row_idx in range(data_start_row, ws.max_row + 1):
        supplier = _normalize_reconciliation_text(ws.cell(row=row_idx, column=1).value)
        brand = _normalize_reconciliation_text(ws.cell(row=row_idx, column=2).value)
        product_code = _normalize_reconciliation_text(
            ws.cell(row=row_idx, column=3).value
        )
        description = _normalize_reconciliation_text(
            ws.cell(row=row_idx, column=4).value
        )
        pack_size = _normalize_reconciliation_text(ws.cell(row=row_idx, column=5).value)
        unit_value = _normalize_reconciliation_text(
            ws.cell(row=row_idx, column=6).value
        )
        current_values = {
            "supplier": supplier,
            "brand": brand,
            "product_code": product_code,
            "description": description,
            "pack_size": pack_size,
            "unit": unit_value,
        }
        for field in carry_fields:
            if current_values[field]:
                carry_fields[field] = current_values[field]
            else:
                current_values[field] = carry_fields[field]
        supplier = current_values["supplier"]
        brand = current_values["brand"]
        product_code = current_values["product_code"]
        description = current_values["description"]
        pack_size = current_values["pack_size"]
        unit_value = current_values["unit"]
        if not any(
            _normalize_reconciliation_text(value)
            for value in [supplier, brand, product_code, description, pack_size]
        ):
            continue
        match_key, match_basis = _build_reconciliation_match_key(
            product_code,
            description,
            pack_size,
            supplier,
            brand,
        )
        for col_idx, parsed_date in date_columns:
            raw_qty = ws.cell(row=row_idx, column=col_idx).value
            qty_text = _normalize_reconciliation_text(raw_qty)
            if not qty_text:
                continue
            stock_ctn, stock_pkt = _parse_stock_outbound_cell(raw_qty, unit_value)
            if stock_ctn == 0 and stock_pkt == 0:
                continue
            rows.append(
                {
                    "match_key": match_key,
                    "match_basis": match_basis,
                    "product_code": _normalize_reconciliation_text(product_code),
                    "description": _normalize_reconciliation_description(description),
                    "pack_size": _normalize_pack_size_value(pack_size),
                    "supplier": _normalize_reconciliation_text(supplier),
                    "brand": _normalize_reconciliation_text(brand),
                    "date": parsed_date,
                    "stock_ctn": int(stock_ctn),
                    "stock_pkt": int(stock_pkt),
                    "stock_qty_text": _format_ctn_pkt_total(stock_ctn, stock_pkt),
                }
            )

    return _aggregate_reconciliation_rows(
        pd.DataFrame(rows),
        qty_ctn_col="stock_ctn",
        qty_pkt_col="stock_pkt",
        qty_text_col="stock_qty_text",
    )


def load_stock_reconciliation_data_from_bytes(
    file_bytes: bytes,
    *,
    preferred_year: Optional[int] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    warns: List[str] = []
    if not file_bytes:
        return pd.DataFrame(), pd.DataFrame(), ["空文件内容，无法读取。"]

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        return pd.DataFrame(), pd.DataFrame(), [f"读取 Excel 失败：{exc}"]

    sheet_name = _find_sheet_name_in_list(workbook.sheetnames, "Stocks report")
    if not sheet_name:
        return pd.DataFrame(), pd.DataFrame(), ["未找到工作表：Stocks report"]

    ws = workbook[sheet_name]
    stock_df, warning = _normalize_reconciliation_stock_rows(ws)
    if warning:
        warns.append(f"Stocks report: {warning}")
    timeline_df = _extract_stock_outbound_timeline_from_sheet(
        ws,
        preferred_year=preferred_year,
    )
    if timeline_df.empty:
        warns.append("Stocks report: 未在第 2 行右侧日期区域读取到任何出货数量。")
    return stock_df, timeline_df, warns


def build_sales_reconciliation_timeline(df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "match_key",
        "match_basis",
        "product_code",
        "description",
        "pack_size",
        "supplier",
        "brand",
        "date",
        "sales_ctn",
        "sales_pkt",
        "sales_qty_text",
    ]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)

    work = df.copy()
    work["Date"] = pd.to_datetime(work.get("Date"), errors="coerce", format="mixed")
    work = work.dropna(subset=["Date"]).copy()
    if work.empty:
        return pd.DataFrame(columns=columns)

    work["product_code"] = work.get("Product Code", pd.Series(dtype="string")).apply(
        _normalize_reconciliation_text
    )
    work["description"] = work.get(
        "Product Description", pd.Series(dtype="string")
    ).apply(_normalize_reconciliation_description)
    work["pack_size"] = work.get("Carton Packing", pd.Series(dtype="string")).apply(
        _normalize_pack_size_value
    )
    work["supplier"] = work.get("Supplier", pd.Series(dtype="string")).apply(
        _normalize_reconciliation_text
    )
    work["brand"] = work.get("Brand/Category", pd.Series(dtype="string")).apply(
        _normalize_reconciliation_text
    )
    match_data = work.apply(
        lambda row: _build_reconciliation_match_key(
            row.get("product_code"),
            row.get("description"),
            row.get("pack_size"),
            row.get("supplier"),
            row.get("brand"),
        ),
        axis=1,
        result_type="expand",
    )
    work[["match_key", "match_basis"]] = match_data
    work["date"] = work["Date"].dt.normalize()
    work["sales_ctn"] = pd.to_numeric(work.get("Qty in Ctns"), errors="coerce").fillna(
        0
    )
    work["sales_pkt"] = pd.to_numeric(work.get("Qty in Pcs"), errors="coerce").fillna(0)

    grouped = _aggregate_reconciliation_rows(
        work,
        qty_ctn_col="sales_ctn",
        qty_pkt_col="sales_pkt",
        qty_text_col="sales_qty_text",
    )
    return grouped[columns]


def build_reconciliation_result(
    stock_timeline: pd.DataFrame,
    sales_timeline: pd.DataFrame,
) -> pd.DataFrame:
    columns = [
        "date",
        "product_code",
        "description",
        "pack_size",
        "supplier",
        "brand",
        "match_basis",
        "stock_qty_text",
        "sales_qty_text",
        "stock_ctn",
        "stock_pkt",
        "sales_ctn",
        "sales_pkt",
        "difference_ctn",
        "difference_pkt",
        "difference_qty_text",
        "is_match",
        "mismatch_reason",
    ]
    stock = _aggregate_reconciliation_rows(
        stock_timeline.copy() if stock_timeline is not None else pd.DataFrame(),
        qty_ctn_col="stock_ctn",
        qty_pkt_col="stock_pkt",
        qty_text_col="stock_qty_text",
    )
    sales = _aggregate_reconciliation_rows(
        sales_timeline.copy() if sales_timeline is not None else pd.DataFrame(),
        qty_ctn_col="sales_ctn",
        qty_pkt_col="sales_pkt",
        qty_text_col="sales_qty_text",
    )
    if stock.empty and sales.empty:
        return pd.DataFrame(columns=columns)

    exact_stock = stock.copy()
    exact_sales = sales.copy()
    exact_merged = exact_stock.merge(
        sales,
        how="outer",
        on=["match_key", "date"],
        suffixes=("_stock", "_sales"),
        indicator=True,
    )

    matched_exact = exact_merged[exact_merged["_merge"] == "both"].copy()
    unmatched_stock = exact_merged[exact_merged["_merge"] == "left_only"].copy()
    unmatched_sales = exact_merged[exact_merged["_merge"] == "right_only"].copy()

    def _finalize_merged(merged: pd.DataFrame) -> pd.DataFrame:
        for field in [
            "product_code",
            "description",
            "pack_size",
            "supplier",
            "brand",
            "match_basis",
        ]:
            stock_series = merged.get(
                f"{field}_stock", pd.Series(index=merged.index, dtype="object")
            )
            sales_series = merged.get(
                f"{field}_sales", pd.Series(index=merged.index, dtype="object")
            )
            merged[field] = stock_series.combine_first(sales_series).fillna("")
        for field in ["stock_ctn", "stock_pkt", "sales_ctn", "sales_pkt"]:
            merged[field] = (
                pd.to_numeric(merged.get(field), errors="coerce")
                .fillna(0)
                .round()
                .astype(int)
            )
        merged["stock_qty_text"] = merged.get(
            "stock_qty_text", pd.Series(index=merged.index, dtype="object")
        ).fillna("0")
        merged["sales_qty_text"] = merged.get(
            "sales_qty_text", pd.Series(index=merged.index, dtype="object")
        ).fillna("0")
        merged["difference_ctn"] = merged["stock_ctn"] - merged["sales_ctn"]
        merged["difference_pkt"] = merged["stock_pkt"] - merged["sales_pkt"]
        merged["difference_qty_text"] = merged.apply(
            lambda row: _format_ctn_pkt_total(
                row["difference_ctn"], row["difference_pkt"]
            ),
            axis=1,
        )
        has_stock = (merged["stock_ctn"] != 0) | (merged["stock_pkt"] != 0)
        has_sales = (merged["sales_ctn"] != 0) | (merged["sales_pkt"] != 0)
        merged["is_match"] = (
            (merged["stock_ctn"] == merged["sales_ctn"])
            & (merged["stock_pkt"] == merged["sales_pkt"])
            & has_stock
            & has_sales
        )
        merged["mismatch_reason"] = "qty_diff"
        merged.loc[merged["is_match"], "mismatch_reason"] = "match"
        merged.loc[has_stock & ~has_sales, "mismatch_reason"] = "missing_in_sales"
        merged.loc[has_sales & ~has_stock, "mismatch_reason"] = "missing_in_stock"
        return merged

    possible_rows: List[Dict[str, Any]] = []
    used_sales_indexes: set[int] = set()
    used_stock_indexes: set[int] = set()
    if not unmatched_stock.empty and not unmatched_sales.empty:
        for stock_row in unmatched_stock.itertuples(index=True):
            best_sales_index: Optional[int] = None
            best_score = -1.0
            best_high_confidence_same_item = False
            stock_supplier = _normalize_reconciliation_text(
                getattr(stock_row, "supplier_stock", "")
            )
            stock_pack = _normalize_pack_size_value(
                getattr(stock_row, "pack_size_stock", "")
            )
            stock_desc = _normalize_reconciliation_description(
                getattr(stock_row, "description_stock", "")
            )
            stock_match_basis = _normalize_reconciliation_text(
                getattr(stock_row, "match_basis_stock", "")
            )
            stock_desc_simple = _simplify_reconciliation_description(stock_desc)
            for sales_row in unmatched_sales.itertuples(index=True):
                if sales_row.Index in used_sales_indexes:
                    continue
                if (
                    pd.Timestamp(getattr(stock_row, "date")).normalize()
                    != pd.Timestamp(getattr(sales_row, "date")).normalize()
                ):
                    continue
                sales_supplier = _normalize_reconciliation_text(
                    getattr(sales_row, "supplier_sales", "")
                )
                supplier_matches = stock_supplier == sales_supplier
                supplier_equivalent = _are_reconciliation_suppliers_equivalent(
                    stock_supplier, sales_supplier
                )
                if not (supplier_matches or supplier_equivalent):
                    continue
                sales_pack = _normalize_pack_size_value(
                    getattr(sales_row, "pack_size_sales", "")
                )
                if stock_pack == sales_pack:
                    pack_matches = True
                else:
                    pack_matches = _are_reconciliation_pack_sizes_close(
                        stock_pack, sales_pack
                    )
                    if not pack_matches:
                        stock_carton_size = _infer_carton_pack_size(stock_pack)
                        sales_carton_size = _infer_carton_pack_size(sales_pack)
                        pack_matches = (
                            stock_carton_size is not None
                            and sales_carton_size is not None
                            and math.isclose(
                                stock_carton_size,
                                sales_carton_size,
                                rel_tol=1e-9,
                                abs_tol=1e-9,
                            )
                        )
                if not pack_matches:
                    continue
                sales_desc = _normalize_reconciliation_description(
                    getattr(sales_row, "description_sales", "")
                )
                sales_match_basis = _normalize_reconciliation_text(
                    getattr(sales_row, "match_basis_sales", "")
                )
                sales_desc_simple = _simplify_reconciliation_description(sales_desc)
                descriptions_match = stock_desc_simple == sales_desc_simple
                high_confidence_same_item = _is_high_confidence_reconciliation_same_item(
                    stock_supplier,
                    stock_desc,
                    stock_pack,
                    sales_supplier,
                    sales_desc,
                    sales_pack,
                ) and (
                    stock_match_basis == sales_match_basis
                    == "fallback_supplier_description_pack"
                )
                if (
                    not high_confidence_same_item
                    and not descriptions_match
                    and not _is_possible_reconciliation_description_match(
                        stock_desc, sales_desc
                    )
                ):
                    continue
                score = SequenceMatcher(
                    None,
                    stock_desc_simple,
                    sales_desc_simple,
                ).ratio()
                if score > best_score:
                    best_score = score
                    best_sales_index = sales_row.Index
                    best_high_confidence_same_item = high_confidence_same_item

            if best_sales_index is None:
                continue

            sales_row = unmatched_sales.loc[best_sales_index]
            used_sales_indexes.add(best_sales_index)
            used_stock_indexes.add(stock_row.Index)
            possible_rows.append(
                {
                    "date": pd.to_datetime(getattr(stock_row, "date"), errors="coerce"),
                    "product_code": _normalize_reconciliation_text(
                        getattr(stock_row, "product_code_stock", "")
                    )
                    or _normalize_reconciliation_text(
                        sales_row.get("product_code_sales", "")
                    ),
                    "description": f"{stock_desc} ↔ {_normalize_reconciliation_description(sales_row.get('description_sales', ''))}",
                    "pack_size": stock_pack
                    or _normalize_pack_size_value(sales_row.get("pack_size_sales", "")),
                    "supplier": stock_supplier,
                    "brand": _normalize_reconciliation_text(
                        getattr(stock_row, "brand_stock", "")
                    )
                    or _normalize_reconciliation_text(sales_row.get("brand_sales", "")),
                    "match_basis": "possible_supplier_pack_description",
                    "_high_confidence_same_item": best_high_confidence_same_item,
                    "stock_qty_text": _normalize_reconciliation_text(
                        getattr(stock_row, "stock_qty_text", "0")
                    )
                    or "0",
                    "sales_qty_text": _normalize_reconciliation_text(
                        sales_row.get("sales_qty_text", "0")
                    )
                    or "0",
                    "stock_ctn": int(getattr(stock_row, "stock_ctn", 0) or 0),
                    "stock_pkt": int(getattr(stock_row, "stock_pkt", 0) or 0),
                    "sales_ctn": int(sales_row.get("sales_ctn", 0) or 0),
                    "sales_pkt": int(sales_row.get("sales_pkt", 0) or 0),
                }
            )

    unmatched_stock_remaining = unmatched_stock.copy()
    unmatched_sales_remaining = unmatched_sales.copy()
    if possible_rows:
        unmatched_stock_remaining = unmatched_stock_remaining.drop(
            index=list(used_stock_indexes),
            errors="ignore",
        )
        unmatched_sales_remaining = unmatched_sales_remaining.drop(
            index=list(used_sales_indexes), errors="ignore"
        )

    exact_result = (
        _finalize_merged(matched_exact.copy())
        if not matched_exact.empty
        else pd.DataFrame(columns=columns)
    )
    remaining_result = (
        _finalize_merged(
            pd.concat(
                [unmatched_stock_remaining, unmatched_sales_remaining],
                ignore_index=True,
            )
        )
        if (not unmatched_stock_remaining.empty or not unmatched_sales_remaining.empty)
        else pd.DataFrame(columns=columns)
    )
    possible_result = pd.DataFrame(possible_rows)
    if not possible_result.empty:
        high_confidence_mask = possible_result.pop(
            "_high_confidence_same_item"
        ).fillna(False)
        possible_result["difference_ctn"] = (
            possible_result["stock_ctn"] - possible_result["sales_ctn"]
        )
        possible_result["difference_pkt"] = (
            possible_result["stock_pkt"] - possible_result["sales_pkt"]
        )
        possible_result["difference_qty_text"] = possible_result.apply(
            lambda row: _format_ctn_pkt_total(
                row["difference_ctn"], row["difference_pkt"]
            ),
            axis=1,
        )
        qty_equal_mask = (
            (possible_result["stock_ctn"] == possible_result["sales_ctn"])
            & (possible_result["stock_pkt"] == possible_result["sales_pkt"])
        )
        possible_result["is_match"] = high_confidence_mask & qty_equal_mask
        possible_result["mismatch_reason"] = "possible_match"
        possible_result.loc[high_confidence_mask, "mismatch_reason"] = "qty_diff"
        possible_result.loc[possible_result["is_match"], "mismatch_reason"] = "match"
        possible_result.loc[high_confidence_mask, "match_basis"] = (
            "supplier_pack_core_description"
        )

    result = pd.concat(
        [
            exact_result[columns]
            if not exact_result.empty
            else pd.DataFrame(columns=columns),
            possible_result[columns]
            if not possible_result.empty
            else pd.DataFrame(columns=columns),
            remaining_result[columns]
            if not remaining_result.empty
            else pd.DataFrame(columns=columns),
        ],
        ignore_index=True,
    )
    result["date"] = pd.to_datetime(result["date"], errors="coerce")
    result = result.sort_values(
        ["date", "product_code", "description", "pack_size"],
        kind="stable",
    ).reset_index(drop=True)
    return result


# ----------------------------- UI 状态回调 -----------------------------


def on_change_expiry_days() -> None:
    value = st.session_state.get("summary_expiry_days", 30)
    try:
        value_int = int(value)
    except (TypeError, ValueError):
        value_int = 30
    if value_int < 1:
        value_int = 1
    st.session_state["summary_expiry_days"] = value_int


def on_change_global_low_stock() -> None:
    for key in ("summary_global_low_ctn", "summary_global_low_pkt"):
        value = st.session_state.get(key, 0)
        try:
            value_int = int(value)
        except (TypeError, ValueError):
            value_int = 0
        if value_int < 0:
            value_int = 0
        st.session_state[key] = value_int


def on_toggle_near_expiry() -> None:
    st.session_state["toggle_only_near"] = bool(
        st.session_state.get("toggle_only_near", False)
    )


def on_toggle_low_stock() -> None:
    st.session_state["toggle_only_low"] = bool(
        st.session_state.get("toggle_only_low", False)
    )


def _clear_text_input_on_backspace(input_key: str, tracker_key: str) -> str:
    current = str(st.session_state.get(input_key, "") or "")
    previous = str(st.session_state.get(tracker_key, "") or "")
    if (
        previous
        and current
        and len(current) < len(previous)
        and previous.startswith(current)
    ):
        st.session_state[input_key] = ""
        current = ""
    st.session_state[tracker_key] = current
    return current


_STOCK_PERSIST_KEYS = [
    "stock_file_name",
    "stock_file_bytes",
    "stock_file_browsed_at",
    "f_wh",
    "f_sup",
    "f_sup_ex",
    "f_brand",
    "f_brand_ex",
    "f_desc",
    "f_code",
    "f_remark",
    "f_remark_ex",
    "use_date_filters",
    "expiry_range",
    "relabel_date_range",
    "summary_expiry_days",
    "summary_global_low_ctn",
    "summary_global_low_pkt",
    "toggle_only_near",
    "toggle_only_low",
    "toggle_show_depleted",
    "product_quick_filter",
    "status_filter_options",
    "batch_mode_radio",
    "stock_show_batch_details",
    "stock_batch_detail_limit",
    "stock_issue_focus_index",
    "stock_focus_active",
    "stock_focus_supplier",
    "stock_focus_product",
    "stock_focus_pack_size",
    "stock_focus_product_code",
]

_SALES_PERSIST_KEYS = [
    "sales_files_payload",
    "sales_files_browsed_at",
    "sales_filter_year",
    "sales_filter_month",
    "sales_filter_customer",
    "sales_filter_customer_exclude",
    "sales_filter_outlet",
    "sales_filter_product_description",
    "sales_filter_supplier",
    "sales_filter_supplier_exclude",
    "sales_filter_brand",
    "sales_filter_product_code",
    "sales_filter_account",
    "sales_filter_invoice",
    "sales_filter_use_exact_date",
    "sales_filter_exact_date",
    "sales_filter_date_from",
    "sales_filter_date_to",
    "sales_filter_apply_count",
    "sales_basic_mode",
    "sales_date_preset",
    "sales_download_package",
    "sales_price_list_account",
]

_RECONCILIATION_PERSIST_KEYS = [
    "reconciliation_stock_file_name",
    "reconciliation_stock_file_bytes",
    "reconciliation_stock_file_browsed_at",
    "reconciliation_sales_files_payload",
    "reconciliation_sales_files_browsed_at",
    "reconciliation_filter_product_code",
    "reconciliation_filter_description",
    "reconciliation_filter_year",
    "reconciliation_filter_month",
    "reconciliation_filter_date",
    "reconciliation_filter_use_specific_date",
    "reconciliation_filter_specific_date",
]


def _restore_persisted_state(persist_key: str) -> None:
    persisted = st.session_state.get(persist_key)
    if not isinstance(persisted, dict):
        return
    for key, value in persisted.items():
        if key not in st.session_state:
            st.session_state[key] = value


def _snapshot_persisted_state(persist_key: str, keys: List[str]) -> None:
    snapshot: Dict[str, Any] = {}
    for key in keys:
        if key in st.session_state:
            snapshot[key] = st.session_state.get(key)
    st.session_state[persist_key] = snapshot


def _coerce_stock_issue_focus_index(value: Any, max_index: int) -> int:
    max_valid = max(int(max_index), 0)
    parsed: Optional[int] = None
    if isinstance(value, bool):
        parsed = None
    elif isinstance(value, int):
        parsed = value
    elif isinstance(value, float) and math.isfinite(value):
        parsed = int(value)
    elif isinstance(value, str):
        text = value.strip()
        if text:
            if re.fullmatch(r"-?\d+", text):
                parsed = int(text)
            else:
                label_match = re.match(r"^(\d+)\.\s+", text)
                if label_match:
                    parsed = int(label_match.group(1)) - 1
    if parsed is None:
        parsed = 0
    return max(0, min(parsed, max_valid))


def _format_browse_timestamp(value: Optional[datetime.datetime]) -> str:
    if value is None:
        return ""
    return value.strftime("%I:%M %p %d/%m/%Y")


def _cache_stock_upload(
    uploaded: Any,
    session_state: Optional[Dict[str, Any]] = None,
    *,
    now: Optional[datetime.datetime] = None,
) -> None:
    if uploaded is None:
        return
    state = session_state if session_state is not None else st.session_state
    state["stock_file_name"] = getattr(uploaded, "name", "uploaded.xlsx")
    state["stock_file_bytes"] = uploaded.getvalue()
    browse_time = now if now is not None else datetime.datetime.now()
    state["stock_file_browsed_at"] = _format_browse_timestamp(browse_time)


def _clear_stock_upload_cache(session_state: Optional[Dict[str, Any]] = None) -> None:
    state = session_state if session_state is not None else st.session_state
    for key in ["stock_file_name", "stock_file_bytes", "stock_file_browsed_at"]:
        state.pop(key, None)


def _handle_stock_uploader_change() -> None:
    uploaded = st.session_state.get("stock_uploader")
    if uploaded is None:
        _clear_stock_upload_cache()
        return
    _cache_stock_upload(uploaded)


def _cache_sales_upload(
    uploaded_files: Any,
    session_state: Optional[Dict[str, Any]] = None,
    *,
    now: Optional[datetime.datetime] = None,
) -> None:
    files = [uploaded for uploaded in (uploaded_files or []) if uploaded is not None]
    if not files:
        return
    state = session_state if session_state is not None else st.session_state
    state["sales_files_payload"] = [
        {"name": uploaded.name, "bytes": uploaded.getvalue()} for uploaded in files
    ]
    browse_time = now if now is not None else datetime.datetime.now()
    state["sales_files_browsed_at"] = _format_browse_timestamp(browse_time)


def _clear_sales_upload_cache(session_state: Optional[Dict[str, Any]] = None) -> None:
    state = session_state if session_state is not None else st.session_state
    for key in ["sales_files_payload", "sales_files_browsed_at"]:
        state.pop(key, None)


def _handle_sales_uploader_change() -> None:
    uploaded_files = st.session_state.get("sales_uploader")
    if not uploaded_files:
        _clear_sales_upload_cache()
        return
    _cache_sales_upload(uploaded_files)


def _cache_reconciliation_stock_upload(
    uploaded: Any,
    session_state: Optional[Dict[str, Any]] = None,
    *,
    now: Optional[datetime.datetime] = None,
) -> None:
    if uploaded is None:
        return
    state = session_state if session_state is not None else st.session_state
    state["reconciliation_stock_file_name"] = getattr(uploaded, "name", "uploaded.xlsx")
    state["reconciliation_stock_file_bytes"] = uploaded.getvalue()
    browse_time = now if now is not None else datetime.datetime.now()
    state["reconciliation_stock_file_browsed_at"] = _format_browse_timestamp(
        browse_time
    )


def _clear_reconciliation_stock_upload_cache(
    session_state: Optional[Dict[str, Any]] = None,
) -> None:
    state = session_state if session_state is not None else st.session_state
    for key in [
        "reconciliation_stock_file_name",
        "reconciliation_stock_file_bytes",
        "reconciliation_stock_file_browsed_at",
    ]:
        state.pop(key, None)


def _handle_reconciliation_stock_uploader_change() -> None:
    uploaded = st.session_state.get("reconciliation_stock_uploader")
    if uploaded is None:
        _clear_reconciliation_stock_upload_cache()
        return
    _cache_reconciliation_stock_upload(uploaded)


def _cache_reconciliation_sales_upload(
    uploaded_files: Any,
    session_state: Optional[Dict[str, Any]] = None,
    *,
    now: Optional[datetime.datetime] = None,
) -> None:
    files = [uploaded for uploaded in (uploaded_files or []) if uploaded is not None]
    if not files:
        return
    state = session_state if session_state is not None else st.session_state
    state["reconciliation_sales_files_payload"] = [
        {"name": uploaded.name, "bytes": uploaded.getvalue()} for uploaded in files
    ]
    browse_time = now if now is not None else datetime.datetime.now()
    state["reconciliation_sales_files_browsed_at"] = _format_browse_timestamp(
        browse_time
    )


def _clear_reconciliation_sales_upload_cache(
    session_state: Optional[Dict[str, Any]] = None,
) -> None:
    state = session_state if session_state is not None else st.session_state
    for key in [
        "reconciliation_sales_files_payload",
        "reconciliation_sales_files_browsed_at",
    ]:
        state.pop(key, None)


def _handle_reconciliation_sales_uploader_change() -> None:
    uploaded_files = st.session_state.get("reconciliation_sales_uploader")
    if not uploaded_files:
        _clear_reconciliation_sales_upload_cache()
        return
    _cache_reconciliation_sales_upload(uploaded_files)


# ----------------------------- 主程序：Stock 页 -----------------------------


def run_stock_page():
    st.title("Stock Dashboard (Stocks DataGrid)")
    st.caption(
        "Upload an Excel file containing the 'Stocks report' and 'Lai Hock Whse' sheets (data starts on row 3)."
    )

    # 朴实无华的文件上传 + 缓存
    uploaded = st.file_uploader(
        "Upload Excel (.xlsx)",
        type=["xlsx"],
        key="stock_uploader",
        on_change=_handle_stock_uploader_change,
    )
    if uploaded is not None and not st.session_state.get("stock_file_bytes"):
        _cache_stock_upload(uploaded)

    stock_file_bytes = st.session_state.get("stock_file_bytes")
    stock_file_name = st.session_state.get("stock_file_name", "cached workbook")
    stock_file_browsed_at = st.session_state.get("stock_file_browsed_at", "")

    if not stock_file_bytes:
        st.info("Upload a source workbook to begin.")
        return

    st.caption(f"Using file: {stock_file_name}")
    if stock_file_browsed_at:
        st.caption(f"Last browsed: {stock_file_browsed_at}")

    try:
        df, warns = load_and_normalize_cached(stock_file_bytes)
    except Exception as e:
        st.error(f"Failed to read Excel: {e}")
        return

    for w in warns:
        st.warning(w)

    display_cols = [
        "warehouse",
        "supplier",
        "brand",
        "product_code",
        "description",
        "pack_size",
        "unit",
        "expiry_date",
        "relabel_to_date",
        "stock_qty",
    ]
    display_cols = [c for c in display_cols if c in df.columns]
    df_display = df[display_cols].copy()

    filtered, selected_descs, _placeholder_due_days, filter_state = apply_filters_v2(
        df_display
    )
    total_rows = len(filtered)
    filter_summary_parts = []
    label_map = {
        "warehouse": "Warehouse",
        "supplier": "Supplier",
        "brand": "Brand",
        "description": "Description",
        "product_code": "Product Code",
        "remark": "Remark",
    }
    for key, label in label_map.items():
        vals = filter_state.get(key)
        if vals is None or vals == "":
            continue
        if isinstance(vals, (list, tuple, set)):
            formatted = ", ".join(str(v) for v in vals if v)
        else:
            formatted = str(vals).strip()
        if formatted:
            filter_summary_parts.append(f"{label}: {formatted}")

    def _fmt_range(value):
        if not value or not isinstance(value, (list, tuple)) or len(value) != 2:
            return ""
        start, end = value
        if start and end:
            return f"{start} ~ {end}"
        return ""

    if filter_state.get("use_date_filters"):
        exp_range = _fmt_range(filter_state.get("expiry_range"))
        relabel_range = _fmt_range(filter_state.get("relabel_range"))
        if exp_range:
            filter_summary_parts.append(f"Expiry: {exp_range}")
        if relabel_range:
            filter_summary_parts.append(f"Relabel: {relabel_range}")

    filter_summary_text = (
        "; ".join(filter_summary_parts)
        if filter_summary_parts
        else "None (showing all rows)"
    )
    st.caption(f"Active filters: {filter_summary_text}")
    notification_holder = st.container()

    summary_df, detail_map = aggregate_summary(filtered)
    summary_df = summary_df.copy()

    if "summary_expiry_days" not in st.session_state:
        st.session_state["summary_expiry_days"] = 30
    if "summary_global_low_ctn" not in st.session_state:
        st.session_state["summary_global_low_ctn"] = 0
    if "summary_global_low_pkt" not in st.session_state:
        st.session_state["summary_global_low_pkt"] = 0
    if "toggle_only_near" not in st.session_state:
        st.session_state["toggle_only_near"] = False
    if "toggle_only_low" not in st.session_state:
        st.session_state["toggle_only_low"] = False
    if "toggle_show_depleted" not in st.session_state:
        st.session_state["toggle_show_depleted"] = True
    if "product_quick_filter" not in st.session_state:
        st.session_state["product_quick_filter"] = ""
    if "stock_show_batch_details" not in st.session_state:
        st.session_state["stock_show_batch_details"] = False
    if "stock_batch_detail_limit" not in st.session_state:
        st.session_state["stock_batch_detail_limit"] = 30
    if "stock_focus_active" not in st.session_state:
        st.session_state["stock_focus_active"] = False
    if "stock_focus_supplier" not in st.session_state:
        st.session_state["stock_focus_supplier"] = ""
    if "stock_focus_product" not in st.session_state:
        st.session_state["stock_focus_product"] = ""
    if "stock_focus_pack_size" not in st.session_state:
        st.session_state["stock_focus_pack_size"] = ""
    if "stock_focus_product_code" not in st.session_state:
        st.session_state["stock_focus_product_code"] = ""

    summary_bar = st.container()
    with summary_bar:
        metric_holder = st.container()
        controls_holder = st.container()

    status_options = ["Expired", "Near-Expiry", "Low-Stock", "Depleted", "OK"]
    if "status_filter_options" not in st.session_state:
        st.session_state["status_filter_options"] = status_options.copy()
    with controls_holder:
        toggle_cols = st.columns([1, 1, 1, 2])
        toggle_cols[0].toggle(
            "Show Near-Expiry Only",
            key="toggle_only_near",
            on_change=on_toggle_near_expiry,
        )
        toggle_cols[1].toggle(
            "Show Low-Stock Only", key="toggle_only_low", on_change=on_toggle_low_stock
        )
        toggle_cols[2].toggle("Show Depleted", key="toggle_show_depleted")
        with toggle_cols[3]:
            threshold_cols = st.columns([1, 1, 1])
            threshold_cols[0].number_input(
                "expiry_days (days)",
                min_value=1,
                max_value=365,
                step=1,
                key="summary_expiry_days",
                on_change=on_change_expiry_days,
                format="%d",
                help="Day threshold applied when tagging Near-Expiry.",
            )
            threshold_cols[1].number_input(
                "global_low_stock ctns",
                min_value=0,
                step=1,
                key="summary_global_low_ctn",
                on_change=on_change_global_low_stock,
                format="%d",
                help="Global carton threshold when a product ROP is missing.",
            )
            threshold_cols[2].number_input(
                "global_low_stock pkts",
                min_value=0,
                step=1,
                key="summary_global_low_pkt",
                on_change=on_change_global_low_stock,
                format="%d",
                help="Global packet threshold when a product ROP is missing.",
            )
            mode_label = st.radio(
                "批次分组显示",
                options=["Expiry only", "Remark only", "Expiry + Remark"],
                horizontal=True,
                key="batch_mode_radio",
                help="选择按到期、按备注，或两者同时分组展示批次",
            )
            mode_map = {
                "Expiry only": "expiry",
                "Remark only": "remark",
                "Expiry + Remark": "both",
            }
            batch_mode = mode_map.get(mode_label, "expiry")

        product_query_raw = st.text_input(
            "Product Quick Filter",
            key="product_quick_filter",
            placeholder="Filter by supplier / brand / product / code",
            help="Client-side filter that applies to Supplier, Brand, Product, and Product Code.",
        )
        product_query_raw = _clear_text_input_on_backspace(
            "product_quick_filter", "__prev_product_quick_filter"
        )
        status_selected = st.multiselect(
            "Status filter",
            options=status_options,
            key="status_filter_options",
            help="Limit the main table by status tags. Clear selection to show all statuses.",
        )

        detail_cols = st.columns([1, 1, 2])
        detail_cols[0].toggle(
            "Show Batch Details",
            key="stock_show_batch_details",
            help="Render per-product batch breakdown cards. Turn off for faster page switching.",
        )
        detail_cols[1].number_input(
            "Detail Rows",
            min_value=5,
            max_value=200,
            step=5,
            key="stock_batch_detail_limit",
            disabled=not bool(st.session_state.get("stock_show_batch_details", False)),
            help="Maximum number of product detail cards to render.",
        )

    expiry_days = int(st.session_state.get("summary_expiry_days", 30))
    global_low_ctn = int(st.session_state.get("summary_global_low_ctn", 0))
    global_low_pkt = int(st.session_state.get("summary_global_low_pkt", 0))
    near_only = bool(st.session_state.get("toggle_only_near", False))
    low_only = bool(st.session_state.get("toggle_only_low", False))
    show_depleted = bool(st.session_state.get("toggle_show_depleted", True))
    product_query = (product_query_raw or "").strip()
    status_selected = status_selected or status_options

    if summary_df.empty:
        with notification_holder:
            st.info("库存通知：当前筛选下没有可用库存数据。")
        with metric_holder:
            metric_cols = st.columns([1, 1, 1, 1])
            metric_cols[0].metric("Filtered Rows", f"{total_rows}")
            metric_cols[1].metric("Totals", "0")
            metric_cols[2].metric("Near-Expiry", "0")
            metric_cols[3].metric("Low-Stock", "0")
        st.info(
            "No data found for Savori Whse / Lai Hock Whse under the current filters."
        )
        return

    from functools import lru_cache

    normalized_detail_map = {
        tuple(k) if not isinstance(k, tuple) else tuple(k): v
        for k, v in (detail_map or {}).items()
        if v is not None
    }

    @lru_cache(maxsize=2048)
    def _cached_split(group_key_tuple, expiry_days, show_depleted, batch_mode):
        key = (
            tuple(group_key_tuple)
            if not isinstance(group_key_tuple, tuple)
            else group_key_tuple
        )
        frame = normalized_detail_map.get(key)
        if frame is None or frame.empty:
            return pd.DataFrame()
        return split_by_expiry(
            frame,
            expiry_days=expiry_days,
            show_depleted=show_depleted,
            batch_mode=batch_mode,
        )

    def _get_expiry_table(group_key_tuple):
        return _cached_split(
            tuple(group_key_tuple)
            if not isinstance(group_key_tuple, tuple)
            else group_key_tuple,
            expiry_days,
            show_depleted,
            batch_mode,
        )

    def _opt_float(val):
        try:
            return float(val)
        except Exception:
            return None

    summary_df["is_low_stock"] = False
    summary_df["low_stock_reason"] = None
    for i, r in summary_df.iterrows():
        total_ctn = _opt_float(r.get("total_ctn")) or 0.0
        total_pkt = _opt_float(r.get("total_pkt")) or 0.0
        rop_ctn = _opt_float(r.get("reorder_point_ctn"))
        rop_pkt = _opt_float(r.get("reorder_point_pkt"))
        if rop_ctn is not None and total_ctn <= rop_ctn + 1e-9:
            summary_df.at[i, "is_low_stock"] = True
            summary_df.at[i, "low_stock_reason"] = "ROP"
        elif rop_pkt is not None and total_pkt <= rop_pkt + 1e-9:
            summary_df.at[i, "is_low_stock"] = True
            summary_df.at[i, "low_stock_reason"] = "ROP"
        else:
            if (global_low_ctn and total_ctn <= float(global_low_ctn) + 1e-9) or (
                global_low_pkt and total_pkt <= float(global_low_pkt) + 1e-9
            ):
                summary_df.at[i, "is_low_stock"] = True
                summary_df.at[i, "low_stock_reason"] = "Global"

    summary_df["has_expired_batch"] = False
    summary_df["has_near_batch"] = False
    summary_df["expired_qty_ctn"] = 0.0
    summary_df["expired_qty_pkt"] = 0.0
    summary_df["near_qty_ctn"] = 0.0
    summary_df["near_qty_pkt"] = 0.0
    summary_df["expired_qty_breakdown"] = [{} for _ in range(len(summary_df))]
    summary_df["near_qty_breakdown"] = [{} for _ in range(len(summary_df))]
    for i, r in summary_df.iterrows():
        tuple_key = (
            tuple(r["group_key"])
            if not isinstance(r["group_key"], tuple)
            else r["group_key"]
        )
        tbl = _get_expiry_table(tuple_key)
        if tbl is None or tbl.empty:
            continue
        pos_qty = (
            tbl["subtotal_ctn"].astype(float) + tbl["subtotal_pkt"].astype(float)
        ) > 0
        if not pos_qty.any():
            continue
        sub = tbl.loc[pos_qty]
        expired_sub = sub[sub["status_batch"] == "Expired"]
        near_sub = sub[sub["status_batch"] == "Near-Expiry"]

        if not expired_sub.empty:
            summary_df.at[i, "has_expired_batch"] = True
            summary_df.at[i, "expired_qty_ctn"] = float(
                expired_sub["subtotal_ctn"].sum()
            )
            summary_df.at[i, "expired_qty_pkt"] = float(
                expired_sub["subtotal_pkt"].sum()
            )
            summary_df.at[i, "expired_qty_breakdown"] = _merge_quantity_breakdowns(
                expired_sub.get("subtotal_qty_breakdown", [])
            )

        if not near_sub.empty:
            summary_df.at[i, "has_near_batch"] = True
            summary_df.at[i, "near_qty_ctn"] = float(near_sub["subtotal_ctn"].sum())
            summary_df.at[i, "near_qty_pkt"] = float(near_sub["subtotal_pkt"].sum())
            summary_df.at[i, "near_qty_breakdown"] = _merge_quantity_breakdowns(
                near_sub.get("subtotal_qty_breakdown", [])
            )

    summary_df["is_depleted"] = (
        summary_df["total_ctn"].fillna(0) + summary_df["total_pkt"].fillna(0)
    ) == 0

    def _primary_status(
        has_expired: bool, has_near: bool, is_low: bool, is_depleted: bool
    ) -> str:
        if is_depleted:
            return "Depleted"
        if has_expired:
            return "Expired"
        if has_near:
            return "Near-Expiry"
        if is_low:
            return "Low-Stock"
        return "OK"

    summary_df["Status Tags"] = [[] for _ in range(len(summary_df))]
    summary_df["status_product"] = ""

    for i, r in summary_df.iterrows():
        tags = []
        if bool(r["is_depleted"]):
            tags.append("Depleted")
        if bool(r["has_expired_batch"]):
            tags.append("Expired")
        if bool(r["has_near_batch"]):
            tags.append("Near-Expiry")
        if bool(r["is_low_stock"]):
            tags.append("Low-Stock")
        if not tags:
            tags = ["OK"]

        summary_df.at[i, "Status Tags"] = tags
        summary_df.at[i, "status_product"] = _primary_status(
            bool(r["has_expired_batch"]),
            bool(r["has_near_batch"]),
            bool(r["is_low_stock"]),
            bool(r["is_depleted"]),
        )

    if not show_depleted:
        summary_df = summary_df[summary_df["status_product"] != "Depleted"]

    summary_df["Savori Whse"] = [
        _format_quantity_breakdown(r.savori_qty_breakdown)
        for r in summary_df.itertuples()
    ]
    summary_df["Lai Hock Whse"] = [
        _format_quantity_breakdown(r.lai_hock_qty_breakdown)
        for r in summary_df.itertuples()
    ]
    summary_df["Total"] = summary_df.apply(
        lambda row: _format_summary_total_qty(
            row.get("total_ctn"),
            row.get("total_pkt"),
            _infer_carton_pack_size(row.get("Pack Size")),
        ),
        axis=1,
    )

    def _format_status(tags, reason, near_breakdown, expired_breakdown):
        if not isinstance(tags, (list, tuple)):
            return str(tags)
        formatted = []
        near_qty_text = _format_quantity_breakdown(near_breakdown)
        expired_qty_text = _format_quantity_breakdown(expired_breakdown)
        for tag in tags:
            if tag == "Low-Stock" and reason:
                if str(reason).upper() == "ROP":
                    formatted.append("Low-Stock(ROP)")
                else:
                    formatted.append("Low-Stock(Global)")
            elif tag == "Near-Expiry":
                formatted.append(f"Near-Expiry({near_qty_text})")
            elif tag == "Expired":
                formatted.append(f"Expired({expired_qty_text})")
            else:
                formatted.append(str(tag))
        return " • ".join(formatted) if formatted else "OK"

    summary_df["Status"] = [
        _format_status(tags, reason, near_breakdown, expired_breakdown)
        for tags, reason, near_breakdown, expired_breakdown in zip(
            summary_df["Status Tags"],
            summary_df["low_stock_reason"],
            summary_df["near_qty_breakdown"],
            summary_df["expired_qty_breakdown"],
        )
    ]

    summary_df["Near-Expiry Qty"] = [
        _format_quantity_breakdown(r.near_qty_breakdown)
        for r in summary_df.itertuples()
    ]
    summary_df["Expired Qty"] = [
        _format_quantity_breakdown(r.expired_qty_breakdown)
        for r in summary_df.itertuples()
    ]

    def _apply_focus_filter(
        supplier: str,
        product: str,
        pack_size: str,
        product_code: str,
    ) -> None:
        st.session_state["stock_focus_active"] = True
        st.session_state["stock_focus_supplier"] = supplier
        st.session_state["stock_focus_product"] = product
        st.session_state["stock_focus_pack_size"] = pack_size
        st.session_state["stock_focus_product_code"] = product_code
        st.session_state["product_quick_filter"] = ""
        st.session_state["status_filter_options"] = status_options.copy()
        st.session_state["toggle_only_near"] = False
        st.session_state["toggle_only_low"] = False

    def _clear_focus_filter() -> None:
        st.session_state["stock_focus_active"] = False
        st.session_state["stock_focus_supplier"] = ""
        st.session_state["stock_focus_product"] = ""
        st.session_state["stock_focus_pack_size"] = ""
        st.session_state["stock_focus_product_code"] = ""

    def _render_stock_notifications(frame: pd.DataFrame) -> None:
        if frame is None or frame.empty:
            with notification_holder:
                st.info("库存通知：当前没有可显示的库存记录。")
            return

        issue_df = frame[
            frame["status_product"].isin(["Expired", "Near-Expiry", "Low-Stock"])
        ].copy()
        expired_count = int((frame["status_product"] == "Expired").sum())
        near_count = int((frame["status_product"] == "Near-Expiry").sum())
        low_count = int((frame["status_product"] == "Low-Stock").sum())

        with notification_holder:
            st.markdown("### 库存通知")
            if issue_df.empty:
                st.success("未发现异常库存，所有商品状态为 OK。")
                return

            st.warning(
                f"检测到 {len(issue_df)} 个商品存在库存异常，请优先处理 Expired / Near-Expiry / Low-Stock。"
            )
            status_cols = st.columns(3)
            status_cols[0].metric("Expired", f"{expired_count}")
            status_cols[1].metric("Near-Expiry", f"{near_count}")
            status_cols[2].metric("Low-Stock", f"{low_count}")

            issue_rows = issue_df.reset_index(drop=True)
            if "stock_issue_focus_index" not in st.session_state:
                st.session_state["stock_issue_focus_index"] = 0
            valid_max_idx = max(len(issue_rows) - 1, 0)
            focus_idx = _coerce_stock_issue_focus_index(
                st.session_state.get("stock_issue_focus_index", 0),
                valid_max_idx,
            )
            st.session_state["stock_issue_focus_index"] = focus_idx
            stock_issue_focus_widget_key = "stock_issue_focus_selectbox"
            st.session_state[stock_issue_focus_widget_key] = _coerce_stock_issue_focus_index(
                st.session_state.get(stock_issue_focus_widget_key, focus_idx),
                valid_max_idx,
            )

            def _focus_label(i: int) -> str:
                row = issue_rows.iloc[int(i)]
                return (
                    f"{int(i) + 1}. {row.get('status_product', '')} | "
                    f"{row.get('Supplier', '')} | {row.get('Product', '')} | {row.get('Pack Size', '')} | {row.get('Product Code', '')}"
                )

            selected_focus_idx = st.selectbox(
                "快速定位异常商品",
                options=list(range(len(issue_rows))),
                key=stock_issue_focus_widget_key,
                format_func=_focus_label,
                help="选择后点击“定位到主表”，将按 Supplier + Product + Pack Size + Product Code 精准定位。",
            )
            st.session_state["stock_issue_focus_index"] = int(selected_focus_idx)
            focus_row = issue_rows.iloc[int(selected_focus_idx)]
            focus_supplier = str(focus_row.get("Supplier") or "").strip()
            focus_product = str(focus_row.get("Product") or "").strip()
            focus_pack_size = str(focus_row.get("Pack Size") or "").strip()
            focus_product_code = str(focus_row.get("Product Code") or "").strip()
            focus_cols = st.columns([1, 1, 2])
            focus_cols[0].button(
                "定位到主表",
                key="stock_focus_apply",
                width="stretch",
                on_click=_apply_focus_filter,
                args=(
                    focus_supplier,
                    focus_product,
                    focus_pack_size,
                    focus_product_code,
                ),
            )
            focus_cols[1].button(
                "清除定位",
                key="stock_focus_clear",
                width="stretch",
                on_click=_clear_focus_filter,
            )

            notify_cols = [
                "Supplier",
                "Brand",
                "Product",
                "Pack Size",
                "Product Code",
                "Near-Expiry Qty",
                "Expired Qty",
                "Total",
                "Status",
            ]
            preview_df = (
                issue_df[notify_cols]
                .head(20)
                .rename(
                    columns={
                        "Near-Expiry Qty": f"Near-Expiry Qty (<= {expiry_days}d)",
                        "Expired Qty": "Expired Qty (< 0d)",
                        "Total": "Overall Stock (All Dates)",
                    }
                )
            )
            st.caption(
                "异常商品预览（最多 20 项）：Near/Expired 是风险库存；Overall Stock 是同商品全部库存（含其他到期日）。"
            )
            st.dataframe(preview_df, width="stretch", hide_index=True)

    _render_stock_notifications(summary_df)

    view_df = summary_df.copy()

    def _has_near(tags):
        return any(t in {"Expired", "Near-Expiry"} for t in (tags or []))

    def _has_low(tags):
        return "Low-Stock" in (tags or [])

    near_mask = view_df["Status Tags"].apply(_has_near)
    low_mask = view_df["Status Tags"].apply(_has_low)

    if near_only and not low_only:
        view_df = view_df[near_mask]
    elif low_only and not near_only:
        view_df = view_df[low_mask]
    elif near_only and low_only:
        view_df = view_df[near_mask | low_mask]

    if product_query:
        mask = (
            view_df["Product"].str.contains(product_query, case=False, na=False)
            | view_df["Supplier"].str.contains(product_query, case=False, na=False)
            | view_df["Brand"].str.contains(product_query, case=False, na=False)
            | view_df["Product Code"].str.contains(product_query, case=False, na=False)
        )
        view_df = view_df[mask]
    if status_selected and set(status_selected) != set(status_options):
        view_df = view_df[view_df["status_product"].isin(status_selected)]

    if bool(st.session_state.get("stock_focus_active", False)):
        focus_supplier = (
            str(st.session_state.get("stock_focus_supplier", "")).strip().lower()
        )
        focus_product = (
            str(st.session_state.get("stock_focus_product", "")).strip().lower()
        )
        focus_pack_size = (
            str(st.session_state.get("stock_focus_pack_size", "")).strip().lower()
        )
        focus_product_code = (
            str(st.session_state.get("stock_focus_product_code", "")).strip().lower()
        )

        focus_mask = pd.Series(True, index=view_df.index)
        if focus_supplier:
            focus_mask = focus_mask & (
                view_df["Supplier"].astype(str).str.strip().str.lower()
                == focus_supplier
            )
        if focus_product:
            focus_mask = focus_mask & (
                view_df["Product"].astype(str).str.strip().str.lower() == focus_product
            )
        if focus_pack_size:
            focus_mask = focus_mask & (
                view_df["Pack Size"].astype(str).str.strip().str.lower()
                == focus_pack_size
            )
        if focus_product_code:
            focus_mask = focus_mask & (
                view_df["Product Code"].astype(str).str.strip().str.lower()
                == focus_product_code
            )
        view_df = view_df[focus_mask]

    view_df = view_df.reset_index(drop=True)

    if view_df.empty:
        with metric_holder:
            metric_cols = st.columns([1, 1, 1, 1])
            metric_cols[0].metric("Filtered Rows", f"{total_rows}")
            metric_cols[1].metric("Totals", "0")
            metric_cols[2].metric("Near-Expiry", "0")
            metric_cols[3].metric("Low-Stock", "0")
        st.info(
            "No rows in the main table match the current view. Adjust filters or thresholds."
        )
        return

    totals_display = _format_stock_scope_total_qty(view_df)

    near_count = int(
        view_df["Status Tags"]
        .apply(lambda tags: any(t in {"Expired", "Near-Expiry"} for t in (tags or [])))
        .sum()
    )
    low_count = int(
        view_df["Status Tags"].apply(lambda tags: "Low-Stock" in (tags or [])).sum()
    )

    with metric_holder:
        metric_cols = st.columns([1, 1, 1, 1])
        metric_cols[0].metric("Filtered Rows", f"{total_rows}")
        metric_cols[1].metric("Totals", totals_display if totals_display else "0")
        metric_cols[2].metric("Near-Expiry", f"{near_count}")
        metric_cols[3].metric("Low-Stock", f"{low_count}")

    nearest_days_map = {}
    for k in view_df["group_key"]:
        key = tuple(k) if not isinstance(k, tuple) else k
        tbl = _get_expiry_table(key)
        if tbl is None or tbl.empty:
            nearest_days_map[key] = 10**9
            continue
        pos = (
            tbl["subtotal_ctn"].astype(float) + tbl["subtotal_pkt"].astype(float)
        ) > 0
        if not pos.any():
            nearest_days_map[key] = 10**9
            continue
        vals = tbl.loc[pos, "days_to_expiry"].dropna()
        try:
            nearest_days_map[key] = int(vals.min()) if not vals.empty else 10**9
        except Exception:
            nearest_days_map[key] = 10**9

    def _status_rank(s: str) -> int:
        return {
            "Depleted": 0,
            "Expired": 1,
            "Near-Expiry": 2,
            "Low-Stock": 3,
            "OK": 4,
        }.get(s, 9)

    view_df["status_priority"] = view_df["status_product"].map(_status_rank)
    view_df["nearest_days"] = [
        nearest_days_map.get(tuple(k) if not isinstance(k, tuple) else k, 10**9)
        for k in view_df["group_key"]
    ]
    view_df = view_df.sort_values(
        by=["status_priority", "nearest_days", "total_ctn", "total_pkt", "Product"],
        kind="stable",
    ).reset_index(drop=True)

    display_columns = [
        "Supplier",
        "Brand",
        "Product",
        "Pack Size",
        "Product Code",
        "Savori Whse",
        "Lai Hock Whse",
        "Total",
        "Status",
    ]
    view_df_display = view_df[display_columns].copy()
    view_df_display_clean = _strip_html_df(view_df_display)
    total_idx = list(view_df_display_clean.columns).index("Total")

    def _tag_code(tags):
        s = set(tags or [])
        if "Expired" in s and "Low-Stock" in s:
            return 3
        if "Near-Expiry" in s and "Low-Stock" in s:
            return 2
        if "Expired" in s:
            return 1
        if "Near-Expiry" in s:
            return 4
        if "Low-Stock" in s:
            return 5
        if "Depleted" in s:
            return 6
        return 0

    view_df["__row_code"] = view_df["Status Tags"].apply(_tag_code)

    def _style_row(row):
        tags = view_df.loc[row.name, "Status Tags"]
        styles = ["" for _ in row]

        if "Expired" in tags and "Low-Stock" in tags:
            styles = ["background-color: rgba(102,51,153,0.25);"] * len(row)
            styles[total_idx] += "font-weight:600; color:#FFF;"
        elif "Near-Expiry" in tags and "Low-Stock" in tags:
            styles = ["background-color: rgba(255,140,0,0.22);"] * len(row)
            styles[total_idx] += "font-weight:600; color:#FFF;"
        elif "Expired" in tags:
            styles = ["background-color: rgba(178,34,34,0.18);"] * len(row)
        elif "Near-Expiry" in tags:
            styles = ["background-color: rgba(255,165,0,0.18);"] * len(row)
        elif "Low-Stock" in tags:
            styles[total_idx] = (
                "background-color: rgba(255, 255, 0, 0.18); font-weight:600;"
            )
        elif "Depleted" in tags:
            styles = ["background-color: rgba(128,128,128,0.12);"] * len(row)
        return styles

    stock_summary_copy_df = _build_stock_summary_copy_df(view_df_display_clean)
    _render_copy_values_button(
        _build_copy_values_tsv(stock_summary_copy_df),
        "stock-summary-main",
        has_rows=not stock_summary_copy_df.empty,
        label="复制当前表格数据",
    )

    styled_view = view_df_display_clean.style.apply(_style_row, axis=1)
    st.dataframe(styled_view, width="stretch", hide_index=True)

    export_summary_cols = [
        "Supplier",
        "Brand",
        "Product",
        "Pack Size",
        "Product Code",
        "Savori Whse",
        "Lai Hock Whse",
        "Total",
        "Status",
        "total_ctn",
        "total_pkt",
        "savori_ctn",
        "savori_pkt",
        "lai_hock_ctn",
        "lai_hock_pkt",
        "reorder_point_ctn",
        "reorder_point_pkt",
        "low_stock_reason",
        "status_product",
    ]
    export_summary = view_df[export_summary_cols].copy()

    expiry_export_frames = []
    for _, row in view_df.iterrows():
        tuple_key = (
            tuple(row["group_key"])
            if not isinstance(row["group_key"], tuple)
            else row["group_key"]
        )
        expiry_table = _get_expiry_table(tuple_key)
        if expiry_table is None or expiry_table.empty:
            continue
        export_block = expiry_table.copy()
        export_block.insert(0, "Supplier", row["Supplier"])
        export_block.insert(1, "Product", row["Product"])
        export_block.insert(2, "Product Code", row["Product Code"])
        export_block.insert(3, "Product Status", row["status_product"])
        export_block.insert(4, "Parent Total", row["Total"])
        expiry_export_frames.append(export_block)

    valid_frames = [f for f in expiry_export_frames if f is not None and not f.empty]

    expected_cols = [
        "Supplier",
        "Product",
        "Product Code",
        "Product Status",
        "Parent Total",
        "Expiry",
        "Remark",
        "Savori Whse",
        "Lai Hock Whse",
        "Subtotal",
        "subtotal_ctn",
        "subtotal_pkt",
        "expiry_date",
        "status_batch",
        "days_to_expiry",
        "Info",
    ]

    normalized = []
    for f in valid_frames:
        g = f.copy()
        for c in expected_cols:
            if c not in g.columns:
                g[c] = pd.Series([pd.NA] * len(g), dtype="string")

        for c in [
            "Supplier",
            "Product",
            "Product Code",
            "Product Status",
            "Parent Total",
            "Expiry",
            "Remark",
            "Savori Whse",
            "Lai Hock Whse",
            "Subtotal",
            "status_batch",
            "Info",
        ]:
            if c in g.columns:
                if g[c].dtype != "string":
                    g[c] = g[c].astype("string")

        for c in ["subtotal_ctn", "subtotal_pkt"]:
            if c in g.columns:
                g[c] = pd.to_numeric(g[c], errors="coerce").astype("float64")

        if "expiry_date" in g.columns:
            g["expiry_date"] = pd.to_datetime(g["expiry_date"], errors="coerce")

        if "days_to_expiry" in g.columns:
            g["days_to_expiry"] = pd.to_numeric(
                g["days_to_expiry"], errors="coerce"
            ).astype("Int64")

        g = g[expected_cols]
        normalized.append(g)

    if normalized:
        expiry_export_df = pd.concat(normalized, ignore_index=True)
    else:
        expiry_export_df = pd.DataFrame(columns=expected_cols)

    detail_frames = []
    if detail_map and not view_df.empty:
        key_order = [
            tuple(k) if not isinstance(k, tuple) else tuple(k)
            for k in view_df["group_key"]
        ]
        for key in dict.fromkeys(key_order):
            frame = detail_map.get(key)
            if frame is not None and not frame.empty:
                detail_frames.append(frame.copy())
    if detail_frames:
        detail_export_df = pd.concat(detail_frames, ignore_index=True, sort=False)
        helper_cols = [c for c in detail_export_df.columns if str(c).startswith("_")]
        detail_export_df = detail_export_df.drop(columns=helper_cols, errors="ignore")
        for col in ["expiry_date", "relabel_to_date"]:
            if col in detail_export_df.columns:
                detail_export_df[col] = pd.to_datetime(
                    detail_export_df[col], errors="coerce"
                ).dt.strftime("%Y-%m-%d")
    else:
        detail_export_df = pd.DataFrame()

    export_summary_clean = _strip_html_df(export_summary)
    expiry_export_clean = _strip_html_df(expiry_export_df)

    stock_with_expiry_copy_tsv = _build_stock_summary_with_expiry_copy_tsv(
        stock_summary_copy_df, expiry_export_clean
    )
    _render_copy_values_button(
        stock_with_expiry_copy_tsv,
        "stock-summary-with-expiry",
        has_rows=not stock_summary_copy_df.empty and not expiry_export_clean.empty,
        label="复制当前表格及Expiry明细",
    )

    issue_summary_df = (
        view_df[view_df["status_product"].isin(["Expired", "Near-Expiry", "Low-Stock"])]
        .groupby(["status_product"], dropna=False)
        .agg(
            Products=("Product", "count"),
            Total_Ctn=("total_ctn", "sum"),
            Total_Pkt=("total_pkt", "sum"),
        )
        .reset_index()
        .rename(columns={"status_product": "Status"})
    )

    mgmt_csv_buffer = io.StringIO()
    export_summary_clean.to_csv(mgmt_csv_buffer, index=False)

    mgmt_excel_buffer = io.BytesIO()
    with pd.ExcelWriter(mgmt_excel_buffer) as writer:
        export_summary_clean.to_excel(writer, sheet_name="Summary", index=False)
        issue_summary_df.to_excel(writer, sheet_name="Issue Summary", index=False)
        pd.DataFrame(
            [{"key": str(k), "value": str(v)} for k, v in (filter_state or {}).items()],
            columns=["key", "value"],
        ).to_excel(writer, sheet_name="Filters", index=False)
    mgmt_excel_buffer.seek(0)
    detail_csv_buffer = io.StringIO()
    detail_export_clean = (
        _strip_html_df(detail_export_df)
        if not detail_export_df.empty
        else detail_export_df
    )
    detail_export_clean.to_csv(detail_csv_buffer, index=False)

    detail_excel_buffer = io.BytesIO()
    with pd.ExcelWriter(detail_excel_buffer) as writer:
        detail_export_clean.to_excel(writer, sheet_name="Stock Report", index=False)
        expiry_export_clean.to_excel(writer, sheet_name="Expiry Breakdown", index=False)
        pd.DataFrame(
            [{"key": str(k), "value": str(v)} for k, v in (filter_state or {}).items()],
            columns=["key", "value"],
        ).to_excel(writer, sheet_name="Filters", index=False)
    detail_excel_buffer.seek(0)
    download_options = {
        "Mgmt CSV": {
            "data": mgmt_csv_buffer.getvalue(),
            "file_name": "stock_management_summary.csv",
            "mime": "text/csv",
        },
        "Mgmt Excel": {
            "data": mgmt_excel_buffer.getvalue(),
            "file_name": "stock_management_summary.xlsx",
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
        "Ops CSV": {
            "data": detail_csv_buffer.getvalue(),
            "file_name": "stock_operations_detail.csv",
            "mime": "text/csv",
        },
        "Ops Excel": {
            "data": detail_excel_buffer.getvalue(),
            "file_name": "stock_operations_report.xlsx",
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    }
    selected_download = st.selectbox(
        "Download package",
        options=list(download_options.keys()),
        key="stock_download_package",
    )
    selected_payload = download_options[selected_download]
    st.download_button(
        "Download",
        data=selected_payload["data"],
        file_name=selected_payload["file_name"],
        mime=selected_payload["mime"],
    )

    if bool(st.session_state.get("stock_show_batch_details", False)):
        detail_limit = int(st.session_state.get("stock_batch_detail_limit", 30))
        detail_source = view_df.head(detail_limit)
        st.caption(
            f"Batch details: showing {len(detail_source)} of {len(view_df)} products"
        )
        for _, row in detail_source.iterrows():
            tuple_key = (
                tuple(row["group_key"])
                if not isinstance(row["group_key"], tuple)
                else row["group_key"]
            )
            expiry_table = _get_expiry_table(tuple_key)
            title = f"{row['Supplier']} | {row['Product']}"
            with st.container():
                st.markdown(f"#### {title}")
                st.caption(f"Product Code: {row['Product Code']}")
                st.caption(f"Status: {row['Status']}")
                if row.get("low_stock_reason"):
                    st.caption(
                        f"Low-stock trigger: {'Product ROP' if row['low_stock_reason'] == 'ROP' else 'Global threshold'}"
                    )

                if expiry_table is None or expiry_table.empty:
                    st.info("No expiry breakdown available.")
                else:
                    if batch_mode == "expiry":
                        display_cols2 = [
                            "Expiry",
                            "Savori Whse",
                            "Lai Hock Whse",
                            "Subtotal",
                            "status_batch",
                            "days_to_expiry",
                            "Info",
                        ]
                    elif batch_mode == "remark":
                        display_cols2 = [
                            "Remark",
                            "Savori Whse",
                            "Lai Hock Whse",
                            "Subtotal",
                            "status_batch",
                            "days_to_expiry",
                            "Info",
                        ]
                    else:
                        display_cols2 = [
                            "Expiry",
                            "Remark",
                            "Savori Whse",
                            "Lai Hock Whse",
                            "Subtotal",
                            "status_batch",
                            "days_to_expiry",
                            "Info",
                        ]

                    table_display = expiry_table[display_cols2].rename(
                        columns={
                            "status_batch": "Batch Status",
                            "days_to_expiry": "DTE",
                        }
                    )

                    def _format_dte(value):
                        if pd.isna(value):
                            return ""
                        try:
                            return f"D{int(float(value)):+d}"
                        except (TypeError, ValueError):
                            return str(value)

                    table_display["DTE"] = table_display["DTE"].apply(_format_dte)

                    def _style_expiry(r):
                        status = table_display.loc[r.name, "Batch Status"]
                        styles = ["" for _ in r]
                        if status == "Expired":
                            styles = ["background-color: rgba(220,20,60,0.16);"] * len(
                                r
                            )
                        elif status == "Near-Expiry":
                            styles = ["background-color: rgba(255,165,0,0.18);"] * len(
                                r
                            )
                        elif status == "Depleted":
                            styles = [
                                "background-color: rgba(128,128,128,0.12);"
                            ] * len(r)

                        dte_text = table_display.loc[r.name, "DTE"]
                        if isinstance(dte_text, str) and dte_text.startswith("D"):
                            try:
                                dte_num = int(dte_text[1:])
                            except Exception:
                                dte_num = None
                            if dte_num is not None:
                                dte_idx = list(table_display.columns).index("DTE")
                                if dte_num < 0:
                                    styles[dte_idx] = (
                                        "background-color: rgba(220,20,60,0.24); font-weight:700;"
                                    )
                                elif dte_num <= 7:
                                    styles[dte_idx] = (
                                        "background-color: rgba(255,140,0,0.22); font-weight:700;"
                                    )
                                elif dte_num <= 30:
                                    styles[dte_idx] = (
                                        "background-color: rgba(255,215,0,0.20);"
                                    )
                        return styles

                    try:
                        styled_expiry = table_display.style.apply(_style_expiry, axis=1)
                        st.dataframe(styled_expiry, width="stretch", hide_index=True)
                    except Exception:
                        st.dataframe(table_display, width="stretch", hide_index=True)

                    subtotal_ctn_sum = float(expiry_table["subtotal_ctn"].sum())
                    subtotal_pkt_sum = float(expiry_table["subtotal_pkt"].sum())
                    parent_ctn = float(row["total_ctn"])
                    parent_pkt = float(row["total_pkt"])
                    if not math.isclose(
                        subtotal_ctn_sum, parent_ctn, abs_tol=1e-6
                    ) or not math.isclose(subtotal_pkt_sum, parent_pkt, abs_tol=1e-6):
                        st.warning(
                            "Expiry breakdown totals do not match the parent totals; please verify the source data."
                        )
                    st.caption(f"Total: {row['Total']}")
                st.divider()
    else:
        st.caption(
            "Batch details are hidden for faster page switching. Turn on 'Show Batch Details' when needed."
        )


# ----------------------------- Sales helpers -----------------------------


def _read_delivery_details_from_bytes(
    content: bytes, label: str
) -> Tuple[pd.DataFrame, Optional[str]]:
    try:
        xls = pd.ExcelFile(io.BytesIO(content), engine="openpyxl")
    except Exception as exc:
        return pd.DataFrame(), f"{label}: 读取工作簿失败：{exc}"

    sheet_name = _find_sheet_name_in_list(xls.sheet_names, "Delivery details")
    if not sheet_name:
        return pd.DataFrame(), f"{label}: 未找到 'Delivery details' 工作表。"

    try:
        frame = xls.parse(sheet_name, header=3, dtype=str).dropna(axis=0, how="all")
        return frame, None
    except Exception as exc:
        return pd.DataFrame(), f"{label}: 读取 '{sheet_name}' 失败：{exc}"


def _normalize_sales_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    trimmed = df.copy().iloc[:, : len(SALES_COLUMNS)]
    trimmed.columns = SALES_COLUMNS[: trimmed.shape[1]]
    for col in SALES_COLUMNS:
        if col not in trimmed.columns:
            trimmed[col] = pd.NA
    trimmed = trimmed[SALES_COLUMNS]
    blank_mask = trimmed.apply(
        lambda col: col.astype("string").str.fullmatch(r"\s*", na=False)
    )
    trimmed = trimmed.mask(blank_mask, pd.NA)
    trimmed = _strip_html_df(trimmed)
    for col in trimmed.select_dtypes(include="object").columns:
        trimmed[col] = trimmed[col].astype("string").str.strip()
    for col in SALES_NUMERIC_COLUMNS:
        trimmed[col] = pd.to_numeric(trimmed[col], errors="coerce")
    for col in SALES_DATE_COLUMNS:
        if not pd.api.types.is_datetime64_any_dtype(trimmed[col]):
            trimmed[col] = pd.to_datetime(trimmed[col], errors="coerce")
    trimmed["carton_packing_numeric"] = trimmed["Carton Packing"].map(
        _infer_carton_pack_size
    )
    trimmed["pcs_to_ctn"] = trimmed["Qty in Pcs"] / trimmed["carton_packing_numeric"]
    valid_pack = trimmed["carton_packing_numeric"].gt(0)
    trimmed["pcs_to_ctn"] = trimmed["pcs_to_ctn"].where(valid_pack, pd.NA)
    trimmed["total_ctn_equivalent"] = pd.to_numeric(
        trimmed["Qty in Ctns"], errors="coerce"
    ).fillna(0.0) + pd.to_numeric(trimmed["pcs_to_ctn"], errors="coerce").fillna(0.0)
    trimmed["total_packets"] = pd.to_numeric(
        trimmed["Qty in Ctns"], errors="coerce"
    ).fillna(0.0) * pd.to_numeric(
        trimmed["carton_packing_numeric"], errors="coerce"
    ).fillna(0.0) + pd.to_numeric(trimmed["Qty in Pcs"], errors="coerce").fillna(0.0)
    trimmed["Unit Price (per pcs)"] = _safe_divide_series(
        trimmed["Total Value"], trimmed["Qty in Pcs"]
    )
    trimmed["Unit Price (per carton)"] = _safe_divide_series(
        trimmed["Total Value"], trimmed["Qty in Ctns"]
    )
    trimmed["Unit Price (per ctn equivalent)"] = _safe_divide_series(
        trimmed["Total Value"], trimmed["total_ctn_equivalent"]
    )
    return trimmed


@st.cache_data(show_spinner=False)
def _cached_sales_dataset(
    inputs: Tuple[Tuple[str, bytes], ...],
) -> Tuple[pd.DataFrame, List[str]]:
    warns: List[str] = []
    frames: List[pd.DataFrame] = []
    for name, content in inputs:
        if not content:
            warns.append(f"{name}: 文件大小为 0，无法读取。")
            continue
        frame, err = _read_delivery_details_from_bytes(content, name)
        if err:
            warns.append(err)
            continue
        if frame.empty:
            warns.append(f"{name}: 'Delivery details' 表为空。")
            continue
        frames.append(frame)
    if not frames:
        return pd.DataFrame(columns=SALES_COLUMNS), warns
    combined = pd.concat(frames, ignore_index=True, sort=False)
    return _normalize_sales_dataframe(combined), warns


def load_sales_data(files) -> Tuple[pd.DataFrame, List[str]]:
    inputs: List[Tuple[str, bytes]] = []
    for uploaded in files or []:
        if uploaded is None:
            continue
        content = uploaded.getvalue()
        inputs.append((uploaded.name, content))
    if not inputs:
        return pd.DataFrame(columns=SALES_COLUMNS), []
    return _cached_sales_dataset(tuple(inputs))


def load_sales_data_from_payload(
    payload: Optional[List[Dict[str, Any]]],
) -> Tuple[pd.DataFrame, List[str]]:
    inputs: List[Tuple[str, bytes]] = []
    for item in payload or []:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "sales.xlsx")
        content = item.get("bytes")
        if isinstance(content, bytes) and content:
            inputs.append((name, content))
    if not inputs:
        return pd.DataFrame(columns=SALES_COLUMNS), []
    return _cached_sales_dataset(tuple(inputs))


def _ensure_multiselect_key_state(
    state_key: str, options: List[str], default: List[str]
):
    # 第一次出现这个 key，用 default 初始化（注意要跟 options 交集一下）
    if state_key not in st.session_state:
        base = [v for v in default if v in options]
        st.session_state[state_key] = base
        return

    # 之后就只做「清理无效选项」，不再强制塞默认值
    cur = st.session_state.get(state_key, [])
    if isinstance(cur, (str, int, float)):
        cur = [str(cur)]
    # 只保留还在 options 里的选项
    cur = [v for v in cur if v in options]
    st.session_state[state_key] = cur


def apply_sales_filters(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, str]]:
    if df.empty:
        return df, {}

    ss = st.session_state
    session_keys = SALES_FILTER_SESSION_KEYS
    customer_exclude_key = SALES_FILTER_CUSTOMER_EXCLUDE_KEY
    supplier_exclude_key = SALES_FILTER_SUPPLIER_EXCLUDE_KEY

    def _series(name: str) -> pd.Series:
        if name not in df.columns:
            return pd.Series([""] * len(df), index=df.index)
        return df[name].astype("string").fillna("").str.strip()

    series_cache = {name: _series(name) for name in session_keys}
    invoice_series = _series("Invoice #")

    filter_columns = {
        "Year": "Year",
        "Month": "Month",
        "Customer": "Customer",
        "Outlet": "Outlet",
        "Product Description": "Product Description",
        "Supplier": "Supplier",
        "Brand/Category": "Brand/Category",
        "Product Code": "Product Code",
        "Account": "Account",
    }

    date_series_raw = None
    default_date_from: Optional[datetime.date] = None
    default_date_to: Optional[datetime.date] = None
    if "Date" in df.columns:
        date_series_raw = pd.to_datetime(df["Date"], errors="coerce")
        valid_dates = date_series_raw.dropna()
        if not valid_dates.empty:
            default_date_from = valid_dates.min().date()
            default_date_to = valid_dates.max().date()

    date_from: Optional[datetime.date] = None
    date_to: Optional[datetime.date] = None
    exact_date: Optional[datetime.date] = None

    def _mask(
        exclude: Optional[str] = None,
        ignore_customer_exclude: bool = False,
    ) -> pd.Series:
        mask = pd.Series(True, index=df.index)
        for key, column in filter_columns.items():
            if key == exclude:
                continue
            sel = ss.get(session_keys[key], [])
            if not sel:
                continue
            selected_mask = series_cache[column].isin(sel)
            if key == "Supplier" and bool(ss.get(supplier_exclude_key, False)):
                mask &= ~selected_mask
            else:
                mask &= selected_mask
        if not ignore_customer_exclude and exclude != "Exclude Customer":
            excluded_customers = ss.get(customer_exclude_key, [])
            if excluded_customers:
                mask &= ~series_cache["Customer"].isin(excluded_customers)
        if exclude != "Invoice contains":
            invoice_query = ss.get("sales_filter_invoice", "")
            if invoice_query:
                mask &= invoice_series.str.contains(invoice_query, case=False, na=False)
        if date_series_raw is not None:
            if exact_date:
                mask &= date_series_raw.dt.date == exact_date
            elif date_from:
                mask &= date_series_raw >= pd.Timestamp(date_from)
                if date_to:
                    mask &= date_series_raw <= pd.Timestamp(date_to)
            elif date_to:
                mask &= date_series_raw <= pd.Timestamp(date_to)
        return mask

    def _options_for(key: str) -> List[str]:
        column = filter_columns[key]
        mask = _mask(
            exclude=key,
            ignore_customer_exclude=(key == "Customer"),
        )
        values = sorted(
            {val for val in series_cache[column][mask] if val},
            key=_month_sort_key if key == "Month" else lambda x: x.upper(),
        )
        return values

    ss.setdefault(session_keys["Year"], [])
    ss.setdefault(session_keys["Month"], [])
    ss.setdefault("sales_basic_mode", True)

    if st.button("Clear sales filters", key="sales_clear_filters_button"):
        _clear_sales_filters_state(ss)

    with st.form("sales_filters_form"):
        basic_sales_mode = st.toggle(
            "Basic mode",
            key="sales_basic_mode",
            help="Show only core sales filters. Turn off to access advanced filters.",
        )

        year_options = _options_for("Year")
        default_year: List[str] = []
        if year_options:
            target_year = "2025"
            if target_year in year_options:
                default_year = [target_year]
            else:
                default_year = [year_options[-1]]

        month_options = _options_for("Month")
        default_month: List[str] = []

        _ensure_multiselect_key_state(session_keys["Year"], year_options, default_year)
        _ensure_multiselect_key_state(
            session_keys["Month"], month_options, default_month
        )

        customer_options = _options_for("Customer")
        _ensure_multiselect_key_state(
            session_keys["Customer"], customer_options, default=[]
        )
        _ensure_multiselect_key_state(
            customer_exclude_key, customer_options, default=[]
        )
        outlet_options = _options_for("Outlet")
        _ensure_multiselect_key_state(
            session_keys["Outlet"], outlet_options, default=[]
        )

        multi_cols = st.columns(4)
        with multi_cols[0]:
            st.multiselect(
                "Year",
                year_options,
                key=session_keys["Year"],
            )
        with multi_cols[1]:
            st.multiselect(
                "Month",
                month_options,
                key=session_keys["Month"],
            )
        with multi_cols[2]:
            st.multiselect(
                "Customer",
                customer_options,
                key=session_keys["Customer"],
            )
        with multi_cols[3]:
            st.multiselect(
                "Outlet",
                outlet_options,
                key=session_keys["Outlet"],
            )

        product_desc_options = _options_for("Product Description")
        _ensure_multiselect_key_state(
            session_keys["Product Description"], product_desc_options, default=[]
        )
        st.multiselect(
            "Product Description",
            product_desc_options,
            key=session_keys["Product Description"],
        )

        if basic_sales_mode:
            ss[customer_exclude_key] = []
            ss[session_keys["Brand/Category"]] = []
            ss["sales_filter_invoice"] = ""
            ss["sales_filter_date_from"] = None
            ss["sales_filter_date_to"] = None

            supplier_options = _options_for("Supplier")
            _ensure_multiselect_key_state(
                session_keys["Supplier"], supplier_options, default=[]
            )
            product_code_options = _options_for("Product Code")
            _ensure_multiselect_key_state(
                session_keys["Product Code"], product_code_options, default=[]
            )
            account_options = _options_for("Account")
            _ensure_multiselect_key_state(
                session_keys["Account"], account_options, default=[]
            )

            basic_extra_cols = st.columns(3)
            with basic_extra_cols[0]:
                st.multiselect(
                    "Supplier",
                    supplier_options,
                    key=session_keys["Supplier"],
                )
                st.checkbox(
                    "Exclude selected (Supplier)",
                    key=supplier_exclude_key,
                    value=bool(ss.get(supplier_exclude_key, False)),
                )
            with basic_extra_cols[1]:
                st.multiselect(
                    "Product Code",
                    product_code_options,
                    key=session_keys["Product Code"],
                )
            with basic_extra_cols[2]:
                st.multiselect(
                    "Account",
                    account_options,
                    key=session_keys["Account"],
                )

            use_exact_date = st.toggle(
                "指定日期",
                key="sales_filter_use_exact_date",
                help="开启后只显示所选当天的销售记录。",
            )
            if use_exact_date:
                exact_date_default = ss.get(
                    "sales_filter_exact_date",
                    default_date_to or default_date_from or datetime.date.today(),
                )
                exact_date = st.date_input(
                    "Sales date",
                    value=exact_date_default,
                    key="sales_filter_exact_date",
                )
            else:
                ss["sales_filter_exact_date"] = None
        else:
            _ensure_multiselect_key_state(
                customer_exclude_key, customer_options, default=[]
            )
            ss["sales_filter_use_exact_date"] = False
            ss["sales_filter_exact_date"] = None

            with st.expander("高级筛选", expanded=False):
                st.multiselect(
                    "Exclude Customer",
                    customer_options,
                    key=customer_exclude_key,
                    help="Selected customers will be excluded from results.",
                )
                supplier_options = _options_for("Supplier")
                _ensure_multiselect_key_state(
                    session_keys["Supplier"], supplier_options, default=[]
                )
                brand_options = _options_for("Brand/Category")
                _ensure_multiselect_key_state(
                    session_keys["Brand/Category"], brand_options, default=[]
                )
                product_code_options = _options_for("Product Code")
                _ensure_multiselect_key_state(
                    session_keys["Product Code"], product_code_options, default=[]
                )
                account_options = _options_for("Account")
                _ensure_multiselect_key_state(
                    session_keys["Account"], account_options, default=[]
                )
                st.multiselect(
                    "Supplier",
                    supplier_options,
                    key=session_keys["Supplier"],
                )
                st.checkbox(
                    "Exclude selected (Supplier)",
                    key=supplier_exclude_key,
                    value=bool(ss.get(supplier_exclude_key, False)),
                    help="Selected suppliers will be excluded from results.",
                )
                st.multiselect(
                    "Brand/Category",
                    brand_options,
                    key=session_keys["Brand/Category"],
                )
                st.multiselect(
                    "Product Code",
                    product_code_options,
                    key=session_keys["Product Code"],
                )
                st.multiselect(
                    "Account",
                    account_options,
                    key=session_keys["Account"],
                )
                st.text_input(
                    "Invoice # contains",
                    key="sales_filter_invoice",
                    placeholder="",
                )
                date_preset_options = [
                    "Default range",
                    "This week",
                    "Last week",
                    "Next week",
                    "Custom range",
                ]
                ss.setdefault("sales_date_preset", date_preset_options[0])
                preset = st.selectbox(
                    "Date shortcut",
                    date_preset_options,
                    key="sales_date_preset",
                )
                today_date = datetime.date.today()

                def _week_bounds(offset: int) -> Tuple[datetime.date, datetime.date]:
                    monday = today_date - datetime.timedelta(days=today_date.weekday())
                    start = monday + datetime.timedelta(weeks=offset)
                    end = start + datetime.timedelta(days=6)
                    return start, end

                preset_from = default_date_from or today_date
                preset_to = default_date_to or today_date
                if preset == "This week":
                    preset_from, preset_to = _week_bounds(0)
                elif preset == "Last week":
                    preset_from, preset_to = _week_bounds(-1)
                elif preset == "Next week":
                    preset_from, preset_to = _week_bounds(1)
                elif preset == "Default range":
                    preset_from = default_date_from or today_date
                    preset_to = default_date_to or today_date

                if preset != "Custom range":
                    st.caption(f"Quick filter '{preset}': {preset_from} to {preset_to}")
                    date_from = preset_from
                    date_to = preset_to
                    ss["sales_filter_date_from"] = date_from
                    ss["sales_filter_date_to"] = date_to
                else:
                    custom_from_default = ss.get(
                        "sales_filter_date_from",
                        default_date_from or today_date,
                    )
                    custom_to_default = ss.get(
                        "sales_filter_date_to",
                        default_date_to or today_date,
                    )
                    date_from = st.date_input(
                        "Date from",
                        value=custom_from_default,
                        key="sales_filter_date_from",
                    )
                    date_to = st.date_input(
                        "Date to",
                        value=custom_to_default,
                        key="sales_filter_date_to",
                    )

        submitted = st.form_submit_button("Apply sales filters")

    _clear_text_input_on_backspace(
        "sales_filter_invoice", "__prev_sales_filter_invoice"
    )

    filtered_df = df.loc[_mask()].reset_index(drop=True)
    selection_values = {
        "Year": ss.get(session_keys["Year"], []),
        "Month": ss.get(session_keys["Month"], []),
        "Customer": ss.get(session_keys["Customer"], []),
        "Exclude Customer": ss.get(customer_exclude_key, []),
        "Outlet": ss.get(session_keys["Outlet"], []),
        "Product Description": ss.get(session_keys["Product Description"], []),
        "Supplier": ss.get(session_keys["Supplier"], []),
        "Exclude Supplier": ["Yes"] if bool(ss.get(supplier_exclude_key, False)) else [],
        "Brand/Category": ss.get(session_keys["Brand/Category"], []),
        "Product Code": ss.get(session_keys["Product Code"], []),
        "Account": ss.get(session_keys["Account"], []),
    }
    filter_state: Dict[str, str] = {}
    for key, vals in selection_values.items():
        filter_state[key] = ", ".join(vals) if vals else "All"
    invoice_query = ss.get("sales_filter_invoice", "")
    filter_state["Invoice contains"] = invoice_query if invoice_query else "All"
    exact_date_state = ss.get("sales_filter_exact_date")
    filter_state["Exact date"] = (
        exact_date_state.isoformat() if exact_date_state else "All"
    )
    date_from_state = ss.get("sales_filter_date_from")
    date_to_state = ss.get("sales_filter_date_to")
    filter_state["Date from"] = (
        date_from_state.isoformat() if date_from_state else "All"
    )
    filter_state["Date to"] = date_to_state.isoformat() if date_to_state else "All"

    return filtered_df, filter_state


def build_sales_summary(
    filtered_df: pd.DataFrame,
    group_by_outlet: bool,
    include_date: bool,
    group_by_week: bool,
) -> pd.DataFrame:
    include_date = include_date and group_by_outlet and "Date" in filtered_df.columns
    has_outlet = group_by_outlet and "Outlet" in filtered_df.columns
    week_enabled = group_by_week and "Date" in filtered_df.columns

    if filtered_df.empty:
        cols = ["Customer", "Total Value", "Total Qty"]
        if week_enabled:
            cols.insert(0, "Week")
        if not has_outlet:
            cols.append("Selling Price (per pkt)")
        return pd.DataFrame(columns=cols)

    working_df = filtered_df.copy()
    week_label_key = "_week_label"

    def _format_week_label(value: pd.Timestamp) -> str:
        if pd.isna(value):
            return ""
        dt = value.date()
        month_start = dt.replace(day=1)
        _, last_day = calendar.monthrange(dt.year, dt.month)
        month_end = dt.replace(day=last_day)
        week_start = dt - datetime.timedelta(days=dt.weekday())
        if week_start < month_start:
            week_start = month_start
        week_end = week_start + datetime.timedelta(days=4)
        if week_end > month_end:
            week_end = month_end
        return f"{week_start.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')}"

    if week_enabled:
        date_series = pd.to_datetime(working_df["Date"], errors="coerce")
        working_df[week_label_key] = date_series.apply(_format_week_label)
    else:
        working_df[week_label_key] = pd.NA

    group_cols: List[str] = []
    if week_enabled:
        group_cols.append(week_label_key)
    group_cols.append("Customer")
    if include_date:
        group_cols.append("Date")
    if has_outlet:
        group_cols.append("Outlet")

    agg_map = {
        "Qty in Pcs": "sum",
        "Qty in Ctns": "sum",
        "Total Value": "sum",
        "carton_packing_numeric": "first",
        "pcs_to_ctn": "sum",
        "total_ctn_equivalent": "sum",
        "total_packets": "sum",
    }
    for optional_metric in ["pcs_to_ctn", "total_ctn_equivalent", "total_packets"]:
        if optional_metric not in working_df.columns:
            working_df[optional_metric] = 0.0
    summary = working_df.groupby(group_cols, dropna=False, as_index=False).agg(agg_map)

    summary["Qty in Pcs"] = summary["Qty in Pcs"].fillna(0)
    summary["Qty in Ctns"] = summary["Qty in Ctns"].fillna(0)
    summary["Total Value"] = summary["Total Value"].fillna(0)
    summary["Qty in Ctns (from pcs)"] = pd.to_numeric(
        summary.get("pcs_to_ctn"), errors="coerce"
    ).fillna(0.0)
    summary["Total Ctns (all)"] = pd.to_numeric(
        summary.get("total_ctn_equivalent"), errors="coerce"
    ).fillna(0.0)
    summary["Total Packets (all)"] = pd.to_numeric(
        summary.get("total_packets"), errors="coerce"
    ).fillna(0.0)

    summary["Total Qty"] = summary.apply(
        lambda row: _format_summary_total_qty(
            row.get("Qty in Ctns"),
            row.get("Qty in Pcs"),
            row.get("carton_packing_numeric"),
        ),
        axis=1,
    )

    if not has_outlet:
        summary["Selling Price (per pkt)"] = _safe_divide_series(
            summary["Total Value"], summary["Total Packets (all)"]
        )
        summary["Selling Price (per ctn eq)"] = _safe_divide_series(
            summary["Total Value"], summary["Total Ctns (all)"]
        )

    if include_date and "Date" in summary.columns:
        date_sort = pd.to_datetime(summary["Date"], errors="coerce")
        summary["_date_sort"] = date_sort
        summary["Date"] = date_sort.dt.strftime("%d/%m/%Y").fillna("")

    sort_keys = []
    if week_enabled:
        sort_keys.append(week_label_key)
    sort_keys.append("Customer")
    if include_date:
        sort_keys.append("_date_sort")
    if has_outlet:
        sort_keys.append("Outlet")
    summary = summary.sort_values(sort_keys, kind="stable")

    if week_label_key in summary.columns:
        summary["Week"] = summary[week_label_key]
    summary = summary.drop(
        columns=[col for col in sort_keys if col.startswith("_")], errors="ignore"
    )
    summary = summary.reset_index(drop=True)
    summary = summary.drop(
        columns=[
            "carton_packing_numeric",
            "pcs_to_ctn",
            "total_ctn_equivalent",
            "total_packets",
            week_label_key,
        ],
        errors="ignore",
    )
    return summary


def _customer_purchase_breakdown(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    group_cols = [
        col
        for col in [
            "Customer",
            "Month",
            "Brand/Category",
            "Supplier",
            "Product Description",
            "Product Code",
        ]
        if col in df.columns
    ]
    agg_map = {
        col: "sum"
        for col in ["Qty in Ctns", "Qty in Pcs", "Total Value"]
        if col in df.columns
    }
    if "carton_packing_numeric" in df.columns:
        agg_map["carton_packing_numeric"] = "first"
    if not group_cols or not agg_map:
        return pd.DataFrame()
    summary = (
        df.groupby(group_cols, dropna=False, sort=False).agg(agg_map).reset_index()
    )
    for qty_col in ["Qty in Ctns", "Qty in Pcs"]:
        if qty_col in summary:
            summary[qty_col] = summary[qty_col].fillna(0)
    if "Total Value" in summary:
        summary["Total Value"] = summary["Total Value"].fillna(0)

    def _row_total_qty(row):
        return _format_summary_total_qty(
            row.get("Qty in Ctns"),
            row.get("Qty in Pcs"),
            row.get("carton_packing_numeric"),
        )

    summary["Total Qty"] = summary.apply(_row_total_qty, axis=1)

    pack_size_series = pd.to_numeric(
        summary.get(
            "carton_packing_numeric",
            pd.Series(0.0, index=summary.index),
        ),
        errors="coerce",
    ).fillna(0.0)
    total_ctns = pd.to_numeric(
        summary.get("Qty in Ctns", pd.Series(0.0, index=summary.index)),
        errors="coerce",
    ).fillna(0.0)
    total_pkts = pd.to_numeric(
        summary.get("Qty in Pcs", pd.Series(0.0, index=summary.index)),
        errors="coerce",
    ).fillna(0.0)
    total_packets = total_ctns * pack_size_series + total_pkts
    if "Total Value" in summary:
        summary["Selling Price (per pkt)"] = _safe_divide_series(
            summary["Total Value"],
            total_packets,
        )
    if "Selling Price (per pkt)" in summary:
        summary["Selling Price (per pkt)"] = summary["Selling Price (per pkt)"].apply(
            _format_price_display
        )

    summary_cols = group_cols + ["Total Qty"]
    if "Selling Price (per pkt)" in summary:
        summary_cols.append("Selling Price (per pkt)")
    return summary[summary_cols]


def build_sales_usage_views(
    df: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    monthly_cols = ["Month", "Total Qty", "Total Value"]
    customer_cols = ["Customer", "Total Qty", "Total Value"]
    matrix_default = pd.DataFrame(columns=["Account", "Customer", "Overall"])

    if df.empty:
        return (
            pd.DataFrame(columns=monthly_cols),
            pd.DataFrame(columns=customer_cols),
            matrix_default,
        )

    working = df.copy()
    for numeric_col in [
        "Qty in Ctns",
        "Qty in Pcs",
        "Total Value",
        "total_ctn_equivalent",
    ]:
        if numeric_col not in working.columns:
            working[numeric_col] = 0.0
        working[numeric_col] = pd.to_numeric(
            working[numeric_col], errors="coerce"
        ).fillna(0.0)

    if "Customer" in working.columns:
        customer_series = working["Customer"].astype("string").fillna("").str.strip()
    else:
        customer_series = pd.Series(
            [""] * len(working), index=working.index, dtype="string"
        )
    working["__customer"] = customer_series.replace("", pd.NA).fillna("Unknown")

    if "Account" in working.columns:
        account_series = working["Account"].astype("string").fillna("").str.strip()
    else:
        account_series = pd.Series(
            [""] * len(working), index=working.index, dtype="string"
        )
    working["__account"] = account_series

    if "Year" in working.columns:
        year_series = working["Year"].astype("string").fillna("").str.strip()
    else:
        year_series = pd.Series(
            [""] * len(working), index=working.index, dtype="string"
        )
    if "Month" in working.columns:
        month_series = working["Month"].astype("string").fillna("").str.strip()
    else:
        month_series = pd.Series(
            [""] * len(working), index=working.index, dtype="string"
        )

    month_label = (
        month_series.apply(_format_month_label).replace("", pd.NA).fillna("Unknown")
    )
    year_clean = year_series.replace("", pd.NA)
    working["__month_label"] = month_label.where(
        year_clean.isna(),
        year_clean.fillna("") + "-" + month_label,
    )

    month_sort = month_series.map(_month_sort_key)
    month_sort = pd.to_numeric(month_sort, errors="coerce").fillna(13)
    month_sort = month_sort.where(month_sort != 999, 13)
    year_sort = pd.to_numeric(year_clean, errors="coerce").fillna(9999)
    working["__month_order"] = year_sort * 100 + month_sort

    def _pack_signature(series: pd.Series) -> Tuple[float, ...]:
        values = pd.to_numeric(series, errors="coerce").dropna()
        values = values[values > 0]
        if values.empty:
            return tuple()
        uniq = sorted({float(v) for v in values.tolist()})
        return tuple(uniq)

    def _format_group_qty(
        ctn_value: float, pcs_value: float, packs: Tuple[float, ...]
    ) -> str:
        if packs and len(packs) == 1:
            pack_size = int(round(float(packs[0])))
            if pack_size > 0:
                total_packets = int(
                    round(float(ctn_value) * pack_size + float(pcs_value))
                )
                return _format_total_qty_text(total_packets, pack_size)
        return _format_quantity_pair(ctn_value, pcs_value)

    monthly_usage = (
        working.groupby(
            ["__month_label", "__month_order"], as_index=False, dropna=False
        )
        .agg(
            {
                "Qty in Ctns": "sum",
                "Qty in Pcs": "sum",
                "Total Value": "sum",
                "carton_packing_numeric": _pack_signature,
            }
        )
        .sort_values(["__month_order", "__month_label"], kind="stable")
    )
    monthly_usage["Total Qty"] = monthly_usage.apply(
        lambda row: _format_group_qty(
            row.get("Qty in Ctns", 0.0),
            row.get("Qty in Pcs", 0.0),
            row.get("carton_packing_numeric", tuple()),
        ),
        axis=1,
    )
    monthly_usage = monthly_usage.rename(columns={"__month_label": "Month"})
    monthly_usage = monthly_usage.drop(
        columns=["carton_packing_numeric"], errors="ignore"
    )
    monthly_usage = monthly_usage[monthly_cols].reset_index(drop=True)

    customer_usage = (
        working.groupby("__customer", as_index=False, dropna=False)[
            ["Qty in Ctns", "Qty in Pcs", "Total Value", "carton_packing_numeric"]
        ]
        .agg(
            {
                "Qty in Ctns": "sum",
                "Qty in Pcs": "sum",
                "Total Value": "sum",
                "carton_packing_numeric": _pack_signature,
            }
        )
        .rename(columns={"__customer": "Customer"})
        .sort_values(
            ["Qty in Ctns", "Qty in Pcs", "Total Value", "Customer"],
            ascending=[False, False, False, True],
            kind="stable",
        )
        .reset_index(drop=True)
    )
    customer_usage["Total Qty"] = customer_usage.apply(
        lambda row: _format_group_qty(
            row.get("Qty in Ctns", 0.0),
            row.get("Qty in Pcs", 0.0),
            row.get("carton_packing_numeric", tuple()),
        ),
        axis=1,
    )
    customer_usage = customer_usage.drop(
        columns=["carton_packing_numeric"], errors="ignore"
    )
    customer_usage = customer_usage[customer_cols]

    customer_month = working.groupby(
        ["__customer", "__month_label", "__month_order"], as_index=False, dropna=False
    ).agg(
        {
            "Qty in Ctns": "sum",
            "Qty in Pcs": "sum",
            "carton_packing_numeric": _pack_signature,
        }
    )
    customer_month["Total Qty"] = customer_month.apply(
        lambda row: _format_group_qty(
            row.get("Qty in Ctns", 0.0),
            row.get("Qty in Pcs", 0.0),
            row.get("carton_packing_numeric", tuple()),
        ),
        axis=1,
    )
    if customer_month.empty:
        customer_month_matrix = matrix_default
    else:
        ordered_months = (
            customer_month[["__month_label", "__month_order"]]
            .drop_duplicates()
            .sort_values(["__month_order", "__month_label"], kind="stable")
        )
        customer_month_matrix = (
            customer_month.pivot(
                index="__customer",
                columns="__month_label",
                values="Total Qty",
            )
            .fillna("-")
            .reindex(columns=ordered_months["__month_label"].tolist())
            .reset_index()
            .rename(columns={"__customer": "Customer"})
        )
        account_map = (
            working.groupby("__customer", dropna=False)["__account"]
            .apply(
                lambda s: ", ".join(
                    sorted({str(v).strip() for v in s if str(v).strip()})
                )
            )
            .to_dict()
        )
        overall_map = customer_usage.set_index("Customer")["Total Qty"]
        customer_month_matrix.insert(
            0,
            "Account",
            customer_month_matrix["Customer"].map(account_map).fillna(""),
        )
        customer_month_matrix.insert(
            2,
            "Overall",
            customer_month_matrix["Customer"].map(overall_map).fillna("0"),
        )

    return monthly_usage, customer_usage, customer_month_matrix


def build_product_monthly_summary(df: pd.DataFrame) -> pd.DataFrame:
    columns = ["年月", "供应商", "产品描述", "产品编码", "箱规", "总销量", "总销售额"]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)

    work = df.copy()
    for col in ["Qty in Ctns", "Qty in Pcs", "Total Value", "carton_packing_numeric"]:
        if col not in work.columns:
            work[col] = 0.0
        work[col] = pd.to_numeric(work[col], errors="coerce").fillna(0.0)

    if "YearMonth" not in work.columns:
        if "Date" in work.columns:
            date_series = pd.to_datetime(work["Date"], errors="coerce")
            work["YearMonth"] = date_series.dt.strftime("%Y-%b")
            work.loc[date_series.isna(), "YearMonth"] = "未知"
        else:
            work["YearMonth"] = "未知"

    for col in ["Supplier", "Product Description", "Product Code", "Carton Packing"]:
        if col not in work.columns:
            work[col] = ""
        work[col] = work[col].astype("string").fillna("").str.strip()
    work["_sales_pack_group_key"] = work["Carton Packing"].map(
        _normalize_sales_pack_group_key
    )

    def _pack_signature(series: pd.Series) -> Tuple[float, ...]:
        values = pd.to_numeric(series, errors="coerce").dropna()
        values = values[values > 0]
        if values.empty:
            return tuple()
        return tuple(sorted({float(v) for v in values.tolist()}))

    def _qty_text(ctn: float, pcs: float, packs: Tuple[float, ...]) -> str:
        if packs and len(packs) == 1:
            size = int(round(float(packs[0])))
            if size > 0:
                total_packets = int(round(float(ctn) * size + float(pcs)))
                return _format_total_qty_text(total_packets, size)
        return _format_quantity_pair(float(ctn), float(pcs))

    grouped = (
        work.groupby(
            [
                "YearMonth",
                "Supplier",
                "Product Description",
                "Product Code",
                "_sales_pack_group_key",
            ],
            dropna=False,
            as_index=False,
        )
        .agg(
            {
                "Carton Packing": "last",
                "Qty in Ctns": "sum",
                "Qty in Pcs": "sum",
                "Total Value": "sum",
                "carton_packing_numeric": _pack_signature,
            }
        )
        .copy()
    )

    month_sort = pd.to_datetime(grouped["YearMonth"], format="%Y-%b", errors="coerce")
    grouped["_month_sort"] = month_sort.fillna(pd.Timestamp.max)
    grouped["总销量"] = grouped.apply(
        lambda row: _qty_text(
            row.get("Qty in Ctns", 0.0),
            row.get("Qty in Pcs", 0.0),
            row.get("carton_packing_numeric", tuple()),
        ),
        axis=1,
    )
    grouped["总销售额"] = grouped["Total Value"].apply(_format_price_display)

    result = (
        grouped.sort_values(
            [
                "_month_sort",
                "YearMonth",
                "Supplier",
                "Product Description",
                "Product Code",
                "_sales_pack_group_key",
                "Carton Packing",
            ],
            kind="stable",
        )
        .rename(
            columns={
                "YearMonth": "年月",
                "Supplier": "供应商",
                "Product Description": "产品描述",
                "Product Code": "产品编码",
                "Carton Packing": "箱规",
            }
        )
        .loc[:, columns]
        .reset_index(drop=True)
    )
    return result


def build_daily_sales_timeline(df: pd.DataFrame) -> pd.DataFrame:
    columns = ["年月", "日期", "总销量", "总销售额"]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)

    work = df.copy()
    for col in ["Qty in Ctns", "Qty in Pcs", "Total Value", "carton_packing_numeric"]:
        if col not in work.columns:
            work[col] = 0.0
        work[col] = pd.to_numeric(work[col], errors="coerce").fillna(0.0)

    if "Date" in work.columns:
        work["Date"] = pd.to_datetime(work["Date"], errors="coerce", format="mixed")
    else:
        work["Date"] = pd.NaT

    work = work.dropna(subset=["Date"]).copy()
    if work.empty:
        return pd.DataFrame(columns=columns)

    work["SaleDay"] = work["Date"].dt.normalize()
    work["YearMonth"] = work["Date"].dt.strftime("%Y-%b")
    work["YearMonthPeriod"] = work["Date"].dt.to_period("M")
    work["__daily_total_packets"] = (
        work["Qty in Ctns"] * work["carton_packing_numeric"] + work["Qty in Pcs"]
    )

    def _pack_signature(series: pd.Series) -> Tuple[float, ...]:
        values = pd.to_numeric(series, errors="coerce").dropna()
        values = values[values > 0]
        if values.empty:
            return tuple()
        return tuple(sorted({float(v) for v in values.tolist()}))

    def _qty_text(
        ctn: float, pcs: float, packs: Tuple[float, ...], total_packets_value: float
    ) -> str:
        if packs and len(packs) == 1:
            size = int(round(float(packs[0])))
            if size > 0:
                total_packets = int(round(float(total_packets_value)))
                return _format_total_qty_text(total_packets, size)
        return f"{_format_qty_display(total_packets_value)} pkts"

    grouped = (
        work.groupby(
            ["YearMonthPeriod", "YearMonth", "SaleDay"], dropna=False, as_index=False
        )
        .agg(
            {
                "Qty in Ctns": "sum",
                "Qty in Pcs": "sum",
                "Total Value": "sum",
                "carton_packing_numeric": _pack_signature,
                "__daily_total_packets": "sum",
            }
        )
        .sort_values(["YearMonthPeriod", "SaleDay"], kind="stable")
        .reset_index(drop=True)
    )

    grouped["总销量"] = grouped.apply(
        lambda row: _qty_text(
            row.get("Qty in Ctns", 0.0),
            row.get("Qty in Pcs", 0.0),
            row.get("carton_packing_numeric", tuple()),
            row.get("__daily_total_packets", 0.0),
        ),
        axis=1,
    )
    grouped["总销售额"] = grouped["Total Value"].apply(_format_price_display)
    grouped["日期"] = grouped["SaleDay"].apply(_format_date_with_month_label)

    return grouped.loc[:, ["YearMonth", "日期", "总销量", "总销售额"]].rename(
        columns={"YearMonth": "年月"}
    )


def compute_sales_amount_kpis(df: pd.DataFrame) -> Tuple[float, float]:
    if df is None or df.empty:
        return 0.0, 0.0

    work = df.copy()
    if "Date" in work.columns:
        work["Date"] = pd.to_datetime(work["Date"], errors="coerce")
    else:
        work["Date"] = pd.NaT

    latest_date = work["Date"].dropna().max()
    if pd.isna(latest_date):
        return 0.0, 0.0

    value_col = (
        "Total Value Inclusive GST"
        if "Total Value Inclusive GST" in work.columns
        else "Total Value"
    )
    if value_col not in work.columns:
        work[value_col] = 0.0
    values = pd.to_numeric(work[value_col], errors="coerce").fillna(0.0)

    this_month_mask = (work["Date"].dt.year == latest_date.year) & (
        work["Date"].dt.month == latest_date.month
    )
    this_year_mask = work["Date"].dt.year == latest_date.year

    month_total = float(values[this_month_mask].sum())
    year_total = float(values[this_year_mask].sum())
    return month_total, year_total


def _render_sales_business_dashboard(filtered_df: pd.DataFrame) -> None:
    if filtered_df is None or filtered_df.empty:
        st.info("当前筛选未产出任何销售数据。")
        return

    work = filtered_df.copy()
    for col in ["Qty in Ctns", "Qty in Pcs", "Total Value", "total_packets"]:
        if col not in work.columns:
            work[col] = 0.0
        work[col] = pd.to_numeric(work[col], errors="coerce").fillna(0.0)

    if "Date" in work.columns:
        work["Date"] = pd.to_datetime(work["Date"], errors="coerce")
    else:
        work["Date"] = pd.NaT

    for col in [
        "Customer",
        "Account",
        "Supplier",
        "Product Description",
        "Product Code",
        "Carton Packing",
    ]:
        if col not in work.columns:
            work[col] = ""
        work[col] = work[col].astype("string").fillna("").str.strip()
    work["_sales_pack_group_key"] = work["Carton Packing"].map(
        _normalize_sales_pack_group_key
    )

    work["YearMonth"] = work["Date"].dt.strftime("%Y-%b")
    work["YearMonthPeriod"] = work["Date"].dt.to_period("M")
    work.loc[work["Date"].isna(), "YearMonth"] = "未知"

    def _pack_signature(series: pd.Series) -> Tuple[float, ...]:
        values = pd.to_numeric(series, errors="coerce").dropna()
        values = values[values > 0]
        if values.empty:
            return tuple()
        return tuple(sorted({float(v) for v in values.tolist()}))

    def _qty_text(ctn: float, pcs: float, packs: Tuple[float, ...]) -> str:
        if packs and len(packs) == 1:
            size = int(round(float(packs[0])))
            if size > 0:
                total_packets = int(round(float(ctn) * size + float(pcs)))
                return _format_total_qty_text(total_packets, size)
        return _format_quantity_pair(ctn, pcs)

    # KPI block: this month, this year, averages
    latest_date = work["Date"].dropna().max()
    if pd.notna(latest_date):
        this_month_mask = (work["Date"].dt.year == latest_date.year) & (
            work["Date"].dt.month == latest_date.month
        )
        this_year_mask = work["Date"].dt.year == latest_date.year
    else:
        this_month_mask = pd.Series(False, index=work.index)
        this_year_mask = pd.Series(False, index=work.index)

    month_ctn = float(work.loc[this_month_mask, "Qty in Ctns"].sum())
    month_pcs = float(work.loc[this_month_mask, "Qty in Pcs"].sum())
    month_val, year_val = compute_sales_amount_kpis(work)

    year_ctn = float(work.loc[this_year_mask, "Qty in Ctns"].sum())
    year_pcs = float(work.loc[this_year_mask, "Qty in Pcs"].sum())

    def _kpi_qty_text(scope_df: pd.DataFrame) -> str:
        if scope_df is None or scope_df.empty:
            return "0"

        if "total_packets" in scope_df.columns:
            total_packets_value = float(
                pd.to_numeric(scope_df["total_packets"], errors="coerce")
                .fillna(0)
                .sum()
            )
        else:
            ctn_series = pd.to_numeric(
                scope_df.get("Qty in Ctns"), errors="coerce"
            ).fillna(0)
            pcs_series = pd.to_numeric(
                scope_df.get("Qty in Pcs"), errors="coerce"
            ).fillna(0)
            pack_series = pd.to_numeric(
                scope_df.get("carton_packing_numeric"), errors="coerce"
            ).fillna(0)
            total_packets_value = float((ctn_series * pack_series + pcs_series).sum())

        if "carton_packing_numeric" in scope_df.columns:
            pack_values = pd.to_numeric(
                scope_df["carton_packing_numeric"], errors="coerce"
            ).dropna()
            pack_values = pack_values[pack_values > 0]
        else:
            pack_values = pd.Series(dtype="float64")

        unique_pack_sizes = sorted(
            {
                int(round(float(value)))
                for value in pack_values.tolist()
                if float(value) > 0
            }
        )

        if len(unique_pack_sizes) == 1:
            return _format_total_qty_text(
                int(round(total_packets_value)), unique_pack_sizes[0]
            )

        return f"{_format_qty_display(total_packets_value)} pkts"

    month_qty_text = _kpi_qty_text(work.loc[this_month_mask])
    year_qty_text = _kpi_qty_text(work.loc[this_year_mask])

    weekly_packets = (
        work.dropna(subset=["Date"])
        .groupby(work["Date"].dt.to_period("W"))["total_packets"]
        .sum()
    )
    monthly_packets = work.groupby("YearMonth", dropna=False)["total_packets"].sum()
    avg_week_packets = float(weekly_packets.mean()) if not weekly_packets.empty else 0.0
    avg_month_packets = (
        float(monthly_packets.mean()) if not monthly_packets.empty else 0.0
    )

    kpi_cols = st.columns(6)
    kpi_cols[0].metric("本月销量", month_qty_text)
    kpi_cols[1].metric("本月销售额(含GST)", _format_price_display(month_val))
    kpi_cols[2].metric("本年销量", year_qty_text)
    kpi_cols[3].metric("本年销售额(含GST)", _format_price_display(year_val))
    kpi_cols[4].metric("周均销量(包)", _format_qty_display(avg_week_packets))
    kpi_cols[5].metric("月均销量(包)", _format_qty_display(avg_month_packets))

    product_monthly_view = build_product_monthly_summary(work)
    st.subheader("产品月度销量（不分客户）")
    _render_copy_values_button(
        _build_copy_values_tsv(product_monthly_view),
        "sales-product-monthly",
        has_rows=not product_monthly_view.empty,
    )
    st.dataframe(product_monthly_view, width="stretch", hide_index=True)

    daily_sales_timeline = build_daily_sales_timeline(work)
    st.subheader("每日销量时间线（不分客户）")
    _render_copy_values_button(
        _build_copy_values_tsv(daily_sales_timeline),
        "sales-daily-timeline",
        has_rows=not daily_sales_timeline.empty,
    )
    st.dataframe(daily_sales_timeline, width="stretch", hide_index=True)

    strict_key_cols = [
        "Customer",
        "Account",
        "Supplier",
        "Product Description",
        "Product Code",
        "_sales_pack_group_key",
    ]

    # Which customer bought how much in which month (strictly split by product/account keys)
    by_customer_month = (
        work.groupby(["YearMonthPeriod", "YearMonth"] + strict_key_cols, dropna=False)
        .agg(
            {
                "Carton Packing": "last",
                "Qty in Ctns": "sum",
                "Qty in Pcs": "sum",
                "Total Value": "sum",
                "carton_packing_numeric": _pack_signature,
            }
        )
        .reset_index()
        .sort_values(
            [
                "YearMonthPeriod",
                "YearMonth",
                "Customer",
                "Account",
                "Supplier",
                "Product Description",
                "Product Code",
                "_sales_pack_group_key",
                "Carton Packing",
            ],
            kind="stable",
        )
    )
    by_customer_month["Total Qty"] = by_customer_month.apply(
        lambda r: _qty_text(
            r.get("Qty in Ctns", 0.0),
            r.get("Qty in Pcs", 0.0),
            r.get("carton_packing_numeric", tuple()),
        ),
        axis=1,
    )
    by_customer_month = _add_sales_price_metrics(by_customer_month)
    by_customer_month["Base Selling Price"] = by_customer_month[
        "Base Selling Price"
    ].apply(_format_price_display)
    by_customer_month["Unit Price per kg"] = by_customer_month[
        "Unit Price per kg"
    ].apply(_format_price_display)
    by_customer_month["Total Value"] = by_customer_month["Total Value"].apply(
        _format_price_display
    )
    st.subheader("客户月度采购明细")
    month_view = by_customer_month[
        [
            "YearMonth",
            "Customer",
            "Account",
            "Supplier",
            "Product Description",
            "Product Code",
            "Carton Packing",
            "Total Qty",
            "Base Selling Price",
            "Unit Price per kg",
        ]
    ].rename(
        columns={
            "YearMonth": "年月",
            "Customer": "客户",
            "Account": "账户",
            "Supplier": "供应商",
            "Product Description": "产品描述",
            "Product Code": "产品编码",
            "Carton Packing": "箱规",
            "Total Qty": "总销量",
            "Base Selling Price": "基础卖价",
            "Unit Price per kg": "每公斤卖价",
        }
    )
    _render_copy_values_button(
        _build_copy_values_tsv(month_view),
        "sales-customer-monthly",
        has_rows=not month_view.empty,
    )
    st.dataframe(month_view, width="stretch", hide_index=True)

    # Which customers buy, when, and how much (strictly split by product/account keys)
    by_customer_date = (
        work.dropna(subset=["Date"])
        .groupby(["Date"] + strict_key_cols, dropna=False)
        .agg(
            {
                "Carton Packing": "last",
                "Qty in Ctns": "sum",
                "Qty in Pcs": "sum",
                "Total Value": "sum",
                "carton_packing_numeric": _pack_signature,
            }
        )
        .reset_index()
        .sort_values(["Date", "Customer"], kind="stable")
    )
    if not by_customer_date.empty:
        by_customer_date["Date"] = by_customer_date["Date"].apply(
            _format_date_with_month_label
        )
        by_customer_date["Qty"] = by_customer_date.apply(
            lambda r: _qty_text(
                r.get("Qty in Ctns", 0.0),
                r.get("Qty in Pcs", 0.0),
                r.get("carton_packing_numeric", tuple()),
            ),
            axis=1,
        )
        by_customer_date = _add_sales_price_metrics(by_customer_date)
        by_customer_date["Base Price"] = by_customer_date["Base Selling Price"].apply(
            _format_price_display
        )
        by_customer_date["Unit Price/kg"] = by_customer_date["Unit Price per kg"].apply(
            _format_price_display
        )
        by_customer_date["Value"] = by_customer_date["Total Value"].apply(
            _format_price_display
        )
    st.subheader("客户采购时间线")
    date_view = (
        by_customer_date[
            [
                "Date",
                "Customer",
                "Account",
                "Supplier",
                "Product Description",
                "Product Code",
                "Carton Packing",
                "Qty",
                "Base Price",
                "Unit Price/kg",
            ]
        ]
        if not by_customer_date.empty
        else pd.DataFrame(
            columns=[
                "Date",
                "Customer",
                "Account",
                "Supplier",
                "Product Description",
                "Product Code",
                "Carton Packing",
                "Qty",
                "Base Price",
                "Unit Price/kg",
            ]
        )
    ).rename(
        columns={
            "Date": "日期",
            "Customer": "客户",
            "Account": "账户",
            "Supplier": "供应商",
            "Product Description": "产品描述",
            "Product Code": "产品编码",
            "Carton Packing": "箱规",
            "Qty": "销量",
            "Base Price": "基础卖价",
            "Unit Price/kg": "每公斤卖价",
        }
    )
    _render_copy_values_button(
        _build_copy_values_tsv(date_view),
        "sales-customer-timeline",
        has_rows=not date_view.empty,
    )
    st.dataframe(date_view, width="stretch", hide_index=True)

    # 下月预测：供应商取最新，历史数量可跨供应商（同产品+同箱规）
    forecast_identity_cols = [
        "Customer",
        "Account",
        "Product Description",
        "Product Code",
        "_sales_pack_group_key",
    ]
    forecast_source = work[work["YearMonthPeriod"].notna()].copy()
    latest_supplier_df = (
        forecast_source.sort_values("Date", kind="stable")
        .drop_duplicates(subset=forecast_identity_cols, keep="last")
        .loc[:, forecast_identity_cols + ["Supplier"]]
    )
    latest_supplier_map = {
        tuple(row[col] for col in forecast_identity_cols): row["Supplier"]
        for _, row in latest_supplier_df.iterrows()
    }

    monthly_key = (
        forecast_source.groupby(
            forecast_identity_cols + ["YearMonthPeriod"], dropna=False
        )
        .agg(
            {
                "Carton Packing": "last",
                "total_packets": "sum",
                "Total Value": "sum",
                "Qty in Ctns": "sum",
                "Qty in Pcs": "sum",
            }
        )
        .reset_index()
    )
    monthly_key["YearMonth"] = monthly_key["YearMonthPeriod"].dt.strftime("%Y-%b")

    forecast_rows: List[Dict[str, Any]] = []
    for _, grp in monthly_key.groupby(forecast_identity_cols, dropna=False):
        identity = {
            col: (grp[col].iloc[0] if not grp.empty else "")
            for col in forecast_identity_cols
        }
        key_tuple = tuple(identity[col] for col in forecast_identity_cols)
        hist = grp.sort_values("YearMonthPeriod", kind="stable")
        last3 = hist.tail(3)
        if last3.empty:
            continue
        avg_packets = float(last3["total_packets"].mean())
        avg_value = float(last3["Total Value"].mean())
        pack_label = str(last3["Carton Packing"].iloc[-1]).strip()
        pack_size = _infer_carton_pack_size(pack_label)
        latest_supplier = latest_supplier_map.get(key_tuple, "")
        if pack_size and pack_size > 0:
            qty_text = _format_total_qty_text(
                int(round(avg_packets)), int(round(pack_size))
            )
        else:
            qty_text = _format_qty_display(avg_packets)
        forecast_rows.append(
            {
                "客户": identity["Customer"],
                "账户": identity["Account"],
                "供应商": latest_supplier,
                "产品描述": identity["Product Description"],
                "产品编码": identity["Product Code"],
                "箱规": pack_label,
                "下月预测销量": qty_text,
                "下月预测销售额": _format_price_display(avg_value),
                "_forecast_packets": avg_packets,
                "_forecast_value": avg_value,
            }
        )

    forecast_df = _append_sales_forecast_total_row(pd.DataFrame(forecast_rows))
    if not forecast_df.empty:
        forecast_df = forecast_df.sort_values(
            by=["客户", "产品描述", "产品编码"],
            kind="stable",
            key=lambda series: series.where(series != "Total", "ZZZ_Total"),
        ).reset_index(drop=True)
    st.subheader("下月预测（近3个月均值基线）")
    st.dataframe(
        forecast_df.drop(
            columns=["_forecast_packets", "_forecast_value"], errors="ignore"
        )
        if not forecast_df.empty
        else pd.DataFrame(
            columns=[
                "客户",
                "账户",
                "供应商",
                "产品描述",
                "产品编码",
                "箱规",
                "下月预测销量",
                "下月预测销售额",
            ]
        ),
        width="stretch",
        hide_index=True,
    )


# ----------------------------- Sales 页 -----------------------------


def run_sales_page():
    st.title("销售看板")
    st.caption(
        "上传一个或多个包含 'Delivery details' 工作表的文件（表头在第4行，A~U列）。"
    )

    uploaded = st.file_uploader(
        "上传销售 Excel (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=True,
        key="sales_uploader",
        on_change=_handle_sales_uploader_change,
    )

    # 朴实无华缓存：以 bytes 持久化，避免页面切换后对象失效/重置
    if uploaded and not st.session_state.get("sales_files_payload"):
        _cache_sales_upload(uploaded)

    sales_payload = st.session_state.get("sales_files_payload", [])
    sales_files_browsed_at = st.session_state.get("sales_files_browsed_at", "")

    if not sales_payload:
        st.info("请至少上传一个 Delivery details 工作簿后继续。")
        return

    if sales_files_browsed_at:
        st.caption(f"Last browsed: {sales_files_browsed_at}")

    try:
        sales_df, warns = load_sales_data_from_payload(sales_payload)
    except Exception as exc:
        st.error(f"读取销售工作簿失败：{exc}")
        return

    for warn in warns:
        st.warning(warn)

    if sales_df.empty:
        st.warning("无法从上传的工作簿读取任何 Sales 数据。")
        return

    filtered_df, filter_state = apply_sales_filters(sales_df)
    filter_order = [
        "Year",
        "Month",
        "Customer",
        "Exclude Customer",
        "Outlet",
        "Product Description",
        "Supplier",
        "Brand/Category",
        "Product Code",
        "Account",
        "Invoice contains",
        "Exact date",
        "Date from",
        "Date to",
    ]
    filter_summary_items = [
        f"{key} = {filter_state.get(key)}"
        for key in filter_order
        if filter_state.get(key) and filter_state.get(key) != "All"
    ]
    st.caption(
        "当前筛选：" + "；".join(filter_summary_items)
        if filter_summary_items
        else "当前筛选：全部数据"
    )

    _render_sales_business_dashboard(filtered_df)

    with st.expander("销售原始明细", expanded=False):
        detail_display = filtered_df.copy()
        if "Date" in detail_display.columns:
            if pd.api.types.is_datetime64_any_dtype(detail_display["Date"]):
                detail_display["Date"] = (
                    detail_display["Date"].dt.strftime("%d-%b-%Y").fillna("")
                )
            else:
                detail_display["Date"] = (
                    detail_display["Date"].astype("string").str.strip()
                )
        for qty_col in ["Qty in Pcs", "Qty in Ctns"]:
            if qty_col in detail_display.columns:
                detail_display[qty_col] = detail_display[qty_col].apply(
                    _format_qty_display
                )
        for conversion_col in ["pcs_to_ctn", "total_ctn_equivalent"]:
            if conversion_col in detail_display.columns:
                detail_display[conversion_col] = detail_display[conversion_col].apply(
                    _format_ctn_equivalent_display
                )
        if "total_packets" in detail_display.columns:
            detail_display["total_packets"] = detail_display["total_packets"].apply(
                _format_qty_display
            )
        for price_col in [
            "Total Value",
            "Unit Price (per pcs)",
            "Unit Price (per carton)",
            "Unit Price (per ctn equivalent)",
        ]:
            if price_col in detail_display.columns:
                detail_display[price_col] = detail_display[price_col].apply(
                    _format_price_display
                )
        if "Month" in detail_display.columns:
            detail_display["Month"] = detail_display["Month"].apply(_format_month_label)

        if detail_display.empty:
            st.info("当前筛选下没有明细数据。")
        else:
            st.dataframe(detail_display, width="stretch", hide_index=True)

        export_df = _strip_html_df(filtered_df)
        csv_buffer = io.StringIO()
        export_df.to_csv(csv_buffer, index=False)
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            export_df.to_excel(writer, sheet_name="Sales Detail", index=False)
            meta_rows = [{"key": k, "value": str(v)} for k, v in filter_state.items()]
            pd.DataFrame(meta_rows, columns=["key", "value"]).to_excel(
                writer, sheet_name="Filters", index=False
            )
        excel_buffer.seek(0)

        sales_download_options = {
            "明细 CSV": {
                "data": csv_buffer.getvalue(),
                "file_name": "sales_detail.csv",
                "mime": "text/csv",
            },
            "明细 Excel": {
                "data": excel_buffer.getvalue(),
                "file_name": "sales_detail.xlsx",
                "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
        }
        sales_selected_download = st.selectbox(
            "下载包",
            options=list(sales_download_options.keys()),
            key="sales_download_package",
        )
        sales_payload = sales_download_options[sales_selected_download]
        st.download_button(
            "下载",
            data=sales_payload["data"],
            file_name=sales_payload["file_name"],
            mime=sales_payload["mime"],
        )

        st.markdown("### Price List by Account")
        account_options = sorted(
            [
                x
                for x in sales_df.get("Account", pd.Series(dtype="string"))
                .astype("string")
                .fillna("")
                .str.strip()
                .unique()
                .tolist()
                if x
            ]
        )
        if not account_options:
            st.info("当前数据没有可用 Account，无法生成 Price List。")
        else:
            selected_account = st.selectbox(
                "选择 Account",
                options=account_options,
                key="sales_price_list_account",
            )
            price_list_df = build_account_price_list(sales_df, selected_account)
            if price_list_df.empty:
                st.info("该 Account 暂无可用于计算 price/kg 的最新记录。")
            else:
                price_list_display = price_list_df.copy()
                price_list_display["Selling Price by kg"] = price_list_display[
                    "Selling Price by kg"
                ].apply(_format_price_display)
                st.dataframe(price_list_display, width="stretch", hide_index=True)

                price_csv = io.StringIO()
                price_list_df.to_csv(price_csv, index=False)

                price_excel = io.BytesIO()
                with pd.ExcelWriter(price_excel, engine="openpyxl") as writer:
                    price_list_df.to_excel(writer, sheet_name="Price List", index=False)
                price_excel.seek(0)

                dl_cols = st.columns(2)
                dl_cols[0].download_button(
                    "下载 Price List CSV",
                    data=price_csv.getvalue(),
                    file_name=f"price_list_{selected_account}.csv",
                    mime="text/csv",
                )
                dl_cols[1].download_button(
                    "下载 Price List Excel",
                    data=price_excel.getvalue(),
                    file_name=f"price_list_{selected_account}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    return


def _highlight_reconciliation_row(row: pd.Series) -> List[str]:
    status = str(row.get("mismatch_reason", ""))
    if status == "match":
        color = "rgba(220, 252, 231, 0.8)"
    elif status in {"missing_in_sales", "missing_in_stock"}:
        color = "rgba(254, 240, 138, 0.8)"
    else:
        color = "rgba(254, 202, 202, 0.8)"
    return [f"background-color: {color};"] * len(row)


def run_reconciliation_page():
    st.title("Reconciliation")
    st.caption(
        "Reconciliation compares dated outbound quantities from Stocks report against sales summary quantities for the same item and day."
    )

    stock_uploaded = st.file_uploader(
        "上传 Stock report Excel (.xlsx)",
        type=["xlsx"],
        key="reconciliation_stock_uploader",
        on_change=_handle_reconciliation_stock_uploader_change,
    )
    if stock_uploaded is not None and not st.session_state.get(
        "reconciliation_stock_file_bytes"
    ):
        _cache_reconciliation_stock_upload(stock_uploaded)

    sales_uploaded = st.file_uploader(
        "上传 Sales summary Excel (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=True,
        key="reconciliation_sales_uploader",
        on_change=_handle_reconciliation_sales_uploader_change,
    )
    if sales_uploaded and not st.session_state.get(
        "reconciliation_sales_files_payload"
    ):
        _cache_reconciliation_sales_upload(sales_uploaded)

    stock_file_bytes = st.session_state.get("reconciliation_stock_file_bytes")
    sales_payload = st.session_state.get("reconciliation_sales_files_payload", [])
    stock_name = st.session_state.get("reconciliation_stock_file_name", "")
    stock_browsed_at = st.session_state.get("reconciliation_stock_file_browsed_at", "")
    sales_browsed_at = st.session_state.get("reconciliation_sales_files_browsed_at", "")

    if stock_name:
        st.caption(f"Stock file: {stock_name}")
    if stock_browsed_at:
        st.caption(f"Stock last browsed: {stock_browsed_at}")
    if sales_browsed_at:
        st.caption(f"Sales last browsed: {sales_browsed_at}")

    if not stock_file_bytes or not sales_payload:
        st.info("请同时上传 Stock report 与 Sales summary 文件后继续。")
        return

    try:
        sales_df, sales_warns = load_sales_data_from_payload(sales_payload)
    except Exception as exc:
        st.error(f"读取销售工作簿失败：{exc}")
        return
    for warn in sales_warns:
        st.warning(warn)
    if sales_df.empty:
        st.warning("无法从上传的销售工作簿读取任何 Sales 数据。")
        return

    preferred_year = _resolve_reconciliation_year(sales_df)
    if preferred_year is not None:
        st.caption(
            f"Using sales year {preferred_year} for stock date headers without year."
        )

    try:
        stock_df, stock_timeline, stock_warns = (
            load_stock_reconciliation_data_from_bytes(
                stock_file_bytes,
                preferred_year=preferred_year,
            )
        )
    except Exception as exc:
        st.error(f"读取 Stock 工作簿失败：{exc}")
        return
    for warn in stock_warns:
        st.warning(warn)

    sales_timeline = build_sales_reconciliation_timeline(sales_df)
    result_df = build_reconciliation_result(stock_timeline, sales_timeline)

    filter_options = _build_reconciliation_filter_options(result_df)
    pre_date_filter_state = {
        "product_code": st.session_state.get("reconciliation_filter_product_code", []),
        "description": st.session_state.get("reconciliation_filter_description", []),
        "year": st.session_state.get("reconciliation_filter_year", []),
        "month": st.session_state.get("reconciliation_filter_month", []),
    }
    specific_date_bounds_filter_state = {
        "product_code": st.session_state.get("reconciliation_filter_product_code", []),
        "description": st.session_state.get("reconciliation_filter_description", []),
        "year": [],
        "month": [],
    }
    min_date, max_date = _get_reconciliation_date_bounds(
        result_df, pre_date_filter_state
    )
    specific_min_date, specific_max_date = _get_reconciliation_date_bounds(
        result_df, specific_date_bounds_filter_state
    )
    with st.sidebar:
        st.subheader("Reconciliation Filters")
        if st.button("Clear Date Filter", key="reconciliation_clear_date_filters_button"):
            for key in [
                "reconciliation_filter_year",
                "reconciliation_filter_month",
                "reconciliation_filter_date",
                "reconciliation_filter_date_widget",
                "reconciliation_filter_use_specific_date",
                "reconciliation_filter_specific_date",
                "reconciliation_filter_specific_date_widget",
            ]:
                if key.endswith(("year", "month")):
                    st.session_state[key] = []
                elif key == "reconciliation_filter_use_specific_date":
                    st.session_state[key] = False
                else:
                    st.session_state[key] = None

        _ensure_multiselect_key_state(
            "reconciliation_filter_product_code",
            filter_options["product_code"],
            [],
        )
        st.multiselect(
            "Product Code",
            filter_options["product_code"],
            key="reconciliation_filter_product_code",
            placeholder="选择产品编码",
        )

        _ensure_multiselect_key_state(
            "reconciliation_filter_description",
            filter_options["description"],
            [],
        )
        st.multiselect(
            "Description",
            filter_options["description"],
            key="reconciliation_filter_description",
            placeholder="选择产品描述",
        )

        _ensure_multiselect_key_state(
            "reconciliation_filter_year",
            filter_options["year"],
            _default_reconciliation_year_selection(filter_options["year"]),
        )
        st.multiselect(
            "Year",
            filter_options["year"],
            key="reconciliation_filter_year",
            placeholder="选择年份",
        )

        _ensure_multiselect_key_state(
            "reconciliation_filter_month",
            filter_options["month"],
            _default_reconciliation_month_selection(filter_options["month"]),
        )
        st.multiselect(
            "Month",
            filter_options["month"],
            key="reconciliation_filter_month",
            placeholder="选择月份",
        )
        current_date_value = _coerce_reconciliation_date_input_value(
            st.session_state.get("reconciliation_filter_date")
        )
        if current_date_value is not None:
            if isinstance(current_date_value, tuple):
                start_date, end_date = current_date_value
                if min_date is not None and start_date < min_date:
                    start_date = min_date
                if max_date is not None and end_date > max_date:
                    end_date = max_date
                current_date_value = (
                    (
                        start_date,
                        end_date,
                    )
                    if start_date <= end_date
                    else None
                )
            else:
                if min_date is not None and current_date_value < min_date:
                    current_date_value = None
                elif max_date is not None and current_date_value > max_date:
                    current_date_value = None
        st.session_state["reconciliation_filter_date_widget"] = current_date_value
        st.date_input(
            "Date Range",
            key="reconciliation_filter_date_widget",
            min_value=min_date,
            max_value=max_date,
            value=st.session_state.get("reconciliation_filter_date_widget"),
            help="Used only when Specific Date Override is empty.",
        )
        st.session_state["reconciliation_filter_date"] = st.session_state.get(
            "reconciliation_filter_date_widget"
        )

        use_specific_date = st.checkbox(
            "Use Specific Date",
            key="reconciliation_filter_use_specific_date",
            value=bool(st.session_state.get("reconciliation_filter_use_specific_date", False)),
            help="When checked, only Specific Date Override is used; when unchecked, Date Range is used.",
        )

        current_specific_date = _coerce_reconciliation_date_input_value(
            st.session_state.get("reconciliation_filter_specific_date")
        )
        if isinstance(current_specific_date, tuple):
            current_specific_date = (
                current_specific_date[0] if current_specific_date else None
            )
        if current_specific_date is not None:
            if specific_min_date is not None and current_specific_date < specific_min_date:
                current_specific_date = None
            elif specific_max_date is not None and current_specific_date > specific_max_date:
                current_specific_date = None
        st.session_state["reconciliation_filter_specific_date_widget"] = (
            current_specific_date
        )
        st.date_input(
            "Specific Date Override",
            key="reconciliation_filter_specific_date_widget",
            min_value=specific_min_date,
            max_value=specific_max_date,
            value=st.session_state.get("reconciliation_filter_specific_date_widget"),
            disabled=not use_specific_date,
            help="Used only when Use Specific Date is checked.",
        )
        st.session_state["reconciliation_filter_specific_date"] = st.session_state.get(
            "reconciliation_filter_specific_date_widget"
        )

        active_specific_date = (
            st.session_state.get("reconciliation_filter_specific_date")
            if use_specific_date
            else None
        )

    filtered_result_df = _apply_reconciliation_filters(
        result_df,
        {
            "product_code": st.session_state.get(
                "reconciliation_filter_product_code", []
            ),
            "description": st.session_state.get(
                "reconciliation_filter_description", []
            ),
            "year": st.session_state.get("reconciliation_filter_year", []),
            "month": st.session_state.get("reconciliation_filter_month", []),
            "date": st.session_state.get("reconciliation_filter_date"),
            "specific_date": active_specific_date,
            "use_specific_date": use_specific_date,
        },
    )

    metric_cols = st.columns(4)
    metric_cols[0].metric("Stock timeline rows", f"{len(stock_timeline):,}")
    metric_cols[1].metric("Sales timeline rows", f"{len(sales_timeline):,}")
    metric_cols[2].metric(
        "Mismatch rows",
        f"{int((filtered_result_df.get('is_match', pd.Series(dtype=bool)) == False).sum()):,}",
    )
    metric_cols[3].metric(
        "Matched rows",
        f"{int(filtered_result_df.get('is_match', pd.Series(dtype=bool)).sum()):,}",
    )

    st.markdown("### Mismatch rows")
    mismatch_df = filtered_result_df[
        filtered_result_df["mismatch_reason"] != "match"
    ].copy()
    display_df = filtered_result_df.copy()
    mismatch_display_df = _build_reconciliation_display_df(mismatch_df)
    all_display_df = _build_reconciliation_display_df(display_df)
    possible_match_display_df = _build_reconciliation_display_df(
        filtered_result_df[
            filtered_result_df["mismatch_reason"] == "possible_match"
        ].copy()
    )

    if mismatch_df.empty:
        st.success("所有已匹配项目的日期数量都一致。")
    else:
        st.dataframe(
            mismatch_display_df.drop(columns=["mismatch_reason"], errors="ignore"),
            width="stretch",
            hide_index=True,
        )

    if not possible_match_display_df.empty:
        with st.expander("Possible match rows", expanded=False):
            st.dataframe(
                possible_match_display_df.drop(
                    columns=["mismatch_reason"], errors="ignore"
                ),
                width="stretch",
                hide_index=True,
            )

    with st.expander("All reconciliation rows", expanded=False):
        if all_display_df.empty:
            st.info("当前没有可显示的对账记录。")
        else:
            st.dataframe(
                all_display_df.drop(columns=["mismatch_reason"], errors="ignore"),
                width="stretch",
                hide_index=True,
            )

    with st.expander("Parsed stock rows", expanded=False):
        if stock_df.empty:
            st.info("当前没有可显示的 Stock 明细。")
        else:
            stock_display = stock_df.copy()
            for col in ["expiry_date", "relabel_to_date"]:
                if col in stock_display.columns:
                    stock_display[col] = pd.to_datetime(
                        stock_display[col], errors="coerce"
                    ).dt.strftime("%d-%b-%Y")
                    stock_display[col] = stock_display[col].fillna("")
            st.dataframe(stock_display, width="stretch", hide_index=True)


def _infer_weight_kg_per_packet(value: Any) -> Optional[float]:
    if pd.isna(value):
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    match = re.search(
        r"(\d+(?:\.\d+)?)\s*(kg|kgs|g|gm|gram|grams|l|lt|ltr|liter|litre|ml)", text
    )
    if not match:
        return None
    number = float(match.group(1))
    unit = match.group(2)
    if unit in {"g", "gm", "gram", "grams", "ml"}:
        return number / 1000.0
    return number


def _normalize_sales_pack_group_key(value: Any) -> tuple:
    if pd.isna(value):
        return ("raw", "")
    text = str(value).strip()
    if not text:
        return ("raw", "")
    pack_count = _infer_carton_pack_size(text)
    weight_kg = _infer_weight_kg_per_packet(text)
    if (
        pack_count is not None
        and weight_kg is not None
        and pack_count > 0
        and weight_kg > 0
    ):
        return ("mass", int(round(float(pack_count))), round(float(weight_kg), 4))
    if pack_count is not None and pack_count > 0:
        return ("count", int(round(float(pack_count))))
    return ("raw", text.lower())


def _add_sales_price_metrics(df: pd.DataFrame) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    result = df.copy()
    if result.empty:
        result["Base Selling Price"] = pd.Series(dtype="float")
        result["Unit Price per kg"] = pd.Series(dtype="float")
        return result

    pack_size = result.get("Carton Packing", pd.Series(dtype="string")).map(
        _infer_carton_pack_size
    )
    weight_kg_per_pkt = result.get("Carton Packing", pd.Series(dtype="string")).map(
        _infer_weight_kg_per_packet
    )
    qty_ctn = pd.to_numeric(result.get("Qty in Ctns", 0.0), errors="coerce").fillna(0.0)
    qty_pcs = pd.to_numeric(result.get("Qty in Pcs", 0.0), errors="coerce").fillna(0.0)
    total_value = pd.to_numeric(result.get("Total Value", 0.0), errors="coerce").fillna(
        0.0
    )

    total_ctn_equivalent = qty_ctn + _safe_divide_series(qty_pcs, pack_size)
    total_packets = (
        qty_ctn * pd.to_numeric(pack_size, errors="coerce").fillna(0.0) + qty_pcs
    )
    total_kg = total_packets * pd.to_numeric(weight_kg_per_pkt, errors="coerce")

    result["Base Selling Price"] = _safe_divide_series(
        total_value, total_ctn_equivalent
    )
    result["Unit Price per kg"] = _safe_divide_series(total_value, total_kg)
    return result


def _append_sales_forecast_total_row(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df if df is not None else pd.DataFrame()

    result = df.copy()
    total_packets = float(
        pd.to_numeric(result.get("_forecast_packets"), errors="coerce")
        .fillna(0.0)
        .sum()
    )
    total_value = float(
        pd.to_numeric(result.get("_forecast_value"), errors="coerce").fillna(0.0).sum()
    )
    pack_labels = (
        result.get("箱规", pd.Series(dtype="string")).astype("string").fillna("")
    )
    normalized_pack_keys = {
        _normalize_sales_pack_group_key(value)
        for value in pack_labels.tolist()
        if str(value).strip()
    }
    total_qty_text = f"{_format_qty_display(total_packets)} pkts"
    if len(normalized_pack_keys) == 1:
        first_pack_label = next(
            (
                str(value).strip()
                for value in pack_labels.tolist()
                if str(value).strip()
            ),
            "",
        )
        pack_size = _infer_carton_pack_size(first_pack_label)
        if pack_size and pack_size > 0:
            total_qty_text = _format_total_qty_text(
                int(round(total_packets)), int(round(pack_size))
            )

    total_row: Dict[str, Any] = {col: "" for col in result.columns}
    total_row["客户"] = "Total"
    total_row["下月预测销量"] = total_qty_text
    total_row["下月预测销售额"] = _format_price_display(total_value)
    if "_forecast_packets" in result.columns:
        total_row["_forecast_packets"] = total_packets
    if "_forecast_value" in result.columns:
        total_row["_forecast_value"] = total_value
    return pd.concat([result, pd.DataFrame([total_row])], ignore_index=True)


def build_account_price_list(df: pd.DataFrame, account: str) -> pd.DataFrame:
    columns = [
        "Customer",
        "Product Code",
        "Product Description",
        "Pack Size",
        "Last Selling Date",
        "Selling Price by kg",
    ]
    if df is None or df.empty or not account:
        return pd.DataFrame(columns=columns)

    work = df.copy()
    for required_col in [
        "Account",
        "Date",
        "Customer",
        "Product Code",
        "Product Description",
        "Carton Packing",
        "Qty in Pcs",
        "Qty in Ctns",
        "Total Value",
    ]:
        if required_col not in work.columns:
            work[required_col] = pd.NA

    work["_account"] = work["Account"].astype("string").fillna("").str.strip()
    work = work[work["_account"] == str(account).strip()]
    if work.empty:
        return pd.DataFrame(columns=columns)

    work["_date"] = pd.to_datetime(work["Date"], errors="coerce")
    work = work[work["_date"].notna()]
    if work.empty:
        return pd.DataFrame(columns=columns)

    work["_customer"] = work["Customer"].astype("string").fillna("").str.strip()
    work["_product_code"] = work["Product Code"].astype("string").fillna("").str.strip()
    work["_product_desc"] = (
        work["Product Description"].astype("string").fillna("").str.strip()
    )
    work["_pack_size"] = work["Carton Packing"].astype("string").fillna("").str.strip()
    work["_sales_pack_group_key"] = work["Carton Packing"].map(
        _normalize_sales_pack_group_key
    )

    work["_carton_pack_size"] = work["Carton Packing"].map(_infer_carton_pack_size)
    work["_weight_kg_per_pkt"] = work["Carton Packing"].map(_infer_weight_kg_per_packet)
    work["_qty_pcs"] = pd.to_numeric(work["Qty in Pcs"], errors="coerce").fillna(0.0)
    work["_qty_ctn"] = pd.to_numeric(work["Qty in Ctns"], errors="coerce").fillna(0.0)
    work["_total_value"] = pd.to_numeric(work["Total Value"], errors="coerce").fillna(
        0.0
    )

    work["_total_packets"] = work["_qty_pcs"] + (
        work["_qty_ctn"]
        * pd.to_numeric(work["_carton_pack_size"], errors="coerce").fillna(0.0)
    )
    work["_total_kg"] = work["_total_packets"] * pd.to_numeric(
        work["_weight_kg_per_pkt"], errors="coerce"
    )

    group_cols = [
        "_customer",
        "_product_code",
        "_product_desc",
        "_sales_pack_group_key",
    ]
    latest_date = work.groupby(group_cols, dropna=False)["_date"].transform("max")
    latest = work[work["_date"] == latest_date].copy()
    if latest.empty:
        return pd.DataFrame(columns=columns)

    summary = (
        latest.groupby(group_cols, dropna=False, as_index=False)
        .agg(
            _pack_size=("_pack_size", "last"),
            _total_value=("_total_value", "sum"),
            _total_kg=("_total_kg", "sum"),
            _last_selling_date=("_date", "max"),
        )
        .rename(
            columns={
                "_customer": "Customer",
                "_product_code": "Product Code",
                "_product_desc": "Product Description",
                "_pack_size": "Pack Size",
            }
        )
    )
    summary["Selling Price by kg"] = _safe_divide_series(
        summary["_total_value"], summary["_total_kg"]
    )
    summary["Last Selling Date"] = (
        pd.to_datetime(summary["_last_selling_date"], errors="coerce")
        .dt.strftime("%Y-%m-%d")
        .fillna("")
    )
    summary = summary.drop(
        columns=["_total_value", "_total_kg", "_last_selling_date"], errors="ignore"
    )
    summary = summary[columns].sort_values(
        by=["Customer", "Product Code", "Product Description", "Pack Size"],
        kind="stable",
    )
    summary = summary.reset_index(drop=True)
    return summary


# ----------------------------- 主入口：使用 tabs 让两个界面独立运行 -----------------------------


def main():
    st.sidebar.title("导航")
    st.sidebar.write("选择页面运行 Sales、Stock 或 Reconciliation。")
    current_page = st.sidebar.radio(
        "Page", ["Sales", "Stock", "Reconciliation"], index=0
    )

    action_for_page = (
        _HOTKEY_ACTION_CLEAR_SALES
        if current_page == "Sales"
        else _HOTKEY_ACTION_CLEAR_STOCK
    )
    _inject_delete_hotkey_listener(action_for_page)

    hotkey_action = _consume_hotkey_action()
    if hotkey_action == _HOTKEY_ACTION_CLEAR_SALES and current_page == "Sales":
        _clear_sales_filters_state(st.session_state)
        st.rerun()
    if hotkey_action == _HOTKEY_ACTION_CLEAR_STOCK and current_page == "Stock":
        _clear_stock_filters_state(st.session_state)
        st.rerun()

    if current_page == "Sales":
        _restore_persisted_state("__sales_persist_state")
        run_sales_page()
        _snapshot_persisted_state("__sales_persist_state", _SALES_PERSIST_KEYS)
    elif current_page == "Stock":
        _restore_persisted_state("__stock_persist_state")
        run_stock_page()
        _snapshot_persisted_state("__stock_persist_state", _STOCK_PERSIST_KEYS)
    else:
        _restore_persisted_state("__reconciliation_persist_state")
        run_reconciliation_page()
        _snapshot_persisted_state(
            "__reconciliation_persist_state", _RECONCILIATION_PERSIST_KEYS
        )

    st.sidebar.markdown("---")
    st.sidebar.caption("Powered by Streamlit")


if __name__ == "__main__":
    main()
