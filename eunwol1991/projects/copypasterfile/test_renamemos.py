import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


MODULE_PATH = Path(__file__).with_name("renamemos.py")
MODULE_SPEC = importlib.util.spec_from_file_location("renamemos", MODULE_PATH)
assert MODULE_SPEC is not None
assert MODULE_SPEC.loader is not None
renamemos = importlib.util.module_from_spec(MODULE_SPEC)
sys.modules["renamemos"] = renamemos
MODULE_SPEC.loader.exec_module(renamemos)


def _make_workbook(path: Path) -> None:
    workbook = Workbook()
    do_sheet = workbook.active
    assert do_sheet is not None
    do_sheet.title = "DO"
    invoice = workbook.create_sheet("Invoice")

    for sheet in (do_sheet, invoice):
        sheet["A5"] = "abcd item"
        sheet["B5"] = "OLD-CODE"
        sheet["C5"] = "Old description"

    do_sheet["G5"] = "old pack"
    do_sheet["H5"] = "old mos pack"
    do_sheet["I5"] = 9
    do_sheet["K5"] = "OLD-UOM"

    invoice["F5"] = "old invoice pack"
    invoice["G5"] = 8
    invoice["H5"] = "OLD-INV-UOM"
    invoice["I5"] = "=1*G5"

    workbook.save(path)
    workbook.close()


class RenameMosReplacementSelectionTest(unittest.TestCase):
    def test_disabled_replacement_fields_keep_existing_cells(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "ABC xx26 - DO & INV.xlsx"
            _make_workbook(workbook_path)

            list(
                renamemos.update_excel_files(
                    directory=temp_dir,
                    debug=False,
                    filename_substring="xx26",
                    find_keyword="abcd",
                    process_do=True,
                    process_invoice=True,
                    product_code="NEW-CODE",
                    product_description="New description",
                    pack_size="new pack",
                    qty=3.0,
                    uom="NEW-UOM",
                    price_formula_template="=3.8*G{row}",
                    replacement_enabled={
                        "product_code": True,
                        "product_description": False,
                        "pack_size": False,
                        "qty": True,
                        "uom": False,
                        "price_formula": False,
                    },
                    log_func=lambda _msg: None,
                )
            )

            workbook = load_workbook(workbook_path, data_only=False)
            try:
                do_sheet = workbook["DO"]
                invoice = workbook["Invoice"]

                self.assertEqual(do_sheet["B5"].value, "NEW-CODE")
                self.assertEqual(do_sheet["C5"].value, "Old description")
                self.assertEqual(do_sheet["G5"].value, "old pack")
                self.assertEqual(do_sheet["H5"].value, "old mos pack")
                self.assertEqual(do_sheet["I5"].value, 3.0)
                self.assertEqual(do_sheet["K5"].value, "OLD-UOM")

                self.assertEqual(invoice["B5"].value, "NEW-CODE")
                self.assertEqual(invoice["C5"].value, "Old description")
                self.assertEqual(invoice["F5"].value, "old invoice pack")
                self.assertEqual(invoice["G5"].value, 3.0)
                self.assertEqual(invoice["H5"].value, "OLD-INV-UOM")
                self.assertEqual(invoice["I5"].value, "=1*G5")
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
