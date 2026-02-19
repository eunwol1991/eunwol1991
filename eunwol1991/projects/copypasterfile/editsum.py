import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl


def update_invoice_formulas_in_file(file_path, filename_substring="xx25", debug=True):
    """
    只处理档名包含 filename_substring 的文件，
    并更新其中 Invoice 表的 Subtotal / GST 9% / Total 公式。
    """
    low_name = os.path.basename(file_path).lower()
    if filename_substring.lower() not in low_name:
        if debug:
            print(f"跳过（不含关键字 '{filename_substring}'）：{file_path}")
        return False

    if not low_name.endswith((".xlsx", ".xlsm")):
        if debug:
            print(f"跳过（非 xlsx/xlsm）：{file_path}")
        return False

    print(f"Processing: {file_path}")

    try:
        # 读档（保留公式）
        workbook = openpyxl.load_workbook(file_path, data_only=False)

        # 确认有 Invoice 工作表（忽略大小写）
        sheetnames_lower = [s.lower() for s in workbook.sheetnames]
        if "invoice" not in sheetnames_lower:
            print(f"Sheet 'Invoice' not found in {file_path}")
            return False

        # 找出真正的表名（防大小写、空格）
        invoice_sheet_name = [
            s for s in workbook.sheetnames if s.strip().lower() == "invoice"
        ][0]
        sheet = workbook[invoice_sheet_name]

        amount_column = 9  # I 栏

        subtotal_row = None
        gst_row = None
        total_row = None

        # 找包含 "Subtotal"、"GST 9%"、"Total" 的行
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_value = cell.value.strip().lower()
                    if "subtotal" in cell_value:
                        subtotal_row = cell.row
                    elif "gst 9%" in cell_value:
                        gst_row = cell.row
                    elif "total" in cell_value:
                        total_row = cell.row

        print(
            f"Subtotal row: {subtotal_row}, GST row: {gst_row}, "
            f"Total row: {total_row}, Amount column: {amount_column}"
        )

        if not subtotal_row or not gst_row or not total_row:
            print(f"Required rows not found in 'Invoice' sheet of {file_path}")
            return False

        # 找金额起始行（Subtotal 上面第一格非空的 I 栏）
        amount_start_row = None
        for r in range(1, subtotal_row):
            if sheet.cell(row=r, column=amount_column).value is not None:
                amount_start_row = r
                break

        # 更新 Subtotal 公式（依你原本逻辑固定从 row 24 开始）
        if amount_start_row:
            subtotal_cell = sheet.cell(row=subtotal_row, column=amount_column)
            subtotal_cell.value = (
                f"=SUM("
                f"{sheet.cell(row=24, column=amount_column).coordinate}:"
                f"{sheet.cell(row=subtotal_row - 1, column=amount_column).coordinate}"
                f")"
            )
            print(f"Updated Subtotal formula to: {subtotal_cell.value}")

        # 更新 GST 公式
        gst_cell = sheet.cell(row=gst_row, column=amount_column)
        gst_cell.value = (
            f"={sheet.cell(row=subtotal_row, column=amount_column).coordinate}*0.09"
        )
        print(f"Updated GST formula to: {gst_cell.value}")

        # 更新 Total 公式
        total_cell = sheet.cell(row=total_row, column=amount_column)
        total_cell.value = (
            f"=SUM("
            f"{sheet.cell(row=subtotal_row, column=amount_column).coordinate}:"
            f"{sheet.cell(row=gst_row, column=amount_column).coordinate}"
            f")"
        )
        print(f"Updated Total formula to: {total_cell.value}")

        # 存档
        workbook.save(file_path)
        print(f"Updated and saved: {file_path}")
        return True

    except Exception as e:
        print(f"Failed to process {file_path}: {e}")
        return False


def batch_update_invoice_formulas(directory, filename_substring="xx25"):
    """
    扫描整个目录，批量调用 update_invoice_formulas_in_file。
    只处理档名包含 filename_substring 的 .xlsx/.xlsm。
    """
    processed = 0
    changed = 0
    failed = 0

    for root, _, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            low = file.lower()
            # 只看 xlsx/xlsm，关键字过滤放在函数里再判断一次
            if low.endswith((".xlsx", ".xlsm")):
                processed += 1
                ok = update_invoice_formulas_in_file(
                    file_path,
                    filename_substring=filename_substring,
                    debug=True
                )
                if ok:
                    changed += 1
                else:
                    # 可能是「没找到 Invoice」/「没关键字」/「错误」
                    # 这里简单算在 failed 里
                    # 如果你要更细拆，可以在函数里返回状态码
                    failed += 0  # 这里先不加，避免把「跳过」也当成失败
    return processed, changed, failed


# -------------------- GUI --------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Invoice Formula Updater")
        self.geometry("650x200")

        self.dir_var = tk.StringVar()
        self.sub_var = tk.StringVar(value="xx25")

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)

        # 目录选择
        ttk.Label(frm, text="目标文件夹：").grid(row=0, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.dir_var, width=60).grid(
            row=0, column=1, padx=6, sticky="w"
        )
        ttk.Button(frm, text="选择…", command=self.choose_dir).grid(
            row=0, column=2, padx=2
        )

        # 文件名过滤子串
        ttk.Label(frm, text="文件名需包含：").grid(
            row=1, column=0, sticky="w", pady=(10, 0)
        )
        ttk.Entry(frm, textvariable=self.sub_var, width=20).grid(
            row=1, column=1, sticky="w", pady=(10, 0)
        )
        ttk.Label(frm, text="（例如：xx25，忽略大小写）").grid(
            row=1, column=2, sticky="w", pady=(10, 0)
        )

        # 开始按钮
        ttk.Button(frm, text="开始更新公式", command=self.run).grid(
            row=2, column=0, columnspan=3, pady=18, sticky="we"
        )

        for i in range(3):
            frm.grid_columnconfigure(i, weight=1)

    def choose_dir(self):
        path = filedialog.askdirectory(title="选择目标文件夹")
        if path:
            self.dir_var.set(path)

    def run(self):
        directory = self.dir_var.get().strip()
        if not directory or not os.path.isdir(directory):
            messagebox.showerror("错误", "请选择有效的文件夹路径")
            return

        sub = self.sub_var.get().strip() or "xx25"

        # 简单确认一下
        if not messagebox.askyesno(
            "确认",
            f"将对\n\n{directory}\n\n下所有档名包含「{sub}」的 xlsx/xlsm 文件更新 Invoice 公式，确定继续？",
        ):
            return

        processed, changed, failed = batch_update_invoice_formulas(
            directory, filename_substring=sub
        )

        messagebox.showinfo(
            "完成",
            f"扫描文件总数（xlsx/xlsm）：{processed}\n"
            f"成功更新公式的文件：{changed}\n"
            f"详细过程请看控制台输出。",
        )


if __name__ == "__main__":
    App().mainloop()
