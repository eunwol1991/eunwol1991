from pathlib import Path
import importlib.util
import sys
import tempfile
import unittest

from openpyxl import Workbook, load_workbook


MODULE_PATH = Path(__file__).with_name("check_do_merges.py")
MODULE_SPEC = importlib.util.spec_from_file_location("check_do_merges", MODULE_PATH)
assert MODULE_SPEC is not None
assert MODULE_SPEC.loader is not None
check_do_merges = importlib.util.module_from_spec(MODULE_SPEC)
sys.modules["check_do_merges"] = check_do_merges
MODULE_SPEC.loader.exec_module(check_do_merges)


def _save_workbook(path: Path, do_merges: list[str], invoice_merges: list[str]) -> None:
    wb = Workbook()
    do_sheet = wb.active
    assert do_sheet is not None
    do_sheet.title = "DO"
    invoice = wb.create_sheet("Invoice")

    do_sheet["A1"] = "DO marker"
    invoice["A1"] = "Invoice marker"

    for merge_range in do_merges:
        do_sheet.merge_cells(merge_range)
    for merge_range in invoice_merges:
        invoice.merge_cells(merge_range)

    wb.save(path)
    wb.close()


def _merged_ranges(path: Path, sheet_name: str) -> set[str]:
    wb = load_workbook(path)
    try:
        return {str(rng) for rng in wb[sheet_name].merged_cells.ranges}
    finally:
        wb.close()


class CheckDoMergesTest(unittest.TestCase):
    def test_extract_outlet_name_uses_last_parenthesized_part(self) -> None:
        outlet = check_do_merges.extract_outlet_name(
            "MOS 0726 - 022 - DO & INV (Bedok Mall).xlsx"
        )

        self.assertEqual(outlet, "Bedok Mall")

    def test_check_workbook_reports_missing_do_ranges_without_saving(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            folder = Path(temp_dir)
            reference = folder / "MOS xx26 - 00x - DO & INV (Bedok Mall).xlsx"
            target = folder / "MOS 0726 - 022 - DO & INV (Bedok Mall).xlsx"
            _save_workbook(reference, ["A7:K7", "A22:G23", "H62:K62"], ["A7:I7"])
            _save_workbook(target, ["A7:K7"], [])
            before_mtime = target.stat().st_mtime_ns

            result = check_do_merges.check_workbook(target, reference)

            self.assertEqual(result.missing_ranges, ["A22:G23", "H62:K62"])
            self.assertEqual(result.skipped_conflicts, [])
            self.assertTrue(result.needs_check)
            self.assertEqual(target.stat().st_mtime_ns, before_mtime)
            self.assertEqual(_merged_ranges(target, "DO"), {"A7:K7"})
            self.assertEqual(_merged_ranges(target, "Invoice"), set())

    def test_check_workbook_reports_conflict_but_does_not_merge(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            folder = Path(temp_dir)
            reference = folder / "MOS xx26 - 00x - DO & INV (Bedok Mall).xlsx"
            target = folder / "MOS 0726 - 022 - DO & INV (Bedok Mall).xlsx"
            _save_workbook(reference, ["H62:K62"], [])
            _save_workbook(target, [], [])

            wb = load_workbook(target)
            wb["DO"]["I62"] = "AI keyed value"
            wb.save(target)
            wb.close()
            before_mtime = target.stat().st_mtime_ns

            result = check_do_merges.check_workbook(target, reference)

            self.assertEqual(result.missing_ranges, [])
            self.assertEqual(result.skipped_conflicts, ["H62:K62"])
            self.assertTrue(result.needs_check)
            self.assertEqual(target.stat().st_mtime_ns, before_mtime)
            self.assertEqual(_merged_ranges(target, "DO"), set())

    def test_check_workbook_reports_signature_range_to_print_area_edge(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            folder = Path(temp_dir)
            reference = folder / "ABC xx26 - 00x - DO & INV (Outlet).xlsx"
            target = folder / "ABC 0726 - 001 - DO & INV (Outlet).xlsx"
            _save_workbook(reference, [], [])
            _save_workbook(target, [], [])

            wb = load_workbook(target)
            wb["DO"].print_area = "A1:J54"
            wb["DO"]["I50"] = "Received In Good Order"
            wb["DO"]["H52"] = "Authorised Signature & Stamp"
            wb.save(target)
            wb.close()

            result = check_do_merges.check_workbook(target, reference)

            self.assertEqual(result.missing_ranges, ["I50:J50", "H52:J52"])
            self.assertTrue(result.needs_check)
            self.assertNotIn("I50:K50", result.missing_ranges)

    def test_check_folder_reports_unmatched_workbooks(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            target_dir = root / "Doc to print - JJ"
            reference_dir = root / "format"
            target_dir.mkdir()
            reference_dir.mkdir()
            unmatched = target_dir / "ABC 0726 - 001 - DO & INV (Unknown Outlet).xlsx"
            _save_workbook(unmatched, [], [])

            results = check_do_merges.check_folder(target_dir, reference_dir)

            self.assertEqual(len(results), 1)
            self.assertEqual(results[0].target_path, unmatched)
            self.assertIsNone(results[0].reference_path)
            self.assertEqual(results[0].skipped_reason, "no matching reference")
            self.assertTrue(results[0].needs_check)

    def test_cli_has_no_apply_option(self) -> None:
        args = check_do_merges.parse_args([])

        self.assertFalse(hasattr(args, "apply"))


if __name__ == "__main__":
    unittest.main()
