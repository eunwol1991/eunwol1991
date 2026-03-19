import os
import re
import threading
import tkinter as tk
import tkinter.font as tkfont
from dataclasses import dataclass
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText

import openpyxl


DEFAULT_OLD_TEXT = "Smoked Chicken Breakfast Rashers 2MM"
DEFAULT_DIRECTORY = "/mnt/c/Users/jhunj/Dropbox/DO & INV/DO & INV 2026/Melvin - MOS Burger/For Customer/3. Mar"
DEFAULT_INVOICES = """047
048
049
050
051
052
053
054
055
056
057
058
059
060
061
063
064
065
066
067
068
069
070
071
072
073
074
075
077
078
081
082
085
086
087
088"""


def _is_wsl() -> bool:
    if os.name == "nt":
        return False
    if os.environ.get("WSL_DISTRO_NAME"):
        return True
    try:
        with open("/proc/version", "r", encoding="utf-8") as f:
            return "microsoft" in f.read().lower()
    except OSError:
        return False


def _platform_drive_root() -> str:
    if os.name == "nt":
        return "c:/"
    if _is_wsl():
        return "/mnt/c"
    return "/"


def _from_c(path_tail: str) -> str:
    tail = (path_tail or "").lstrip("/")
    root = _platform_drive_root()
    if root.endswith("/"):
        return f"{root}{tail}"
    return f"{root}/{tail}"


def configure_cjk_font(root: tk.Tk) -> None:
    try:
        families = {name.lower(): name for name in tkfont.families(root)}
    except Exception:
        families = {}

    candidates = [
        "Microsoft YaHei UI",
        "Microsoft YaHei",
        "PingFang SC",
        "Noto Sans CJK SC",
        "WenQuanYi Micro Hei",
        "SimHei",
        "Arial Unicode MS",
    ]
    chosen = None
    for name in candidates:
        key = name.lower()
        if key in families:
            chosen = families[key]
            break

    if not chosen:
        return

    for font_name in ("TkDefaultFont", "TkTextFont", "TkMenuFont", "TkHeadingFont"):
        try:
            tkfont.nametofont(font_name).configure(family=chosen, size=10)
        except Exception:
            pass


def suggest_default_directory() -> str:
    candidates = [
        DEFAULT_DIRECTORY,
        _from_c("Users/jhunj/Dropbox/DO & INV/DO & INV 2026"),
        _from_c("Work/Savori-WorkSpace"),
        _from_c("Users"),
        "/mnt/c",
        os.path.expanduser("~"),
    ]
    for path in candidates:
        if path and os.path.isdir(path):
            return path
    return ""


def parse_invoice_tokens(text: str) -> list[str]:
    tokens = []
    for raw in re.split(r"[\s,;]+", text or ""):
        item = raw.strip()
        if not item:
            continue
        match = re.search(r"(\d{3})$", item)
        if match:
            item = match.group(1)
        if item not in tokens:
            tokens.append(item)
    return tokens


def matches_invoice_token(path: Path, invoice_tokens: list[str]) -> bool:
    if not invoice_tokens:
        return True
    name = path.name.lower()
    for token in invoice_tokens:
        if f"mos 0326 - {token.lower()}" in name:
            return True
    return False


def iter_target_workbooks(directory: str, invoice_tokens: list[str]):
    for root, _, files in os.walk(directory):
        for file_name in files:
            lower = file_name.lower()
            if not lower.endswith((".xlsx", ".xlsm")):
                continue
            if "mos 0326 - " not in lower or "do & inv" not in lower:
                continue
            path = Path(root) / file_name
            if matches_invoice_token(path, invoice_tokens):
                yield path


@dataclass
class ReplaceResult:
    file_path: str
    changed_cells: list[str]
    saved: bool


def replace_description_in_workbook(
    workbook_path: str,
    old_text: str,
    new_text: str,
    *,
    dry_run: bool,
) -> ReplaceResult:
    workbook = openpyxl.load_workbook(workbook_path)
    changed_cells: list[str] = []
    old_text_norm = old_text.strip()
    target_columns = {"DO": "C", "Invoice": "C"}
    try:
        for sheet_name in ("DO", "Invoice"):
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            desc_col = target_columns[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str):
                        continue
                    if old_text_norm not in value:
                        continue
                    target_cell = sheet[f"{desc_col}{cell.row}"]
                    target_value = target_cell.value
                    if (
                        not isinstance(target_value, str)
                        or old_text_norm not in target_value
                    ):
                        continue
                    target_cell.value = target_value.replace(old_text_norm, new_text)
                    coord = f"{sheet_name}!{target_cell.coordinate}"
                    if coord not in changed_cells:
                        changed_cells.append(coord)
        if changed_cells and not dry_run:
            workbook.save(workbook_path)
        return ReplaceResult(
            file_path=workbook_path,
            changed_cells=changed_cells,
            saved=bool(changed_cells) and not dry_run,
        )
    finally:
        workbook.close()


def batch_replace_descriptions(
    directory: str,
    invoice_tokens: list[str],
    old_text: str,
    new_text: str,
    *,
    dry_run: bool,
    log_func=print,
):
    candidates = sorted(
        iter_target_workbooks(directory, invoice_tokens), key=lambda p: p.name
    )
    total = len(candidates)
    log_func(f"Found {total} target workbook(s).")
    for index, path in enumerate(candidates, start=1):
        result = replace_description_in_workbook(
            str(path),
            old_text,
            new_text,
            dry_run=dry_run,
        )
        if result.changed_cells:
            action = "DRY-RUN" if dry_run else "UPDATED"
            log_func(f"[{action}] {path.name} -> {', '.join(result.changed_cells)}")
        else:
            log_func(f"[SKIP] {path.name} -> no match")
        yield index, total


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        configure_cjk_font(self)
        self.title("MOS Description Batch Replace")
        self.geometry("900x720")

        top = ttk.Frame(self, padding=10)
        top.pack(fill=tk.X)
        ttk.Label(top, text="Root directory:").pack(side=tk.LEFT)
        self.dir_var = tk.StringVar(value=suggest_default_directory())
        ttk.Entry(top, textvariable=self.dir_var, width=88).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Browse...", command=self.choose_dir).pack(side=tk.LEFT)

        config = ttk.LabelFrame(self, text="Settings", padding=10)
        config.pack(fill=tk.X, padx=10, pady=6)

        ttk.Label(config, text="Find text:").grid(row=0, column=0, sticky="e")
        self.old_text_var = tk.StringVar(value=DEFAULT_OLD_TEXT)
        ttk.Entry(config, textvariable=self.old_text_var, width=60).grid(
            row=0, column=1, sticky="we", padx=6, pady=2
        )

        ttk.Label(config, text="Replace with:").grid(row=1, column=0, sticky="e")
        self.new_text_var = tk.StringVar()
        ttk.Entry(config, textvariable=self.new_text_var, width=60).grid(
            row=1, column=1, sticky="we", padx=6, pady=2
        )

        self.dry_run_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(config, text="Dry run only", variable=self.dry_run_var).grid(
            row=2, column=1, sticky="w", padx=6, pady=2
        )
        config.columnconfigure(1, weight=1)

        invoice_frame = ttk.LabelFrame(
            self, text="MOS numbers (one per line or spaced)", padding=10
        )
        invoice_frame.pack(fill=tk.BOTH, expand=False, padx=10, pady=6)
        self.invoice_text = ScrolledText(invoice_frame, height=10)
        self.invoice_text.pack(fill=tk.BOTH, expand=True)
        self.invoice_text.insert("1.0", DEFAULT_INVOICES)

        ops = ttk.Frame(self, padding=10)
        ops.pack(fill=tk.X)
        self.run_button = ttk.Button(ops, text="Start", command=self.on_run)
        self.run_button.pack(side=tk.LEFT)
        self.progress = ttk.Progressbar(ops, mode="determinate", length=420)
        self.progress.pack(side=tk.LEFT, padx=10)
        self.progress_label = ttk.Label(ops, text="0/0")
        self.progress_label.pack(side=tk.LEFT)

        log_frame = ttk.LabelFrame(self, text="Logs", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        self.log = ScrolledText(log_frame, height=18)
        self.log.pack(fill=tk.BOTH, expand=True)

        self.worker = None

    def choose_dir(self):
        start_dir = self.dir_var.get().strip() or suggest_default_directory()
        path = filedialog.askdirectory(
            title="Select MOS folder",
            mustexist=True,
            initialdir=start_dir if start_dir else "/mnt/c",
        )
        if path:
            self.dir_var.set(path)

    def append_log(self, message: str):
        if threading.current_thread() is not threading.main_thread():
            self.after(0, self.append_log, message)
            return
        self.log.insert(tk.END, message + "\n")
        self.log.see(tk.END)
        self.update_idletasks()

    def set_progress(self, done: int, total: int):
        if threading.current_thread() is not threading.main_thread():
            self.after(0, self.set_progress, done, total)
            return
        self.progress["maximum"] = max(total, 1)
        self.progress["value"] = done
        self.progress_label.config(text=f"{done}/{total}")

    def set_run_enabled(self, enabled: bool):
        if threading.current_thread() is not threading.main_thread():
            self.after(0, self.set_run_enabled, enabled)
            return
        self.run_button.config(state=tk.NORMAL if enabled else tk.DISABLED)

    def on_run(self):
        if self.worker and self.worker.is_alive():
            messagebox.showinfo("Notice", "A task is already running.")
            return

        directory = self.dir_var.get().strip()
        if not directory or not os.path.isdir(directory):
            messagebox.showerror("Error", "Please choose a valid directory.")
            return

        old_text = self.old_text_var.get().strip()
        new_text = self.new_text_var.get().strip()
        if not old_text:
            messagebox.showerror("Error", "Find text cannot be empty.")
            return
        if not new_text:
            messagebox.showerror("Error", "Replace-with text cannot be empty.")
            return

        invoice_tokens = parse_invoice_tokens(self.invoice_text.get("1.0", tk.END))
        if not invoice_tokens:
            messagebox.showerror("Error", "Please provide at least one MOS number.")
            return

        dry_run = self.dry_run_var.get()
        self.log.delete("1.0", tk.END)
        self.set_progress(0, 0)

        def task():
            try:
                gen = batch_replace_descriptions(
                    directory,
                    invoice_tokens,
                    old_text,
                    new_text,
                    dry_run=dry_run,
                    log_func=self.append_log,
                )
                for done, total in gen:
                    self.set_progress(done, total)
            finally:
                self.set_run_enabled(True)

        self.set_run_enabled(False)
        self.worker = threading.Thread(target=task, daemon=True)
        self.worker.start()


if __name__ == "__main__":
    app = App()
    app.mainloop()
