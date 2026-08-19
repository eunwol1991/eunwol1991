import os
import re
import threading
import traceback
import tkinter as tk
import tkinter.font as tkfont
from collections.abc import Mapping
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText

import openpyxl




def _is_wsl() -> bool:
    if os.name == "nt":
        return False
    if os.environ.get("WSL_DISTRO_NAME"):
        return True
    try:
        with open("/proc/version", "r", encoding="utf-8") as f:
            return "microsoft" in f.read().lower()
    except OSError:
        return False


def _platform_drive_root() -> str:
    if os.name == "nt":
        return "c:/"
    if _is_wsl():
        return "/mnt/c"
    return "/"


def _from_c(path_tail: str) -> str:
    tail = (path_tail or "").lstrip("/")
    root = _platform_drive_root()
    if root.endswith("/"):
        return f"{root}{tail}"
    return f"{root}/{tail}"
def configure_cjk_font(root: tk.Tk):
    try:
        families = {name.lower(): name for name in tkfont.families(root)}
    except Exception:
        families = {}

    candidates = [
        "Microsoft YaHei UI",
        "Microsoft YaHei",
        "PingFang SC",
        "Noto Sans CJK SC",
        "WenQuanYi Micro Hei",
        "SimHei",
        "Arial Unicode MS",
    ]
    chosen = None
    for name in candidates:
        key = name.lower()
        if key in families:
            chosen = families[key]
            break

    if not chosen:
        return

    for font_name in ("TkDefaultFont", "TkTextFont", "TkMenuFont", "TkHeadingFont"):
        try:
            tkfont.nametofont(font_name).configure(family=chosen, size=10)
        except Exception:
            pass


def _windows_path_to_wsl(path: str) -> str:
    text = (path or "").strip()
    match = re.match(r"^([A-Za-z]):\\(.*)$", text)
    if not match:
        return text
    drive = match.group(1).lower()
    rest = match.group(2).replace("\\", "/")
    return f"/mnt/{drive}/{rest}"


def suggest_default_directory() -> str:
    preferred_dropbox_dir = _from_c("Users/jhunj/Dropbox/DO & INV/DO & INV 2026")
    candidates = []
    candidates.append(preferred_dropbox_dir)

    userprofile = os.environ.get("USERPROFILE", "").strip()
    if userprofile:
        wsl_userprofile = _windows_path_to_wsl(userprofile)
        candidates.extend(
            [
                f"{wsl_userprofile}/Downloads",
                f"{wsl_userprofile}/Desktop",
                wsl_userprofile,
            ]
        )

    candidates.extend(
        [
            _from_c("Work/Savori-WorkSpace"),
            _from_c("Users"),
            "/mnt/c",
            os.path.expanduser("~"),
        ]
    )

    for path in candidates:
        if path and os.path.isdir(path):
            return path
    return ""


def update_excel_files(
    directory: str,
    debug: bool,
    filename_substring: str,
    find_keyword: str,
    process_do: bool,
    process_invoice: bool,
    product_code: str,
    product_description: str,
    pack_size: str,
    qty: float,
    uom: str,
    price_formula_template: str,
    replacement_enabled: Mapping[str, bool] | None = None,
    log_func=print,
):
    """
    批量更新 Excel 文件（.xlsx/.xlsm），在 DO / Invoice 两张表中查找包含关键字的单元格并修改对应列。
    - filename_substring: 仅处理文件名中包含该子串（忽略大小写）的文件，如 'xx25'
    - find_keyword: 查找关键字（忽略大小写、去两端空格）
    - price_formula_template: 价格单元格写入的公式模板，支持 {row} 占位符，例如 '=3.8*G{row}'
    """
    processed_count = 0
    error_files = []
    matched_files = []

    # 先统计需要处理的文件数量，便于进度条
    for root, _, files in os.walk(directory):
        for file in files:
            low = file.lower()
            if low.endswith((".xlsx", ".xlsm")) and filename_substring.lower() in low:
                matched_files.append(os.path.join(root, file))

    total = len(matched_files)
    log_func(f"Found {total} candidate files.")

    find_kw = (find_keyword or "").strip().lower()

    def should_replace(field_name: str) -> bool:
        if replacement_enabled is None:
            return True
        return replacement_enabled.get(field_name, True)

    for idx, file_path in enumerate(matched_files, start=1):
        if debug:
            log_func(f"[Processing] [{idx}/{total}] {file_path}")

        is_mos_file = "mos" in os.path.basename(file_path).lower()

        workbook = None
        modified = False
        try:
            workbook = openpyxl.load_workbook(file_path)
            if debug:
                log_func(f"[Loaded] {os.path.basename(file_path)}")
                log_func(f"[Sheets] {workbook.sheetnames}")

            # 根据勾选处理的表，生成顺序列表
            sheets_to_process = []
            if process_do and "DO" in workbook.sheetnames:
                sheets_to_process.append("DO")
            if process_invoice and "Invoice" in workbook.sheetnames:
                sheets_to_process.append("Invoice")

            for sheet_name in sheets_to_process:
                sheet = workbook[sheet_name]
                if debug:
                    log_func(f"[Processing sheet] {sheet_name}")

                # 逐格扫描
                for row in sheet.iter_rows():
                    for cell in row:
                        cell_value = (
                            str(cell.value).strip().lower()
                            if cell.value is not None
                            else ""
                        )
                        if find_kw and find_kw in cell_value:
                            row_number = cell.row
                            updated_cells = []

                            # 价格公式：将 {row} 替换为行号
                            price_formula = (price_formula_template or "").replace(
                                "{row}", str(row_number)
                            )

                            if debug:
                                log_func(
                                    f"[Keyword match] {find_kw} -> Sheet: {sheet_name}, Row: {row_number}"
                                )

                            if sheet_name == "DO":
                                if should_replace("product_code"):
                                    sheet[f"B{row_number}"].value = product_code
                                    updated_cells.append(f"B{row_number}")
                                if should_replace("product_description"):
                                    sheet[f"C{row_number}"].value = product_description
                                    updated_cells.append(f"C{row_number}")
                                if should_replace("pack_size"):
                                    if is_mos_file:
                                        sheet[f"H{row_number}"].value = pack_size
                                        sheet[f"G{row_number}"].value = None
                                        pack_col = "H"
                                    else:
                                        sheet[f"G{row_number}"].value = pack_size
                                        sheet[f"H{row_number}"].value = None
                                        pack_col = "G"
                                    updated_cells.append(f"{pack_col}{row_number}")
                                if should_replace("qty"):
                                    sheet[f"I{row_number}"].value = qty
                                    updated_cells.append(f"I{row_number}")
                                if should_replace("uom"):
                                    sheet[f"K{row_number}"].value = uom
                                    updated_cells.append(f"K{row_number}")
                                if debug:
                                    if updated_cells:
                                        log_func(f"[DO updated] {', '.join(updated_cells)}")
                                    else:
                                        log_func(f"[DO skipped] Row {row_number}: no replacement fields selected")

                            elif sheet_name == "Invoice":
                                if should_replace("product_code"):
                                    sheet[f"B{row_number}"].value = product_code
                                    updated_cells.append(f"B{row_number}")
                                if should_replace("product_description"):
                                    sheet[f"C{row_number}"].value = product_description
                                    updated_cells.append(f"C{row_number}")
                                if should_replace("pack_size"):
                                    sheet[f"F{row_number}"].value = pack_size
                                    updated_cells.append(f"F{row_number}")
                                if should_replace("qty"):
                                    sheet[f"G{row_number}"].value = qty
                                    updated_cells.append(f"G{row_number}")
                                if should_replace("uom"):
                                    sheet[f"H{row_number}"].value = uom
                                    updated_cells.append(f"H{row_number}")
                                if should_replace("price_formula"):
                                    sheet[f"I{row_number}"].value = (
                                        price_formula if price_formula else None
                                    )
                                    updated_cells.append(f"I{row_number}")
                                if debug:
                                    if updated_cells:
                                        log_func(f"[Invoice updated] {', '.join(updated_cells)}")
                                    else:
                                        log_func(f"[Invoice skipped] Row {row_number}: no replacement fields selected")

                            if updated_cells:
                                modified = True
                            # 不跳出，继续处理所有匹配

            if modified:
                workbook.save(file_path)
                processed_count += 1
                if debug:
                    log_func(f"[Saved] {file_path}")

        except Exception as e:
            traceback.print_exc()
            error_files.append(file_path)
            log_func(f"[Failed] {file_path}, error: {e}")
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass

        yield idx, total  # 进度更新

    log_func(f"[Done] Updated {processed_count} file(s).")
    if error_files:
        log_func("[Warning] The following files failed:")
        for ef in error_files:
            log_func(f"  [Failed] {ef}")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        configure_cjk_font(self)
        self.title("Excel Batch Update Tool")
        self.geometry("860x620")

        # 路径选择
        frm_top = ttk.Frame(self, padding=10)
        frm_top.pack(fill=tk.X)
        ttk.Label(frm_top, text="Root directory:").pack(side=tk.LEFT)
        self.dir_var = tk.StringVar(value=suggest_default_directory())
        ttk.Entry(frm_top, textvariable=self.dir_var, width=80).pack(
            side=tk.LEFT, padx=6
        )
        ttk.Button(frm_top, text="Browse...", command=self.choose_dir).pack(
            side=tk.LEFT
        )
        ttk.Button(frm_top, text="Use default path", command=self.use_default_dir).pack(
            side=tk.LEFT, padx=(6, 0)
        )

        # 参数区
        frm_cfg = ttk.LabelFrame(self, text="Settings", padding=10)
        frm_cfg.pack(fill=tk.X, padx=10, pady=6)

        # 左列
        left = ttk.Frame(frm_cfg)
        left.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.filename_substring_var = tk.StringVar(value="xx26")
        self.find_keyword_var = tk.StringVar(value="abcd")
        self.debug_var = tk.BooleanVar(value=True)
        self.process_do_var = tk.BooleanVar(value=True)
        self.process_invoice_var = tk.BooleanVar(value=True)

        ttk.Label(left, text="Filename contains (case-insensitive):").grid(
            row=0, column=0, sticky="w"
        )
        ttk.Entry(left, textvariable=self.filename_substring_var, width=30).grid(
            row=0, column=1, padx=6, pady=2, sticky="w"
        )

        ttk.Label(left, text="Find keyword (case-insensitive):").grid(
            row=1, column=0, sticky="w"
        )
        ttk.Entry(left, textvariable=self.find_keyword_var, width=30).grid(
            row=1, column=1, padx=6, pady=2, sticky="w"
        )

        ttk.Checkbutton(
            left, text="Process DO sheet", variable=self.process_do_var
        ).grid(row=2, column=0, sticky="w", pady=2)
        ttk.Checkbutton(
            left, text="Process Invoice sheet", variable=self.process_invoice_var
        ).grid(row=2, column=1, sticky="w", pady=2)

        ttk.Checkbutton(left, text="Debug logs", variable=self.debug_var).grid(
            row=3, column=0, sticky="w", pady=2
        )

        # 右列：替换内容
        right = ttk.LabelFrame(frm_cfg, text="Replacement values", padding=10)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))

        self.product_code_var = tk.StringVar(value="NA")
        self.product_desc_var = tk.StringVar(value="Toasted Coconut Flakes")
        self.pack_size_var = tk.StringVar(value="(36 x 200g)")
        self.qty_var = tk.StringVar(value="1.00")
        self.uom_var = tk.StringVar(value="PKT")
        self.price_tmpl_var = tk.StringVar(value="=3.8*G{row}")
        self.replace_product_code_var = tk.BooleanVar(value=True)
        self.replace_product_desc_var = tk.BooleanVar(value=True)
        self.replace_pack_size_var = tk.BooleanVar(value=True)
        self.replace_qty_var = tk.BooleanVar(value=True)
        self.replace_uom_var = tk.BooleanVar(value=True)
        self.replace_price_tmpl_var = tk.BooleanVar(value=True)

        r = 0
        ttk.Checkbutton(right, variable=self.replace_product_code_var).grid(
            row=r, column=0, sticky="w"
        )
        ttk.Label(right, text="Product code:").grid(row=r, column=1, sticky="e")
        ttk.Entry(right, textvariable=self.product_code_var, width=28).grid(
            row=r, column=2, padx=6, pady=2, sticky="w"
        )
        r += 1

        ttk.Checkbutton(right, variable=self.replace_product_desc_var).grid(
            row=r, column=0, sticky="w"
        )
        ttk.Label(right, text="Product description:").grid(row=r, column=1, sticky="e")
        ttk.Entry(right, textvariable=self.product_desc_var, width=28).grid(
            row=r, column=2, padx=6, pady=2, sticky="w"
        )
        r += 1

        ttk.Checkbutton(right, variable=self.replace_pack_size_var).grid(
            row=r, column=0, sticky="w"
        )
        ttk.Label(right, text="Pack size:").grid(row=r, column=1, sticky="e")
        ttk.Entry(right, textvariable=self.pack_size_var, width=28).grid(
            row=r, column=2, padx=6, pady=2, sticky="w"
        )
        r += 1

        ttk.Checkbutton(right, variable=self.replace_qty_var).grid(
            row=r, column=0, sticky="w"
        )
        ttk.Label(right, text="Quantity (Qty):").grid(row=r, column=1, sticky="e")
        ttk.Entry(right, textvariable=self.qty_var, width=28).grid(
            row=r, column=2, padx=6, pady=2, sticky="w"
        )
        r += 1

        ttk.Checkbutton(right, variable=self.replace_uom_var).grid(
            row=r, column=0, sticky="w"
        )
        ttk.Label(right, text="Unit (UOM):").grid(row=r, column=1, sticky="e")
        ttk.Entry(right, textvariable=self.uom_var, width=28).grid(
            row=r, column=2, padx=6, pady=2, sticky="w"
        )
        r += 1

        ttk.Checkbutton(right, variable=self.replace_price_tmpl_var).grid(
            row=r, column=0, sticky="w"
        )
        ttk.Label(right, text="Price formula template:").grid(
            row=r, column=1, sticky="e"
        )
        ttk.Entry(right, textvariable=self.price_tmpl_var, width=28).grid(
            row=r, column=2, padx=6, pady=2, sticky="w"
        )
        ttk.Label(
            right, text="Example: =3.8*G{row}, {row} will be replaced by row number"
        ).grid(row=r, column=3, sticky="w")
        r += 1

        # 操作区
        frm_ops = ttk.Frame(self, padding=10)
        frm_ops.pack(fill=tk.X)

        self.btn_run = ttk.Button(frm_ops, text="Start", command=self.on_run)
        self.btn_run.pack(side=tk.LEFT)

        self.prog = ttk.Progressbar(frm_ops, mode="determinate", length=400)
        self.prog.pack(side=tk.LEFT, padx=10)

        self.prog_label = ttk.Label(frm_ops, text="0/0")
        self.prog_label.pack(side=tk.LEFT)

        # 日志区
        frm_log = ttk.LabelFrame(self, text="Logs", padding=8)
        frm_log.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        self.log = ScrolledText(frm_log, height=18)
        try:
            self.log.configure(font=tkfont.nametofont("TkTextFont"))
        except Exception:
            pass
        self.log.pack(fill=tk.BOTH, expand=True)

        # 状态
        self.worker_thread = None

    def choose_dir(self):
        start_dir = self.dir_var.get().strip() or suggest_default_directory()
        if start_dir and not os.path.isdir(start_dir):
            start_dir = suggest_default_directory()
        path = filedialog.askdirectory(
            title="Select root directory",
            mustexist=True,
            initialdir=start_dir if start_dir else "/mnt/c",
        )
        if path:
            self.dir_var.set(path)

    def use_default_dir(self):
        default_dir = suggest_default_directory()
        if default_dir:
            self.dir_var.set(default_dir)
        else:
            messagebox.showwarning(
                "Notice",
                "No default directory found. Please enter a /mnt/c/... path manually.",
            )

    def append_log(self, msg: str):
        if threading.current_thread() is not threading.main_thread():
            self.after(0, self.append_log, msg)
            return
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.update_idletasks()

    def set_progress(self, done: int, total: int):
        if threading.current_thread() is not threading.main_thread():
            self.after(0, self.set_progress, done, total)
            return
        self.prog["maximum"] = max(total, 1)
        self.prog["value"] = done
        self.prog_label.config(text=f"{done}/{total}")

    def set_run_button_enabled(self, enabled: bool):
        if threading.current_thread() is not threading.main_thread():
            self.after(0, self.set_run_button_enabled, enabled)
            return
        self.btn_run.config(state=tk.NORMAL if enabled else tk.DISABLED)

    def on_task_finished(self):
        self.worker_thread = None
        self.set_run_button_enabled(True)

    def on_run(self):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showinfo("Notice", "A task is already running.")
            return

        directory = self.dir_var.get().strip()
        if not directory or not os.path.isdir(directory):
            messagebox.showerror("Error", "Please select a valid root directory.")
            return

        # 读取参数
        filename_substring = self.filename_substring_var.get().strip()
        find_keyword = self.find_keyword_var.get().strip()
        process_do = self.process_do_var.get()
        process_invoice = self.process_invoice_var.get()
        debug = self.debug_var.get()

        product_code = self.product_code_var.get()
        product_desc = self.product_desc_var.get()
        pack_size = self.pack_size_var.get()
        uom = self.uom_var.get()
        price_tmpl = self.price_tmpl_var.get()
        replacement_enabled = {
            "product_code": self.replace_product_code_var.get(),
            "product_description": self.replace_product_desc_var.get(),
            "pack_size": self.replace_pack_size_var.get(),
            "qty": self.replace_qty_var.get(),
            "uom": self.replace_uom_var.get(),
            "price_formula": self.replace_price_tmpl_var.get(),
        }

        # 校验数量
        try:
            qty = float(self.qty_var.get())
        except ValueError:
            messagebox.showerror("Error", "Quantity (Qty) must be a number.")
            return

        if not (process_do or process_invoice):
            messagebox.showerror(
                "Error", "Please select at least one sheet type (DO or Invoice)."
            )
            return

        # 清空进度与日志
        self.log.delete("1.0", tk.END)
        self.set_progress(0, 0)

        def run_task():
            try:
                # 先统计文件数量，设置进度上限
                matched_files = []
                for root, _, files in os.walk(directory):
                    for file in files:
                        low = file.lower()
                        if (
                            low.endswith((".xlsx", ".xlsm"))
                            and filename_substring.lower() in low
                        ):
                            matched_files.append(os.path.join(root, file))
                total = len(matched_files)
                self.set_progress(0, total)

                # 执行处理
                gen = update_excel_files(
                    directory=directory,
                    debug=debug,
                    filename_substring=filename_substring,
                    find_keyword=find_keyword,
                    process_do=process_do,
                    process_invoice=process_invoice,
                    product_code=product_code,
                    product_description=product_desc,
                    pack_size=pack_size,
                    qty=qty,
                    uom=uom,
                    price_formula_template=price_tmpl,
                    replacement_enabled=replacement_enabled,
                    log_func=self.append_log,
                )
                for done, tot in gen:
                    self.set_progress(done, tot)
            except Exception as e:
                self.append_log(f"[GUI] Task exception: {e}")
            finally:
                self.after(0, self.on_task_finished)

        self.set_run_button_enabled(False)
        self.worker_thread = threading.Thread(target=run_task, daemon=True)
        self.worker_thread.start()


if __name__ == "__main__":
    app = App()
    app.mainloop()
