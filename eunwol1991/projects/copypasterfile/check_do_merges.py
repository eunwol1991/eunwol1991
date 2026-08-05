import argparse
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Sequence

from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.cell import range_boundaries


def _is_wsl() -> bool:
    if os.name == "nt":
        return False
    if os.environ.get("WSL_DISTRO_NAME"):
        return True
    try:
        with open("/proc/version", "r", encoding="utf-8") as f:
            return "microsoft" in f.read().lower()
    except OSError:
        return False


def _platform_drive_root() -> str:
    if os.name == "nt":
        return "c:/"
    if _is_wsl():
        return "/mnt/c"
    return "/"


def _from_c(path_tail: str) -> Path:
    tail = (path_tail or "").lstrip("/")
    root = _platform_drive_root()
    if root.endswith("/"):
        return Path(f"{root}{tail}")
    return Path(f"{root}/{tail}")


DEFAULT_TARGET_DIR = _from_c("Users/jhunj/Dropbox/for jj/Doc to print - JJ")
DEFAULT_REFERENCE_DIR = _from_c("Users/jhunj/Dropbox/DO & INV/DO & INV 2026")
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
DO_SIGNATURE_LABELS = {
    "received in good order",
    "authorised signature & stamp",
}
SIGNATURE_END_COLUMN = 11


@dataclass
class CheckResult:
    target_path: Path
    reference_path: Path | None
    missing_ranges: list[str]
    skipped_conflicts: list[str]
    skipped_reason: str = ""

    @property
    def needs_check(self) -> bool:
        return bool(self.missing_ranges or self.skipped_conflicts or self.skipped_reason)


def _norm(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip().lower())


def extract_outlet_name(file_name: str | Path) -> str:
    matches = re.findall(r"\(([^()]*)\)", Path(file_name).name)
    if not matches:
        return ""
    return matches[-1].strip()


def iter_excel_files(folder: Path, *, recursive: bool = False):
    pattern = "**/*" if recursive else "*"
    for path in sorted(folder.glob(pattern)):
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        if path.suffix.lower() not in EXCEL_EXTENSIONS:
            continue
        if "do & inv" not in path.name.lower():
            continue
        yield path


def _select_reference(target_path: Path, candidates: list[Path]) -> Path | None:
    if not candidates:
        return None
    target_prefix = target_path.name.split()[0].lower()
    same_prefix = [p for p in candidates if p.name.lower().startswith(target_prefix)]
    if same_prefix:
        candidates = same_prefix
    format_candidates = [p for p in candidates if "xx26" in p.name.lower()]
    if format_candidates:
        candidates = format_candidates
    return sorted(candidates, key=lambda p: (len(p.parts), p.name.lower()))[0]


ProgressFunc = Callable[[str], None]


def build_reference_index(reference_dir: Path) -> dict[str, list[Path]]:
    index: dict[str, list[Path]] = {}
    for path in reference_dir.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in EXCEL_EXTENSIONS:
            continue
        name_lower = path.name.lower()
        if path.name.startswith("~$"):
            continue
        if "do & inv" not in name_lower:
            continue
        outlet = extract_outlet_name(path)
        if not outlet:
            continue
        index.setdefault(_norm(outlet), []).append(path)
    return index


def find_reference_workbook(target_path: Path, reference_dir: Path) -> Path | None:
    outlet = extract_outlet_name(target_path)
    if not outlet:
        return None
    reference_index = build_reference_index(reference_dir)
    return _select_reference(target_path, reference_index.get(_norm(outlet), []))


def _range_has_non_anchor_value(sheet, merge_range: str) -> bool:
    cell_range = CellRange(merge_range)
    for row in range(cell_range.min_row, cell_range.max_row + 1):
        for col in range(cell_range.min_col, cell_range.max_col + 1):
            if row == cell_range.min_row and col == cell_range.min_col:
                continue
            if sheet.cell(row=row, column=col).value not in (None, ""):
                return True
    return False


def _is_merged(sheet, merge_range: str) -> bool:
    return merge_range in {str(rng) for rng in sheet.merged_cells.ranges}


def _print_area_end_column(sheet) -> int:
    print_area = str(sheet.print_area or "").strip()
    if not print_area:
        return SIGNATURE_END_COLUMN
    first_area = print_area.split(",", 1)[0]
    if "!" in first_area:
        first_area = first_area.split("!", 1)[1]
    first_area = first_area.replace("'", "").replace("$", "")
    try:
        _min_col, _min_row, max_col, _max_row = range_boundaries(first_area)
    except ValueError:
        return SIGNATURE_END_COLUMN
    if max_col is None:
        return SIGNATURE_END_COLUMN
    return max_col


def _missing_signature_label_rows(sheet) -> tuple[list[str], list[str]]:
    missing: list[str] = []
    conflicts: list[str] = []
    end_column = _print_area_end_column(sheet)
    for row in sheet.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            label = _norm(cell.value)
            if label not in DO_SIGNATURE_LABELS:
                continue
            if cell.column >= end_column:
                continue
            start_coord = sheet.cell(row=cell.row, column=cell.column).coordinate
            end_coord = sheet.cell(row=cell.row, column=end_column).coordinate
            merge_range = f"{start_coord}:{end_coord}"
            if _is_merged(sheet, merge_range):
                continue
            if _range_has_non_anchor_value(sheet, merge_range):
                conflicts.append(merge_range)
                continue
            missing.append(merge_range)
    return missing, conflicts


def check_workbook(target_path: Path, reference_path: Path) -> CheckResult:
    target_wb = load_workbook(target_path)
    reference_wb = load_workbook(reference_path, read_only=False, data_only=False)
    missing_ranges: list[str] = []
    skipped_conflicts: list[str] = []
    try:
        if "DO" not in target_wb.sheetnames:
            return CheckResult(target_path, reference_path, [], [], "missing DO sheet")
        if "DO" not in reference_wb.sheetnames:
            return CheckResult(target_path, reference_path, [], [], "reference missing DO sheet")

        target_sheet = target_wb["DO"]
        reference_sheet = reference_wb["DO"]
        existing = {str(rng) for rng in target_sheet.merged_cells.ranges}

        for ref_range in sorted(str(rng) for rng in reference_sheet.merged_cells.ranges):
            if ref_range in existing:
                continue
            if _range_has_non_anchor_value(target_sheet, ref_range):
                skipped_conflicts.append(ref_range)
                continue
            missing_ranges.append(ref_range)

        label_ranges, label_conflicts = _missing_signature_label_rows(target_sheet)
        for merge_range in label_ranges:
            if merge_range not in missing_ranges:
                missing_ranges.append(merge_range)
        for merge_range in label_conflicts:
            if merge_range not in skipped_conflicts:
                skipped_conflicts.append(merge_range)

        return CheckResult(target_path, reference_path, missing_ranges, skipped_conflicts)
    finally:
        target_wb.close()
        reference_wb.close()


def check_folder(
    target_dir: Path,
    reference_dir: Path,
    *,
    recursive: bool = False,
    progress: ProgressFunc | None = None,
) -> list[CheckResult]:
    results: list[CheckResult] = []
    if progress:
        progress(f"Scanning target folder: {target_dir.name}")
    target_paths = list(iter_excel_files(target_dir, recursive=recursive))
    if progress:
        progress(f"Found {len(target_paths)} target workbook(s).")
        progress(f"Indexing reference folder: {reference_dir.name}")
    reference_index = build_reference_index(reference_dir)
    reference_count = sum(len(paths) for paths in reference_index.values())
    if progress:
        progress(f"Found {reference_count} reference workbook(s).")
    for index, target_path in enumerate(target_paths, start=1):
        if progress:
            progress(f"[{index}/{len(target_paths)}] Checking {target_path.name}")
        outlet = extract_outlet_name(target_path)
        reference_path = None
        if outlet:
            reference_path = _select_reference(
                target_path, reference_index.get(_norm(outlet), [])
            )
        if reference_path is None:
            results.append(
                CheckResult(target_path, None, [], [], "no matching reference")
            )
            continue
        results.append(check_workbook(target_path, reference_path))
    return results


def _print_results(results: list[CheckResult]) -> None:
    print("Mode: CHECK ONLY (no workbook changes)")
    for result in results:
        rel_ref = result.reference_path.name if result.reference_path else "-"
        if result.skipped_reason:
            print(f"SKIP {result.target_path.name}: {result.skipped_reason}")
            continue
        status = "CHECK" if result.needs_check else "OK"
        print(
            f"{status} {result.target_path.name}: reference={rel_ref}; "
            f"missing={len(result.missing_ranges)} {result.missing_ranges}; "
            f"conflict={len(result.skipped_conflicts)} {result.skipped_conflicts}"
        )


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Check DO sheets for missing merged cells without modifying workbooks."
    )
    parser.add_argument("--target-dir", type=Path, default=DEFAULT_TARGET_DIR)
    parser.add_argument("--reference-dir", type=Path, default=DEFAULT_REFERENCE_DIR)
    parser.add_argument("--recursive", action="store_true")
    return parser.parse_args(argv)


def main() -> None:
    args = parse_args()

    results = check_folder(
        args.target_dir,
        args.reference_dir,
        recursive=args.recursive,
        progress=lambda message: print(message, flush=True),
    )
    _print_results(results)


if __name__ == "__main__":
    main()
