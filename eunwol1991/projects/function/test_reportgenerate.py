import importlib.util
import os
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook


MODULE_PATH = Path(__file__).with_name("reportgenerate.py")


def _load_module():
    spec = importlib.util.spec_from_file_location("reportgenerate", MODULE_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load module from {MODULE_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class ReportGenerateTests(unittest.TestCase):
    def _workbook_without_default_sheet(self):
        wb = Workbook()
        default = wb.active
        if default is not None:
            wb.remove(default)
        return wb

    def _source_data_sheet(self, rows):
        module = _load_module()
        wb = self._workbook_without_default_sheet()
        ws_data = wb.create_sheet(module.SHEETS["data"])
        ws_data.append(
            [
                "Date",
                "Supplier",
                "Product Code",
                "Product Description",
                "Carton Packing",
                "Customer",
                "Outlet",
                "Qty in Ctns",
                "Qty in Pcs",
                "Invoice #",
                "Total Value",
                "GST",
                "Total Value Inclusive GST",
                "Account",
                "Customer PO#",
                "Total Qty in Pcs",
                "Total Qty in Ctns",
            ]
        )
        for row in rows:
            ws_data.append(row)
        return module, wb, ws_data

    def test_find_latest_source_file_picks_newest_matching_file(self):
        module = _load_module()

        with tempfile.TemporaryDirectory() as temp_dir:
            folder = Path(temp_dir)
            old_file = folder / "Savori Sales Summary old.xlsx"
            new_file = folder / "Savori Sales Summary new.xlsx"
            old_file.write_text("old")
            new_file.write_text("new")
            os.utime(old_file, (1, 1))
            os.utime(new_file, (2, 2))

            self.assertEqual(module.find_latest_source_file(folder), new_file)

    def test_parse_report_date_accepts_iso_date(self):
        module = _load_module()

        self.assertEqual(module.parse_report_date("2026-05-07").isoformat(), "2026-05-07")

    def test_output_file_for_uses_selected_month_and_year(self):
        module = _load_module()

        self.assertEqual(module.output_file_for("2026", "June").name, "Ebi June 2026 Report.xlsx")

    def test_normalize_month_accepts_numeric_month(self):
        module = _load_module()

        self.assertEqual(module.normalize_month("5"), "May")

    def test_resolve_source_input_ignores_shell_cd_command(self):
        module = _load_module()

        with tempfile.TemporaryDirectory() as temp_dir:
            folder = Path(temp_dir)
            source = folder / "Savori Sales Summary latest.xlsx"
            source.write_text("source")

            self.assertEqual(module.resolve_source_input(f"cd {folder}", folder=folder), source)

    def test_build_pivot_like_ignores_formula_quantities_without_crashing(self):
        module = _load_module()

        wb = self._workbook_without_default_sheet()
        ws_data = wb.create_sheet(module.SHEETS["data"])
        ws_data.append(["Customer", "Product Description", "Total Qty in Pcs"])
        ws_data.append(["Customer A", "Product A", "=K70750+6*L70750"])

        wb_out = self._workbook_without_default_sheet()

        module.build_pivot_like(ws_data, wb_out)

        ws_pivot = wb_out[module.SHEETS["pivot"]]
        self.assertEqual(ws_pivot.cell(6, 1).value, "Customer A")
        self.assertEqual(ws_pivot.cell(7, 1).value, "Product A")
        self.assertEqual(ws_pivot.cell(7, 2).value, 0)

    def test_build_pivot_like_uses_monthly_sales_grouped_layout(self):
        module = _load_module()
        wb = self._workbook_without_default_sheet()
        ws_data = wb.create_sheet(module.SHEETS["data"])
        ws_data.append(["Customer", "Product Description", "Total Qty in Pcs"])
        ws_data.append(["Customer B", "Product B", 3])
        ws_data.append(["Customer A", "Product A", 2])
        ws_data.append(["Customer A", "Product B", 4])

        wb_out = self._workbook_without_default_sheet()

        module.build_pivot_like(ws_data, wb_out)

        ws_pivot = wb_out[module.SHEETS["pivot"]]
        self.assertEqual(ws_pivot.title, "EB Monthly Sales")
        self.assertEqual(ws_pivot["A1"].value, "Year")
        self.assertEqual(ws_pivot["B1"].value, module.FILTERS["Year"])
        self.assertEqual(ws_pivot["A2"].value, "Month")
        self.assertEqual(ws_pivot["B2"].value, module.FILTERS["Month"])
        self.assertEqual(ws_pivot["A3"].value, "Supplier")
        self.assertEqual(ws_pivot["B3"].value, module.FILTERS["Supplier"])
        self.assertEqual(ws_pivot["A5"].value, "Row Labels")
        self.assertEqual(ws_pivot["B5"].value, "Sum of Total Qty in Pcs")
        self.assertEqual(ws_pivot["A6"].value, "Customer A")
        self.assertTrue(ws_pivot["A6"].font.bold)
        self.assertEqual(ws_pivot["A7"].value, "Product A")
        self.assertEqual(ws_pivot["B7"].value, 2)
        self.assertEqual(ws_pivot["A8"].value, "Product B")
        self.assertEqual(ws_pivot["B8"].value, 4)
        self.assertEqual(ws_pivot["A9"].value, "Customer B")
        self.assertEqual(ws_pivot["A10"].value, "Product B")
        self.assertEqual(ws_pivot["B10"].value, 3)
        self.assertEqual(ws_pivot["A11"].value, "Grand Total")
        self.assertEqual(ws_pivot["B11"].value, 9)

    def test_visible_sheet_names_are_business_names(self):
        module, wb_out, ws_data = self._source_data_sheet(
            [["2026-05-01", "Ebi", "P001", "Product A", "12 pcs", "MOS Burger", "MOS Outlet", 2, 24, "INV-001", 10, 0.9, 10.9, "Account A", "PO-001", 24, 2]]
        )
        module.build_pivot_like(ws_data, wb_out)
        module.build_weekly(ws_data, wb_out)
        module.build_detail(ws_data, wb_out)
        module.build_upload(wb_out)

        self.assertIn("EB Monthly Sales", wb_out.sheetnames)
        self.assertIn("MOS Trade", wb_out.sheetnames)
        self.assertIn("EB Sell Out to MOS", wb_out.sheetnames)

    def test_weekly_report_filters_date_range_and_leaves_value_columns_blank(self):
        module, wb_out, ws_data = self._source_data_sheet(
            [
                ["2026-05-01", "Ebi", "P001", "Product A", "12 pcs", "Customer A", "Outlet A", 2, 24, "INV-001", 10, 0.9, 10.9, "Account A", "PO-001", 24, 2],
                ["2026-05-08", "Ebi", "P002", "Product B", "12 pcs", "Customer B", "Outlet B", 3, 36, "INV-002", 20, 1.8, 21.8, "Account B", "PO-002", 36, 3],
            ]
        )

        module.build_weekly(ws_data, wb_out, start_date=date(2026, 5, 1), end_date=date(2026, 5, 7))

        ws_weekly = wb_out[module.SHEETS["weekly"]]
        self.assertEqual(ws_weekly.max_row, 2)
        self.assertEqual(ws_weekly.cell(2, 1).value, "2026-05-01")
        self.assertIsNotNone(ws_weekly.cell(2, 5).value)
        self.assertIsNone(ws_weekly.cell(2, 9).value)
        self.assertIsNone(ws_weekly.cell(2, 10).value)
        self.assertIsNone(ws_weekly.cell(2, 11).value)

    def test_weekly_report_splits_same_product_by_order_unit(self):
        module, wb_out, ws_data = self._source_data_sheet(
            [
                ["2026-05-04", "Ebi", "P001", "Mozzerella Cheese Stick", "6 X 850G", "Canadian Pizza", "Outlet A", 0, 11, "INV-001", 10, 0.9, 10.9, "Account A", "PO-001", 11, 1.8333333333],
                ["2026-05-04", "Ebi", "P001", "Mozzerella Cheese Stick", "6 X 850G", "JJ Café", "Outlet B", 1, 0, "INV-002", 20, 1.8, 21.8, "Account B", "PO-002", 6, 1],
            ]
        )

        module.build_weekly(ws_data, wb_out, start_date=date(2026, 5, 4), end_date=date(2026, 5, 4))

        ws_weekly = wb_out[module.SHEETS["weekly"]]
        self.assertEqual(ws_weekly.max_row, 3)
        self.assertEqual(ws_weekly.cell(2, 3).value, "Mozzerella Cheese Stick")
        self.assertEqual(ws_weekly.cell(3, 3).value, "Mozzerella Cheese Stick")
        self.assertIn("_SourceData!I:I", str(ws_weekly.cell(2, 5).value))
        self.assertIn("_SourceData!H:H", str(ws_weekly.cell(3, 6).value))
        self.assertIsNone(ws_weekly.cell(2, 6).value)
        self.assertIsNone(ws_weekly.cell(3, 5).value)

    def test_main_accepts_non_interactive_runtime_options(self):
        module = _load_module()
        self.assertTrue(
            callable(
                lambda: module.main(
                    source_file=Path("sample.xlsx"),
                    year="2026",
                    month="May",
                    weekly_start=date(2026, 5, 1),
                    weekly_end=date(2026, 5, 7),
                    prompt=False,
                )
            )
        )

    def test_detail_and_upload_include_only_mos_rows_and_keep_formulas(self):
        module, wb_out, ws_data = self._source_data_sheet(
            [
                [
                    "2026-05-01",
                    "Ebi",
                    "P001",
                    "Product A",
                    "12 pcs",
                    "MOS Burger",
                    "MOS Outlet",
                    2,
                    24,
                    "INV-001",
                    "=J_SOURCE",
                    "=K_SOURCE",
                    "=L_SOURCE",
                    "Account A",
                    "PO-001",
                    24,
                    "=H2*12",
                ],
                [
                    "2026-05-02",
                    "Ebi",
                    "P002",
                    "Product B",
                    "24 pcs",
                    "Other",
                    "Other Outlet",
                    3,
                    72,
                    "INV-002",
                    20,
                    1.8,
                    21.8,
                    "Account B",
                    "PO-002",
                    72,
                    3,
                ],
            ]
        )

        module.build_detail(ws_data, wb_out)
        module.build_upload(wb_out)

        ws_detail = wb_out[module.SHEETS["detail"]]
        ws_upload = wb_out[module.SHEETS["upload"]]

        self.assertEqual(ws_detail.max_row, 2)
        self.assertEqual(ws_upload.max_row, 2)
        self.assertEqual(ws_detail.cell(2, 6).value, f'={module.SHEETS["data"]}!F2')
        self.assertEqual(ws_detail.cell(2, 10).value, f'={module.SHEETS["data"]}!K2')
        self.assertEqual(ws_upload.cell(2, 4).value, f'={module.SHEETS["data"]}!G2')
        self.assertEqual(ws_upload.cell(2, 7).value, f'={module.SHEETS["data"]}!Q2')

    def test_apply_workbook_font_sets_all_cells_to_calibri(self):
        module = _load_module()
        wb = self._workbook_without_default_sheet()
        ws = wb.create_sheet("Sample")
        ws["A1"] = "Header"
        ws["A2"] = "Value"

        module.apply_workbook_font(wb)

        self.assertEqual(ws["A1"].font.name, "Calibri")
        self.assertEqual(ws["A2"].font.name, "Calibri")


if __name__ == "__main__":
    unittest.main()
