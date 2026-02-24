from copy import copy

from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter


def _normalize_user_values(user_values: dict) -> dict[int, object]:
    out: dict[int, object] = {}
    for key, value in (user_values or {}).items():
        if isinstance(key, int):
            out[key] = value
        else:
            out[column_index_from_string(str(key))] = value
    return out


def _snapshot_row(ws, row: int, min_col: int, max_col: int) -> dict:
    cells = []
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        value = cell.value
        is_formula = isinstance(value, str) and value.startswith("=")
        cells.append(
            {
                "col": col,
                "value": value,
                "is_formula": is_formula,
                "style": copy(cell._style),
                "number_format": cell.number_format,
            }
        )
    return {"cells": cells, "height": ws.row_dimensions[row].height, "row": row}


def _formula_for_row(formula: str, src_row: int, src_col: int, dst_row: int) -> str:
    src_coord = f"{get_column_letter(src_col)}{src_row}"
    dst_coord = f"{get_column_letter(src_col)}{dst_row}"
    return Translator(formula, origin=src_coord).translate_formula(dst_coord)


def insert_with_template(
    ws,
    insert_row: int,
    template_row: int,
    min_col: int,
    max_col: int,
    user_values: dict,
):
    if template_row > insert_row:
        template_row -= 1
    snapshot = _snapshot_row(ws, template_row, min_col, max_col)
    ws.insert_rows(insert_row, 1)

    formula_cols: set[int] = set()
    for item in snapshot["cells"]:
        col = item["col"]
        dst = ws.cell(row=insert_row, column=col)
        dst._style = copy(item["style"])
        dst.number_format = item["number_format"]
        if item["is_formula"]:
            formula_cols.add(col)
            dst.value = _formula_for_row(
                item["value"], snapshot["row"], col, insert_row
            )
        else:
            dst.value = None

    ws.row_dimensions[insert_row].height = snapshot["height"]

    input_map = _normalize_user_values(user_values)
    for col, val in input_map.items():
        if col in formula_cols:
            continue
        ws.cell(row=insert_row, column=col).value = val

    return {"formula_cols": formula_cols}
