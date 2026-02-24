from datetime import datetime
from pathlib import Path
import shutil

from openpyxl import load_workbook


def open_workbook(path: str):
    keep_vba = str(path).lower().endswith(".xlsm")
    return load_workbook(path, data_only=False, keep_vba=keep_vba)


def open_workbook_for_read(path: str):
    keep_vba = str(path).lower().endswith(".xlsm")
    return load_workbook(path, data_only=True, read_only=True, keep_vba=keep_vba)


def make_backup(path: str) -> str:
    src = Path(path)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = src.with_name(f"{src.stem}.backup.{stamp}{src.suffix}")
    shutil.copy2(src, dst)
    return str(dst)
