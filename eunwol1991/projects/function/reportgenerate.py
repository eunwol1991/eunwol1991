from copy import copy
from datetime import datetime
from pathlib import Path
from time import perf_counter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

SOURCE_FOLDER = Path("/mnt/c/Users/jhunj/Dropbox/DO & INV")
SOURCE_FILE = SOURCE_FOLDER / "Savori Sales Summary (2024 - 2026) - As of 28 May'26.xlsx"
SOURCE_SHEET = "Delivery details"

OUTPUT_FOLDER = Path("/mnt/c/Users/jhunj/Dropbox/for jj")
OUTPUT_FILE = OUTPUT_FOLDER / "Ebi May 2026 Report.xlsx"

HEADER_ROW = 4
FIRST_COL = 1
LAST_COL = 24

FILTERS = {
    "Year": "2026",
    "Month": "May",
    "Supplier": "Ebi",
}

SHEETS = {
    "pivot": "EB Monthly Sales",
    "weekly": "Weekly_Report",
    "detail": "MOS Trade",
    "upload": "EB Sell Out to MOS",
    "data": "_SourceData",
}

PROGRESS_EVERY_ROWS = 500
DETAIL_CUSTOMER = "mos"


def log(message):
    print(message, flush=True)


def log_step(step, total_steps, message):
    log(f"[{step}/{total_steps}] {message}")


def log_row_progress(label, current, total):
    if total <= 0:
        return
    if current == total or current % PROGRESS_EVERY_ROWS == 0:
        percent = current / total * 100
        log(f"    {label}: {current}/{total} rows ({percent:.1f}%)")


def find_latest_source_file(folder=SOURCE_FOLDER):
    files = list(Path(folder).glob("Savori Sales Summary*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No Savori Sales Summary*.xlsx found in {folder}")
    return max(files, key=lambda path: path.stat().st_mtime)


def resolve_source_input(value, folder=SOURCE_FOLDER):
    text = str(value).strip()
    if not text or text.startswith("cd "):
        return find_latest_source_file(folder)
    return Path(text)


def normalize_month(value):
    text = str(value).strip()
    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    short_names = {name[:3].lower(): name for name in month_names}
    if text.isdigit():
        month_number = int(text)
        if 1 <= month_number <= 12:
            return month_names[month_number - 1]
    return short_names.get(text[:3].lower(), text)


def parse_report_date(value):
    return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()


def output_file_for(year, month):
    return OUTPUT_FOLDER / f"Ebi {normalize_month(month)} {year} Report.xlsx"


def cell_date(value):
    if hasattr(value, "date"):
        return value.date()
    return parse_report_date(value)


def in_date_range(value, start_date=None, end_date=None):
    if start_date is None and end_date is None:
        return True
    actual = cell_date(value)
    if start_date is not None and actual < start_date:
        return False
    if end_date is not None and actual > end_date:
        return False
    return True


def weekly_order_unit(qty_pcs, qty_ctns):
    if to_number(qty_ctns) and not to_number(qty_pcs):
        return "CTN"
    return "PCS"


def get_runtime_config():
    source_input = input("Source file [Enter = latest Savori Sales Summary*.xlsx]: ").strip()
    source_file = resolve_source_input(source_input)

    year = input(f"Year [{FILTERS['Year']}]: ").strip() or FILTERS["Year"]
    month = normalize_month(input(f"Month [{FILTERS['Month']}]: ").strip() or FILTERS["Month"])

    weekly_start_input = input("Weekly start date YYYY-MM-DD [Enter = whole month]: ").strip()
    weekly_start = parse_report_date(weekly_start_input) if weekly_start_input else None
    weekly_end = None
    if weekly_start is not None:
        weekly_end_input = input("Weekly end date YYYY-MM-DD [Enter = same as start]: ").strip()
        weekly_end = parse_report_date(weekly_end_input) if weekly_end_input else weekly_start

    return source_file, year, month, weekly_start, weekly_end


def to_number(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text or text.startswith("="):
            return 0
        return float(text)
    return float(value)


def sheet_ref(sheet_name, cell_ref):
    safe_name = sheet_name.replace("'", "''")
    if any(ch.isspace() for ch in sheet_name):
        return f"'{safe_name}'!{cell_ref}"
    return f"{safe_name}!{cell_ref}"


def is_detail_customer(value):
    return DETAIL_CUSTOMER in norm(value)


def with_calibri(font):
    updated = copy(font)
    updated.name = "Calibri"
    return updated


def apply_workbook_font(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = with_calibri(cell.font)


def norm(v):
    if v is None:
        return ""
    return str(v).replace("\n", " ").replace("\r", " ").replace("\xa0", " ").strip().lower()


def get_headers(ws):
    headers = {}
    for col in range(FIRST_COL, LAST_COL + 1):
        name = norm(ws.cell(HEADER_ROW, col).value)
        if name:
            headers[name] = col
    return headers


def require(headers, name):
    key = norm(name)
    if key not in headers:
        raise ValueError(f"Missing header: {name}")
    return headers[key]


def copy_cell(src, dst):
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def style_header(ws, max_col):
    fill = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1][:max_col]:
        cell.font = Font(name="Calibri", bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.auto_filter.ref = ws.dimensions


def autofit(ws, max_width=38):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                width = max(width, min(len(str(value)) + 2, max_width))
        ws.column_dimensions[letter].width = width


def build_source_data(ws_src, wb_out):
    ws_data = wb_out.create_sheet(SHEETS["data"])
    headers = get_headers(ws_src)

    filter_cols = {k: require(headers, k) for k in FILTERS}

    for col in range(FIRST_COL, LAST_COL + 1):
        copy_cell(ws_src.cell(HEADER_ROW, col), ws_data.cell(1, col))

    out_row = 2
    max_row = ws_src.max_row

    total_rows = max_row - HEADER_ROW
    for row in range(HEADER_ROW + 1, max_row + 1):
        log_row_progress("Filtering source data", row - HEADER_ROW, total_rows)
        matched = True
        for field, value in FILTERS.items():
            actual = ws_src.cell(row, filter_cols[field]).value
            if str(actual).strip() != value:
                matched = False
                break

        if matched:
            for col in range(FIRST_COL, LAST_COL + 1):
                copy_cell(ws_src.cell(row, col), ws_data.cell(out_row, col))
            out_row += 1

    if out_row == 2:
        raise ValueError("No data matched filters.")

    ws_data.sheet_state = "hidden"
    return ws_data


def build_pivot_like(ws_data, wb_out):
    ws = wb_out.create_sheet(SHEETS["pivot"])

    headers = {norm(ws_data.cell(1, col).value): col for col in range(1, ws_data.max_column + 1)}
    c_customer = require(headers, "Customer")
    c_product = require(headers, "Product Description")
    c_qty = require(headers, "Total Qty in Pcs")

    ws["A1"] = "Year"
    ws["B1"] = FILTERS["Year"]
    ws["A2"] = "Month"
    ws["B2"] = FILTERS["Month"]
    ws["A3"] = "Supplier"
    ws["B3"] = FILTERS["Supplier"]

    summary = {}

    total_rows = ws_data.max_row - 1
    for row in range(2, ws_data.max_row + 1):
        log_row_progress("Building pivot summary", row - 1, total_rows)
        customer = ws_data.cell(row, c_customer).value
        product = ws_data.cell(row, c_product).value
        qty = ws_data.cell(row, c_qty).value or 0

        key = (customer, product)
        summary[key] = summary.get(key, 0) + to_number(qty)

    ws["A5"] = "Row Labels"
    ws["B5"] = "Sum of Total Qty in Pcs"

    out_row = 6
    grand_total = 0
    current_customer = object()
    for (customer, product), qty in sorted(summary.items(), key=lambda x: (str(x[0][0]), str(x[0][1]))):
        if customer != current_customer:
            current_customer = customer
            ws.cell(out_row, 1).value = customer
            ws.cell(out_row, 1).font = Font(name="Calibri", bold=True)
            out_row += 1
        ws.cell(out_row, 1).value = product
        ws.cell(out_row, 2).value = qty
        grand_total += qty
        out_row += 1

    ws.cell(out_row, 1).value = "Grand Total"
    ws.cell(out_row, 2).value = grand_total
    for cell in ws[out_row][:2]:
        cell.font = Font(name="Calibri", bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")

    for row in range(1, 4):
        for cell in ws[row][:2]:
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
    fill = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[5][:2]:
        cell.font = Font(name="Calibri", bold=True)
        cell.fill = fill
        cell.border = border
    ws.auto_filter.ref = ws.dimensions
    autofit(ws)


def build_weekly(ws_data, wb_out, start_date=None, end_date=None):
    ws = wb_out.create_sheet(SHEETS["weekly"])

    headers_out = [
        "Date", "Supplier", "Product Description", "Carton Packing",
        "Qty in Pcs", "Qty in Ctns", "Total Qty in Pcs", "Total Qty in Ctns",
        "Total Value", "GST", "Total Value Inclusive GST"
    ]

    for i, h in enumerate(headers_out, 1):
        ws.cell(1, i).value = h

    headers = {norm(ws_data.cell(1, col).value): col for col in range(1, ws_data.max_column + 1)}

    c_date = require(headers, "Date")
    c_supplier = require(headers, "Supplier")
    c_product = require(headers, "Product Description")
    c_pack = require(headers, "Carton Packing")
    c_qty_pcs = require(headers, "Qty in Pcs")
    c_qty_ctns = require(headers, "Qty in Ctns")

    sum_fields_by_unit = {
        "PCS": {5: "Qty in Pcs", 7: "Total Qty in Pcs"},
        "CTN": {6: "Qty in Ctns", 8: "Total Qty in Ctns"},
    }

    keys = []
    seen = set()

    total_rows = ws_data.max_row - 1
    for row in range(2, ws_data.max_row + 1):
        log_row_progress("Collecting weekly keys", row - 1, total_rows)
        if not in_date_range(ws_data.cell(row, c_date).value, start_date, end_date):
            continue
        key = (
            ws_data.cell(row, c_date).value,
            ws_data.cell(row, c_supplier).value,
            ws_data.cell(row, c_product).value,
            ws_data.cell(row, c_pack).value,
            weekly_order_unit(ws_data.cell(row, c_qty_pcs).value, ws_data.cell(row, c_qty_ctns).value),
        )
        if key not in seen:
            seen.add(key)
            keys.append(key)

    total_keys = len(keys)
    for r, key in enumerate(keys, 2):
        log_row_progress("Writing weekly rows", r - 1, total_keys)
        ws.cell(r, 1).value = key[0]
        ws.cell(r, 2).value = key[1]
        ws.cell(r, 3).value = key[2]
        ws.cell(r, 4).value = key[3]
        unit = key[4]

        for idx, field in sum_fields_by_unit[unit].items():
            sum_col = get_column_letter(require(headers, field))
            date_col = get_column_letter(c_date)
            supplier_col = get_column_letter(c_supplier)
            product_col = get_column_letter(c_product)
            pack_col = get_column_letter(c_pack)
            unit_col = get_column_letter(c_qty_pcs if unit == "PCS" else c_qty_ctns)

            ws.cell(r, idx).value = (
                f'=SUMIFS({SHEETS["data"]}!{sum_col}:{sum_col},'
                f'{SHEETS["data"]}!{date_col}:{date_col},A{r},'
                f'{SHEETS["data"]}!{supplier_col}:{supplier_col},B{r},'
                f'{SHEETS["data"]}!{product_col}:{product_col},C{r},'
                f'{SHEETS["data"]}!{pack_col}:{pack_col},D{r},'
                f'{SHEETS["data"]}!{unit_col}:{unit_col},"<>0")'
            )

    style_header(ws, len(headers_out))
    autofit(ws)


def build_detail(ws_data, wb_out):
    ws = wb_out.create_sheet(SHEETS["detail"])

    headers_out = [
        "Date", "Supplier", "Product Code", "Product Description", "Carton Packing",
        "Customer", "Outlet", "Qty in Ctns", "Invoice #",
        "Total Value", "GST", "Total Value Inclusive GST",
        "Account", "Customer PO#"
    ]

    for i, h in enumerate(headers_out, 1):
        ws.cell(1, i).value = h

    headers = {norm(ws_data.cell(1, col).value): col for col in range(1, ws_data.max_column + 1)}
    c_customer = require(headers, "Customer")

    total_rows = ws_data.max_row - 1
    out_row = 2
    for source_row in range(2, ws_data.max_row + 1):
        log_row_progress("Writing detail rows", source_row - 1, total_rows)
        if not is_detail_customer(ws_data.cell(source_row, c_customer).value):
            continue
        for c, h in enumerate(headers_out, 1):
            src_col = require(headers, h)
            src_letter = get_column_letter(src_col)
            ws.cell(out_row, c).value = f'={sheet_ref(SHEETS["data"], f"{src_letter}{source_row}")}'
        out_row += 1

    style_header(ws, len(headers_out))
    autofit(ws)


def build_upload(wb_out):
    ws_detail = wb_out[SHEETS["detail"]]
    ws_data = wb_out[SHEETS["data"]]
    ws = wb_out.create_sheet(SHEETS["upload"])

    headers_out = [
        "Date", "Doc No", "Reference NO", "Customer", "Product Code",
        "Product Description", "Sales Qty", "UOM", "Price Per PCS",
        "Total Value", "Pack Size (PCS)"
    ]

    for i, h in enumerate(headers_out, 1):
        ws.cell(1, i).value = h

    detail_headers = {norm(ws_detail.cell(1, col).value): col for col in range(1, ws_detail.max_column + 1)}

    mapping = {
        1: "Date",
        2: "Supplier",
        3: "Invoice #",
        5: "Product Code",
        6: "Product Description",
    }

    source_data_headers = {norm(ws_data.cell(1, col).value): col for col in range(1, ws_data.max_column + 1)}
    customer_col = require(source_data_headers, "Customer")
    outlet_col = require(source_data_headers, "Outlet")
    qty_ctn_col = require(source_data_headers, "Total Qty in Ctns")

    total_rows = ws_detail.max_row - 1
    out_row = 2
    for detail_row in range(2, ws_detail.max_row + 1):
        log_row_progress("Writing upload rows", detail_row - 1, total_rows)
        source_row = None
        detail_customer_formula = ws_detail.cell(detail_row, require(detail_headers, "Customer")).value
        if isinstance(detail_customer_formula, str) and detail_customer_formula.startswith(f'={SHEETS["data"]}!'):
            source_row = int("".join(ch for ch in detail_customer_formula.split("!", 1)[1] if ch.isdigit()))
        if source_row is None or not is_detail_customer(ws_data.cell(source_row, customer_col).value):
            continue
        for out_col, source_header in mapping.items():
            src_col = require(detail_headers, source_header)
            src_letter = get_column_letter(src_col)
            ws.cell(out_row, out_col).value = f'={sheet_ref(SHEETS["detail"], f"{src_letter}{detail_row}")}'

        ws.cell(out_row, 4).value = f'={sheet_ref(SHEETS["data"], f"{get_column_letter(outlet_col)}{source_row}")}'
        ws.cell(out_row, 7).value = f'={sheet_ref(SHEETS["data"], f"{get_column_letter(qty_ctn_col)}{source_row}")}'

        # H-K 留空，给你自己填
        for col in range(8, 12):
            ws.cell(out_row, col).value = None
        out_row += 1

    style_header(ws, len(headers_out))
    autofit(ws)


def main(source_file=None, year=None, month=None, weekly_start=None, weekly_end=None, prompt=True):
    if prompt and source_file is None and year is None and month is None and weekly_start is None and weekly_end is None:
        source_file, year, month, weekly_start, weekly_end = get_runtime_config()

    if source_file is None:
        source_file = SOURCE_FILE
    if year is not None:
        FILTERS["Year"] = str(year)
    if month is not None:
        FILTERS["Month"] = normalize_month(month)

    output_file = output_file_for(FILTERS["Year"], FILTERS["Month"])

    source_path = Path(source_file)
    if not source_path.exists():
        raise FileNotFoundError(f"Source file not found: {source_path}")

    start = perf_counter()
    log_step(1, 7, f"Loading source workbook: {source_path}")
    wb_src = load_workbook(source_path, data_only=True)
    ws_src = wb_src[SOURCE_SHEET]

    wb_out = Workbook()
    default = wb_out.active
    if default is not None:
        wb_out.remove(default)

    log_step(2, 7, "Filtering matching source rows...")
    ws_data = build_source_data(ws_src, wb_out)
    log_step(3, 7, "Building pivot report...")
    build_pivot_like(ws_data, wb_out)
    log_step(4, 7, "Building weekly report...")
    build_weekly(ws_data, wb_out, weekly_start, weekly_end)
    log_step(5, 7, "Building detail report...")
    build_detail(ws_data, wb_out)
    log_step(6, 7, "Building upload format...")
    build_upload(wb_out)

    wb_out.move_sheet(wb_out[SHEETS["data"]], offset=4)
    apply_workbook_font(wb_out)

    log_step(7, 7, "Saving output workbook...")
    wb_out.save(output_file)
    elapsed = perf_counter() - start
    log(f"Done in {elapsed:.1f}s: {output_file}")


if __name__ == "__main__":
    main()
