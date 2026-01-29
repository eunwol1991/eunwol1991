import os
import re
import threading
import traceback
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText

import openpyxl


def update_excel_files(
    directory: str,
    debug: bool,
    filename_substring: str,
    find_keyword: str,
    process_do: bool,
    process_invoice: bool,
    product_code: str,
    product_description: str,
    pack_size: str,
    qty: float,
    uom: str,
    price_formula_template: str,
    log_func=print,
):
    """
    批量更新 Excel 文件（.xlsx/.xlsm），在 DO / Invoice 两张表中查找包含关键字的单元格并修改对应列。
    - filename_substring: 仅处理文件名中包含该子串（忽略大小写）的文件，如 'xx25'
    - find_keyword: 查找关键字（忽略大小写、去两端空格）
    - price_formula_template: 价格单元格写入的公式模板，支持 {row} 占位符，例如 '=3.8*G{row}'
    """
    processed_count = 0
    error_files = []
    matched_files = []

    # 先统计需要处理的文件数量，便于进度条
    for root, _, files in os.walk(directory):
        for file in files:
            low = file.lower()
            if low.endswith((".xlsx", ".xlsm")) and filename_substring.lower() in low:
                matched_files.append(os.path.join(root, file))

    total = len(matched_files)
    log_func(f"共发现 {total} 个候选文件。")

    find_kw = (find_keyword or "").strip().lower()

    for idx, file_path in enumerate(matched_files, start=1):
        if debug:
            log_func(f"🔄 正在处理 [{idx}/{total}]: {file_path}")

        workbook = None
        modified = False
        try:
            workbook = openpyxl.load_workbook(file_path)
            if debug:
                log_func(f"✅ 已加载: {os.path.basename(file_path)}")
                log_func(f"📄 工作表: {workbook.sheetnames}")

            # 根据勾选处理的表，生成顺序列表
            sheets_to_process = []
            if process_do and "DO" in workbook.sheetnames:
                sheets_to_process.append("DO")
            if process_invoice and "Invoice" in workbook.sheetnames:
                sheets_to_process.append("Invoice")

            for sheet_name in sheets_to_process:
                sheet = workbook[sheet_name]
                if debug:
                    log_func(f"📌 处理工作表: {sheet_name}")

                # 逐格扫描
                for row in sheet.iter_rows():
                    for cell in row:
                        cell_value = str(cell.value).strip(
                        ).lower() if cell.value is not None else ""
                        if find_kw and find_kw in cell_value:
                            row_number = cell.row

                            # 价格公式：将 {row} 替换为行号
                            price_formula = (price_formula_template or "").replace(
                                "{row}", str(row_number))

                            if debug:
                                log_func(
                                    f"🔍 匹配到 {find_kw} → Sheet: {sheet_name}, Row: {row_number}")

                            if sheet_name == "DO":
                                sheet[f"B{row_number}"].value = product_code
                                sheet[f"C{row_number}"].value = product_description
                                sheet[f"G{row_number}"].value = pack_size
                                # sheet[f"H{row_number}"].value = None
                                sheet[f"I{row_number}"].value = qty
                                sheet[f"K{row_number}"].value = uom
                                if debug:
                                    log_func(
                                        f"✅ DO 修改: B{row_number}, C{row_number}, H{row_number}, I{row_number}, K{row_number}")

                            elif sheet_name == "Invoice":
                                sheet[f"B{row_number}"].value = product_code
                                sheet[f"C{row_number}"].value = product_description
                                sheet[f"F{row_number}"].value = pack_size
                                sheet[f"G{row_number}"].value = qty
                                sheet[f"H{row_number}"].value = uom
                                sheet[f"I{row_number}"].value = price_formula if price_formula else None
                                if debug:
                                    log_func(
                                        f"✅ Invoice 修改: B{row_number}, C{row_number}, F{row_number}, G{row_number}, H{row_number}, I{row_number}")

                            modified = True
                            # 不跳出，继续处理所有匹配

            if modified:
                workbook.save(file_path)
                processed_count += 1
                if debug:
                    log_func(f"💾 已保存修改: {file_path}")

        except Exception as e:
            traceback.print_exc()
            error_files.append(file_path)
            log_func(f"❌ 处理失败: {file_path}，错误: {e}")
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass

        yield idx, total  # 进度更新

    log_func(f"✅ 处理完成，共修改 {processed_count} 个文件。")
    if error_files:
        log_func("⚠️ 以下文件处理失败：")
        for ef in error_files:
            log_func(f"  ❌ {ef}")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel 批量更新工具")
        self.geometry("860x620")

        # 路径选择
        frm_top = ttk.Frame(self, padding=10)
        frm_top.pack(fill=tk.X)
        ttk.Label(frm_top, text="根目录：").pack(side=tk.LEFT)
        self.dir_var = tk.StringVar()
        ttk.Entry(frm_top, textvariable=self.dir_var,
                  width=80).pack(side=tk.LEFT, padx=6)
        ttk.Button(frm_top, text="浏览…",
                   command=self.choose_dir).pack(side=tk.LEFT)

        # 参数区
        frm_cfg = ttk.LabelFrame(self, text="参数设置", padding=10)
        frm_cfg.pack(fill=tk.X, padx=10, pady=6)

        # 左列
        left = ttk.Frame(frm_cfg)
        left.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.filename_substring_var = tk.StringVar(value="xx26")
        self.find_keyword_var = tk.StringVar(value="abcd")
        self.debug_var = tk.BooleanVar(value=True)
        self.process_do_var = tk.BooleanVar(value=True)
        self.process_invoice_var = tk.BooleanVar(value=True)

        ttk.Label(left, text="文件名包含（忽略大小写）：").grid(row=0, column=0, sticky="w")
        ttk.Entry(left, textvariable=self.filename_substring_var, width=30).grid(
            row=0, column=1, padx=6, pady=2, sticky="w")

        ttk.Label(left, text="查找关键字（忽略大小写）：").grid(row=1, column=0, sticky="w")
        ttk.Entry(left, textvariable=self.find_keyword_var, width=30).grid(
            row=1, column=1, padx=6, pady=2, sticky="w")

        ttk.Checkbutton(left, text="处理 DO 表", variable=self.process_do_var).grid(
            row=2, column=0, sticky="w", pady=2)
        ttk.Checkbutton(left, text="处理 Invoice 表", variable=self.process_invoice_var).grid(
            row=2, column=1, sticky="w", pady=2)

        ttk.Checkbutton(left, text="Debug 日志", variable=self.debug_var).grid(
            row=3, column=0, sticky="w", pady=2)

        # 右列：替换内容
        right = ttk.LabelFrame(frm_cfg, text="替换内容", padding=10)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))

        self.product_code_var = tk.StringVar(value="NA")
        self.product_desc_var = tk.StringVar(value="Toasted Coconut Flakes")
        self.pack_size_var = tk.StringVar(value="(36 x 200g)")
        self.qty_var = tk.StringVar(value="1.00")
        self.uom_var = tk.StringVar(value="PKT")
        self.price_tmpl_var = tk.StringVar(value="=3.8*G{row}")

        r = 0
        ttk.Label(right, text="产品代码：").grid(row=r, column=0, sticky="e")
        ttk.Entry(right, textvariable=self.product_code_var, width=28).grid(
            row=r, column=1, padx=6, pady=2, sticky="w")
        r += 1

        ttk.Label(right, text="产品描述：").grid(row=r, column=0, sticky="e")
        ttk.Entry(right, textvariable=self.product_desc_var, width=28).grid(
            row=r, column=1, padx=6, pady=2, sticky="w")
        r += 1

        ttk.Label(right, text="规格（Pack Size）：").grid(
            row=r, column=0, sticky="e")
        ttk.Entry(right, textvariable=self.pack_size_var, width=28).grid(
            row=r, column=1, padx=6, pady=2, sticky="w")
        r += 1

        ttk.Label(right, text="数量（Qty）：").grid(row=r, column=0, sticky="e")
        ttk.Entry(right, textvariable=self.qty_var, width=28).grid(
            row=r, column=1, padx=6, pady=2, sticky="w")
        r += 1

        ttk.Label(right, text="单位（UOM）：").grid(row=r, column=0, sticky="e")
        ttk.Entry(right, textvariable=self.uom_var, width=28).grid(
            row=r, column=1, padx=6, pady=2, sticky="w")
        r += 1

        ttk.Label(right, text="单价公式模板：").grid(row=r, column=0, sticky="e")
        ttk.Entry(right, textvariable=self.price_tmpl_var, width=28).grid(
            row=r, column=1, padx=6, pady=2, sticky="w")
        ttk.Label(
            right, text="例如：=3.8*G{row}，{row} 将被替换为行号").grid(row=r, column=2, sticky="w")
        r += 1

        # 操作区
        frm_ops = ttk.Frame(self, padding=10)
        frm_ops.pack(fill=tk.X)

        self.btn_run = ttk.Button(frm_ops, text="开始处理", command=self.on_run)
        self.btn_run.pack(side=tk.LEFT)

        self.prog = ttk.Progressbar(frm_ops, mode="determinate", length=400)
        self.prog.pack(side=tk.LEFT, padx=10)

        self.prog_label = ttk.Label(frm_ops, text="0/0")
        self.prog_label.pack(side=tk.LEFT)

        # 日志区
        frm_log = ttk.LabelFrame(self, text="日志", padding=8)
        frm_log.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        self.log = ScrolledText(frm_log, height=18)
        self.log.pack(fill=tk.BOTH, expand=True)

        # 状态
        self.worker_thread = None

    def choose_dir(self):
        path = filedialog.askdirectory(title="选择根目录")
        if path:
            self.dir_var.set(path)

    def append_log(self, msg: str):
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.update_idletasks()

    def on_run(self):
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showinfo("提示", "任务正在进行中…")
            return

        directory = self.dir_var.get().strip()
        if not directory or not os.path.isdir(directory):
            messagebox.showerror("错误", "请选择有效的根目录")
            return

        # 读取参数
        filename_substring = self.filename_substring_var.get().strip()
        find_keyword = self.find_keyword_var.get().strip()
        process_do = self.process_do_var.get()
        process_invoice = self.process_invoice_var.get()
        debug = self.debug_var.get()

        product_code = self.product_code_var.get()
        product_desc = self.product_desc_var.get()
        pack_size = self.pack_size_var.get()
        uom = self.uom_var.get()
        price_tmpl = self.price_tmpl_var.get()

        # 校验数量
        try:
            qty = float(self.qty_var.get())
        except ValueError:
            messagebox.showerror("错误", "数量（Qty）必须是数字")
            return

        if not (process_do or process_invoice):
            messagebox.showerror("错误", "请至少勾选一种工作表（DO 或 Invoice）")
            return

        # 清空进度与日志
        self.log.delete("1.0", tk.END)
        self.prog["value"] = 0
        self.prog_label.config(text="0/0")

        def run_task():
            try:
                # 先统计文件数量，设置进度上限
                matched_files = []
                for root, _, files in os.walk(directory):
                    for file in files:
                        low = file.lower()
                        if low.endswith((".xlsx", ".xlsm")) and filename_substring.lower() in low:
                            matched_files.append(os.path.join(root, file))
                total = len(matched_files)
                self.prog["maximum"] = max(total, 1)
                self.prog_label.config(text=f"0/{total}")

                # 执行处理
                gen = update_excel_files(
                    directory=directory,
                    debug=debug,
                    filename_substring=filename_substring,
                    find_keyword=find_keyword,
                    process_do=process_do,
                    process_invoice=process_invoice,
                    product_code=product_code,
                    product_description=product_desc,
                    pack_size=pack_size,
                    qty=qty,
                    uom=uom,
                    price_formula_template=price_tmpl,
                    log_func=self.append_log,
                )
                for done, tot in gen:
                    self.prog["value"] = done
                    self.prog_label.config(text=f"{done}/{tot}")
            except Exception as e:
                self.append_log(f"[GUI] 任务异常: {e}")
            finally:
                self.btn_run.config(state=tk.NORMAL)

        self.btn_run.config(state=tk.DISABLED)
        self.worker_thread = threading.Thread(target=run_task, daemon=True)
        self.worker_thread.start()


if __name__ == "__main__":
    app = App()
    app.mainloop()
