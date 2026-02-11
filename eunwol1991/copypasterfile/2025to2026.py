import os
from openpyxl import load_workbook

def replace_in_cell(cell):
    if cell.value is None:
        return

    # 如果是字符串
    if isinstance(cell.value, str):
        new_value = cell.value.replace("xx25", "xx26").replace("2025", "2026")
        cell.value = new_value

    # 如果是数字，专门处理 2025
    elif isinstance(cell.value, int) and cell.value == 2025:
        cell.value = 2026


def process_excel_file(file_path):
    try:
        wb = load_workbook(file_path)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    replace_in_cell(cell)

        wb.save(file_path)
        print(f"Updated content: {file_path}")

    except Exception as e:
        print(f"Failed to process {file_path}: {e}")


def rename_file_if_needed(file_path):
    folder, filename = os.path.split(file_path)
    if "xx25" in filename:
        new_filename = filename.replace("xx25", "xx26")
        new_path = os.path.join(folder, new_filename)
        os.rename(file_path, new_path)
        print(f"Renamed file: {filename} → {new_filename}")
        return new_path
    return file_path


def main():
    root_dir = input("请输入要处理的文件夹路径: ").strip()

    if not os.path.isdir(root_dir):
        print("路径无效")
        return

    for root, _, files in os.walk(root_dir):
        for file in files:
            if file.lower().endswith((".xlsx", ".xlsm", ".xls")):
                full_path = os.path.join(root, file)
                process_excel_file(full_path)
                rename_file_if_needed(full_path)

    print("处理完成 ✅")


if __name__ == "__main__":
    main()
