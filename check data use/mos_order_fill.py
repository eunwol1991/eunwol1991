"""
MOS Order → Excel  自动填表脚本（Savori三行版 + 关键字门店匹配）
依赖：pip install pymupdf openpyxl regex
"""

from pathlib import Path
import re, sys, shutil
import fitz                     # PyMuPDF
import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

# === 1. 配置区 ============================================================ #
EXCEL_PATH   = Path(r"C:\Users\User\Dropbox\for jj\mos order\Order Summary For MOS - JJ.xlsx")
SHEET_NAME   = "MOS Format"
PDF_DIR      = Path(r"C:\Users\User\Dropbox\for jj\mos_pdfs")  # 待处理 PDF 文件夹
PO_COL       = "AI"              # PO 号所在列
ROW_START    = 39                # B39-B73
ROW_END      = 73
COL_START    = "C"               # C38-AD38
COL_END      = "AD"
HIGHLIGHT    = PatternFill(fill_type="solid", fgColor="00FFFCD7")
ARCHIVE_DIR  = Path(r"C:\Users\User\Dropbox\for jj\mos order")

# ======== ① 门店别名（键已全部小写） =====================================
STORE_ALIAS = {
    "mos burger 100 am (38)":                 "MOS - 100AM",
    "mos burger 18 tai seng (67)":            "MOS - 18 Tai Seng",
    "mos burger amk hub (24)":                "MOS - Ang Mo Kio Hub",
    "mos burger bedok mall (42)":             "MOS - Bedok Mall",
    "mos burger jewel changi airport (59)":   "MOS - Café Jewel Changi Airport*",
    "mos burger causeway point (33)":         "MOS - Causeway Point",
    "mos burger changi city point (80)":      "MOS - Changi City Point",
    "mos burger city square mall (69)":       "MOS - City Square Mall",
    "mos burger compass one (49)":            "MOS - Compass One",
    "mos burger eastpoint (64)":              "MOS - Eastpoint Mall",
    "mos burger bukit gombak mrt (71)":       "MOS - Express Bukit Gombak MRT",
    "mos burger holland village mrt (70)":    "MOS - Express Holland Village MRT",
    "mos burger fusionopolis one (55)":       "MOS - Fusionopolis One",
    "mos burger harbourfront centre (31)":    "MOS - Harbourfront Centre",
    "mos burger heartland mall kovan (58)":   "MOS - Heartland Mall Kovan",
    "mos burger hillion mall (56)":           "MOS - Hillion Mall",
    "mos burger hougang mall (63)":           "MOS - Hougang Mall",
    "mos burger ion orchard (79)":            "MOS - ION Orchard",
    "mos burger jem (39)":                    "MOS - JEM",
    "mos burger jurong point (14)":           "MOS - Jurong Point",
    "mos burger kampung admiralty (53)":      "MOS - Kampung Admiralty",
    "mos burger marina bay link mall (81)":   "MOS - Marina Bay Link Mall",
    "mos burger merlion park (72)":           "MOS - Merlion Park",
    "mos burger millenia walk (30)":          "MOS - Millenia Walk",
    "mos burger nex (36)":                    "MOS - NEX",
    "mos burger northpoint city (50)":        "MOS - Northpoint",
    "mos burger our tampines hub (57)":       "MOS - Our Tampines Hub",
    "mos burger raffles city (20)":           "MOS - Raffles City",
    "mos burger seletar mall (75)":           "MOS - Seletar Mall",
    "mos burger tampines mall (35)":          "MOS - Tampines Mall",
    "mos burger the centrepoint (51)":        "MOS - The Centrepoint",
    "mos burger tiong bahru plaza (48)":      "MOS - Tiong Bahru Plaza",
    "mos burger toa payoh hdb hub (12)":      "MOS - Toa Payoh HDB Hub",
    "mos burger waterway point (62)":         "MOS - Waterway Point",
    "mos burger white sands (47)":            "MOS - White Sands",
}

# ===== ② 物品别名（键已全部小写、尽量去括号） ============================
ITEM_ALIAS = {
    "ikeda japanese chicken cutlet":                "Ikeda Japanese Chicken Cutlet (6 x 1.1kg)",
    "japanese chicken katsu 70g":                   "CS TAY Japanese Chicken Katsu 70G (6 x 1kg)",
    "crispy fried 3 joint wing":                    "CS Tay Crispy Fried Chicken Wings 3 Joints (1132) (6 x 1.1kg)",
    "cs tay crispy chicken patty":                  "CS Tay Crispy Chicken Patty (4 x 2.5kg)",
    "eb kranch alaska pollock fingers":             "Kranch Alaska Pollock Finger (5 x 900g)",
    "battered natural onion rings":                 "Salud Battered Natural Onion Ring (6 x 1kg)",
    "rosti hashbrown":                              "Rosti Hashbrown (10 x 1kg)",
    "prawn noodle soup":                            "Prawn Noodle Soup (10 x 1kg)",
    "prawn noodle sauce":                           "Prawn Noodle Sauce (10 x 1kg)",
    "tomato capsicum soup":                         "Tomato Capsicum Soup (10 x 1kg)",
    "par cooked beef patty 120g":                   "Cooked Hamburg Beef Patty 120G (10 x 720g)",
    "demi glace sauce":                             "Demi Glace Sauce (10 x 1kg)",
    "hoisin sauce":                                 "Hoisin Sauce (10 x 1kg)",
    "seaweed egg drop soup":                        "Seaweed Egg Drop Soup (10 x 1kg)",
    "chilli crab sauce":                            "Chilli Crab Sauce (10 x 1kg)",
    "coconut sauce":                                "Coconut Sauce (12 x 500g)",
    "coleslaw":                                     "Coleslaw (5 x 2kg)",
    "anchor minidish unsalted butter 7g":           "Anchor Minidish Unsalted Butter (1 x 144 x 7g)",
    "mini mantou, plain":                           "Mini Mantou Plain (24 x 180g)",
    "bobo sliced fish cake 1kg":                    "BoBo Sliced Fish Cake (1kg)",
    "chicken honey baked ham sliced":               "Chicken Honey Baked Ham Sliced (12 x 1kg)",
    "mozzarella cheese 5g pearl iqf":               "Mozzerella Cheese 5G Pearls IQF (8 x 1kg)",
    "premium smoked salmon sliced r trout":         "Premium Smoked Salmon Sliced (R Trout) (10 x 1kg)",
    "hard boiled egg 10pcs only":                   "Hard Boiled Egg (18 x 10pcs)",
    "fried shallot 1kg":                            "Fried Shallot (10 x 1kg)",
    "jalapenos nachos sliced":                      "Jalapeno Pepper Sliced (6 x 2.95kg)",
    "toasted coconut flakes":                       "Toasted Coconut Flakes (36 x 200g)",
    "parsley flakes":                               "Parsley Leaves (6 x 5gm)",
}

# ========================================================================== #

UNIT_RE = r"(?:ctn|ctns|pkt|pkts|tin|tins|can|cans|box|boxes|btl|btls|pc|pcs)"

# UOM pluralization helpers
_PLURAL_MAP = {
    "ctn": "ctns",
    "pkt": "pkts",
    "tin": "tins",
    "can": "cans",
    "box": "boxes",
    "btl": "btls",
    "pc":  "pcs",
}
_SINGULAR_MAP = {v: k for k, v in _PLURAL_MAP.items()}

def _normalize_uom(uom: str) -> str:
    u = (uom or "").strip().lower()
    # If already plural form, convert to singular base
    if u in _SINGULAR_MAP:
        return _SINGULAR_MAP[u]
    return u

def _format_qty_uom(qty: int, uom: str) -> str:
    base = _normalize_uom(uom)
    if qty and qty > 1:
        plural = _PLURAL_MAP.get(base, base + "s")
        return f"{qty} {plural}"
    # qty == 0 or 1 -> singular
    return f"{qty} {base}"

def _norm_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()

def _norm_item_key(desc: str) -> str:
    s = desc.lower()
    s = re.sub(r"\(.*?\)", "", s)          # 去括号内容
    s = re.sub(r"[^a-z0-9\s,.-]+", " ", s) # 去奇异符号
    s = _norm_spaces(s)
    return s

def _map_item(desc: str) -> str:
    key = _norm_item_key(desc)
    if key in ITEM_ALIAS:
        return ITEM_ALIAS[key]
    k2 = key.replace(",", "")
    if k2 in ITEM_ALIAS:
        return ITEM_ALIAS[k2]
    for k in ITEM_ALIAS:
        if k in key or key in k:
            return ITEM_ALIAS[k]
    return desc

def _tokenize_item(s: str) -> set[str]:
    """Tokenize item text for fuzzy matching: lowercased words, length>=3."""
    norm = _norm_item_key(s)
    toks = [w for w in norm.split() if len(w) >= 3]
    return set(toks)

def find_best_item_col(desc: str, item_col: dict[str, int]) -> tuple[int | None, str | None]:
    """
    Fuzzy match: choose Excel header with maximum token overlap.
    - Accept if unique max overlap >= 1.
    - If tie with max==1 (e.g., only 'japanese' overlaps), treat as ambiguous -> None.
    - If tie with max>=2, pick the one with greater total overlap character length; if still tie, ambiguous.
    Returns (column_index, matched_header) or (None, None) if not resolvable.
    """
    pdf_tokens = _tokenize_item(desc)
    if not pdf_tokens:
        return None, None

    scored: list[tuple[int, int, str, int]] = []  # (score, overlap_chars, header, col)
    for header, col in item_col.items():
        header_tokens = _tokenize_item(header)
        overlap = pdf_tokens & header_tokens
        score = len(overlap)
        if score > 0:
            overlap_chars = sum(len(t) for t in overlap)
            scored.append((score, overlap_chars, header, col))

    if not scored:
        return None, None

    scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
    top_score = scored[0][0]
    top = [s for s in scored if s[0] == top_score]

    if len(top) == 1 and top_score >= 1:
        return top[0][3], top[0][2]

    if top_score >= 2:
        # Use overlap_chars to break ties
        top.sort(key=lambda x: x[1], reverse=True)
        if len(top) == 1 or top[0][1] > top[1][1]:
            return top[0][3], top[0][2]
        # Still ambiguous
        return None, None

    # top_score == 1 and multiple candidates -> ambiguous, require one more word
    return None, None

# ---------- 关键字/部分匹配：门店归一化与解析 ----------
def _norm_store_text(s: str) -> str:
    """把门店文本归一化，用于“包含关系”判断。"""
    s = s.lower()
    s = re.sub(r"\(.*?\)", "", s)                 # 去掉 (12)
    s = s.replace("mos", " ").replace("burger", " ")
    s = s.replace("mosburger", " ")
    s = s.replace("mos -", " ").replace("mos–", " ")
    s = s.replace("-", " ")
    s = re.sub(r"[^a-z0-9\s]", " ", s)            # 去符号
    s = _norm_spaces(s)
    return s

def resolve_store(pdf_store_line: str, store_row_keys: list[str]) -> str | None:
    """
    返回 Excel 里的标准门店名称（store_row 的键），支持关键字/部分匹配。
    匹配顺序：
      1) 走别名表（严格）；
      2) 归一化后做“包含关系”匹配（pdf_norm in excel_norm 或 excel_norm in pdf_norm）。
    """
    # 1) 别名（原始行直接查）
    alias_hit = STORE_ALIAS.get(pdf_store_line.lower())
    if alias_hit:
        # 找到别名对应的 Excel 标准名（store_row 的键应该是这个样子）
        # store_row 的键是 Excel 的显示文本，通常已经是标准名；这里直接返回别名目标文本
        return alias_hit

    # 2) 关键字/部分匹配
    pdf_norm = _norm_store_text(pdf_store_line)
    # 预先归一化 Excel 的门店键
    excel_norm_map = {}  # norm -> original key
    for k in store_row_keys:
        excel_norm_map[_norm_store_text(k)] = k

    # 2.1 直接包含（pdf_norm 是 excel_norm 的子串，或反之）
    candidates = []
    for excel_norm, orig in excel_norm_map.items():
        if not pdf_norm or not excel_norm:
            continue
        if pdf_norm in excel_norm or excel_norm in pdf_norm:
            candidates.append((len(excel_norm), orig))
    if candidates:
        # 选择 excel_norm 最长的（更具体）
        candidates.sort(reverse=True)
        return candidates[0][1]

    # 2.2 把 pdf_norm 拆词做关键词 contains（例如“toa payoh”）
    words = [w for w in pdf_norm.split() if len(w) >= 3]
    for excel_norm, orig in excel_norm_map.items():
        if all(w in excel_norm for w in words):
            return orig

    return None

def _ensure_dir(p: Path):
    try:
        p.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass

def _safe_move_file(src: Path, dest_dir: Path):
    try:
        _ensure_dir(dest_dir)
        target = dest_dir / src.name
        if target.exists():
            stem = src.stem
            suffix = src.suffix
            i = 1
            while True:
                cand = dest_dir / f"{stem} ({i}){suffix}"
                if not cand.exists():
                    target = cand
                    break
                i += 1
        shutil.move(str(src), str(target))
    except Exception as e:
        print(f"[!] 无法移动文件 {src} -> {dest_dir}: {e}", file=sys.stderr)

def _extract_po_from_lines(lines: list[str]) -> str:
    """稳健抽取：优先 Approved by 附近，允许数字被空白/标点分隔；回退 To 规则。"""
    joined = " \n ".join(lines)
    m = re.search(r"(?is)approved\s*by\s*[:\-]?[^\d]{0,80}?(\d\D?\d\D?\d\D?\d\D?\d\D?\d)", joined)
    if m:
        digits = re.sub(r"\D", "", m.group(1))
        if len(digits) >= 6 and int(digits[:6]) >= 600000:
            return digits[:6]
    for i, ln in enumerate(lines):
        if re.search(r"(?i)approved\s*by", ln):
            window = " ".join(lines[i:i+4])
            m2 = re.search(r"(?is)\b(\d\D?\d\D?\d\D?\d\D?\d\D?\d)\b", window)
            if m2:
                digits = re.sub(r"\D", "", m2.group(1))
                if len(digits) >= 6 and int(digits[:6]) >= 600000:
                    return digits[:6]
    for i in range(len(lines) - 1):
        if lines[i].strip().lower() == "to":
            nxt = " ".join(lines[i+1:i+3])
            m3 = re.search(r"\b(\d\D?\d\D?\d\D?\d\D?\d\D?\d)\b", nxt)
            if m3:
                digits = re.sub(r"\D", "", m3.group(1))
                if len(digits) >= 6 and int(digits[:6]) >= 600000:
                    return digits[:6]
    approved_pos = re.search(r"(?i)approved\s*by", joined)
    pos = approved_pos.start() if approved_pos else 0
    candidates = [(m.start(), re.sub(r"\D", "", m.group(0))) for m in re.finditer(r"\b\d{6}\b", joined)]
    candidates = [c for c in candidates if int(c[1][:6]) >= 600000]
    if candidates:
        candidates.sort(key=lambda t: abs(t[0] - pos))
        return candidates[0][1]
    return ""

def parse_pdf(pdf_path: Path):
    """
    适配 Savori 的“3行配对”版式：
      行1：纯数字货号
      行2：描述
      行3：数量+单位（如 '2 ctn' / '1 pkt' / '1 can'）
    门店：独立一行 'MOS Burger ... (nn)'（也支持关键字匹配）
    PO：上一行是 'To'，下一行是纯数字（如 201024）
    """
    pages = []
    with fitz.open(pdf_path) as doc:
        for page in doc:
            raw = page.get_text("text") or ""
            pages.append([ln.strip() for ln in raw.splitlines() if ln.strip()])
    return pages

def extract_orders_from_lines(lines: list[str]) -> dict:
    """从单页 lines 中抽取: store_line, po, items(list[{desc, qty}])"""
    # 门店行（原样）
    store_line = None
    for ln in lines:
        if re.fullmatch(r"(?i)MOS Burger.+\(\d+\)", ln, flags=re.I):
            store_line = ln
            break
    # PO 优先从 “Approved by: <Name>” 后面抓取；找不到再回退到 'To' 规则
    po = ""
    for i, ln in enumerate(lines):
        low = ln.lower()
        if "approved by:" in low:
            # 场景1：同一行后面直接跟 6 开头的 6 位数字
            m = re.search(r"\b(\d{6})\b", ln)
            if m and int(m.group(1)) >= 600000:
                po = m.group(1)
                break
            # 场景2：下一行是 6 开头的 6 位数字
            if i + 1 < len(lines) and re.fullmatch(r"\d{6}", lines[i + 1].strip()):
                if int(lines[i + 1].strip()) >= 600000:
                    po = lines[i + 1].strip()
                    break
    if not po:
        for i in range(len(lines) - 1):
            if lines[i].strip().lower() == "to" and re.fullmatch(r"\d{6}", lines[i + 1].strip()):
                if int(lines[i + 1].strip()) >= 600000:
                    po = lines[i + 1].strip()
                    break
    # 最后再用跨行/灵活匹配加强一次（若上面未取到或被错误取值）
    po2 = _extract_po_from_lines(lines)
    if po2:
        po = po2

    # 明细：三行配对
    items = []
    # Capture both quantity and UOM
    qty_re = re.compile(rf"^(\d+)\s+({UNIT_RE})$", re.I)
    i = 0
    while i < len(lines) - 2:
        if re.fullmatch(r"\d{5,}", lines[i]):
            desc = lines[i + 1]
            m_qty = qty_re.fullmatch(lines[i + 2])
            if m_qty:
                items.append({
                    "desc": desc,
                    "qty": int(m_qty.group(1)),
                    "uom": m_qty.group(2).lower(),
                })
                i += 3
                continue
        i += 1
    return {"store_line": store_line, "po": po, "items": items}

def build_mappings(ws):
    """生成 {店名: 行号}, {物品名: 列号}"""
    def _norm_store_for_excel(s: str) -> str:
        s = _norm_spaces(str(s))
        s = re.sub(r"\*+$", "", s).strip()
        return s

    store_row = {}
    for r in range(ROW_START, ROW_END + 1):
        v = ws[f"B{r}"].value
        if v is None:
            continue
        store_row[_norm_store_for_excel(v)] = r

    c1 = column_index_from_string(COL_START)
    c2 = column_index_from_string(COL_END)
    item_col = {}
    for c in range(c1, c2 + 1):
        v = ws.cell(38, c).value
        if v:
            item_col[_norm_spaces(str(v))] = c

    return store_row, item_col

def clear_target_cells(ws):
    # 清空目标范围内的历史值与高亮
    c1 = column_index_from_string(COL_START)
    c2 = column_index_from_string(COL_END)
    for r in range(ROW_START, ROW_END + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill()
        # 清空 PO 列
        po_cell = ws[f"{PO_COL}{r}"]
        po_cell.value = None

def _merge_po(orig: str, new_po: str) -> str:
    if not new_po:
        return orig
    if not orig:
        return new_po
    parts = [p.strip() for p in str(orig).split(",") if p.strip()]
    if new_po not in parts:
        parts.append(new_po)
    return ", ".join(parts)

def main():
    wb = xl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    clear_target_cells(ws)
    store_row, item_col = build_mappings(ws)
    not_found_store, not_found_item, updated = [], [], []

    pdf_list = sorted(PDF_DIR.glob("*.pdf"))
    for pdf_file in pdf_list:
        pdf_had_update = False
        try:
            pages = parse_pdf(pdf_file)
            for lines in pages:
                page_info = extract_orders_from_lines(lines)
                # 检测是否为 amend 页（包含 amend/amended/amendment 等关键字）
                page_has_amend = any(re.search(r"\bamend", ln, flags=re.I) for ln in lines)
                if not page_info["items"]:
                    continue  # 该页无明细

                # 门店解析：优先使用 store_line；支持关键字/部分匹配
                resolved_store_name = None
                if page_info["store_line"]:
                    resolved_store_name = resolve_store(page_info["store_line"], list(store_row.keys()))
                else:
                    # 如果 PDF 没有清晰门店行，也可以在 lines 里找任何包含 MOS 的行来尝试
                    for ln in lines:
                        if "mos" in ln.lower():
                            resolved_store_name = resolve_store(ln, list(store_row.keys()))
                            if resolved_store_name:
                                break

                if not resolved_store_name:
                    # 门店还匹配不到，就记录未匹配（给出提要关键词）
                    not_found_store.append({"store": page_info["store_line"] or "(missing)", "po": page_info["po"]})
                    continue

                r = store_row.get(resolved_store_name)
                if r is None:
                    not_found_store.append({"store": page_info["store_line"], "po": page_info["po"]})
                    continue

                # 写入每个明细
                for it in page_info["items"]:
                    item_std = _map_item(it["desc"])
                    c = item_col.get(item_std)
                    matched_header = item_std
                    if c is None:
                        # Fuzzy fallback by token overlap
                        c, matched_header = find_best_item_col(it["desc"], item_col)
                    if c is None:
                        not_found_item.append({"store": resolved_store_name, "item": matched_header or item_std, "qty": it.get("qty"), "uom": it.get("uom", ""), "po": page_info["po"]})
                        continue

                    # 写入“数字 + UOM”，并根据数量进行单复数调整
                    if page_has_amend and int(it.get("qty", 0)) == 0:
                        cell = ws.cell(r, c)
                        cell.value = None
                        cell.fill = PatternFill()
                    else:
                        ws.cell(r, c).value = _format_qty_uom(it["qty"], it.get("uom", ""))
                        ws.cell(r, c).fill  = HIGHLIGHT
                    pdf_had_update = True

                # PO 写入为数字类型；如有多次写入，后写覆盖前写
                po_cell = ws[f"{PO_COL}{r}"]
                po_val = page_info["po"].strip()
                if re.fullmatch(r"\d{1,}\Z", po_val):
                    try:
                        po_cell.value = int(po_val)
                    except Exception:
                        po_cell.value = po_val  # fallback to text if conversion fails
                else:
                    po_cell.value = po_val
                if po_val:
                    pdf_had_update = True

                updated.append((resolved_store_name, page_info["po"]))
        except Exception as e:
            print(f"[!] 解析 {pdf_file.name} 出错：{e}", file=sys.stderr)
        # 已写入则移动 PDF 到归档目录
        if pdf_had_update:
            _safe_move_file(pdf_file, ARCHIVE_DIR)

    # 覆盖保存（只保存一次）
    wb.save(EXCEL_PATH)
    wb.close()

    print(f"✅ 已更新 {len(updated)} 条记录，处理 {len(pdf_list)} 份 PDF → {EXCEL_PATH}")
    if not_found_store or not_found_item:
        print("\n⚠️ 未匹配成功：")
        if not_found_store:
            print("  [门店未匹配]")
            for od in not_found_store:
                print(f"   - store={od.get('store')} / PO {od.get('po')}")
        if not_found_item:
            print("  [物品未匹配]")
            for od in not_found_item:
                qty_uom = _format_qty_uom(od.get('qty') or 0, od.get('uom', '')) if od.get('qty') is not None else ''
                print(f"   - {od['store']} / {od['item']} / {qty_uom} / PO {od['po']}")

if __name__ == "__main__":
    main()
