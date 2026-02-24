from openpyxl import Workbook, load_workbook


def _make_file(tmp_path):
    p = tmp_path / "delivery.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Delivery Details"

    headers = [
        "Date",
        "Month",
        "Year",
        "Product Description",
        "Customer",
        "Outlet",
        "Qty in Pcs",
        "Qty in Ctns",
        "Total Qty in Pcs",
        "GST",
        "Invoice #",
    ]
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=3, column=idx).value = h

    ws.cell(row=4, column=1).value = "19/02/2026"
    ws.cell(row=4, column=2).value = 2
    ws.cell(row=4, column=3).value = 2026
    ws.cell(row=4, column=4).value = "Chicken Nugget"
    ws.cell(row=4, column=5).value = "AEON"
    ws.cell(row=4, column=6).value = "Tampines"
    ws.cell(row=4, column=7).value = 12
    ws.cell(row=4, column=8).value = 1
    ws.cell(row=4, column=9).value = "=G4+H4*24"
    ws.cell(row=4, column=10).value = "=I4*0.09"
    ws.cell(row=4, column=11).value = "INV 2602-001"

    ws.cell(row=5, column=1).value = "02/01/2024"
    ws.cell(row=5, column=4).value = "Chicken Nugget"
    ws.cell(row=5, column=5).value = "AEON"
    ws.cell(row=5, column=6).value = "Tampines Mall"
    ws.cell(row=5, column=7).value = 8
    ws.cell(row=5, column=8).value = 1
    ws.cell(row=5, column=9).value = "=G5+H5*24"
    ws.cell(row=5, column=10).value = "=I5*0.09"
    ws.cell(row=5, column=11).value = "INV 2401-004"

    wb.save(p)
    return str(p)


def test_service_suggest_and_insert(tmp_path):
    from eunwol1991.projects.function.delivery_assistant.service import (
        apply_insert,
        load_context,
        preview_insert,
        suggest,
    )

    path = _make_file(tmp_path)
    ctx = load_context(path)
    ranked = suggest(
        ctx,
        {"customer": "aeon", "outlet": "tampines", "description": "chicken nugget"},
        limit=2,
    )
    assert ranked[0]["record"]["row_idx"] == 4

    plan = preview_insert(
        ctx,
        {
            "description": "Chicken Nugget",
            "customer": "AEON",
            "outlet": "Bedok",
            "qty_pcs": 10,
            "qty_ctns": 2,
            "invoice": "INV 2602-002",
        },
    )
    backup = apply_insert(ctx, plan)
    assert backup.endswith(".xlsx")

    wb2 = load_workbook(path, data_only=False)
    ws = wb2["Delivery Details"]
    assert ws.cell(row=4, column=4).value == "Chicken Nugget"
    assert ws.cell(row=4, column=5).value == "AEON"
    assert ws.cell(row=4, column=6).value == "Bedok"
    assert ws.cell(row=4, column=7).value == 10
    assert ws.cell(row=4, column=8).value == 2
    assert ws.cell(row=4, column=9).value == "=G4+H4*24"
    assert ws.cell(row=4, column=10).value == "=I4*0.09"
    assert ws.cell(row=4, column=11).value == "INV 2602-002"
    wb2.close()
