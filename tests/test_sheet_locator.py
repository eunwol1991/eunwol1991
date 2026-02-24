from openpyxl import Workbook


def test_find_sheet_ci_case_insensitive():
    from eunwol1991.projects.function.delivery_assistant.sheet_locator import (
        find_sheet_ci,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Details"

    found = find_sheet_ci(wb, "delivery details")
    assert found.title == "Delivery Details"


def test_detect_header_row_best_match():
    from eunwol1991.projects.function.delivery_assistant.sheet_locator import (
        detect_header_row,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery details"

    ws["A1"] = "random"
    ws["A3"] = "Date"
    ws["B3"] = "Month"
    ws["C3"] = "Customer"
    ws["D3"] = "Outlet"
    ws["E3"] = "Product Description"
    ws["F3"] = "Qty in Pcs"
    ws["G3"] = "Qty in Ctns"
    ws["H3"] = "Invoice #"

    header_row = detect_header_row(
        ws,
        expected_headers={
            "date",
            "month",
            "customer",
            "outlet",
            "product description",
            "qty in pcs",
            "qty in ctns",
            "invoice #",
        },
        scan_rows=10,
    )

    assert header_row == 3
