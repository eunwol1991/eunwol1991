from copy import copy

from openpyxl import Workbook
from openpyxl.styles import PatternFill


def _build_ws():
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery details"

    ws["A3"] = "Date"
    ws["B3"] = "Customer"
    ws["C3"] = "Outlet"
    ws["D3"] = "Qty in Pcs"
    ws["E3"] = "Qty in Ctns"
    ws["F3"] = "Total Qty in Pcs"
    ws["G3"] = "GST"
    ws["H3"] = "Invoice #"

    ws["A4"] = "19/02/2026"
    ws["B4"] = "AEON"
    ws["C4"] = "Tampines"
    ws["D4"] = 12
    ws["E4"] = 1
    ws["F4"] = "=D4+E4*24"
    ws["G4"] = "=F4*0.09"
    ws["H4"] = "INV 2602-001"

    ws["D4"].fill = PatternFill(fill_type="solid", fgColor="00FFFCD7")
    ws["F4"].number_format = "#,##0"
    ws["G4"].number_format = "#,##0.00"
    ws.row_dimensions[4].height = 24
    return wb, ws


def test_insert_row_copies_formula_and_style():
    from eunwol1991.projects.function.delivery_assistant.row_insert import (
        insert_with_template,
    )

    _, ws = _build_ws()
    insert_with_template(
        ws=ws,
        insert_row=4,
        template_row=5,
        min_col=1,
        max_col=8,
        user_values={
            "A": "20/02/2026",
            "B": "AEON",
            "C": "Bedok",
            "D": 10,
            "E": 2,
            "H": "INV 2602-002",
        },
    )

    assert ws["F4"].value == "=D4+E4*24"
    assert ws["G4"].value == "=F4*0.09"
    assert ws["D4"].fill.fgColor.rgb == ws["D5"].fill.fgColor.rgb
    assert ws.row_dimensions[4].height == 24


def test_insert_does_not_overwrite_formula_columns():
    from eunwol1991.projects.function.delivery_assistant.row_insert import (
        insert_with_template,
    )

    _, ws = _build_ws()
    insert_with_template(
        ws=ws,
        insert_row=4,
        template_row=5,
        min_col=1,
        max_col=8,
        user_values={
            "F": 12345,
            "G": 888,
            "D": 1,
            "E": 1,
        },
    )

    assert ws["F4"].value == "=D4+E4*24"
    assert ws["G4"].value == "=F4*0.09"
