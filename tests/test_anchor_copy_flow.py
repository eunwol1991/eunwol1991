from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


def _make_anchor_file(tmp_path):
    p = tmp_path / "delivery_anchor.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery details"

    headers = [
        "Date",
        "Month",
        "Year",
        "Product Code",
        "Product Description",
        "Customer",
        "Outlet",
        "Qty in Pcs",
        "Qty in Ctns",
        "Total Qty in Pcs",
        "GST",
        "Invoice #",
    ]
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=3, column=idx).value = h

    ws.cell(row=4, column=1).value = "19/02/2026"
    ws.cell(row=4, column=4).value = "PC-A"
    ws.cell(row=4, column=5).value = "Chicken Nugget"
    ws.cell(row=4, column=6).value = "AEON"
    ws.cell(row=4, column=7).value = "Tampines"
    ws.cell(row=4, column=8).value = 12
    ws.cell(row=4, column=9).value = 1
    ws.cell(row=4, column=10).value = "=H4+I4*24"
    ws.cell(row=4, column=11).value = "=J4*0.09"
    ws.cell(row=4, column=12).value = "INV-001"

    ws.cell(row=5, column=1).value = "18/02/2026"
    ws.cell(row=5, column=4).value = "PC-B"
    ws.cell(row=5, column=5).value = "Chicken Nugget"
    ws.cell(row=5, column=6).value = "AEON"
    ws.cell(row=5, column=7).value = "Bedok"
    ws.cell(row=5, column=8).value = 20
    ws.cell(row=5, column=9).value = 2
    ws.cell(row=5, column=10).value = "=H5+I5*24"
    ws.cell(row=5, column=11).value = "=J5*0.09"
    ws.cell(row=5, column=12).value = "INV-002"

    fill = PatternFill(fill_type="solid", fgColor="00FFF2CC")
    for col in range(1, 13):
        ws.cell(row=8, column=col).fill = fill
        ws.cell(row=8, column=col).value = "anchor"

    wb.save(p)
    return str(p)


def test_preview_uses_anchor_and_product_code_source(tmp_path):
    from eunwol1991.projects.function.delivery_assistant.service import (
        load_context,
        preview_insert,
    )

    path = _make_anchor_file(tmp_path)
    ctx = load_context(path)
    plan = preview_insert(
        ctx,
        {
            "description": "Chicken Nugget",
            "product_code": "PC-B",
            "customer": "AEON",
            "outlet": "Bedok",
            "qty_pcs": 11,
            "qty_ctns": 1,
            "invoice": "INV-NEW",
        },
    )
    assert plan["anchor_row"] == 8
    assert plan["source_row"] == 5
    assert plan["insert_row"] == 9
    assert plan["template_row"] == 5


def test_apply_insert_copies_source_formula_and_updates_inputs(tmp_path):
    from eunwol1991.projects.function.delivery_assistant.service import (
        apply_insert,
        load_context,
        preview_insert,
    )

    path = _make_anchor_file(tmp_path)
    ctx = load_context(path)
    plan = preview_insert(
        ctx,
        {
            "description": "Chicken Nugget",
            "product_code": "PC-B",
            "customer": "AEON",
            "outlet": "Bedok North",
            "qty_pcs": 9,
            "qty_ctns": 3,
            "invoice": "INV-NEW-01",
        },
    )
    apply_insert(ctx, plan)

    wb = load_workbook(path, data_only=False)
    ws = wb["Delivery details"]
    assert ws.cell(row=9, column=4).value == "PC-B"
    assert ws.cell(row=9, column=7).value == "Bedok North"
    assert ws.cell(row=9, column=8).value == 9
    assert ws.cell(row=9, column=9).value == 3
    assert ws.cell(row=9, column=10).value == "=H9+I9*24"
    assert ws.cell(row=9, column=11).value == "=J9*0.09"
    wb.close()
