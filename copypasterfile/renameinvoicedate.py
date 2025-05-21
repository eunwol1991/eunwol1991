import os
import openpyxl
from tkinter import Tk, filedialog

def update_excel_files():
    """
    Search for Excel files in the user-selected directory and its subdirectories. In each file, update cells with values
    containing 'xx24' to 'xx25' and dates formatted as 'xx/xx/2024' to 'xx/xx/2025'.
    """
    # Open a file dialog to select the directory
    Tk().withdraw()  # Hide the main Tkinter window
    directory = filedialog.askdirectory(title="Select Directory")

    if not directory:
        print("No directory selected. Exiting.")
        return

    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith(('.xlsx', '.xlsm')):
                file_path = os.path.join(root, file)
                print(f"Processing: {file_path}")

                try:
                    workbook = openpyxl.load_workbook(file_path)

                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]

                        for row in sheet.iter_rows():
                            for cell in row:
                                if isinstance(cell.value, str):
                                    # Replace 'xx24' with 'xx25'
                                    if 'xx24' in cell.value:
                                        cell.value = cell.value.replace('xx24', 'xx25')

                                    # Replace 'xx/xx/2024' with 'xx/xx/2025'
                                    if '/2024' in cell.value:
                                        cell.value = cell.value.replace('/2024', '/2025')

                    # Save changes back to the file
                    workbook.save(file_path)
                    print(f"Updated and saved: {file_path}")

                except Exception as e:
                    print(f"Failed to process {file_path}: {e}")

# Example usage
if __name__ == "__main__":
    update_excel_files()
