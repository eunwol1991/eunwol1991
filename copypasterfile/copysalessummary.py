import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import openpyxl
from openpyxl import Workbook
import datetime
import os

def main():
    root = tk.Tk()
    root.title("订单自动生成 - 示例：修复 row_in_excel 类型问题")
    root.geometry("600x550")

    # =================== 全局变量 ===================
    df_global = None
    wb_global = None
    source_sheet_name = "Delivery details"
    target_sheet_name = "Sheet1"
    TARGET_FILE_PATH = r"C:\Users\User\Dropbox\for jj\For Sales Summary(py)\tmr sales data1.xlsx"

    def select_file():
        file_path = filedialog.askopenfilename(
            title="选择源 Sales Summary 文件",
            filetypes=[("Excel文件", "*.xlsx *.xls")]
        )
        file_var.set(file_path)
        if file_path:
            load_data(file_path)

    def load_data(excel_file):
        nonlocal df_global, wb_global

        # ========== 读取Excel ==========
        try:
            # 如果第4行为表头，header=3
            df = pd.read_excel(
                excel_file,
                sheet_name=source_sheet_name,
                header=3
            )
        except Exception as e:
            messagebox.showerror("错误", f"无法读取Excel：\n{e}")
            return

        # ========== 强制把Year转成int，避免出现 2022.0 ==========
        if "Year" not in df.columns:
            messagebox.showerror("错误", "缺少 'Year' 列，请检查表头。")
            return
        df["Year"] = pd.to_numeric(df["Year"], errors="coerce")
        df.dropna(subset=["Year"], inplace=True)
        df["Year"] = df["Year"].astype(int)

        df_global = df

        # ========== 打开源工作簿(复制格式需要) ==========
        try:
            wb = openpyxl.load_workbook(excel_file)
            wb_global = wb
        except Exception as e:
            messagebox.showerror("错误", f"无法打开源工作簿：\n{e}")
            return

        # ========== 初始化下拉框 Year ==========
        years = df_global["Year"].dropna().unique().tolist()
        year_cb['values'] = sorted(years)
        year_var.set("")

        # 其余下拉框重置
        customer_cb['values'] = []
        customer_var.set("")
        outlet_cb['values'] = []
        outlet_var.set("")
        product_cb['values'] = []
        product_var.set("")

    # ========== 链式下拉回调 ==========
    def on_year_selected(event):
        if df_global is None:
            return
        try:
            selected_year = int(year_var.get())
        except:
            return
        df_year = df_global[df_global["Year"] == selected_year]
        if "Customer" not in df_year.columns:
            messagebox.showerror("错误", "缺少 'Customer' 列。")
            return
        customers = df_year["Customer"].dropna().unique().tolist()
        customer_cb['values'] = sorted(customers)
        customer_var.set("")
        outlet_cb['values'] = []
        outlet_var.set("")
        product_cb['values'] = []
        product_var.set("")

    def on_customer_selected(event):
        if df_global is None:
            return
        try:
            selected_year = int(year_var.get())
        except:
            return
        selected_customer = customer_var.get()

        df_filtered = df_global[
            (df_global["Year"] == selected_year) &
            (df_global["Customer"] == selected_customer)
        ]
        if "Outlet" not in df_filtered.columns:
            messagebox.showerror("错误", "缺少 'Outlet' 列。")
            return
        outlets = df_filtered["Outlet"].dropna().unique().tolist()
        outlet_cb['values'] = sorted(outlets)
        outlet_var.set("")
        product_cb['values'] = []
        product_var.set("")

    def on_outlet_selected(event):
        if df_global is None:
            return
        try:
            selected_year = int(year_var.get())
        except:
            return
        selected_customer = customer_var.get()
        selected_outlet = outlet_var.get()

        df_filtered = df_global[
            (df_global["Year"] == selected_year) &
            (df_global["Customer"] == selected_customer) &
            (df_global["Outlet"] == selected_outlet)
        ]
        if "Product Description" not in df_filtered.columns:
            messagebox.showerror("错误", "缺少 'Product Description' 列。")
            return
        products = df_filtered["Product Description"].dropna().unique().tolist()
        product_cb['values'] = sorted(products)
        product_var.set("")

    # ========== 核心：生成订单(复制行+更新字段) ==========
    def generate_order():
        excel_file = file_var.get()
        if not excel_file:
            messagebox.showerror("错误", "请先选择源文件。")
            return

        if df_global is None or wb_global is None:
            messagebox.showerror("错误", "数据未正确载入。")
            return

        # 获取界面输入
        try:
            selected_year = int(year_var.get())
        except:
            messagebox.showerror("错误", "Year选择有误。")
            return

        selected_customer = customer_var.get()
        selected_outlet = outlet_var.get()
        selected_product = product_var.get()
        target_date_str = target_date_var.get()
        invoice_num = invoice_var.get()
        qty_pcs = qty_pcs_var.get()
        qty_ctns = qty_ctns_var.get()
        po_number = po_var.get()

        if not (selected_customer and selected_outlet and selected_product
                and target_date_str and invoice_num and qty_pcs and qty_ctns):
            messagebox.showerror("错误", "请完整填写必填项！（日期、Invoice、件数、箱数）")
            return

        try:
            target_date = pd.to_datetime(target_date_str)
        except ValueError:
            messagebox.showerror("错误", "目标日期格式错误（YYYY-MM-DD）。")
            return

        df_copy = df_global.copy()
        if "Date" not in df_copy.columns:
            messagebox.showerror("错误", "缺少 'Date' 列。")
            return
        df_copy["Date"] = pd.to_datetime(df_copy["Date"], errors="coerce")
        df_copy.sort_values(by="Date", ascending=False, inplace=True)

        filtered = df_copy[
            (df_copy["Year"] == selected_year) &
            (df_copy["Customer"] == selected_customer) &
            (df_copy["Outlet"] == selected_outlet) &
            (df_copy["Product Description"] == selected_product)
        ]
        if filtered.empty:
            messagebox.showinfo("提示", "未找到符合条件的记录。")
            return

        latest_record = filtered.iloc[0]
        global_index = latest_record.name

        # ========== 修正：确保 row_in_excel 是 int，而不是 bytes/str ==========
        row_in_excel = int(global_index) + 5
        print("DEBUG: row_in_excel =", row_in_excel, type(row_in_excel))
        # 这里应打印出一个纯 <class 'int'>, 如 row_in_excel=10, <class 'int'>

        source_sheet = wb_global[source_sheet_name]

        # 打开目标文件
        try:
            target_wb = openpyxl.load_workbook(TARGET_FILE_PATH)
        except FileNotFoundError:
            target_wb = openpyxl.Workbook()
            target_wb.create_sheet(target_sheet_name, 0)
        except Exception as e:
            messagebox.showerror("错误", f"无法打开目标文件：\n{e}")
            return

        if target_sheet_name not in target_wb.sheetnames:
            target_ws = target_wb.create_sheet(target_sheet_name)
        else:
            target_ws = target_wb[target_sheet_name]

        target_last_row = target_ws.max_row
        paste_row = target_last_row + 1

        # ========== 复制源行到目标行 ==========
        # 注意：source_sheet[row_in_excel] 要求 row_in_excel 是 int
        source_row = source_sheet[row_in_excel]
        for col_idx, source_cell in enumerate(source_row, start=1):
            target_cell = target_ws.cell(row=paste_row, column=col_idx, value=source_cell.value)
            if source_cell.has_style:
                target_cell._style = source_cell._style

        # 更新字段，列索引要匹配真实Excel
        date_col_idx = 2
        invoice_col_idx = 15
        qty_pcs_col_idx = 11
        qty_ctns_col_idx = 12
        po_col_idx = 20

        target_ws.cell(row=paste_row, column=date_col_idx).value = target_date
        target_ws.cell(row=paste_row, column=invoice_col_idx).value = invoice_num
        target_ws.cell(row=paste_row, column=qty_pcs_col_idx).value = qty_pcs
        target_ws.cell(row=paste_row, column=qty_ctns_col_idx).value = qty_ctns

        if po_number:
            target_ws.cell(row=paste_row, column=po_col_idx).value = po_number
        else:
            target_ws.cell(row=paste_row, column=po_col_idx).value = ""

        try:
            target_wb.save(TARGET_FILE_PATH)
            messagebox.showinfo("成功", f"已复制并更新到：\n{TARGET_FILE_PATH}\n第 {paste_row} 行。")
        except Exception as e:
            messagebox.showerror("错误", f"保存文件失败：\n{e}")

    # ========== 界面布局 ==========
    file_var = tk.StringVar()
    tk.Label(root, text="源文件：").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    tk.Entry(root, textvariable=file_var, width=50, state='readonly').grid(row=0, column=1, padx=5, pady=5)
    tk.Button(root, text="浏览...", command=select_file).grid(row=0, column=2, padx=5, pady=5)

    # Year
    tk.Label(root, text="Year：").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    year_var = tk.StringVar()
    year_cb = ttk.Combobox(root, textvariable=year_var, state="readonly", width=47)
    year_cb.grid(row=1, column=1, columnspan=2, padx=5, pady=5)
    year_cb.bind("<<ComboboxSelected>>", on_year_selected)

    # Customer
    tk.Label(root, text="Customer：").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    customer_var = tk.StringVar()
    customer_cb = ttk.Combobox(root, textvariable=customer_var, state="readonly", width=47)
    customer_cb.grid(row=2, column=1, columnspan=2, padx=5, pady=5)
    customer_cb.bind("<<ComboboxSelected>>", on_customer_selected)

    # Outlet
    tk.Label(root, text="Outlet：").grid(row=3, column=0, padx=5, pady=5, sticky="e")
    outlet_var = tk.StringVar()
    outlet_cb = ttk.Combobox(root, textvariable=outlet_var, state="readonly", width=47)
    outlet_cb.grid(row=3, column=1, columnspan=2, padx=5, pady=5)
    outlet_cb.bind("<<ComboboxSelected>>", on_outlet_selected)

    # Product
    tk.Label(root, text="Product：").grid(row=4, column=0, padx=5, pady=5, sticky="e")
    product_var = tk.StringVar()
    product_cb = ttk.Combobox(root, textvariable=product_var, state="readonly", width=47)
    product_cb.grid(row=4, column=1, columnspan=2, padx=5, pady=5)

    # 目标日期
    tk.Label(root, text="目标日期(YYYY-MM-DD)：").grid(row=5, column=0, padx=5, pady=5, sticky="e")
    target_date_var = tk.StringVar()
    default_date = (datetime.date.today() + datetime.timedelta(days=7)).strftime("%Y-%m-%d")
    target_date_var.set(default_date)
    tk.Entry(root, textvariable=target_date_var, width=50).grid(row=5, column=1, columnspan=2, padx=5, pady=5)

    # Invoice #
    tk.Label(root, text="Invoice #：").grid(row=6, column=0, padx=5, pady=5, sticky="e")
    invoice_var = tk.StringVar()
    tk.Entry(root, textvariable=invoice_var, width=50).grid(row=6, column=1, columnspan=2, padx=5, pady=5)

    # Qty in Pcs
    tk.Label(root, text="Qty in Pcs：").grid(row=7, column=0, padx=5, pady=5, sticky="e")
    qty_pcs_var = tk.StringVar()
    tk.Entry(root, textvariable=qty_pcs_var, width=50).grid(row=7, column=1, columnspan=2, padx=5, pady=5)

    # Qty in Ctns
    tk.Label(root, text="Qty in Ctns：").grid(row=8, column=0, padx=5, pady=5, sticky="e")
    qty_ctns_var = tk.StringVar()
    tk.Entry(root, textvariable=qty_ctns_var, width=50).grid(row=8, column=1, columnspan=2, padx=5, pady=5)

    # PO
    tk.Label(root, text="PO Number(可空)：").grid(row=9, column=0, padx=5, pady=5, sticky="e")
    po_var = tk.StringVar()
    tk.Entry(root, textvariable=po_var, width=50).grid(row=9, column=1, columnspan=2, padx=5, pady=5)

    # 生成按钮
    tk.Button(
        root, text="生成订单", command=generate_order,
        bg="#4CAF50", fg="white"
    ).grid(row=10, column=1, pady=20)

    root.mainloop()

if __name__ == "__main__":
    main()
