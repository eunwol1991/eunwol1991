import os
import openpyxl

def copy_format(source_cell, target_cell):
    """Copy formatting from source_cell to target_cell."""
    if source_cell.has_style:
        # Copy font
        target_cell.font = source_cell.font
        # Copy fill
        target_cell.fill = source_cell.fill
        # Copy border
        target_cell.border = source_cell.border
        # Copy alignment
        target_cell.alignment = source_cell.alignment
        # Copy number format
        target_cell.number_format = source_cell.number_format

def update_excel_files(directory):
    for root, _, files in os.walk(directory):
        for file in files:
            if "xx25" in file.lower() and file.lower().endswith(('.xlsx', '.xlsm')):
                file_path = os.path.join(root, file)
                print(f"Processing: {file_path}")

                try:
                    workbook = openpyxl.load_workbook(file_path)
                    for sheet_name in workbook.sheetnames:
                        if sheet_name.lower() in ["do", "invoice"]:
                            sheet = workbook[sheet_name]

                            # Remove color for all cells
                            for row in sheet.iter_rows():
                                for cell in row:
                                    cell.fill = openpyxl.styles.PatternFill(fill_type=None)

                    workbook.save(file_path)
                    print(f"Updated and saved: {file_path}")

                except Exception as e:
                    print(f"Failed to process {file_path}: {e}")

# Example usage
directory_path = r"C:\\Users\\User\\Dropbox\\DO & INV\\DO & INV 2025\\Melvin - MOS Burger\\For Customer\\MOS DOC (OTL) - Format"
update_excel_files(directory_path)
