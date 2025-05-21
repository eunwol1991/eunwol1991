import os
from copy import copy
import openpyxl
from openpyxl.styles import PatternFill

def copy_format(source_cell, target_cell):
    """Copy formatting from source_cell to target_cell with .copy() to avoid StyleProxy errors."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.number_format = source_cell.number_format  # 字符串格式无需 copy()

def update_excel_files(directory):
    ALLOWED_NAMES = ["parsley leaves", "rosti hashbrown"]
    ITEM_NAME_COL = 3  # 假设商品名称在第 3 列

    file_count = 0
    match_count = 0
    copy_count = 0

    for root, _, files in os.walk(directory):
        for file in files:
            if "xx25" in file.lower() and file.lower().endswith(('.xlsx', '.xlsm')):
                file_path = os.path.join(root, file)
                print(f"\n[INFO] Processing: {file_path}")
                file_count += 1

                try:
                    workbook = openpyxl.load_workbook(file_path)
                    for sheet_name in workbook.sheetnames:
                        if sheet_name.lower() in ["do", "invoice"]:
                            print(f"[INFO]  -> Checking sheet: {sheet_name}")
                            sheet = workbook[sheet_name]

                            for row in sheet.iter_rows():
                                for cell in row:
                                    # 匹配关键字
                                    if cell.value and isinstance(cell.value, str) \
                                        and "sweet and spicy seafood sauce" in cell.value.strip().lower():
                                        match_count += 1
                                        location = cell.coordinate
                                        row_number = cell.row
                                        print(f"[MATCH] Found '{cell.value}' at {location} (Row {row_number}) in sheet '{sheet_name}'")

                                        # 给源行加上一个标记填充色
                                        #cell.fill = PatternFill(
                                        #    start_color='FFC000',
                                        #    end_color='FFC000',
                                        #    fill_type='solid'
                                        #)

                                        # 从找到的行开始往下复制格式
                                        for r in range(row_number + 1, sheet.max_row + 1):
                                            item_name_cell = sheet.cell(row=r, column=ITEM_NAME_COL)
                                            item_name_value = ""
                                            if item_name_cell.value and isinstance(item_name_cell.value, str):
                                                item_name_value = item_name_cell.value.strip().lower()

                                            print(f"      [DEBUG] Row {r} item name: '{item_name_cell.value}'")

                                            # 如果该行商品名不是 Parsley Leaves 或 Rosti Hashbrown 就停止
                                            if not item_name_value or item_name_value not in ALLOWED_NAMES:
                                                print(f"      [DEBUG] Stopping at row {r} due to unmatched item name.")
                                                break

                                            # 复制格式
                                            for col_idx, source_cell in enumerate(sheet[row_number], start=1):
                                                target_cell = sheet.cell(row=r, column=col_idx)
                                                copy_format(source_cell, target_cell)
                                                copy_count += 1

                    workbook.save(file_path)
                    print(f"[INFO] Updated and saved: {file_path}")

                except Exception as e:
                    print(f"[ERROR] Failed to process {file_path}: {e}")

    print("\n========== Summary ==========")
    print(f"Total files processed: {file_count}")
    print(f"Total matches found:   {match_count}")
    print(f"Total format copied:   {copy_count} times")

if __name__ == "__main__":
    directory_path = r"C:\Users\User\Dropbox\DO & INV\DO & INV 2025\Melvin - MOS Burger\For Customer\MOS DOC (OTL) - Format"
    update_excel_files(directory_path)
