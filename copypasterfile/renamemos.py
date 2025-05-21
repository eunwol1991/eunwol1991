import os
import openpyxl
import traceback

def update_excel_files(directory, debug=False):
    """
    æ‰¹é‡æ›´æ–° Excel æ–‡ä»¶ï¼Œæé«˜æ•ˆç‡å¹¶å¢å¼ºç¨³å®šæ€§ã€‚
    """
    processed_count = 0
    error_files = []

    for root, _, files in os.walk(directory):
        for file in files:
            if "xx25" in file.lower() and file.lower().endswith(('.xlsx', '.xlsm')):
                file_path = os.path.join(root, file)
                if debug:
                    print(f"ğŸ”„ æ­£åœ¨å¤„ç†: {file_path}")

                workbook = None
                try:
                    workbook = openpyxl.load_workbook(file_path)
                    if debug:
                        print(f"âœ… å·²åŠ è½½: {file_path}")
                        print(f"ğŸ“„ å·¥ä½œè¡¨: {workbook.sheetnames}")

                    modified = False

                    # éå† "do" å’Œ "invoice" å·¥ä½œè¡¨ï¼Œä¸è·³å‡ºå¾ªç¯ï¼Œå¤„ç†æ‰€æœ‰åŒ¹é…è¡Œ
                    for sheet_name in ["DO", "Invoice"]:
                        if sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            if debug:
                                print(f"ğŸ“Œ å¤„ç†å·¥ä½œè¡¨: {sheet_name}")
                            for row in sheet.iter_rows():
                                for cell in row:
                                    cell_value = str(cell.value).strip().lower() if cell.value is not None else ""
                                    # æ¨¡ç³ŠåŒ¹é… "tartar sauce"
                                    cell_name = "honey mustard sauce"
                                    
                                    if cell_name in cell_value:
                                        row_number = cell.row
                                        product_code = "SD2136"
                                        product_description = "Honey Mustard Sauce"
                                        pack_size = "(10 x 1kg)"
                                        qty = 7.00
                                        UOM = "PKTS"
                                        selling_price = f"=7.5*G{row_number}"
                                        if debug:
                                            print(f"ğŸ” åŒ¹é…åˆ° {cell_name} â†’ Sheet: {sheet_name}, Row: {row_number}")
                                        if sheet_name == "DO":
                                            sheet[f"B{row_number}"].value = product_code
                                            sheet[f"C{row_number}"].value = product_description
                                            sheet[f"G{row_number}"].value = pack_size
                                            sheet[f"I{row_number}"].value = qty
                                            sheet[f"K{row_number}"].value = UOM
                                            if debug:
                                                print(f"âœ… DO ä¿®æ”¹æˆåŠŸ: B{row_number}, C{row_number}, G{row_number}, I{row_number}, K{row_number}")
                                        elif sheet_name == "Invoice":
                                            sheet[f"B{row_number}"].value = product_code
                                            sheet[f"C{row_number}"].value = product_description
                                            sheet[f"F{row_number}"].value = pack_size
                                            sheet[f"G{row_number}"].value = qty
                                            sheet[f"H{row_number}"].value = UOM
                                            sheet[f"I{row_number}"].value = selling_price
                                            if debug:
                                                print(f"âœ… Invoice ä¿®æ”¹æˆåŠŸ: B{row_number}, C{row_number}, F{row_number}, G{row_number}, H{row_number}, I{row_number}")
                                        modified = True
                                        # ä¸è·³å‡ºå¾ªç¯ï¼Œç»§ç»­å¤„ç†æ‰€æœ‰åŒ¹é…å•å…ƒæ ¼
                    if modified:
                        workbook.save(file_path)
                        processed_count += 1
                        if debug:
                            print(f"ğŸ’¾ å·²ä¿å­˜ä¿®æ”¹: {file_path}")
                    workbook.close()
                except Exception as e:
                    traceback.print_exc()
                    error_files.append(file_path)
                    if debug:
                        print(f"âŒ å¤„ç†å¤±è´¥: {file_path}ï¼Œé”™è¯¯: {e}")
                finally:
                    if workbook is not None:
                        workbook.close()

    print(f"âœ… å¤„ç†å®Œæˆï¼Œå…±ä¿®æ”¹ {processed_count} ä¸ªæ–‡ä»¶ã€‚")
    if error_files:
        print("âš ï¸ ä»¥ä¸‹æ–‡ä»¶å¤„ç†å¤±è´¥ï¼š")
        for error_file in error_files:
            print(f"  âŒ {error_file}")

# ç¤ºä¾‹è°ƒç”¨
directory_path = r"C:\Users\User\Dropbox\DO & INV\DO & INV 2025\Anthony - Select(Waker Chicken)"
update_excel_files(directory_path, debug=True)
