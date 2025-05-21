import os
import openpyxl

def update_Invoice_formulas(directory):
    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith(('.xlsx', '.xlsm')):
                file_path = os.path.join(root, file)
                print(f"Processing: {file_path}")

                try:
                    workbook = openpyxl.load_workbook(file_path)

                    # Ensure the "Invoice" sheet exists
                    if "invoice" not in [sheet.lower() for sheet in workbook.sheetnames]:
                        print(f"Sheet 'Invoice' not found in {file_path}")
                        continue

                    # Match "Invoice" sheet case-insensitively
                    sheet = workbook[[sheet for sheet in workbook.sheetnames if sheet.strip().lower() == "invoice"][0]]

                    # Set "Amount" column to column I (Excel column index starts from 1, so I = 9)
                    amount_column = 9

                    # Identify rows with "Subtotal", "GST 9%", and "Total"
                    subtotal_row = None
                    gst_row = None
                    total_row = None

                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                cell_value = cell.value.strip().lower()
                                if "subtotal" in cell_value:
                                    subtotal_row = cell.row
                                elif "gst 9%" in cell_value:
                                    gst_row = cell.row
                                elif "total" in cell_value:
                                    total_row = cell.row

                    # Debugging logs
                    print(f"Subtotal row: {subtotal_row}, GST row: {gst_row}, Total row: {total_row}, Amount column: {amount_column}")

                    if not subtotal_row or not gst_row or not total_row:
                        print(f"Required rows not found in 'Invoice' sheet of {file_path}")
                        continue

                    # Update Subtotal formula
                    amount_start_row = None
                    for row in range(1, subtotal_row):
                        if sheet.cell(row=row, column=amount_column).value is not None:
                            amount_start_row = row
                            break

                    if amount_start_row:
                        subtotal_cell = sheet.cell(row=subtotal_row, column=amount_column)
                        subtotal_cell.value = f"=SUM({sheet.cell(row=24, column=amount_column).coordinate}:{sheet.cell(row=subtotal_row - 1, column=amount_column).coordinate})"
                        print(f"Updated Subtotal formula to: {subtotal_cell.value}")

                    # Update GST formula
                    gst_cell = sheet.cell(row=gst_row, column=amount_column)
                    gst_cell.value = f"={sheet.cell(row=subtotal_row, column=amount_column).coordinate}*0.09"
                    print(f"Updated GST formula to: {gst_cell.value}")

                    # Update Total formula
                    total_cell = sheet.cell(row=total_row, column=amount_column)
                    total_cell.value = f"=sum({sheet.cell(row=subtotal_row, column=amount_column).coordinate}:{sheet.cell(row=gst_row, column=amount_column).coordinate})"
                    print(f"Updated Total formula to: {total_cell.value}")

                    # Save the workbook
                    workbook.save(file_path)
                    print(f"Updated and saved: {file_path}")

                except Exception as e:
                    print(f"Failed to process {file_path}: {e}")

# Example usage
directory_path = r"C:\\Users\\User\\Dropbox\\DO & INV\\DO & INV 2025\\Melvin - MOS Burger\\For Customer\\MOS DOC (OTL) - Format"
update_Invoice_formulas(directory_path)
