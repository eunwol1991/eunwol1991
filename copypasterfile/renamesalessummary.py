import os
import openpyxl
from tkinter import Tk, filedialog

def update_sales_summary():
    """
    Search for Excel files with 'sales summary' in their name within a user-selected directory and its subdirectories.
    Replace all occurrences of the value 2024 with 2025 in their sheets.
    """
    try:
        # Open a file dialog to select the directory
        Tk().withdraw()  # Hide the root window
        directory = filedialog.askdirectory(title="Select Directory")

        if not directory:
            print("No directory selected. Exiting.")
            return

        for root, _, files in os.walk(directory):
            for filename in files:
                # Check if the file name contains 'sales summary' and is an Excel file
                if 'sales summary' in filename.lower() and filename.endswith(('.xlsx', '.xlsm')):
                    file_path = os.path.join(root, filename)

                    try:
                        # Open the workbook
                        workbook = openpyxl.load_workbook(file_path)
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value == 2024:
                                        cell.value = 2025

                        # Save the changes
                        workbook.save(file_path)
                        print(f"Updated: {file_path}")
                    except Exception as e:
                        print(f"Failed to process {file_path}: {e}")

    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
update_sales_summary()
