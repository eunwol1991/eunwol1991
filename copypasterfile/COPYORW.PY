import os
import sys
from copy import copy
import openpyxl
from openpyxl.styles import PatternFill, Border, Alignment, Font

def unmerge_and_fill(sheet, start_row, end_row):
    """
    取消指定行范围内所有合并单元格，
    并将原合并区域左上角的值填充到该区域所有单元格。
    """
    merges_to_process = []
    for merged_range in list(sheet.merged_cells.ranges):
        min_row, min_col, max_row, max_col = merged_range.bounds
        # 如果该合并区域和指定行段有重叠，就处理
        if not (max_row < start_row or min_row > end_row):
            merges_to_process.append(merged_range)

    for merged_range in merges_to_process:
        sheet.merged_cells.remove(merged_range)
        min_row, min_col, max_row, max_col = merged_range.bounds
        # 获取左上角单元格的值
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        # 填充该合并区域
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                sheet.cell(row=r, column=c).value = top_left_value

def copy_cell_format(source_cell, target_cell):
    """
    将 source_cell 的格式复制到 target_cell。
    使用 copy() 避免部分 openpyxl 版本的 'unhashable type: StyleProxy' 问题。
    """
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def copy_row_data_and_format(source_sheet, target_sheet, source_row, target_row):
    """
    将 source_sheet 中的第 source_row 行复制(数据+格式)到 target_sheet 的第 target_row 行（覆盖）。
    注意：列数根据源表的实际使用列数而定。
    """
    for col_idx, source_cell in enumerate(source_sheet[source_row], start=1):
        target_cell = target_sheet.cell(row=target_row, column=col_idx)
        # 复制数据
        target_cell.value = source_cell.value
        # 复制格式
        copy_cell_format(source_cell, target_cell)

def copy_rows_block(source_sheet, target_sheet, src_start, src_end, tgt_start):
    """
    将 source_sheet 的 [src_start ~ src_end] 行“覆盖”复制到 target_sheet，
    从第 tgt_start 行开始逐行覆盖。
    """
    row_count = (src_end - src_start + 1)
    for i in range(row_count):
        s_row = src_start + i
        t_row = tgt_start + i
        copy_row_data_and_format(source_sheet, target_sheet, s_row, t_row)

def process_file(source_wb, target_file, do_row_range, invoice_row_range):
    """
    处理单个目标文件：
      - 如果存在 DO 工作表，覆盖 DO 指定行
      - 如果存在 Invoice 工作表，覆盖 Invoice 指定行
    """
    print(f"Processing: {target_file}")
    try:
        target_wb = openpyxl.load_workbook(target_file)
    except Exception as e:
        print(f"  [ERROR] Cannot open {target_file}: {e}")
        return

    # 覆盖 DO 工作表
    if "DO" in target_wb.sheetnames:
        try:
            src_sheet_do = source_wb["DO"]
            tgt_sheet_do = target_wb["DO"]
            src_start, src_end = do_row_range  # e.g. (44, 48)

            # 覆盖前先取消合并，避免冲突
            unmerge_and_fill(tgt_sheet_do, src_start, src_end)

            copy_rows_block(
                source_sheet = src_sheet_do,
                target_sheet = tgt_sheet_do,
                src_start = src_start,
                src_end = src_end,
                tgt_start = src_start
            )
            print(f"  [INFO] DO rows {src_start}~{src_end} covered.")

        except Exception as e:
            print(f"  [ERROR] DO copy failed in {target_file}: {e}")
    else:
        print("  [SKIP] No DO sheet found.")

    # 覆盖 Invoice 工作表
    if "Invoice" in target_wb.sheetnames:
        try:
            src_sheet_inv = source_wb["Invoice"]
            tgt_sheet_inv = target_wb["Invoice"]
            src_start, src_end = invoice_row_range  # e.g. (48, 50)

            unmerge_and_fill(tgt_sheet_inv, src_start, src_end)

            copy_rows_block(
                source_sheet = src_sheet_inv,
                target_sheet = tgt_sheet_inv,
                src_start = src_start,
                src_end = src_end,
                tgt_start = src_start
            )
            print(f"  [INFO] Invoice rows {src_start}~{src_end} covered.")

        except Exception as e:
            print(f"  [ERROR] Invoice copy failed in {target_file}: {e}")
    else:
        print("  [SKIP] No Invoice sheet found.")

    # 保存目标文件
    try:
        target_wb.save(target_file)
        print(f"Updated and saved: {target_file}")
    except Exception as e:
        print(f"  [ERROR] Save failed for {target_file}: {e}")

def main():
    # ======== 配置区 ========
    source_file = r"C:\Users\User\Dropbox\DO & INV\DO & INV 2025\Melvin - MOS Burger\For Customer\MOS DOC (OTL) - Format\MOS xx25 - 00x - DO & INV (18 Tai Seng).xlsx"
    target_directory = r"C:\Users\User\Dropbox\DO & INV\DO & INV 2025\Melvin - MOS Burger\For Customer\MOS DOC (OTL) - Format"

    # DO 表要复制的行段 (示例: 第 44 到 48 行)
    do_row_range = (44, 48)
    # Invoice 表要复制的行段 (示例: 第 48 到 50 行)
    invoice_row_range = (48, 50)
    # ======== 配置结束 ========

    # 打开源文件
    try:
        source_wb = openpyxl.load_workbook(source_file)
    except Exception as e:
        print(f"[FATAL] Cannot open source file: {source_file}\n{e}")
        sys.exit(1)

    # 遍历目标文件夹及子文件夹中的所有 .xlsx / .xlsm
    for root, dirs, files in os.walk(target_directory):
        for filename in files:
            if filename.lower().endswith(('.xlsx', '.xlsm')):
                target_path = os.path.join(root, filename)
                # 跳过源文件本身，避免覆盖自己
                if os.path.abspath(target_path) == os.path.abspath(source_file):
                    continue
                process_file(source_wb, target_path, do_row_range, invoice_row_range)

    print("Done.")

if __name__ == "__main__":
    main()
